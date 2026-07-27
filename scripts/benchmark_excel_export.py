"""Benchmark scalable Excel export through the production document worker."""

from __future__ import annotations

import argparse
from io import BytesIO
import json
import math
from pathlib import Path
from statistics import median
import sys
import time

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.documents.document_worker import run_document_job
from backend.documents.excel_workbook_builder import create_phanlo_excel


DEFAULT_ROW_COUNTS = (10, 100, 1_000, 10_000)
DEFAULT_ITERATIONS = 3
MAX_BENCHMARK_ROWS = 50_000


def nearest_rank_percentile(values: list[float], percentile: float) -> float:
    if not values:
        raise ValueError("Không có số đo để tính percentile.")
    if not 0 < percentile <= 1:
        raise ValueError("Percentile phải nằm trong khoảng (0, 1].")
    ordered = sorted(float(value) for value in values)
    return ordered[max(0, math.ceil(len(ordered) * percentile) - 1)]


def build_lot_rows(row_count: int) -> list[dict]:
    if row_count < 1 or row_count > MAX_BENCHMARK_ROWS:
        raise ValueError(
            f"Số dòng benchmark phải từ 1 đến {MAX_BENCHMARK_ROWS}."
        )
    return [
        {
            "maPhanLo": f"L{index + 1:05d}",
            "tenPhanLo": f"Lô hàng hóa số {index + 1}",
            "giaTriPhanLo": (index + 1) * 1_000_000,
            "baoDamDuThau": (index + 1) * 10_000,
            "thoiGianThucHien": 30 + index % 90,
        }
        for index in range(row_count)
    ]


def _summary(values: list[float]) -> dict[str, float]:
    return {
        "medianMs": round(median(values), 3),
        "p95Ms": round(nearest_rank_percentile(values, 0.95), 3),
    }


def benchmark_case(row_count: int, iterations: int, mode: str = "both") -> dict:
    if iterations < 1 or iterations > 50:
        raise ValueError("Số lượt benchmark phải từ 1 đến 50.")
    if mode not in {"direct", "worker", "both"}:
        raise ValueError("Chế độ benchmark không hợp lệ.")

    rows = build_lot_rows(row_count)
    output = b""
    result: dict = {"rows": row_count, "iterations": iterations}

    if mode in {"direct", "both"}:
        durations = []
        for _ in range(iterations):
            started_at = time.perf_counter()
            workbook = create_phanlo_excel(rows)
            stream = BytesIO()
            workbook.save(stream)
            durations.append((time.perf_counter() - started_at) * 1_000)
            output = stream.getvalue()
        result["directExport"] = _summary(durations)

    if mode in {"worker", "both"}:
        durations = []
        for _ in range(iterations):
            started_at = time.perf_counter()
            output = run_document_job(
                "export_excel",
                {"function": "create_phanlo_excel", "args": [rows]},
                timeout_seconds=60,
            )
            durations.append((time.perf_counter() - started_at) * 1_000)
        result["isolatedWorker"] = _summary(durations)

    workbook = load_workbook(BytesIO(output), read_only=True, data_only=True)
    try:
        rendered_rows = workbook["Phan Lo"].max_row - 1
    finally:
        workbook.close()
    if rendered_rows != row_count:
        raise RuntimeError(
            f"XLSX kết quả có {rendered_rows} dòng, dự kiến {row_count}."
        )
    result["outputBytes"] = len(output)
    result["renderedRows"] = rendered_rows
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--rows",
        nargs="+",
        type=int,
        default=list(DEFAULT_ROW_COUNTS),
    )
    parser.add_argument("--iterations", type=int, default=DEFAULT_ITERATIONS)
    parser.add_argument(
        "--mode",
        choices=("direct", "worker", "both"),
        default="both",
    )
    args = parser.parse_args()
    results = [
        benchmark_case(row_count, args.iterations, args.mode)
        for row_count in args.rows
    ]
    print(
        json.dumps(
            {
                "mode": args.mode,
                "iterations": args.iterations,
                "results": results,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

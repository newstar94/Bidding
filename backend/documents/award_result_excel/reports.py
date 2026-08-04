"""Create a standalone reconciliation report from server-derived match data."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook

from backend.documents.spreadsheet_security import safe_spreadsheet_text


def reconciliation_filename(source_filename: Any) -> str:
    stem = Path(str(source_filename or "ket_qua.xlsx")).stem
    stem = re.sub(r"[^0-9A-Za-zÀ-ỹ._ -]+", "_", stem).strip(" ._")
    return f"{stem or 'ket_qua'}_bao_cao_doi_chieu.xlsx"


def _cell(value: Any) -> Any:
    if isinstance(value, (int, float)) or value is None:
        return value
    if isinstance(value, (dict, list, tuple)):
        import json

        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    return safe_spreadsheet_text(value)


def build_reconciliation_workbook(payload: dict[str, Any]) -> bytes:
    """Render bounded plain data; the caller owns authorization and recomputation."""

    workbook = Workbook(write_only=True)
    overview = workbook.create_sheet("Tổng quan")
    overview.append(["Trường", "Giá trị"])
    metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    for label, key in (
        ("SHA-256 file nguồn", "sourceSha256"),
        ("Gói thầu", "packageCode"),
        ("Người xuất", "userId"),
        ("Thời gian tạo", "generatedAt"),
    ):
        overview.append([label, _cell(metadata.get(key))])
    for label, key in (
        ("Tổng dòng", "totalRows"),
        ("Khớp mã định danh", "exactMatches"),
        ("Khớp mã số thuế", "fallbackMatches"),
        ("Không khớp", "unmatchedRows"),
        ("Dòng sẽ cập nhật", "updatedRows"),
    ):
        overview.append([label, _cell(summary.get(key, 0))])

    detail = workbook.create_sheet("Đối chiếu")
    detail.append([
        "Dòng Excel",
        "Phần/lô",
        "Nhà thầu",
        "Phương pháp khớp",
        "Trạng thái",
        "Cột",
        "Giá trị cũ",
        "Giá trị mới",
        "Nguồn dữ liệu",
        "Cảnh báo",
    ])
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    if len(rows) > 10_000:
        raise ValueError("Báo cáo đối chiếu vượt quá giới hạn 10.000 dòng.")
    for row in rows:
        if not isinstance(row, dict):
            continue
        changes = row.get("changes") if isinstance(row.get("changes"), list) else []
        if len(changes) > 20:
            raise ValueError("Một dòng đối chiếu có quá nhiều thay đổi.")
        if not changes:
            changes = [{}]
        warning_codes = ", ".join(
            str(item.get("code") or "")
            for item in (row.get("warnings") or [])
            if isinstance(item, dict) and item.get("code")
        )
        for change in changes:
            if not isinstance(change, dict):
                change = {}
            detail.append([
                _cell(row.get("excelRow")),
                _cell(row.get("lotCode")),
                _cell(row.get("bidderName")),
                _cell(row.get("matchMethod")),
                _cell("Sẽ cập nhật" if row.get("writable") else row.get("status")),
                _cell(change.get("field")),
                _cell(change.get("oldValue")),
                _cell(change.get("newValue")),
                _cell(change.get("source")),
                _cell(warning_codes),
            ])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

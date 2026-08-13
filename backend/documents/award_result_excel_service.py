"""Validation, matching and loss-minimising export for muasamcong award workbooks."""

from __future__ import annotations

import asyncio
import base64
from datetime import date, datetime
from decimal import Decimal
import hashlib
import hmac
from io import BytesIO
import json
import os
from pathlib import Path
import re
import shutil
import time
import unicodedata
import uuid
from typing import Any, Iterable
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.documents.award_result_mapping import (
    LotApprovedOutcome,
    map_bidder_award,
    parse_lot_outcome,
)
from backend.documents.award_result_excel.templates import (
    EXPECTED_HEADERS,  # noqa: F401 - compatibility re-export
    MEDICINE_EXPECTED_HEADERS,  # noqa: F401 - compatibility re-export
    MEDICINE_TEMPLATE,  # noqa: F401 - compatibility re-export
    STANDARD_TEMPLATE,
    TEMPLATE_DEFINITIONS,
    WorkbookTemplateDefinition,
)
from backend.documents.award_result_excel.types import (
    AwardRecord,
    AwardResultExcelError,
    RowMatch,
)
from backend.documents.award_result_excel.normalization import (
    normalize_code,
    normalize_tax_code,
    normalize_text as _normalised_text,
)
from backend.documents.spreadsheet_security import safe_spreadsheet_text
from backend.documents.workbook_preservation import patch_worksheet_cells
from backend.shared.paths import resolve_runtime_path


XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
MAX_AWARD_RESULT_ROWS = 10_000
MAX_WORKBOOK_CELLS = 1_000_000
VALIDATION_TTL_SECONDS = 15 * 60
VALIDATION_EXPORT_LEASE_SECONDS = 2 * 60
_VALIDATION_ID = re.compile(r"[a-f0-9]{32}")
_ARTIFACT_CLEANUP_FAILURES = 0
_ARTIFACT_QUOTA_REJECTIONS = 0

def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else format(value, "f")
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _row_fingerprint(cells: Iterable[Any]) -> str:
    payload = [
        {"value": _json_value(cell.value), "dataType": str(cell.data_type or "")}
        for cell in cells
    ]
    encoded = json.dumps(
        payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _issue(code: str, message: str, *, excel_row: int | None = None, **extra):
    value: dict[str, Any] = {"code": code, "message": message}
    if excel_row is not None:
        value["excelRow"] = int(excel_row)
    value.update(extra)
    return value


def _header_matches(
    definition: WorkbookTemplateDefinition, index: int, value: Any
) -> bool:
    accepted = {definition.headers[index], *definition.aliases.get(index, set())}
    return _normalised_text(value) in {_normalised_text(item) for item in accepted}


def _candidate_headers(workbook):
    candidates = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        for row_index in range(1, min(20, worksheet.max_row) + 1):
            for definition in TEMPLATE_DEFINITIONS:
                columns_by_header: dict[int, list[int]] = {}
                for actual_column in range(1, min(256, worksheet.max_column) + 1):
                    value = worksheet.cell(row_index, actual_column).value
                    for header_index in range(len(definition.headers)):
                        if _header_matches(definition, header_index, value):
                            columns_by_header.setdefault(header_index, []).append(actual_column)
                matches = [
                    index in columns_by_header
                    for index in range(len(definition.headers))
                ]
                source_matches = sum(
                    matches[index] for index in definition.source_indices
                )
                if source_matches >= len(definition.source_indices) - 1:
                    candidates.append(
                        (
                            sum(matches), source_matches, len(definition.headers),
                            -sheet_index, definition, worksheet, row_index,
                            columns_by_header,
                        )
                    )
    return sorted(candidates, key=lambda item: item[:4], reverse=True)


def inspect_award_result_workbook(content: bytes) -> dict[str, Any]:
    """Read only the template contract and matching fields from an XLSX."""

    try:
        workbook = load_workbook(
            BytesIO(content), data_only=False, keep_links=False, rich_text=True
        )
    except Exception as exc:  # openpyxl error classes vary across malformed inputs
        raise AwardResultExcelError(
            "WORKBOOK_INVALID", "Tệp không phải workbook Excel hợp lệ."
        ) from exc

    if sum(sheet.max_row * sheet.max_column for sheet in workbook.worksheets) > MAX_WORKBOOK_CELLS:
        raise AwardResultExcelError(
            "WORKBOOK_CELL_LIMIT_EXCEEDED",
            "Workbook có phạm vi ô sử dụng vượt quá giới hạn an toàn.",
        )

    blocking_errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    candidates = _candidate_headers(workbook)
    if not candidates:
        return {
            "sheetName": None,
            "headerRow": None,
            "totalRows": 0,
            "rows": [],
            "blockingErrors": [
                _issue(
                    "UNSUPPORTED_TEMPLATE_VERSION",
                    "Không nhận diện được phiên bản biểu mẫu kết quả muasamcong.",
                )
            ],
            "warnings": [],
        }

    best = candidates[0]
    tied = [item for item in candidates if item[:3] == best[:3]]
    if len(tied) > 1:
        blocking_errors.append(
            _issue(
                "WORKSHEET_AMBIGUOUS",
                "Có nhiều vị trí tiêu đề phù hợp; không thể xác định sheet cần xử lý.",
            )
        )
    _, _, _, _, definition, worksheet, header_row, columns_by_header = best
    with zipfile.ZipFile(BytesIO(content)) as archive:
        unsupported_parts = sorted(
            name
            for name in archive.namelist()
            if name in definition.unsupported_entries
            or any(
                name.startswith(prefix)
                for prefix in definition.unsupported_entry_prefixes
            )
        )
    if unsupported_parts:
        blocking_errors.append(
            _issue(
                "UNSUPPORTED_TEMPLATE_PART",
                "Workbook chứa thành phần OOXML không được hỗ trợ cho luồng xuất kết quả.",
                parts=unsupported_parts[:20],
            )
        )
    for index in range(len(definition.headers)):
        columns = columns_by_header.get(index, [])
        if not columns:
            blocking_errors.append(
                _issue(
                    "REQUIRED_HEADER_MISSING",
                    f"Thiếu tiêu đề bắt buộc: {definition.headers[index]}.",
                    excel_row=header_row,
                    expectedHeader=definition.headers[index],
                )
            )
        elif len(columns) > 1:
            blocking_errors.append(
                _issue(
                    "REQUIRED_HEADER_DUPLICATED",
                    f"Tiêu đề xuất hiện nhiều lần: {definition.headers[index]}.",
                    excel_row=header_row,
                    expectedHeader=definition.headers[index],
                    columns=columns,
                )
            )
    if worksheet.sheet_state != "visible":
        warnings.append(
            _issue(
                "TARGET_SHEET_HIDDEN",
                "Sheet chứa danh sách nhà thầu đang bị ẩn.",
            )
        )

    rows = []
    existing_result_rows = 0
    source_columns = [
        columns_by_header[index][0]
        for index in definition.source_indices
        if len(columns_by_header.get(index, [])) == 1
    ]
    output_columns = [
        columns_by_header[index][0]
        for index in definition.output_indices
        if len(columns_by_header.get(index, [])) == 1
    ]
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        if len(source_columns) != len(definition.source_indices):
            break
        cells_by_header = {
            index: worksheet.cell(row_index, columns_by_header[index][0])
            for index in definition.source_indices
        }
        source_cells = [cells_by_header[index] for index in definition.source_indices]
        if all(cell.value in (None, "") for cell in source_cells):
            continue
        if len(rows) >= MAX_AWARD_RESULT_ROWS:
            raise AwardResultExcelError(
                "WORKBOOK_ROW_LIMIT_EXCEEDED",
                f"Workbook vượt quá giới hạn {MAX_AWARD_RESULT_ROWS} dòng dữ liệu.",
            )
        output_cells = [worksheet.cell(row_index, column) for column in output_columns]
        has_existing_result = any(cell.value not in (None, "") for cell in output_cells)
        existing_result_rows += int(has_existing_result)
        rows.append(
            {
                "excelRow": row_index,
                "lotCode": _json_value(cells_by_header[definition.lot_index].value),
                "bidderIdentifier": _json_value(
                    cells_by_header[definition.bidder_identifier_index].value
                ),
                "taxCode": _json_value(
                    cells_by_header[definition.tax_code_index].value
                ),
                "bidderName": _json_value(
                    cells_by_header[definition.bidder_name_index].value
                ),
                "goodsSequence": _json_value(cells_by_header[0].value)
                if definition.template_type == "medicine"
                else None,
                "goodsName": _json_value(cells_by_header[2].value)
                if definition.template_type == "medicine"
                else None,
                "sourceFingerprint": _row_fingerprint(source_cells),
                "hasExistingResult": has_existing_result,
                "existingOutputValues": [
                    _json_value(cell.value) for cell in output_cells
                ],
                "existingOutputDataTypes": [
                    str(cell.data_type or "") for cell in output_cells
                ],
            }
        )

    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.max_row > header_row
            and merged_range.min_row <= worksheet.max_row
            and any(
                merged_range.min_col <= column <= merged_range.max_col
                for column in output_columns
            )
        ):
            blocking_errors.append(
                _issue(
                    "MERGED_WRITE_RANGE",
                    f"Vùng dữ liệu được phép ghi giao với ô gộp {merged_range}.",
                )
            )

    return {
        "sheetName": worksheet.title,
        "templateType": definition.template_type,
        "templateVersion": definition.version,
        "templateFingerprint": definition.fingerprint,
        "headerRow": header_row,
        "columnMap": {
            "source": source_columns,
            "output": output_columns,
            "headers": {
                definition.headers[index]: columns[0]
                for index, columns in columns_by_header.items()
                if len(columns) == 1
            },
            "outputRoles": list(definition.output_roles),
        },
        "dataStartRow": header_row + 1,
        "maxRow": worksheet.max_row,
        "totalRows": len(rows),
        "existingResultRows": existing_result_rows,
        "rows": rows,
        "blockingErrors": blocking_errors,
        "warnings": warnings,
    }


def _award_record(value: AwardRecord | dict[str, Any]) -> AwardRecord:
    return value if isinstance(value, AwardRecord) else AwardRecord(**value)


def _append_index(index, key, record):
    if key[1]:
        index.setdefault(key, []).append(record)


def _output_values_equal(old: Any, new: Any) -> bool:
    if isinstance(new, dict) and set(new) == {"decimal"}:
        new = Decimal(str(new["decimal"]))
    if isinstance(old, (int, float, Decimal)) and isinstance(
        new, (int, float, Decimal)
    ):
        return Decimal(str(old)) == Decimal(str(new))
    return old == new


def match_award_result_rows(
    inspection: dict[str, Any],
    records: Iterable[AwardRecord | dict[str, Any]],
    *,
    known_lot_codes: Iterable[Any] = (),
    foreign_lot_codes: Iterable[Any] = (),
) -> dict[str, Any]:
    """Match sanitized workbook rows without using bidder names as keys."""

    award_records = [_award_record(item) for item in records]
    definition = next(
        (
            item
            for item in TEMPLATE_DEFINITIONS
            if item.template_type == inspection.get("templateType")
        ),
        STANDARD_TEMPLATE,
    )
    medicine_template = definition.template_type == "medicine"
    primary_index: dict[tuple[str, ...], list[AwardRecord]] = {}
    fallback_index: dict[tuple[str, ...], list[AwardRecord]] = {}
    blocking_errors = list(inspection.get("blockingErrors") or [])
    warnings = list(inspection.get("warnings") or [])
    for record in award_records:
        lot_key = normalize_code(record.lot_code)
        item_key = normalize_code(record.goods_sequence) if medicine_template else ""
        if medicine_template and not str(record.goods_item_id or "").strip():
            blocking_errors.append(
                _issue(
                    "MEDICINE_GOODS_ID_MISSING",
                    "Hàng thuốc thiếu mã hàng hóa ổn định.",
                    openingId=record.opening_id,
                )
            )
        if medicine_template and not item_key:
            blocking_errors.append(
                _issue(
                    "MEDICINE_GOODS_SEQUENCE_MISSING",
                    "Hàng thuốc thiếu STT nguồn để đối chiếu.",
                    openingId=record.opening_id,
                    goodsItemId=record.goods_item_id,
                )
            )
            continue
        primary_key = (lot_key, normalize_code(record.bidder_identifier))
        fallback_key = (lot_key, normalize_tax_code(record.tax_code))
        if medicine_template:
            primary_key += (item_key,)
            fallback_key += (item_key,)
        _append_index(
            primary_index,
            primary_key,
            record,
        )
        _append_index(
            fallback_index,
            fallback_key,
            record,
        )

    known_lots = {normalize_code(item) for item in known_lot_codes}
    known_lots.update(normalize_code(item.lot_code) for item in award_records)
    foreign_lots = {normalize_code(item) for item in foreign_lot_codes}
    matched_rows: list[RowMatch] = []
    matched_goods_rows: dict[tuple[str, str], int] = {}
    exact_matches = fallback_matches = unmatched_rows = 0
    duplicate_rows = conflict_rows = 0
    missing_lot_rows = missing_identity_rows = 0

    for source in inspection.get("rows") or []:
        excel_row = int(source["excelRow"])
        lot_key = normalize_code(source.get("lotCode"))
        identifier_key = normalize_code(source.get("bidderIdentifier"))
        tax_key = normalize_tax_code(source.get("taxCode"))
        item_key = normalize_code(source.get("goodsSequence")) if medicine_template else ""
        row = RowMatch(
            excel_row=excel_row,
            lot_code=str(source.get("lotCode") or ""),
            bidder_identifier=str(source.get("bidderIdentifier") or ""),
            tax_code=str(source.get("taxCode") or ""),
            bidder_name=str(source.get("bidderName") or ""),
            goods_sequence=source.get("goodsSequence"),
            goods_name=str(source.get("goodsName") or ""),
            source_fingerprint=str(source.get("sourceFingerprint") or ""),
            status="unmatched",
        )
        if not lot_key:
            missing_lot_rows += 1
            row.warnings.append(
                _issue("LOT_CODE_MISSING", "Dòng thiếu mã phần/lô.", excel_row=excel_row)
            )
        elif lot_key not in known_lots and lot_key in foreign_lots:
            issue = _issue(
                "LOT_BELONGS_TO_OTHER_PACKAGE",
                "Mã phần/lô thuộc gói thầu khác trong cùng tổ chức.",
                excel_row=excel_row,
            )
            row.errors.append(issue)
            blocking_errors.append(issue)

        if not identifier_key and not tax_key:
            missing_identity_rows += 1
            row.warnings.append(
                _issue(
                    "BIDDER_IDENTITY_MISSING",
                    "Dòng thiếu cả mã định danh và mã số thuế nhà thầu.",
                    excel_row=excel_row,
                )
            )

        if medicine_template and not item_key:
            issue = _issue(
                "MEDICINE_GOODS_SEQUENCE_MISSING",
                "Dòng thuốc thiếu STT để đối chiếu hàng hóa.",
                excel_row=excel_row,
            )
            row.errors.append(issue)
            blocking_errors.append(issue)

        primary_key = (lot_key, identifier_key)
        fallback_key = (lot_key, tax_key)
        if medicine_template:
            primary_key += (item_key,)
            fallback_key += (item_key,)
        primary = primary_index.get(primary_key, []) if identifier_key and item_key else []
        fallback = fallback_index.get(fallback_key, []) if tax_key and item_key else []
        if not medicine_template:
            primary = primary_index.get(primary_key, []) if identifier_key else []
            fallback = fallback_index.get(fallback_key, []) if tax_key else []
        if len(primary) > 1 or len(fallback) > 1:
            duplicate_rows += 1
            issue = _issue(
                "MEDICINE_GOODS_SEQUENCE_DUPLICATED"
                if medicine_template
                else "DUPLICATE_MATCH_KEY",
                "STT hàng thuốc cho ra nhiều hàng hóa trong dữ liệu ứng dụng."
                if medicine_template
                else "Khóa đối chiếu cho ra nhiều kết quả trong dữ liệu ứng dụng.",
                excel_row=excel_row,
            )
            row.errors.append(issue)
            blocking_errors.append(issue)
        elif primary and fallback and (
            primary[0].opening_id,
            str(primary[0].goods_item_id or ""),
        ) != (
            fallback[0].opening_id,
            str(fallback[0].goods_item_id or ""),
        ):
            conflict_rows += 1
            issue = _issue(
                "IDENTIFIER_TAX_CONFLICT",
                "Mã định danh và mã số thuế khớp với hai nhà thầu khác nhau.",
                excel_row=excel_row,
            )
            row.errors.append(issue)
            blocking_errors.append(issue)
        else:
            record = primary[0] if primary else (fallback[0] if fallback else None)
            goods_identity = (
                (record.opening_id, str(record.goods_item_id or ""))
                if record
                else ("", "")
            )
            if (
                medicine_template
                and record is not None
                and goods_identity[1]
                and goods_identity in matched_goods_rows
            ):
                duplicate_rows += 1
                issue = _issue(
                    "MEDICINE_GOODS_MATCHED_MULTIPLE_ROWS",
                    "Một hàng thuốc trong dữ liệu ứng dụng khớp với nhiều dòng Excel.",
                    excel_row=excel_row,
                    firstExcelRow=matched_goods_rows[goods_identity],
                    goodsItemId=goods_identity[1],
                )
                row.errors.append(issue)
                blocking_errors.append(issue)
                record = None
            if record is None:
                if not row.errors:
                    unmatched_rows += 1
                    row.warnings.append(
                        _issue(
                            "RESULT_NOT_FOUND",
                            "Không tìm thấy kết quả phù hợp cho dòng Excel.",
                            excel_row=excel_row,
                        )
                    )
            else:
                row.record = record
                row.status = "matched"
                if medicine_template and goods_identity[1]:
                    matched_goods_rows[goods_identity] = excel_row
                if primary:
                    row.match_method = "lot_code_and_bidder_identifier"
                    exact_matches += 1
                else:
                    row.match_method = "lot_code_and_tax_code"
                    fallback_matches += 1
                if (
                    _normalised_text(source.get("bidderName"))
                    and _normalised_text(record.bidder_name)
                    and _normalised_text(source.get("bidderName"))
                    != _normalised_text(record.bidder_name)
                ):
                    row.warnings.append(
                        _issue(
                            "BIDDER_NAME_DIFFERS",
                            "Tên nhà thầu trong Excel khác dữ liệu ứng dụng nhưng mã vẫn khớp.",
                            excel_row=excel_row,
                        )
                    )
                if (
                    medicine_template
                    and _normalised_text(source.get("goodsName"))
                    and _normalised_text(record.goods_name)
                    and _normalised_text(source.get("goodsName"))
                    != _normalised_text(record.goods_name)
                ):
                    row.warnings.append(
                        _issue(
                            "MEDICINE_GOODS_NAME_DIFFERS",
                            "Tên hoạt chất trong Excel khác hàng hóa đã liên kết; STT và mã nhà thầu vẫn khớp.",
                            excel_row=excel_row,
                            goodsItemId=record.goods_item_id,
                        )
                    )
                if record.status is None:
                    row.warnings.append(
                        _issue(
                            "RESULT_NOT_APPROVED",
                            "Nhà thầu khớp nhưng chưa có kết quả được phê duyệt để điền.",
                            excel_row=excel_row,
                        )
                    )
                missing_fields = []
                if record.status == "Trúng thầu" and record.award_price is None:
                    missing_fields.append("Giá trúng thầu")
                if record.status == "Không trúng thầu" and not record.rejection_reason:
                    missing_fields.append("Lý do không đáp ứng")
                if missing_fields:
                    row.warnings.append(
                        _issue(
                            "RESULT_FIELDS_MISSING",
                            "Một số trường kết quả chưa có dữ liệu.",
                            excel_row=excel_row,
                            fields=missing_fields,
                        )
                    )
                if record.status is not None:
                    existing_values = list(source.get("existingOutputValues") or [])
                    existing_types = list(source.get("existingOutputDataTypes") or [])
                    new_values = record.output_values(definition.output_roles)
                    for index, (role, new_value) in enumerate(
                        zip(definition.output_roles, new_values, strict=True)
                    ):
                        old_value = (
                            existing_values[index]
                            if index < len(existing_values)
                            else None
                        )
                        old_type = (
                            existing_types[index]
                            if index < len(existing_types)
                            else ""
                        )
                        safe_new_value = safe_spreadsheet_text(new_value)
                        if old_type == "f" or not _output_values_equal(
                            old_value, safe_new_value
                        ):
                            row.changes.append(
                                {
                                    "field": role,
                                    "oldValue": old_value,
                                    "newValue": _json_value(safe_new_value),
                                    "source": f"approved_result.{role}",
                                }
                            )
                    row.writable = bool(row.changes)
        if source.get("hasExistingResult"):
            row.warnings.append(
                _issue(
                    "EXISTING_RESULT_WILL_BE_OVERWRITTEN",
                    "Dòng đã có dữ liệu ở các trường kết quả; các ô tương ứng của dòng khớp sẽ được ghi đè.",
                    excel_row=excel_row,
                )
            )
        warnings.extend(row.warnings)
        matched_rows.append(row)

    matched_count = exact_matches + fallback_matches
    approved_count = sum(
        1
        for row in matched_rows
        if not row.errors and row.record is not None and row.record.status is not None
    )
    writable_count = sum(int(row.writable and not row.errors) for row in matched_rows)
    if writable_count == 0 and not blocking_errors:
        blocking_errors.append(
            _issue(
                "NO_APPROVED_RESULT_TO_EXPORT",
                "Không có dòng kết quả đã phê duyệt có thể ghi vào workbook.",
            )
        )

    return {
        "sheetName": inspection.get("sheetName"),
        "templateType": inspection.get("templateType"),
        "templateVersion": inspection.get("templateVersion"),
        "templateFingerprint": inspection.get("templateFingerprint"),
        "headerRow": inspection.get("headerRow"),
        "totalRows": int(inspection.get("totalRows") or 0),
        "matchedRows": matched_count,
        "approvedRows": approved_count,
        "writableRows": writable_count,
        "updatedRows": writable_count,
        "exactMatches": exact_matches,
        "fallbackMatches": fallback_matches,
        "unmatchedRows": unmatched_rows,
        "duplicateRows": duplicate_rows,
        "conflictRows": conflict_rows,
        "missingLotRows": missing_lot_rows,
        "missingBidderIdentityRows": missing_identity_rows,
        "existingResultRows": int(inspection.get("existingResultRows") or 0),
        "blockingErrors": blocking_errors,
        "warnings": warnings,
        "canExport": not blocking_errors and writable_count > 0,
        "rows": [row.public_dict() for row in matched_rows],
        "_rowMatches": matched_rows,
    }


def export_updates_from_match(match_result: dict[str, Any]) -> list[dict[str, Any]]:
    template_type = str(match_result.get("templateType") or "standard")
    definition = next(
        (item for item in TEMPLATE_DEFINITIONS if item.template_type == template_type),
        STANDARD_TEMPLATE,
    )
    updates = []
    for row in match_result.get("_rowMatches") or []:
        if (
            row.errors
            or row.record is None
            or row.record.status is None
            or not row.writable
        ):
            continue
        updates.append(
            {
                "excelRow": row.excel_row,
                "sourceFingerprint": row.source_fingerprint,
                "values": row.record.output_values(definition.output_roles),
            }
        )
    return updates


def public_validation_result(
    match_result: dict[str, Any],
    *,
    page: int = 1,
    page_size: int = 100,
    status: str | None = None,
    warning: str | None = None,
    match_method: str | None = None,
    writable: bool | None = None,
) -> dict[str, Any]:
    rows = list(match_result.get("rows") or [])
    if status:
        rows = [item for item in rows if item.get("status") == status]
    if warning:
        rows = [
            item
            for item in rows
            if any(issue.get("code") == warning for issue in item.get("warnings") or [])
        ]
    if match_method:
        rows = [item for item in rows if item.get("matchMethod") == match_method]
    if writable is not None:
        rows = [item for item in rows if bool(item.get("writable")) is writable]
    bounded_page_size = max(1, min(int(page_size), 200))
    bounded_page = max(1, int(page))
    warning_summary: dict[str, int] = {}
    for issue in match_result.get("warnings") or []:
        code = str(issue.get("code") or "UNKNOWN")
        warning_summary[code] = warning_summary.get(code, 0) + 1
    blocking_error_summary: dict[str, int] = {}
    for issue in match_result.get("blockingErrors") or []:
        code = str(issue.get("code") or "UNKNOWN")
        blocking_error_summary[code] = blocking_error_summary.get(code, 0) + 1
    result = {
        key: value
        for key, value in match_result.items()
        if not key.startswith("_")
        and key not in {"rows", "warnings", "blockingErrors"}
    }
    filtered_count = len(rows)
    total_pages = max(1, (filtered_count + bounded_page_size - 1) // bounded_page_size)
    bounded_page = min(bounded_page, total_pages)
    start = (bounded_page - 1) * bounded_page_size
    result.update(
        {
            "rows": rows[start : start + bounded_page_size],
            "warnings": list(match_result.get("warnings") or [])[:bounded_page_size],
            "blockingErrors": list(match_result.get("blockingErrors") or [])[
                :bounded_page_size
            ],
            "warningSummary": warning_summary,
            "blockingErrorSummary": blocking_error_summary,
            "page": bounded_page,
            "pageSize": bounded_page_size,
            "filteredRows": filtered_count,
            "remainingRows": max(0, filtered_count - start - bounded_page_size),
            "totalPages": total_pages,
            "hasPreviousPage": bounded_page > 1,
            "hasNextPage": bounded_page < total_pages,
        }
    )
    return result


def _color_payload(color):
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb if color.type == "rgb" else None,
        "indexed": color.indexed if color.type == "indexed" else None,
        "theme": color.theme if color.type == "theme" else None,
        "tint": color.tint,
    }


def _side_payload(side):
    return {"style": side.style, "color": _color_payload(side.color)}


def _cell_style_payload(cell):
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    protection = cell.protection
    border = cell.border
    return {
        "numberFormat": cell.number_format,
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "strike": font.strike,
            "color": _color_payload(font.color),
        },
        "fill": {
            "type": fill.fill_type,
            "fg": _color_payload(fill.fgColor),
            "bg": _color_payload(fill.bgColor),
        },
        "border": {
            "left": _side_payload(border.left),
            "right": _side_payload(border.right),
            "top": _side_payload(border.top),
            "bottom": _side_payload(border.bottom),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap": alignment.wrap_text,
            "rotation": alignment.text_rotation,
            "shrink": alignment.shrink_to_fit,
            "indent": alignment.indent,
        },
        "protection": {"locked": protection.locked, "hidden": protection.hidden},
    }


def _dimension_payload(dimensions):
    return {
        str(key): {
            "width": getattr(value, "width", None),
            "height": getattr(value, "height", None),
            "hidden": value.hidden,
            "outline": value.outlineLevel,
            "collapsed": value.collapsed,
            "style": value.style,
        }
        for key, value in dimensions.items()
    }


def _validation_payload(worksheet):
    if not worksheet.data_validations:
        return []
    return [
        {
            "sqref": str(item.sqref),
            "type": item.type,
            "operator": item.operator,
            "formula1": item.formula1,
            "formula2": item.formula2,
            "allowBlank": item.allow_blank,
            "error": item.error,
            "prompt": item.prompt,
        }
        for item in worksheet.data_validations.dataValidation
    ]


def _preservation_digest(
    workbook,
    target_sheet: str,
    allowed_rows: set[int],
    allowed_columns: set[int],
) -> str:
    digest = hashlib.sha256()

    def update(value):
        digest.update(
            json.dumps(
                value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
                default=str,
            ).encode("utf-8")
        )
        digest.update(b"\n")

    properties = workbook.properties
    update(
        {
            "sheets": workbook.sheetnames,
            "definedNames": [
                {
                    "name": item.name,
                    "text": item.attr_text,
                    "localSheetId": item.localSheetId,
                    "hidden": item.hidden,
                }
                for item in workbook.defined_names.values()
            ],
            "properties": {
                name: _json_value(getattr(properties, name, None))
                for name in (
                    "title", "subject", "creator", "keywords", "description",
                    "lastModifiedBy", "category", "contentStatus", "identifier",
                    "language", "version", "created", "modified",
                )
            },
            "calculation": {
                "calcMode": workbook.calculation.calcMode,
                "fullCalcOnLoad": workbook.calculation.fullCalcOnLoad,
                "forceFullCalc": workbook.calculation.forceFullCalc,
            },
        }
    )
    for worksheet in workbook.worksheets:
        update(
            {
                "sheet": worksheet.title,
                "state": worksheet.sheet_state,
                "maxRow": worksheet.max_row,
                "maxColumn": worksheet.max_column,
                "merged": sorted(str(item) for item in worksheet.merged_cells.ranges),
                "freeze": str(worksheet.freeze_panes or ""),
                "gridlines": worksheet.sheet_view.showGridLines,
                "autoFilter": str(worksheet.auto_filter.ref or ""),
                "printArea": str(worksheet.print_area or ""),
                "printTitleRows": worksheet.print_title_rows,
                "printTitleCols": worksheet.print_title_cols,
                "columns": _dimension_payload(worksheet.column_dimensions),
                "rows": _dimension_payload(worksheet.row_dimensions),
                "validations": _validation_payload(worksheet),
                "pageSetup": {
                    "orientation": worksheet.page_setup.orientation,
                    "paperSize": worksheet.page_setup.paperSize,
                    "fitToWidth": worksheet.page_setup.fitToWidth,
                    "fitToHeight": worksheet.page_setup.fitToHeight,
                    "scale": worksheet.page_setup.scale,
                },
                "margins": {
                    name: getattr(worksheet.page_margins, name)
                    for name in ("left", "right", "top", "bottom", "header", "footer")
                },
                "tables": sorted(worksheet.tables.keys()),
                "charts": len(worksheet._charts),
                "images": len(worksheet._images),
            }
        )
        for row in worksheet.iter_rows():
            for cell in row:
                allowed_value = (
                    worksheet.title == target_sheet
                    and cell.row in allowed_rows
                    and cell.column in allowed_columns
                )
                original_value = _json_value(cell.value)
                semantically_blank = original_value in (None, "")
                update(
                    {
                        "sheet": worksheet.title,
                        "cell": cell.coordinate,
                        "value": None if allowed_value or semantically_blank else original_value,
                        "dataType": None if allowed_value or semantically_blank else cell.data_type,
                        "style": _cell_style_payload(cell),
                        "comment": (
                            {"text": cell.comment.text, "author": cell.comment.author}
                            if cell.comment else None
                        ),
                        "hyperlink": (
                            {
                                "target": cell.hyperlink.target,
                                "location": cell.hyperlink.location,
                            }
                            if cell.hyperlink else None
                        ),
                    }
                )
    return digest.hexdigest()


def _restore_document_properties(original: bytes, generated: bytes) -> bytes:
    """Keep OOXML document properties byte-for-byte from the uploaded workbook."""

    preserved_parts = {"docProps/core.xml", "docProps/app.xml"}
    with zipfile.ZipFile(BytesIO(original)) as source, zipfile.ZipFile(
        BytesIO(generated)
    ) as produced:
        available = preserved_parts.intersection(source.namelist())
        if not available:
            return generated
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as destination:
            for info in produced.infolist():
                data = source.read(info.filename) if info.filename in available else produced.read(info.filename)
                destination.writestr(info, data)
        return output.getvalue()


def _materialize_excel_value(value: Any) -> Any:
    if isinstance(value, dict) and set(value) == {"decimal"}:
        text = str(value["decimal"])
        if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
            raise AwardResultExcelError(
                "DECIMAL_VALUE_INVALID", "Giá trị số thập phân không hợp lệ."
            )
        return Decimal(text)
    return safe_spreadsheet_text(value)


def write_award_result_workbook(content: bytes, updates: list[dict[str, Any]]) -> bytes:
    """Write only header-detected result cells and prove all other state is stable."""

    if not updates:
        raise AwardResultExcelError(
            "NO_APPROVED_RESULT_TO_EXPORT",
            "Không có dòng kết quả đã phê duyệt có thể ghi vào workbook.",
        )

    inspection = inspect_award_result_workbook(content)
    if inspection.get("blockingErrors"):
        raise AwardResultExcelError(
            "WORKBOOK_STRUCTURE_INVALID",
            "Workbook có lỗi cấu trúc nên không thể xuất.",
        )
    sheet_name = str(inspection["sheetName"])
    output_columns = [int(value) for value in inspection["columnMap"]["output"]]
    output_roles = list(inspection["columnMap"].get("outputRoles") or [])
    if not output_roles or len(output_columns) != len(output_roles):
        raise AwardResultExcelError(
            "WORKBOOK_STRUCTURE_INVALID",
            "Không xác định đủ các cột kết quả để xuất.",
        )
    rows_by_number = {int(item["excelRow"]): item for item in inspection["rows"]}
    updates_by_row = {int(item["excelRow"]): item for item in updates}
    if len(updates_by_row) != len(updates):
        raise AwardResultExcelError(
            "DUPLICATE_UPDATE_ROW", "Danh sách cập nhật chứa dòng trùng nhau."
        )
    for row_number, update in updates_by_row.items():
        source = rows_by_number.get(row_number)
        if not source or not hmac.compare_digest(
            str(source.get("sourceFingerprint") or ""),
            str(update.get("sourceFingerprint") or ""),
        ):
            raise AwardResultExcelError(
                "WORKBOOK_CHANGED_AFTER_VALIDATION",
                "File đã thay đổi sau bước kiểm tra; vui lòng tải lên và kiểm tra lại.",
                status_code=409,
            )
        values = update.get("values")
        if not isinstance(values, list) or len(values) != len(output_columns):
            raise AwardResultExcelError(
                "UPDATE_VALUES_INVALID", "Dữ liệu cập nhật workbook không hợp lệ."
            )

    values_by_coordinate = {}
    for row_number, update in updates_by_row.items():
        for column, value in zip(output_columns, update["values"], strict=True):
            values_by_coordinate[
                f"{get_column_letter(column)}{row_number}"
            ] = _materialize_excel_value(value)
    try:
        result, _worksheet_part = patch_worksheet_cells(
            content, sheet_name, values_by_coordinate
        )
        reopened = load_workbook(
            BytesIO(result), data_only=False, keep_links=False, rich_text=True
        )
        reopened.close()
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        raise AwardResultExcelError(
            "WORKBOOK_PRESERVATION_FAILED",
            "Không thể chứng minh workbook được bảo toàn ngoài các trường kết quả.",
        ) from exc
    return result


def _row_mapping(row, columns: tuple[str, ...]):
    if row is None:
        return None
    if hasattr(row, "keys"):
        row_keys = set(row.keys())
        if all(name in row_keys for name in columns):
            return {name: row[name] for name in columns}
    return dict(zip(columns, row, strict=False))


def _number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return _exact_number(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    return _exact_number(value)


def _exact_number(value):
    if value is None or value == "":
        return None
    decimal_value = Decimal(str(value))
    if not decimal_value.is_finite():
        raise AwardResultExcelError(
            "NUMBER_INVALID", "Dữ liệu số trong kết quả không hợp lệ."
        )
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return {"decimal": format(decimal_value, "f")}


def _decimal_from_exact(value):
    if value is None:
        return None
    if isinstance(value, dict) and set(value) == {"decimal"}:
        return Decimal(str(value["decimal"]))
    return Decimal(str(value))


def load_award_result_dataset(package_id: str, organization_id: str, *, database_obj=None):
    """Load package, lots, bidders and bidder goods in bounded queries (never N+1)."""

    if database_obj is None:
        from backend.shared.helpers import database as database_obj

    connection = database_obj.get_connection()
    try:
        cursor = connection.cursor()
        package_columns = (
            "id", "ma_goi_thau", "ten_goi_thau", "phan_lo", "trang_thai",
            "nha_thau_trung_thau_id", "gia_trung_thau", "thoi_gian_goi_thau",
            "thoi_gian_hop_dong", "phuong_phap_danh_gia", "is_thuoc",
            "so_quyet_dinh_ket_qua", "ngay_quyet_dinh_ket_qua",
        )
        package = _row_mapping(
            cursor.execute(
                """SELECT id, ma_goi_thau, ten_goi_thau, phan_lo, trang_thai,
                          nha_thau_trung_thau_id, gia_trung_thau,
                          thoi_gian_goi_thau, thoi_gian_hop_dong,
                          phuong_phap_danh_gia, is_thuoc,
                          so_quyet_dinh_ket_qua, ngay_quyet_dinh_ket_qua
                   FROM goi_thau
                   WHERE id = ? AND organization_id = ?""",
                (package_id, organization_id),
            ).fetchone(),
            package_columns,
        )
        if not package:
            raise AwardResultExcelError(
                "PACKAGE_NOT_FOUND", "Không tìm thấy gói thầu.", status_code=404
            )

        lot_columns = (
            "id", "ma_phan_lo", "ten_phan_lo", "nha_thau_trung_thau_id",
            "gia_trung_thau", "thoi_gian_goi_thau", "thoi_gian_hop_dong",
            "outcome", "sequence_no",
        )
        lot_rows = cursor.execute(
            """SELECT lot.id, lot.ma_phan_lo, lot.ten_phan_lo,
                      lot.nha_thau_trung_thau_id, lot.gia_trung_thau,
                      lot.thoi_gian_goi_thau, lot.thoi_gian_hop_dong,
                      detail.outcome, batch.sequence_no
               FROM goi_thau_phan_lo AS lot
               LEFT JOIN dot_xu_ly_phan_lo_chi_tiet AS detail
                 ON detail.organization_id = lot.organization_id
                AND detail.lot_id = lot.id
                AND detail.current_stage = 'RESULT_APPROVED'
               LEFT JOIN dot_xu_ly_phan_lo AS batch
                 ON batch.organization_id = detail.organization_id
                AND batch.id = detail.batch_id
               WHERE lot.goi_thau_id = ? AND lot.organization_id = ?
                 AND lot.archived_at IS NULL
               ORDER BY lot.sort_order, lot.id, batch.sequence_no DESC""",
            (package_id, organization_id),
        ).fetchall()
        lots = {}
        for raw_row in lot_rows:
            row = _row_mapping(raw_row, lot_columns)
            lots.setdefault(str(row["id"]), row)
        lots_by_code = {
            normalize_code(row["ma_phan_lo"]): row
            for row in lots.values()
            if normalize_code(row["ma_phan_lo"])
        }

        opening_columns = (
            "opening_id", "ma_phan_lo", "ma_dinh_danh", "nha_thau_id",
            "ten_nha_thau", "ma_nha_thau", "ma_so_thue", "gia_du_thau",
            "gia_sau_giam_gia", "gia_danh_gia_sau_uu_dai",
            "diem", "ly_do_loai", "danh_gia_ket_luan",
        )
        opening_rows = cursor.execute(
            """SELECT opening.id, opening.ma_phan_lo, opening.ma_dinh_danh,
                      opening.nha_thau_id,
                      COALESCE(NULLIF(opening.ten_nha_thau, ''), bidder.ten_nha_thau),
                      bidder.ma_nha_thau, bidder.ma_so_thue,
                      opening.gia_du_thau, opening.gia_sau_giam_gia,
                      opening.gia_danh_gia_sau_uu_dai,
                      result.diem, result.ly_do_loai,
                      result.danh_gia_ket_luan
               FROM thong_tin_mo_thau AS opening
               JOIN nha_thau AS bidder
                 ON bidder.organization_id = opening.organization_id
                AND bidder.id = opening.nha_thau_id
               LEFT JOIN ket_qua_danh_gia_nha_thau AS result
                 ON result.organization_id = opening.organization_id
                AND result.thong_tin_mo_thau_id = opening.id
               WHERE opening.goi_thau_id = ? AND opening.organization_id = ?
                 AND opening.archived_at IS NULL
               ORDER BY opening.ma_phan_lo, opening.id""",
            (package_id, organization_id),
        ).fetchall()
        goods_columns = (
            "offered_goods_id", "thong_tin_mo_thau_id", "khoi_luong",
            "don_gia_du_thau", "gia_tri_co_so_sau_giam_gia",
            "goi_thau_hang_hoa_id", "stt_nguon", "danh_muc_hang_hoa",
            "offered_unit", "mapping_status", "requirement_code",
            "requirement_name", "requirement_unit", "sort_order",
        )
        goods_rows = cursor.execute(
            """SELECT offered.id, offered.thong_tin_mo_thau_id,
                      offered.khoi_luong, offered.don_gia_du_thau,
                      offered.gia_tri_co_so_sau_giam_gia,
                      offered.goi_thau_hang_hoa_id, offered.stt_nguon,
                      offered.danh_muc_hang_hoa, offered.don_vi_tinh,
                      offered.mapping_status, requirement.ma_hang_hoa,
                      requirement.ten_hang_hoa, requirement.don_vi_tinh,
                      offered.sort_order
               FROM hang_hoa_du_thau_nha_thau AS offered
               LEFT JOIN goi_thau_hang_hoa AS requirement
                 ON requirement.organization_id = offered.organization_id
                AND requirement.goi_thau_id = offered.goi_thau_id
                AND requirement.id = offered.goi_thau_hang_hoa_id
               WHERE offered.goi_thau_id = ?
                 AND offered.organization_id = ? AND offered.is_draft = 0
               ORDER BY offered.thong_tin_mo_thau_id,
                        offered.sort_order, offered.id""",
            (package_id, organization_id),
        ).fetchall()
    finally:
        connection.close()

    goods_by_opening: dict[str, list[dict[str, Any]]] = {}
    for raw_row in goods_rows:
        goods = _row_mapping(raw_row, goods_columns)
        goods_by_opening.setdefault(str(goods["thong_tin_mo_thau_id"]), []).append(goods)

    records = []
    dataset_errors: list[dict[str, Any]] = []
    package_is_lotted = str(package["phan_lo"] or "") == "Có"
    package_is_medicine = bool(int(package["is_thuoc"] or 0))
    package_status = str(package["trang_thai"] or "")
    edit_path = f"/packages/{package_id}/award-result"
    opening_bidder_ids_by_lot: dict[str, set[str]] = {}
    all_opening_bidder_ids: set[str] = set()
    for raw_row in opening_rows:
        opening = _row_mapping(raw_row, opening_columns)
        bidder_id = str(opening.get("nha_thau_id") or "")
        all_opening_bidder_ids.add(bidder_id)
        opening_bidder_ids_by_lot.setdefault(
            normalize_code(opening.get("ma_phan_lo")), set()
        ).add(bidder_id)

    approved_lots = [
        lot for lot in lots.values() if str(lot.get("outcome") or "").strip()
    ]
    has_approved_result = package_status in {"AWARDED", "PARTIALLY_AWARDED"} or bool(
        approved_lots
    )
    if has_approved_result and (
        not str(package.get("so_quyet_dinh_ket_qua") or "").strip()
        or not str(package.get("ngay_quyet_dinh_ket_qua") or "").strip()
    ):
        dataset_errors.append(
            _issue(
                "AWARD_DECISION_NOT_READY",
                "Kết quả đã duyệt nhưng thiếu số hoặc ngày quyết định phê duyệt.",
                editPath=edit_path,
            )
        )
    if package_is_lotted:
        for lot in approved_lots:
            parsed_lot_outcome = parse_lot_outcome(lot.get("outcome"))
            if parsed_lot_outcome is not LotApprovedOutcome.AWARDED:
                continue
            winner = str(lot.get("nha_thau_trung_thau_id") or "")
            lot_code = normalize_code(lot.get("ma_phan_lo"))
            if not winner or winner not in opening_bidder_ids_by_lot.get(lot_code, set()):
                dataset_errors.append(
                    _issue(
                        "AWARD_WINNER_NOT_READY",
                        "Phần/lô đã duyệt trúng thầu nhưng nhà thầu trúng không hợp lệ.",
                        lotCode=lot.get("ma_phan_lo"),
                        editPath=edit_path,
                    )
                )
            if lot.get("gia_trung_thau") is None:
                dataset_errors.append(
                    _issue(
                        "AWARD_PRICE_NOT_READY",
                        "Phần/lô đã duyệt trúng thầu nhưng thiếu giá trúng thầu.",
                        lotCode=lot.get("ma_phan_lo"),
                        editPath=edit_path,
                    )
                )
    elif package_status == "AWARDED":
        package_winner = str(package.get("nha_thau_trung_thau_id") or "")
        if not package_winner or package_winner not in all_opening_bidder_ids:
            dataset_errors.append(
                _issue(
                    "AWARD_WINNER_NOT_READY",
                    "Gói thầu đã duyệt nhưng nhà thầu trúng không hợp lệ.",
                    editPath=edit_path,
                )
            )
        if package.get("gia_trung_thau") is None:
            dataset_errors.append(
                _issue(
                    "AWARD_PRICE_NOT_READY",
                    "Gói thầu đã duyệt nhưng thiếu giá trúng thầu.",
                    editPath=edit_path,
                )
            )
    for raw_row in opening_rows:
        opening = _row_mapping(raw_row, opening_columns)
        lot = lots_by_code.get(normalize_code(opening["ma_phan_lo"]))
        winner_id = None
        award_price = package_duration = contract_duration = None
        outcome = None
        if package_is_lotted and lot:
            winner_id = lot["nha_thau_trung_thau_id"]
            award_price = lot["gia_trung_thau"]
            package_duration = lot["thoi_gian_goi_thau"]
            contract_duration = lot["thoi_gian_hop_dong"]
            outcome = str(lot["outcome"] or "") or None
        elif not package_is_lotted:
            winner_id = package["nha_thau_trung_thau_id"]
            award_price = package["gia_trung_thau"]
            package_duration = package["thoi_gian_goi_thau"]
            contract_duration = package["thoi_gian_hop_dong"]

        status = None
        is_winner = False
        rejection_reason = None
        lot_cancelled = False
        parsed_outcome = (
            parse_lot_outcome(outcome)
            if package_is_lotted
            else (
                LotApprovedOutcome.AWARDED
                if package_status == "AWARDED"
                else LotApprovedOutcome.CANCELLED_LOT
                if package_status == "CANCELLED"
                else None
            )
        )
        if outcome and parsed_outcome is None:
            dataset_errors.append(
                _issue(
                    "LOT_APPROVED_OUTCOME_UNKNOWN",
                    "Kết quả phê duyệt phần/lô không được hỗ trợ để xuất Excel.",
                    outcome=str(outcome),
                )
            )
        if parsed_outcome is not None:
            is_winner = bool(
                parsed_outcome is LotApprovedOutcome.AWARDED
                and winner_id
                and str(winner_id) == str(opening["nha_thau_id"])
            )
            mapped = map_bidder_award(
                parsed_outcome,
                is_winner=is_winner,
                evaluation_reason=opening["ly_do_loai"],
            )
            status = mapped.status.value
            rejection_reason = mapped.reason
            lot_cancelled = not mapped.lot_has_award
        corrected_price = (
            opening["gia_sau_giam_gia"]
            if opening["gia_sau_giam_gia"] is not None
            else opening["gia_du_thau"]
        )
        if parsed_outcome is not None and corrected_price is None:
            dataset_errors.append(
                _issue(
                    "CORRECTED_PRICE_NOT_READY",
                    "Kết quả đã duyệt nhưng thiếu giá dự thầu sau hiệu chỉnh/giảm giá.",
                    openingId=str(opening["opening_id"]),
                    editPath=edit_path,
                )
            )
        if (
            parsed_outcome is LotApprovedOutcome.AWARDED
            and not is_winner
            and not rejection_reason
        ):
            dataset_errors.append(
                _issue(
                    "NON_WINNER_REASON_NOT_READY",
                    "Nhà thầu không trúng trong phần/lô có người trúng nhưng thiếu lý do.",
                    openingId=str(opening["opening_id"]),
                    editPath=edit_path,
                )
            )
        if is_winner and (
            not str(package_duration or "").strip()
            or not str(contract_duration or "").strip()
        ):
            dataset_errors.append(
                _issue(
                    "AWARD_DURATION_NOT_READY",
                    "Kết quả trúng thầu thiếu thời gian thực hiện gói thầu hoặc hợp đồng.",
                    openingId=str(opening["opening_id"]),
                    editPath=edit_path,
                )
            )
        goods_for_opening = goods_by_opening.get(str(opening["opening_id"]), [])
        if package_is_medicine and not goods_for_opening:
            dataset_errors.append(
                _issue(
                    "MEDICINE_GOODS_MISSING",
                    "Nhà thầu chưa có hàng thuốc chính thức để đối chiếu.",
                    openingId=str(opening["opening_id"]),
                )
            )
        record_goods = goods_for_opening if package_is_medicine else [None]
        seen_sequences: set[str] = set()
        winner_item_totals: list[Decimal] = []
        for goods in record_goods:
            award_quantity = award_unit_price = item_award_price = None
            item_corrected_price = corrected_price
            if goods is not None:
                goods_id = str(goods.get("goi_thau_hang_hoa_id") or "").strip()
                sequence = str(goods.get("stt_nguon") or "").strip()
                sequence_key = normalize_code(sequence)
                issue_context = {
                    "openingId": str(opening["opening_id"]),
                    "offeredGoodsId": str(goods.get("offered_goods_id") or ""),
                    "goodsItemId": goods_id or None,
                }
                if not goods_id:
                    dataset_errors.append(
                        _issue(
                            "MEDICINE_GOODS_ID_MISSING",
                            "Hàng thuốc thiếu mã hàng hóa ổn định.",
                            **issue_context,
                        )
                    )
                if not sequence_key:
                    dataset_errors.append(
                        _issue(
                            "MEDICINE_GOODS_SEQUENCE_MISSING",
                            "Hàng thuốc thiếu STT nguồn để đối chiếu.",
                            **issue_context,
                        )
                    )
                elif sequence_key in seen_sequences:
                    dataset_errors.append(
                        _issue(
                            "MEDICINE_GOODS_SEQUENCE_DUPLICATED",
                            "Một nhà thầu có nhiều hàng thuốc cùng STT nguồn trong một phần/lô.",
                            goodsSequence=sequence,
                            **issue_context,
                        )
                    )
                else:
                    seen_sequences.add(sequence_key)
                if str(goods.get("mapping_status") or "") != "matched":
                    dataset_errors.append(
                        _issue(
                            "MEDICINE_GOODS_MAPPING_NOT_READY",
                            "Hàng thuốc chưa được liên kết chính thức với danh mục yêu cầu.",
                            mappingStatus=str(goods.get("mapping_status") or ""),
                            **issue_context,
                        )
                    )
                offered_unit = _normalised_text(goods.get("offered_unit"))
                requirement_unit = _normalised_text(goods.get("requirement_unit"))
                if offered_unit and requirement_unit and offered_unit != requirement_unit:
                    dataset_errors.append(
                        _issue(
                            "MEDICINE_GOODS_UNIT_MISMATCH",
                            "Đơn vị tính của hàng dự thầu khác danh mục hàng hóa đã liên kết.",
                            offeredUnit=str(goods.get("offered_unit") or ""),
                            requirementUnit=str(goods.get("requirement_unit") or ""),
                            **issue_context,
                        )
                    )
                allocated_total = goods.get("gia_tri_co_so_sau_giam_gia")
                if allocated_total is not None:
                    item_corrected_price = allocated_total
                if is_winner:
                    award_quantity = _exact_number(goods.get("khoi_luong"))
                    quantity_decimal = _decimal_from_exact(award_quantity)
                    if not quantity_decimal or quantity_decimal <= 0:
                        dataset_errors.append(
                            _issue(
                                "MEDICINE_AWARD_QUANTITY_MISSING",
                                "Hàng thuốc trúng thầu chưa có số lượng hợp lệ.",
                                **issue_context,
                            )
                        )
                    else:
                        if allocated_total is not None:
                            unit_decimal = Decimal(str(allocated_total)) / quantity_decimal
                            award_unit_price = _exact_number(unit_decimal)
                        else:
                            award_unit_price = _exact_number(goods.get("don_gia_du_thau"))
                        unit_decimal = _decimal_from_exact(award_unit_price)
                        if unit_decimal is None:
                            dataset_errors.append(
                                _issue(
                                    "MEDICINE_AWARD_UNIT_PRICE_MISSING",
                                    "Hàng thuốc trúng thầu chưa có đơn giá hợp lệ.",
                                    **issue_context,
                                )
                            )
                        else:
                            computed_total = quantity_decimal * unit_decimal
                            winner_item_totals.append(computed_total)
                            item_award_price = _exact_number(computed_total)
            records.append(
                AwardRecord(
                    opening_id=str(opening["opening_id"]),
                    lot_code=str(opening["ma_phan_lo"] or ""),
                    bidder_identifier=str(opening["ma_dinh_danh"] or ""),
                    tax_code=str(opening["ma_so_thue"] or ""),
                    bidder_name=str(opening["ten_nha_thau"] or ""),
                    status=status,
                    goods_item_id=(
                        str(goods.get("goi_thau_hang_hoa_id") or "").strip() or None
                    ) if goods is not None else None,
                    goods_sequence=(
                        str(goods.get("stt_nguon") or "").strip() or None
                    ) if goods is not None else None,
                    goods_code=(
                        str(goods.get("requirement_code") or "").strip() or None
                    ) if goods is not None else None,
                    goods_name=(
                        str(
                            goods.get("requirement_name")
                            or goods.get("danh_muc_hang_hoa")
                            or ""
                        ).strip() or None
                    ) if goods is not None else None,
                    goods_unit=(
                        str(goods.get("requirement_unit") or "").strip() or None
                    ) if goods is not None else None,
                    corrected_price=_number(item_corrected_price),
                    technical_score=_number(opening["diem"]),
                    evaluated_price=_number(opening["gia_danh_gia_sau_uu_dai"]),
                    award_quantity=award_quantity,
                    award_unit_price=award_unit_price,
                    award_price=item_award_price,
                    rejection_reason=rejection_reason,
                    package_duration=(str(package_duration or "").strip() or None) if is_winner else None,
                    contract_duration=(str(contract_duration or "").strip() or None) if is_winner else None,
                    other_content=(
                        str(opening["danh_gia_ket_luan"] or "").strip() or None
                    ),
                    lot_cancelled=lot_cancelled,
                )
            )
        if (
            package_is_medicine
            and is_winner
            and award_price is not None
            and len(winner_item_totals) == len(goods_for_opening)
            and sum(winner_item_totals, Decimal("0")) != Decimal(str(award_price))
        ):
            dataset_errors.append(
                _issue(
                    "MEDICINE_AWARD_VALUE_CONFLICT",
                    "Tổng số lượng nhân đơn giá theo hàng không khớp giá trúng thầu đã phê duyệt.",
                    openingId=str(opening["opening_id"]),
                    computedTotal=format(sum(winner_item_totals, Decimal("0")), "f"),
                    approvedTotal=format(Decimal(str(award_price)), "f"),
                )
            )
    return {
        "package": package,
        "records": records,
        "lotCodes": [row["ma_phan_lo"] for row in lots.values()],
        "blockingErrors": dataset_errors,
    }


def _winning_unit_price(row: dict[str, Any]) -> Decimal:
    quantity = Decimal(str(row.get("khoi_luong") or "0"))
    if quantity <= 0:
        raise AwardResultExcelError(
            "WINNING_GOODS_QUANTITY_INVALID",
            "Hàng hóa trúng thầu có khối lượng không hợp lệ.",
            status_code=409,
        )
    allocated = row.get("gia_tri_co_so_sau_giam_gia")
    if allocated is not None:
        return Decimal(str(allocated)) / quantity
    unit_price = row.get("don_gia_du_thau")
    if unit_price is None:
        raise AwardResultExcelError(
            "WINNING_GOODS_PRICE_MISSING",
            "Hàng hóa trúng thầu thiếu đơn giá chính thức.",
            status_code=409,
        )
    return Decimal(str(unit_price))


def _decimal_text(value: Decimal) -> str:
    return format(value.normalize(), "f") if value else "0"


def load_winning_goods_export_model(
    package_id: str,
    organization_id: str,
    expected_revision: int,
    *,
    database_obj=None,
):
    """Load only committed winner goods at one authoritative package revision."""

    if database_obj is None:
        from backend.shared.helpers import database as database_obj
    connection = database_obj.get_connection()
    try:
        cursor = connection.cursor()
        package = cursor.execute(
            """SELECT id, ma_goi_thau, ten_goi_thau, linh_vuc, phan_lo,
                      nha_thau_trung_thau_id, row_version
               FROM goi_thau
               WHERE id = ? AND organization_id = ? AND archived_at IS NULL""",
            (package_id, organization_id),
        ).fetchone()
        if not package:
            raise AwardResultExcelError(
                "PACKAGE_NOT_FOUND", "Không tìm thấy gói thầu.", status_code=404
            )
        package = _row_mapping(
            package,
            ("id", "ma_goi_thau", "ten_goi_thau", "linh_vuc", "phan_lo",
             "nha_thau_trung_thau_id", "row_version"),
        )
        actual_revision = int(package.get("row_version") or 1)
        if actual_revision != expected_revision:
            raise AwardResultExcelError(
                "PACKAGE_REVISION_CONFLICT",
                "Gói thầu đã thay đổi; vui lòng tải lại trước khi xuất.",
                status_code=409,
            )
        if str(package.get("linh_vuc") or "").strip() not in {"Hàng hóa", "Hỗn hợp"}:
            raise AwardResultExcelError(
                "WINNING_GOODS_EXPORT_UNSUPPORTED",
                "Chức năng chỉ áp dụng cho gói Hàng hóa hoặc Hỗn hợp.",
                status_code=409,
            )
        lot_rows = cursor.execute(
            """SELECT id, ma_phan_lo, ten_phan_lo, nha_thau_trung_thau_id,
                      sort_order
               FROM goi_thau_phan_lo
               WHERE goi_thau_id = ? AND organization_id = ?
                 AND archived_at IS NULL
               ORDER BY sort_order, id""",
            (package_id, organization_id),
        ).fetchall()
        lots = [
            _row_mapping(row, ("id", "ma_phan_lo", "ten_phan_lo",
                               "nha_thau_trung_thau_id", "sort_order"))
            for row in lot_rows
        ]
        opening_rows = cursor.execute(
            """SELECT opening.id, opening.nha_thau_id, opening.ma_phan_lo,
                      COALESCE(NULLIF(opening.ten_nha_thau, ''), bidder.ten_nha_thau)
               FROM thong_tin_mo_thau AS opening
               JOIN nha_thau AS bidder
                 ON bidder.organization_id = opening.organization_id
                AND bidder.id = opening.nha_thau_id
               WHERE opening.goi_thau_id = ? AND opening.organization_id = ?
                 AND opening.archived_at IS NULL
               ORDER BY opening.ma_phan_lo, opening.id""",
            (package_id, organization_id),
        ).fetchall()
        openings = [
            _row_mapping(row, ("id", "nha_thau_id", "ma_phan_lo", "ten_nha_thau"))
            for row in opening_rows
        ]
        goods_rows = cursor.execute(
            """SELECT id, thong_tin_mo_thau_id, phan_lo_id, stt_nguon,
                      danh_muc_hang_hoa, ky_ma_hieu, nhan_hieu, nam_san_xuat,
                      xuat_xu, hang_san_xuat, cau_hinh_tinh_nang_ky_thuat,
                      don_vi_tinh, khoi_luong, ma_hs, don_gia_du_thau,
                      gia_tri_co_so_sau_giam_gia, sort_order
               FROM hang_hoa_du_thau_nha_thau
               WHERE goi_thau_id = ? AND organization_id = ? AND is_draft = 0
               ORDER BY thong_tin_mo_thau_id, sort_order, id""",
            (package_id, organization_id),
        ).fetchall()
    finally:
        connection.close()

    goods_columns = (
        "id", "thong_tin_mo_thau_id", "phan_lo_id", "stt_nguon",
        "danh_muc_hang_hoa", "ky_ma_hieu", "nhan_hieu", "nam_san_xuat",
        "xuat_xu", "hang_san_xuat", "cau_hinh_tinh_nang_ky_thuat",
        "don_vi_tinh", "khoi_luong", "ma_hs", "don_gia_du_thau",
        "gia_tri_co_so_sau_giam_gia", "sort_order",
    )
    goods_by_opening: dict[str, list[dict[str, Any]]] = {}
    for raw in goods_rows:
        row = _row_mapping(raw, goods_columns)
        goods_by_opening.setdefault(str(row["thong_tin_mo_thau_id"]), []).append(row)

    is_lotted = str(package.get("phan_lo") or "") == "Có"
    scopes = []
    if is_lotted:
        scopes = [
            (lot, str(lot.get("nha_thau_trung_thau_id") or ""))
            for lot in lots if str(lot.get("nha_thau_trung_thau_id") or "")
        ]
    else:
        winner = str(package.get("nha_thau_trung_thau_id") or "")
        if winner:
            scopes = [(None, winner)]
    if not scopes:
        raise AwardResultExcelError(
            "WINNING_GOODS_RESULT_MISSING",
            "Gói thầu chưa có kết quả trúng thầu chính thức để xuất.",
            status_code=409,
        )

    grouped: dict[str, dict[str, Any]] = {}
    for lot, winner_id in scopes:
        lot_code = str((lot or {}).get("ma_phan_lo") or "").strip()
        matches = [
            opening for opening in openings
            if str(opening.get("nha_thau_id") or "") == winner_id
            and str(opening.get("ma_phan_lo") or "").strip() == lot_code
        ]
        if len(matches) != 1:
            raise AwardResultExcelError(
                "WINNING_GOODS_OPENING_AMBIGUOUS",
                "Không xác định duy nhất hồ sơ mở thầu của nhà thầu trúng thầu.",
                status_code=409,
            )
        opening = matches[0]
        official = [
            row for row in goods_by_opening.get(str(opening["id"]), [])
            if (not lot or str(row.get("phan_lo_id") or "") == str(lot["id"]))
        ]
        if not official:
            raise AwardResultExcelError(
                "WINNING_GOODS_COMMITTED_DATA_MISSING",
                "Nhà thầu trúng thầu chưa có hàng hóa chính thức để xuất.",
                status_code=409,
            )
        group = grouped.setdefault(winner_id, {
            "contractorId": winner_id,
            "contractorName": str(opening.get("ten_nha_thau") or ""),
            "lots": [],
        })
        projected = []
        for index, row in enumerate(official, start=1):
            projected.append({
                "stt": str(row.get("stt_nguon") or index),
                "danhMucHangHoa": str(row.get("danh_muc_hang_hoa") or ""),
                "kyMaHieu": str(row.get("ky_ma_hieu") or ""),
                "nhanHieu": str(row.get("nhan_hieu") or ""),
                "namSanXuat": str(row.get("nam_san_xuat") or ""),
                "xuatXu": str(row.get("xuat_xu") or ""),
                "hangSanXuat": str(row.get("hang_san_xuat") or ""),
                "cauHinhTinhNangKyThuat": str(
                    row.get("cau_hinh_tinh_nang_ky_thuat") or ""
                ),
                "donViTinh": str(row.get("don_vi_tinh") or ""),
                "khoiLuong": str(row.get("khoi_luong") or ""),
                "maHs": str(row.get("ma_hs") or ""),
                "donGiaTrungThau": _decimal_text(_winning_unit_price(row)),
            })
        group["lots"].append({
            "lotId": str((lot or {}).get("id") or ""),
            "lotCode": lot_code,
            "lotName": str((lot or {}).get("ten_phan_lo") or ""),
            "rows": projected,
        })
    return {
        "packageCode": str(package.get("ma_goi_thau") or ""),
        "packageName": str(package.get("ten_goi_thau") or ""),
        "isLotted": is_lotted,
        "revision": actual_revision,
        "groups": list(grouped.values()),
    }


def find_foreign_lot_codes(
    lot_codes: Iterable[Any], package_id: str, organization_id: str, *, database_obj=None
) -> set[str]:
    if database_obj is None:
        from backend.shared.helpers import database as database_obj

    codes = sorted({normalize_code(item) for item in lot_codes if normalize_code(item)})
    if not codes:
        return set()
    found = set()
    connection = database_obj.get_connection()
    try:
        cursor = connection.cursor()
        placeholders = ", ".join("?" for _ in codes)
        rows = cursor.execute(
            f"""SELECT ma_phan_lo_normalized FROM goi_thau_phan_lo
                WHERE organization_id = ? AND goi_thau_id != ?
                  AND archived_at IS NULL
                  AND ma_phan_lo_normalized IN ({placeholders})""",  # noqa: S608 - placeholders only
            (organization_id, package_id, *codes),
        ).fetchall()
        found.update(str(row[0]) for row in rows)
    finally:
        connection.close()
    return found


def _validation_root() -> Path:
    root = (resolve_runtime_path("DOCUMENT_WORKER_TEMP_DIR") / "award-result-validations").resolve()
    root.mkdir(parents=True, exist_ok=True, mode=0o700)
    return root


def validate_artifact_store_configuration(environ=None) -> None:
    environment = os.environ if environ is None else environ
    production = str(environment.get("APP_ENV", "development")).casefold() in {
        "prod",
        "production",
    }
    try:
        instances = max(1, int(environment.get("APP_INSTANCE_COUNT", "1")))
    except (TypeError, ValueError):
        instances = 1
    shared_confirmed = str(
        environment.get("AWARD_RESULT_ARTIFACT_SHARED_STORAGE_CONFIRMED", "false")
    ).casefold() in {"1", "true", "yes", "on"}
    if production and instances > 1 and not shared_confirmed:
        raise RuntimeError(
            "AWARD_RESULT_ARTIFACT_SHARED_STORAGE_CONFIRMED=true is required "
            "when APP_INSTANCE_COUNT is greater than one."
        )


def _signing_key(environ=None) -> bytes:
    environment = os.environ if environ is None else environ
    configured = str(
        environment.get("AWARD_RESULT_EXCEL_TOKEN_KEY")
        or environment.get("SYNC_CURSOR_SIGNING_KEY")
        or ""
    )
    production = str(environment.get("APP_ENV", "development")).casefold() in {
        "prod", "production"
    }
    if production and len(configured.encode("utf-8")) < 32:
        raise RuntimeError(
            "AWARD_RESULT_EXCEL_TOKEN_KEY or SYNC_CURSOR_SIGNING_KEY must be at least 32 bytes."
        )
    if not configured:
        configured = "biddingflow-development-award-result-excel-token"
    return hashlib.sha256(("award-result-excel\0" + configured).encode("utf-8")).digest()


def _encode_signature(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def _token(validation_id: str) -> str:
    signature = hmac.new(_signing_key(), validation_id.encode("ascii"), hashlib.sha256).digest()
    return f"{validation_id}.{_encode_signature(signature)}"


def _validation_path(validation_id: str) -> Path:
    root = _validation_root()
    path = (root / f"validation-{validation_id}").resolve()
    if path.parent != root:
        raise AwardResultExcelError("VALIDATION_TOKEN_INVALID", "Validation token không hợp lệ.")
    return path


def _configured_limit(name: str, default: int) -> int:
    try:
        return max(1, int(os.environ.get(name, default)))
    except (TypeError, ValueError):
        return default


def _artifact_inventory(*, now: int | None = None) -> list[dict[str, Any]]:
    current = int(time.time() if now is None else now)
    inventory = []
    for path in _validation_root().glob("validation-*"):
        try:
            if path.is_symlink() or path.resolve().parent != _validation_root():
                continue
            metadata = json.loads(
                (path / "metadata.json").read_text(encoding="utf-8")
            )
            inventory.append(
                {
                    "path": path,
                    "userId": str(metadata.get("userId") or ""),
                    "organizationId": str(metadata.get("organizationId") or ""),
                    "sizeBytes": int(metadata.get("sizeBytes") or 0),
                    "expiresAt": int(metadata.get("expiresAt") or 0),
                    "expired": int(metadata.get("expiresAt") or 0) <= current,
                }
            )
        except (OSError, ValueError, json.JSONDecodeError):
            continue
    return inventory


def validation_artifact_metrics(*, now: int | None = None) -> dict[str, int]:
    inventory = _artifact_inventory(now=now)
    return {
        "count": len(inventory),
        "totalBytes": sum(item["sizeBytes"] for item in inventory),
        "expiredCount": sum(int(item["expired"]) for item in inventory),
        "cleanupFailures": _ARTIFACT_CLEANUP_FAILURES,
        "quotaRejections": _ARTIFACT_QUOTA_REJECTIONS,
    }


def _enforce_artifact_quota(
    *, user_id: str, organization_id: str, size_bytes: int, now: int
) -> None:
    global _ARTIFACT_QUOTA_REJECTIONS

    active = [item for item in _artifact_inventory(now=now) if not item["expired"]]
    user_items = [item for item in active if item["userId"] == str(user_id)]
    organization_items = [
        item
        for item in active
        if item["organizationId"] == str(organization_id)
    ]
    exceeded = (
        len(user_items) + 1
        > _configured_limit("AWARD_RESULT_ARTIFACT_MAX_PER_USER", 20)
        or sum(item["sizeBytes"] for item in user_items) + size_bytes
        > _configured_limit("AWARD_RESULT_ARTIFACT_MAX_BYTES_PER_USER", 100 * 1024 * 1024)
        or len(organization_items) + 1
        > _configured_limit("AWARD_RESULT_ARTIFACT_MAX_PER_ORGANIZATION", 200)
        or sum(item["sizeBytes"] for item in organization_items) + size_bytes
        > _configured_limit(
            "AWARD_RESULT_ARTIFACT_MAX_BYTES_PER_ORGANIZATION",
            1024 * 1024 * 1024,
        )
        or sum(item["sizeBytes"] for item in active) + size_bytes
        > _configured_limit("AWARD_RESULT_ARTIFACT_MAX_GLOBAL_BYTES", 2 * 1024 * 1024 * 1024)
    )
    if exceeded:
        _ARTIFACT_QUOTA_REJECTIONS += 1
        raise AwardResultExcelError(
            "VALIDATION_ARTIFACT_QUOTA_EXCEEDED",
            "Đã đạt giới hạn tệp Excel chờ xuất; vui lòng hoàn tất hoặc hủy tệp cũ.",
            status_code=429,
        )


def cleanup_expired_validation_artifacts(*, now: int | None = None, limit: int = 128) -> int:
    global _ARTIFACT_CLEANUP_FAILURES

    current = int(time.time() if now is None else now)
    removed = 0
    root = _validation_root()
    candidates = []
    for path in root.glob("validation-*"):
        try:
            if path.is_symlink() or path.resolve().parent != root:
                continue
            metadata_path = path / "metadata.json"
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            lock_path = path / "export.lock"
            if lock_path.exists():
                try:
                    claimed_at = int(lock_path.read_text(encoding="ascii") or 0)
                except (OSError, ValueError):
                    claimed_at = current
                if claimed_at + VALIDATION_EXPORT_LEASE_SECONDS > current:
                    continue
            candidates.append((int(metadata.get("expiresAt") or 0), path))
        except (OSError, ValueError, json.JSONDecodeError):
            _ARTIFACT_CLEANUP_FAILURES += 1
            try:
                if (
                    not (path / "export.lock").exists()
                    and int(path.stat().st_mtime) + VALIDATION_TTL_SECONDS
                    <= current
                    and removed < max(1, min(limit, 512))
                ):
                    shutil.rmtree(path)
                    removed += 1
            except OSError:
                _ARTIFACT_CLEANUP_FAILURES += 1
            continue
    removal_limit = max(1, min(limit, 512))
    for expires_at, path in sorted(candidates, key=lambda item: (item[0], item[1].name)):
        if expires_at > current or removed >= removal_limit:
            break
        try:
            shutil.rmtree(path)
            removed += 1
        except OSError:
            _ARTIFACT_CLEANUP_FAILURES += 1
            continue
    return removed


def create_validation_artifact(
    content: bytes,
    inspection: dict[str, Any],
    *,
    user_id: str,
    organization_id: str,
    package_id: str,
    original_filename: str,
    now: int | None = None,
) -> tuple[str, dict[str, Any]]:
    cleanup_expired_validation_artifacts(now=now)
    created_at = int(time.time() if now is None else now)
    _enforce_artifact_quota(
        user_id=str(user_id),
        organization_id=str(organization_id),
        size_bytes=len(content),
        now=created_at,
    )
    validation_id = uuid.uuid4().hex
    path = _validation_path(validation_id)
    temporary_path = _validation_root() / f".tmp-validation-{validation_id}"
    temporary_path.mkdir(mode=0o700)
    workbook_path = temporary_path / "workbook.xlsx"
    metadata_path = temporary_path / "metadata.json"
    digest = hashlib.sha256(content).hexdigest()
    metadata = {
        "version": 1,
        "validationId": validation_id,
        "userId": str(user_id),
        "organizationId": str(organization_id),
        "packageId": str(package_id),
        "originalFilename": os.path.basename(str(original_filename or "workbook.xlsx")),
        "sha256": digest,
        "sizeBytes": len(content),
        "createdAt": created_at,
        "expiresAt": created_at + VALIDATION_TTL_SECONDS,
        "inspection": inspection,
    }
    try:
        with workbook_path.open("xb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        workbook_path.chmod(0o600)
        with metadata_path.open("x", encoding="utf-8") as handle:
            json.dump(metadata, handle, ensure_ascii=False, separators=(",", ":"))
            handle.flush()
            os.fsync(handle.fileno())
        metadata_path.chmod(0o600)
        os.replace(temporary_path, path)
    except Exception:
        shutil.rmtree(temporary_path, ignore_errors=True)
        raise
    return _token(validation_id), metadata


def load_validation_artifact(
    token: str,
    *,
    user_id: str,
    organization_id: str,
    package_id: str,
    now: int | None = None,
    claim: bool = False,
) -> tuple[dict[str, Any], bytes]:
    try:
        validation_id, supplied_signature = str(token or "").split(".", 1)
    except ValueError as exc:
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_INVALID", "Validation token không hợp lệ.", status_code=409
        ) from exc
    if not _VALIDATION_ID.fullmatch(validation_id):
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_INVALID", "Validation token không hợp lệ.", status_code=409
        )
    expected = _token(validation_id).split(".", 1)[1]
    if not hmac.compare_digest(expected, supplied_signature):
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_INVALID", "Validation token không hợp lệ.", status_code=409
        )
    path = _validation_path(validation_id)
    try:
        metadata = json.loads((path / "metadata.json").read_text(encoding="utf-8"))
        content = (path / "workbook.xlsx").read_bytes()
    except (OSError, json.JSONDecodeError) as exc:
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_NOT_FOUND", "Validation token không còn tồn tại.", status_code=410
        ) from exc
    current = int(time.time() if now is None else now)
    if int(metadata.get("expiresAt") or 0) <= current:
        shutil.rmtree(path, ignore_errors=True)
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_EXPIRED", "Validation token đã hết hạn.", status_code=410
        )
    binding = (
        str(metadata.get("userId")),
        str(metadata.get("organizationId")),
        str(metadata.get("packageId")),
    )
    if binding != (str(user_id), str(organization_id), str(package_id)):
        raise AwardResultExcelError(
            "VALIDATION_TOKEN_SCOPE_MISMATCH",
            "Validation token không thuộc user, tổ chức hoặc gói thầu hiện tại.",
            status_code=403,
        )
    if not hmac.compare_digest(
        hashlib.sha256(content).hexdigest(), str(metadata.get("sha256") or "")
    ):
        raise AwardResultExcelError(
            "WORKBOOK_CHANGED_AFTER_VALIDATION",
            "File đã thay đổi sau bước kiểm tra.",
            status_code=409,
        )
    if claim:
        try:
            with (path / "export.lock").open("x", encoding="ascii") as handle:
                handle.write(str(current))
        except FileExistsError as exc:
            raise AwardResultExcelError(
                "VALIDATION_TOKEN_IN_USE",
                "Validation token đang được dùng để xuất file.",
                status_code=409,
            ) from exc
    return metadata, content


def consume_validation_artifact(token: str) -> None:
    try:
        validation_id, supplied_signature = str(token or "").split(".", 1)
    except ValueError:
        return
    if _VALIDATION_ID.fullmatch(validation_id):
        expected = _token(validation_id).split(".", 1)[1]
        if not hmac.compare_digest(expected, supplied_signature):
            return
        shutil.rmtree(_validation_path(validation_id), ignore_errors=True)


def release_validation_artifact(token: str) -> None:
    try:
        validation_id, supplied_signature = str(token or "").split(".", 1)
    except ValueError:
        return
    if _VALIDATION_ID.fullmatch(validation_id):
        expected = _token(validation_id).split(".", 1)[1]
        if not hmac.compare_digest(expected, supplied_signature):
            return
        try:
            (_validation_path(validation_id) / "export.lock").unlink()
        except FileNotFoundError:
            pass


async def run_validation_artifact_janitor() -> None:
    interval = _configured_limit("AWARD_RESULT_ARTIFACT_CLEANUP_INTERVAL_SECONDS", 60)
    while True:
        await asyncio.to_thread(cleanup_expired_validation_artifacts)
        await asyncio.sleep(interval)


def output_filename(original_filename: str) -> str:
    name = os.path.basename(str(original_filename or "workbook.xlsx"))
    stem = Path(name).stem
    stem = unicodedata.normalize("NFKC", stem)
    stem = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", stem)
    stem = re.sub(r"\s+", " ", stem).strip(" ._")[:120] or "workbook"
    return f"{stem}_da_dien_ket_qua.xlsx"

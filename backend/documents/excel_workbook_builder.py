"""Pure in-memory Excel builders that do not import application/database state."""

from __future__ import annotations

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from backend.documents.spreadsheet_security import safe_spreadsheet_text


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid",
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
BORDER_SIDE = Side(border_style="thin", color="D9D9D9")
THIN_BORDER = Border(
    left=BORDER_SIDE,
    right=BORDER_SIDE,
    top=BORDER_SIDE,
    bottom=BORDER_SIDE,
)

TIMELINE_STATUS_LABELS = {
    "PENDING": "Chưa thực hiện",
    "IN_PROGRESS": "Đang thực hiện",
    "DONE": "Đã hoàn thành",
    "NOT_APPLICABLE": "Không áp dụng",
}

_safe_spreadsheet_text = safe_spreadsheet_text


def _add_dropdown_sheet(workbook, options_map):
    ranges = {}
    if not options_map:
        return ranges
    sheet = workbook.create_sheet(title="Dropdowns")
    sheet.sheet_state = "hidden"
    for option_index, (header, values) in enumerate(options_map.items(), start=1):
        column_letter = get_column_letter(option_index)
        for value_index, value in enumerate(values, start=1):
            sheet.cell(row=value_index, column=option_index, value=_safe_spreadsheet_text(value))
        ranges[header] = (
            f"Dropdowns!${column_letter}$1:${column_letter}${len(values)}"
        )
    return ranges


def _format_cell(cell, field_format):
    if field_format == "currency":
        cell.number_format = "#,##0"
    elif field_format == "date":
        cell.number_format = "dd/mm/yyyy"
    elif field_format == "datetime":
        cell.number_format = 'hh:mm "ngày" dd/mm/yyyy'
    elif field_format == "decimal":
        cell.number_format = "0.##########"


def _finalize_widths(sheet):
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = max(
            max_len + 3,
            15,
        )


def _build_configured_workbook(
    title,
    headers,
    rows=None,
    options_map=None,
    formats_map=None,
    numeric_constraints=None,
    empty_rows=0,
    validation_padding=10,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    rows = list(rows or [])
    formats_map = formats_map or {}
    numeric_constraints = numeric_constraints or {}
    option_ranges = _add_dropdown_sheet(workbook, options_map or {})

    sheet.append(headers)
    sheet.row_dimensions[1].height = 28
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    total_rows = max(len(rows), empty_rows)
    for row_index in range(2, total_rows + 2):
        sheet.row_dimensions[row_index].height = 22
        values = rows[row_index - 2] if row_index - 2 < len(rows) else []
        for column_index, header in enumerate(headers, start=1):
            value = (
                values[column_index - 1]
                if column_index - 1 < len(values)
                else None
            )
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=_safe_spreadsheet_text(value),
            )
            cell.border = THIN_BORDER
            _format_cell(cell, formats_map.get(header))

    validation_end = max(2 + len(rows) + validation_padding, 1 + empty_rows)
    for column_index, header in enumerate(headers, start=1):
        formula = option_ranges.get(header)
        if not formula:
            continue
        column_letter = get_column_letter(column_index)
        validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
        )
        validation.error = "Giá trị nhập không hợp lệ, vui lòng chọn từ danh sách"
        validation.errorTitle = "Lỗi nhập dữ liệu"
        validation.prompt = "Vui lòng chọn giá trị từ danh sách"
        validation.promptTitle = header
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{validation_end}")

    for column_index, header in enumerate(headers, start=1):
        constraint = numeric_constraints.get(header)
        if not constraint:
            continue
        minimum = constraint.get("minimum")
        if not isinstance(minimum, (int, float)) or isinstance(minimum, bool):
            raise ValueError("Excel numeric constraint minimum must be a number.")
        column_letter = get_column_letter(column_index)
        validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=str(minimum),
            allow_blank=True,
        )
        validation.error = "Giá trị nhập không hợp lệ."
        validation.errorTitle = "Lỗi nhập dữ liệu"
        validation.prompt = f"Vui lòng nhập số lớn hơn hoặc bằng {minimum}."
        validation.promptTitle = header
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{validation_end}")

    _finalize_widths(sheet)
    return workbook


def create_excel_from_spec(spec):
    """Build a workbook from a JSON-compatible, data-only export contract."""
    if not isinstance(spec, dict):
        raise ValueError("Excel export spec must be an object.")
    allowed_keys = {
        "title",
        "headers",
        "rows",
        "options_map",
        "formats_map",
        "numeric_constraints",
        "empty_rows",
        "validation_padding",
    }
    if set(spec) - allowed_keys:
        raise ValueError("Excel export spec contains unsupported fields.")
    title = spec.get("title")
    headers = spec.get("headers")
    if not isinstance(title, str) or not title.strip():
        raise ValueError("Excel export spec requires a title.")
    if not isinstance(headers, list) or not headers or not all(
        isinstance(header, str) and header for header in headers
    ):
        raise ValueError("Excel export spec requires headers.")
    return _build_configured_workbook(**spec)


WINNING_GOODS_HEADERS = (
    "STT", "Danh mục hàng hóa", "Ký mã hiệu", "Nhãn hiệu",
    "Năm sản xuất", "Xuất xứ", "Hãng sản xuất",
    "Cấu hình, tính năng kỹ thuật cơ bản", "Đơn vị tính", "Khối lượng",
    "Mã HS", "Đơn giá trúng thầu",
)


def create_winning_goods_excel(export_model):
    """Build an official winning-goods workbook from a server-owned model."""

    if not isinstance(export_model, dict):
        raise ValueError("Dữ liệu xuất hàng hóa trúng thầu không hợp lệ.")
    groups = export_model.get("groups")
    if not isinstance(groups, list) or not groups:
        raise ValueError("Không có hàng hóa trúng thầu chính thức để xuất.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HangHoaTrungThau"
    sheet.sheet_view.showGridLines = False
    last_column = len(WINNING_GOODS_HEADERS)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(1, 1, "DANH SÁCH HÀNG HÓA TRÚNG THẦU").font = Font(bold=True, size=14)
    sheet.cell(1, 1).alignment = CENTER_ALIGN
    row_index = 2
    package_label = " - ".join(
        str(export_model.get(key) or "").strip()
        for key in ("packageCode", "packageName")
        if str(export_model.get(key) or "").strip()
    )
    if package_label:
        sheet.merge_cells(
            start_row=row_index, start_column=1, end_row=row_index, end_column=last_column
        )
        sheet.cell(row_index, 1, safe_spreadsheet_text(f"GÓI THẦU: {package_label}"))
        sheet.cell(row_index, 1).font = Font(bold=True)
        row_index += 1

    for group in groups:
        contractor_name = str((group or {}).get("contractorName") or "").strip()
        lots = (group or {}).get("lots")
        if not contractor_name or not isinstance(lots, list) or not lots:
            raise ValueError("Nhóm nhà thầu trúng thầu không hợp lệ.")
        sheet.merge_cells(
            start_row=row_index, start_column=1, end_row=row_index, end_column=last_column
        )
        sheet.cell(row_index, 1, safe_spreadsheet_text(f"NHÀ THẦU: {contractor_name}"))
        sheet.cell(row_index, 1).font = Font(bold=True)
        row_index += 1
        for lot in lots:
            if export_model.get("isLotted"):
                lot_label = " - ".join(
                    str((lot or {}).get(key) or "").strip()
                    for key in ("lotCode", "lotName")
                    if str((lot or {}).get(key) or "").strip()
                )
                sheet.merge_cells(
                    start_row=row_index, start_column=1,
                    end_row=row_index, end_column=last_column,
                )
                sheet.cell(row_index, 1, safe_spreadsheet_text(f"PHẦN (LÔ): {lot_label}"))
                sheet.cell(row_index, 1).font = Font(bold=True)
                row_index += 1
            rows = (lot or {}).get("rows")
            if not isinstance(rows, list) or not rows:
                raise ValueError("Phần trúng thầu không có hàng hóa chính thức.")
            for column_index, header in enumerate(WINNING_GOODS_HEADERS, start=1):
                cell = sheet.cell(row_index, column_index, header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            row_index += 1
            for row in rows:
                values = (
                    row.get("stt"), row.get("danhMucHangHoa"), row.get("kyMaHieu"),
                    row.get("nhanHieu"), row.get("namSanXuat"), row.get("xuatXu"),
                    row.get("hangSanXuat"), row.get("cauHinhTinhNangKyThuat"),
                    row.get("donViTinh"), row.get("khoiLuong"), row.get("maHs"),
                    row.get("donGiaTrungThau"),
                )
                for column_index, value in enumerate(values, start=1):
                    cell = sheet.cell(
                        row_index, column_index,
                        safe_spreadsheet_text(value),
                    )
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                row_index += 1

    sheet.freeze_panes = "A3"
    widths = (10, 32, 20, 18, 15, 18, 24, 45, 14, 14, 16, 22)
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    return workbook


def _timeline_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    raw = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw[:10], pattern)
        except ValueError:
            continue
    return None


def _timeline_metadata_value(record, code_key, name_key):
    code = str(record.get(code_key) or "").strip()
    name = str(record.get(name_key) or "").strip()
    if code and name:
        return f"{code} - {name}"
    return code or name


def create_timeline_excel(context):
    """Build an editable package Timeline workbook from a data-only context."""
    if not isinstance(context, dict):
        raise ValueError("Timeline Excel context must be an object.")

    package = context.get("goi_thau") or {}
    plan = context.get("ke_hoach") or {}
    organization = context.get("to_chuc") or {}
    sections = context.get("timeline_sections") or []
    if not isinstance(package, dict) or not isinstance(sections, list):
        raise ValueError("Timeline Excel context is invalid.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timeline"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A10"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:9"

    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = "TIMELINE GÓI THẦU"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    sheet.row_dimensions[1].height = 32

    metadata = [
        ("Mã gói thầu", package.get("ma_goi_thau") or ""),
        ("Tên gói thầu", package.get("ten_goi_thau") or ""),
        (
            "Kế hoạch LCNT",
            _timeline_metadata_value(plan, "ma_ke_hoach", "ten_ke_hoach"),
        ),
        ("Đơn vị", organization.get("ten_to_chuc") or ""),
        (
            "Thông tin xuất",
            "Phiên bản {version} - Ngày {generated}".format(
                version=context.get("timeline_template_version") or "-",
                generated=context.get("generated_date") or "-",
            ),
        ),
    ]
    for row_index, (label, value) in enumerate(metadata, start=3):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(
            row=row_index,
            column=2,
            value=_safe_spreadsheet_text(value),
        )
        sheet.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=7,
        )
        sheet.cell(row=row_index, column=2).alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
        sheet.row_dimensions[row_index].height = 22

    headers = [
        "STT",
        "Công việc",
        "Đơn vị ban hành",
        "Số văn bản",
        "Thời gian",
        "Trạng thái",
        "Nguồn",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=9, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    sheet.row_dimensions[9].height = 28

    status_validation = DataValidation(
        type="list",
        formula1='"Chưa thực hiện,Đang thực hiện,Đã hoàn thành,Không áp dụng"',
        allow_blank=False,
    )
    status_validation.error = "Vui lòng chọn một trạng thái trong danh sách."
    status_validation.errorTitle = "Trạng thái không hợp lệ"
    sheet.add_data_validation(status_validation)

    current_row = 10
    for section in sections:
        if not isinstance(section, dict):
            continue
        code = str(section.get("code") or "").strip()
        title = str(section.get("title") or "").strip()
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=7,
        )
        section_cell = sheet.cell(
            row=current_row,
            column=1,
            value=_safe_spreadsheet_text(f"{code}. {title}".strip(". ")),
        )
        section_cell.font = Font(bold=True, color="1F1F1F")
        section_cell.fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid",
        )
        section_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[current_row].height = 24
        current_row += 1

        for item in section.get("items") or []:
            if not isinstance(item, dict):
                continue
            date_value = _timeline_date(
                item.get("ngay_thuc_te") or item.get("ngay_du_kien")
            )
            values = [
                item.get("display_code") or "",
                item.get("cong_viec") or "",
                item.get("don_vi_ban_hanh") or "",
                item.get("so_van_ban") or "",
                date_value,
                TIMELINE_STATUS_LABELS.get(
                    str(item.get("trang_thai") or "").upper(),
                    item.get("trang_thai") or "",
                ),
                "Thủ công"
                if str(item.get("source_mode") or "").upper() == "MANUAL"
                else "Tự động",
            ]
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(
                    row=current_row,
                    column=column_index,
                    value=value if column_index == 5 else _safe_spreadsheet_text(value),
                )
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if column_index in {1, 5, 6, 7} else "left",
                    vertical="center",
                    wrap_text=column_index in {2, 3, 4},
                )
            date_cell = sheet.cell(row=current_row, column=5)
            date_cell.number_format = "dd/mm/yyyy"
            if item.get("is_planned_date"):
                date_cell.font = Font(color="C00000")
            status_validation.add(sheet.cell(row=current_row, column=6))
            sheet.row_dimensions[current_row].height = 30
            current_row += 1

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 48
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 22
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 14
    return workbook


def create_excel_template(import_type):
    """Create a schema-based import template without application state."""
    from backend.documents.excel_handler import (
        _schema_to_formats,
        _schema_to_headers,
        _schema_to_options,
    )

    headers = _schema_to_headers(import_type)
    if not headers:
        raise ValueError(f"Invalid type: {import_type}")
    return _build_configured_workbook(
        "Nhap Lieu",
        headers,
        options_map=_schema_to_options(import_type),
        formats_map=_schema_to_formats(import_type),
        empty_rows=50,
    )


OPENING_TEMPLATE_HEADERS = {
    "TU_VAN": [
        "Loại nhà thầu",
        "Mã nhà thầu",
        "Tên nhà thầu (Nhập chính xác)",
        "Hiệu lực E-HSĐXKT (ngày)",
        "Thời gian thực hiện (ngày)",
    ],
    "1G2T_NO_LOT": [
        "Loại nhà thầu",
        "Mã nhà thầu",
        "Tên nhà thầu (Nhập chính xác)",
        "Đảm bảo dự thầu (VND)",
        "Hiệu lực đảm bảo (ngày)",
        "Hiệu lực E-HSĐXKT (ngày)",
    ],
    "1G2T_WITH_LOT": [
        "Loại nhà thầu",
        "Mã phần lô",
        "Tên phần lô (Tự động điền)",
        "Mã nhà thầu",
        "Tên nhà thầu (Nhập chính xác)",
        "Đảm bảo dự thầu (VND)",
        "Hiệu lực đảm bảo (ngày)",
        "Hiệu lực E-HSĐXKT (ngày)",
    ],
    "1G1T_NO_LOT": [
        "Loại nhà thầu",
        "Mã nhà thầu",
        "Tên nhà thầu (Nhập chính xác)",
        "Giá dự thầu (VND)",
        "Tỷ lệ giảm giá (%)",
        "Giá sau giảm giá (nếu có)",
        "Hiệu lực E-HSDT (ngày)",
        "Giá trị ĐB DT (VND)",
        "Hiệu lực ĐB (ngày)",
        "Thời gian thực hiện (ngày)",
    ],
    "1G1T_WITH_LOT": [
        "Loại nhà thầu",
        "Mã phần lô",
        "Tên phần lô (Tự động điền)",
        "Mã nhà thầu",
        "Tên nhà thầu (Nhập chính xác)",
        "Giá dự thầu (VND)",
        "Tỷ lệ giảm (%)",
        "Giá sau giảm giá (nếu có)",
        "Hiệu lực E-HSDT (ngày)",
        "Giá trị ĐB (VND)",
        "Hiệu lực ĐB",
        "Thời gian thực hiện (ngày)",
    ],
}


def create_mothau_template(case_type, lot_codes):
    """Create an opening workbook without importing database state."""
    headers = OPENING_TEMPLATE_HEADERS.get(case_type)
    if not headers:
        raise ValueError(f"Invalid opening template type: {case_type}")
    options_map = {"Loại nhà thầu": ["Độc lập", "Liên danh"]}
    if "_WITH_LOT" in case_type and lot_codes:
        options_map["Mã phần lô"] = lot_codes
    currency_headers = {
        header: "currency"
        for header in headers
        if "(VND)" in header or header.startswith("Giá sau giảm")
    }
    return _build_configured_workbook(
        "Mo Thau",
        headers,
        options_map=options_map,
        formats_map=currency_headers,
        empty_rows=50,
    )


def create_phanlo_excel(phan_lo_list):
    """Export package-lot rows without importing database state."""
    headers = [
        "Mã phần lô",
        "Tên phần lô",
        "Giá trị phần lô (VND)",
        "Bảo đảm dự thầu (VND)",
        "Thời gian thực hiện (ngày)",
    ]
    rows = [
        [
            item.get("maPhanLo", ""),
            item.get("tenPhanLo", ""),
            item.get("giaTriPhanLo", 0),
            item.get("baoDamDuThau", 0),
            item.get("thoiGianThucHien", 0),
        ]
        for item in phan_lo_list
    ]
    return _build_configured_workbook(
        "Phan Lo",
        headers,
        rows=rows,
        formats_map={
            "Giá trị phần lô (VND)": "currency",
            "Bảo đảm dự thầu (VND)": "currency",
        },
    )


def create_tuychonmuathem_excel(tuy_chon_list):
    """Export optional-purchase rows without importing database state."""
    headers = [
        "Hạng mục",
        "Đơn vị",
        "Số lượng",
        "Tỷ lệ phần trăm (%)",
        "Giá trị ước tính",
    ]
    rows = [
        [
            item.get("hangMuc", ""),
            item.get("donVi", ""),
            item.get("soLuong", 0),
            item.get("tyLe", 0),
            item.get("giaTriUocTinh", 0),
        ]
        for item in tuy_chon_list
    ]
    return _build_configured_workbook(
        "Tuy Chon Mua Them",
        headers,
        rows=rows,
        formats_map={"Giá trị ước tính": "currency"},
    )

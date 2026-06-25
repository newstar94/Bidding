import os
import json
import secrets
from datetime import datetime
from io import BytesIO
from starlette.responses import StreamingResponse, JSONResponse

from helpers import (
    database,
    verify_session,
    clean_id,
    VietnameseFloat,
    SCHEMA_DINH_NGHIA,
    to_camel_case,
    get_active_org,
    load_base64_image,
    OrgPermissionError
)

ENTITY_SCHEMA = {
    'chudautu': [
        {'field': 'maChuDauTu',           'label': 'Mã chủ đầu tư',              'aliases': ['Mã chủ đầu tư', 'Mã CĐT', 'maChuDauTu']},
        {'field': 'tenChuDauTu',           'label': 'Tên chủ đầu tư',             'aliases': ['Tên chủ đầu tư', 'Tên CĐT', 'tenChuDauTu']},
        {'field': 'maSoThue',              'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'chucVuNguoiDungDau',   'label': 'Chức vụ người đứng đầu',      'aliases': ['Chức vụ người đứng đầu', 'Chức vụ', 'chucVuNguoiDungDau']},
        {'field': 'nguoiKyQuyetDinh',     'label': 'Người ký QĐ',          'aliases': ['Người ký QĐ', 'Người ký', 'nguoiKyQuyetDinh']},
        {'field': 'chucVuNguoiKy',        'label': 'Chức vụ người ký',             'aliases': ['Chức vụ người ký', 'chucVuNguoiKy']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'Ông/Bà', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'diaChi',               'label': 'Địa chỉ trụ sở',              'aliases': ['Địa chỉ', 'Địa chỉ trụ sở', 'diaChi']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'STK', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'Ngân hàng', 'noiMoTaiKhoan']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'Địa chỉ email', 'email']},
        {'field': 'maQHNS',               'label': 'Mã QHNS',                    'aliases': ['Mã QHNS', 'maQHNS']},
        {'field': 'coQuanChuQuan',        'label': 'Cơ quan chủ quản',            'aliases': ['Cơ quan chủ quản', 'coQuanChuQuan']},
    ],
    'kehoach': [
        {'field': 'maKeHoach',            'label': 'Mã kế hoạch',                'aliases': ['Mã kế hoạch', 'maKeHoach']},
        {'field': 'maDuan',               'label': 'Mã dự án',                   'aliases': ['Mã dự án', 'maDuan']},
        {'field': 'tenKeHoach',           'label': 'Tên kế hoạch',               'aliases': ['Tên kế hoạch', 'tenKeHoach']},
        {'field': 'loaiHinhMuaSam',       'label': 'Loại hình',                   'aliases': ['Loại hình', 'Loại hình mua sắm', 'loaiHinhMuaSam'],
                                           'options': ['Dự án', 'Dự toán mua sắm']},
        {'field': 'tenDuAnDuToan',        'label': 'Tên dự án',                  'aliases': ['Tên dự án', 'Dự án', 'Dự án / Dự toán', 'tenDuAnDuToan']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'Mã chủ đầu tư', 'chuDauTuId']},
        {'field': 'tongMucDauTu',         'label': 'Tổng mức đầu tư',             'aliases': ['Tổng giá trị', 'Tổng mức đầu tư', 'tongMucDauTu']},
        {'field': 'ngayPheDuyet',         'label': 'Ngày phê duyệt',               'aliases': ['Ngày phê duyệt', 'ngayPheDuyet']},
        {'field': 'quyetDinhPheDuyet',    'label': 'Số QĐ',               'aliases': ['Số QĐ', 'QĐ phê duyệt', 'quyetDinhPheDuyet']},
        {'field': 'thoiGianDangMa',       'label': 'Thời gian đăng',               'aliases': ['Thời gian đăng', 'Thời gian đăng mã', 'thoiGianDangMa']},
        {'field': 'soQdPheDuyetDuAn',     'label': 'Số QĐ phê duyệt dự án',       'aliases': ['Số QĐ phê duyệt dự án', 'soQdPheDuyetDuAn']},
        {'field': 'ngayQdPheDuyetDuAn',   'label': 'Ngày QĐ phê duyệt dự án',     'aliases': ['Ngày QĐ phê duyệt dự án', 'ngayQdPheDuyetDuAn']},
        {'field': 'coQuanPheDuyetDuAn',   'label': 'Cơ quan phê duyệt dự án',     'aliases': ['Cơ quan phê duyệt dự án', 'coQuanPheDuyetDuAn']},
        {'field': 'nguonVon',             'label': 'Nguồn vốn',                   'aliases': ['Nguồn vốn', 'nguonVon']},
        {'field': 'thoiGianDuAn',         'label': 'Thời gian dự án',             'aliases': ['Thời gian dự án', 'thoiGianDuAn']},
        {'field': 'diaDiemQuyMo',         'label': 'Địa điểm quy mô',             'aliases': ['Địa điểm quy mô', 'diaDiemQuyMo']},
        {'field': 'thongTinKhac',         'label': 'Thông tin khác',              'aliases': ['Thông tin khác', 'thongTinKhac']},
    ],
    'goithau': [
        {'field': 'maGoiThau',            'label': 'Mã gói thầu',                 'aliases': ['Mã gói thầu', 'maGoiThau']},
        {'field': 'tenGoiThau',           'label': 'Tên gói thầu',                'aliases': ['Tên gói thầu', 'tenGoiThau']},
        {'field': 'keHoachId',            'label': 'Kế hoạch',                    'aliases': ['Kế hoạch', 'Kế hoạch liên kết', 'keHoachId']},
        {'field': 'giaGoiThau',           'label': 'Giá gói thầu',               'aliases': ['Giá gói thầu', 'Giá gói', 'giaGoiThau']},
        {'field': 'hinhThucLuaChon',      'label': 'Hình thức',                   'aliases': ['Hình thức', 'Hình thức lựa chọn', 'hinhThucLuaChon'],
                                           'options': ['Đấu thầu rộng rãi', 'Đấu thầu hạn chế', 'Chỉ định thầu', 'Chỉ định thầu rút gọn', 'Chào hàng cạnh tranh', 'Lựa chọn nhà thầu trong trường hợp đặc biệt']},
        {'field': 'phuongThucLuaChon',    'label': 'Phương thức',                 'aliases': ['Phương thức', 'Phương thức lựa chọn', 'phuongThucLuaChon'],
                                           'options': ['Một giai đoạn một túi hồ sơ', 'Một giai đoạn hai túi hồ sơ', 'Hai giai đoạn một túi hồ sơ', 'Hai giai đoạn hai túi hồ sơ', 'Không có']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
        {'field': 'trangThai',            'label': 'Trạng thái',                   'aliases': ['Trạng thái', 'trangThai'],
                                           'options': ['Chuẩn bị', 'Đang mời thầu', 'Đã mở thầu', 'Đang chấm thầu', 'Đã có kết quả', 'Hủy thầu']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'nguonVon',             'label': 'Nguồn vốn',                   'aliases': ['Nguồn vốn', 'nguonVon']},
        {'field': 'linhVuc',              'label': 'Lĩnh vực',                    'aliases': ['Lĩnh vực', 'linhVuc'],
                                           'options': ['Tư vấn', 'Phi tư vấn', 'Xây lắp', 'Hỗn hợp', 'Hàng hóa']},
        {'field': 'tuyChonMuaThem',       'label': 'Tùy chọn mua thêm',           'aliases': ['Tùy chọn mua thêm', 'tuyChonMuaThem'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianToChuc',       'label': 'Thời gian tổ chức',            'aliases': ['Thời gian tổ chức', 'thoiGianToChuc']},
        {'field': 'thoiGianBatDauToChuc', 'label': 'Thời gian bắt đầu tổ chức',  'aliases': ['Thời gian bắt đầu tổ chức', 'thoiGianBatDauToChuc']},
        {'field': 'quaMang',              'label': 'Qua mạng / Trực tiếp',        'aliases': ['Qua mạng / Trực tiếp', 'Qua mạng', 'quaMang'],
                                           'options': ['Qua mạng', 'Trực tiếp']},
        {'field': 'trongNuocQuocTe',      'label': 'Trong nước / Quốc tế',       'aliases': ['Trong nước / Quốc tế', 'Trong nước', 'trongNuocQuocTe'],
                                           'options': ['Trong nước', 'Quốc tế']},
        {'field': 'phanLo',               'label': 'Phân lô / Không phân lô',     'aliases': ['Phân lô / Không phân lô', 'Phân lô', 'phanLo'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianDangTai',      'label': 'Thời gian đăng tải',          'aliases': ['Thời gian đăng tải', 'thoiGianDangTai']},
        {'field': 'thoiGianDongThau',     'label': 'Thời gian đóng thầu',         'aliases': ['Thời gian đóng thầu', 'thoiGianDongThau']},
        {'field': 'thoiGianMoThau',       'label': 'Thời gian mở thầu',           'aliases': ['Thời gian mở thầu', 'thoiGianMoThau']},
        {'field': 'soQuyetDinh',          'label': 'Số QĐ phê duyệt',     'aliases': ['Số QĐ phê duyệt', 'Số QĐ', 'soQuyetDinh']},
        {'field': 'ngayQuyetDinh',        'label': 'Ngày QĐ phê duyệt',   'aliases': ['Ngày QĐ phê duyệt', 'Ngày QĐ', 'ngayQuyetDinh']},
        {'field': 'soQuyetDinhKetQua',    'label': 'Số QĐ phê duyệt kết quả LCNT', 'aliases': ['Số QĐ phê duyệt kết quả LCNT', 'Số QĐ kết quả', 'soQuyetDinhKetQua']},
        {'field': 'ngayQuyetDinhKetQua',  'label': 'Ngày ký QĐ kết quả LCNT',     'aliases': ['Ngày ký QĐ kết quả LCNT', 'Ngày QĐ kết quả', 'ngayQuyetDinhKetQua']},
        {'field': 'nhaThauTrungThauId',   'label': 'Nhà thầu trúng thầu',          'aliases': ['Nhà thầu trúng thầu', 'Nhà thầu', 'nhaThauTrungThauId']},
        {'field': 'giaTrungThau',         'label': 'Giá trúng thầu',               'aliases': ['Giá trúng thầu', 'giaTrungThau']},
        {'field': 'thoiGianGoiThau',      'label': 'Thời gian gói thầu',          'aliases': ['Thời gian gói thầu', 'thoiGianGoiThau']},
        {'field': 'thoiGianHopDong',      'label': 'Thời gian hợp đồng',          'aliases': ['Thời gian hợp đồng', 'thoiGianHopDong']},
        {'field': 'giaHanList',           'label': 'Gia hạn thời điểm đóng thầu',  'aliases': ['Gia hạn thời điểm đóng thầu', 'giaHanList']},
        {'field': 'yeuCauLamRoList',      'label': 'Yêu cầu làm rõ HSMT',          'aliases': ['Yêu cầu làm rõ HSMT', 'yeuCauLamRoList']},
        {'field': 'traLoiLamRoList',      'label': 'Trả lời làm rõ HSMT',          'aliases': ['Trả lời làm rõ HSMT', 'traLoiLamRoList']},
        {'field': 'giaToDamBaoDuThau',    'label': 'Giá trị bảo đảm dự thầu',      'aliases': ['Giá trị bảo đảm dự thầu', 'giaToDamBaoDuThau']},
        {'field': 'hieuLucHsdtGoiThau',   'label': 'Hiệu lực HSDT',                'aliases': ['Hiệu lực HSDT', 'hieuLucHsdtGoiThau']},
        {'field': 'hieuLucDamBaoDuThau',  'label': 'Hiệu lực bảo đảm dự thầu',     'aliases': ['Hiệu lực bảo đảm dự thầu', 'hieuLucDamBaoDuThau']},
        {'field': 'awardedPhanLoList',    'label': 'Danh sách phân lô trúng thầu', 'aliases': ['Danh sách phân lô trúng thầu', 'awardedPhanLoList']},
    ],
    'nhathau': [
        {'field': 'loaiNhaThau',          'label': 'Loại nhà thầu',               'aliases': ['Loại nhà thầu', 'loaiNhaThau'],
                                           'options': ['Độc lập', 'Liên danh']},
        {'field': 'maNhaThau',            'label': 'Mã nhà thầu',                 'aliases': ['Mã nhà thầu', 'Mã định danh', 'Mã nhà thầu', 'maNhaThau']},
        {'field': 'tenNhaThau',           'label': 'Tên nhà thầu',                'aliases': ['Tên nhà thầu', 'tenNhaThau']},
        {'field': 'maSoThue',             'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'nguoiDaiDien',         'label': 'Người đại diện',              'aliases': ['Người đại diện', 'nguoiDaiDien']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'email']},
        {'field': 'diaChi',               'label': 'Địa chỉ',                    'aliases': ['Địa chỉ', 'diaChi']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'noiMoTaiKhoan']},
        {'field': 'thanhVienLienDanh',    'label': 'Thành viên liên danh',        'aliases': ['Thành viên liên danh', 'thanhVienLienDanh']},
        {'field': 'maNganHang',           'label': 'Mã ngân hàng',                'aliases': ['Mã ngân hàng', 'maNganHang']},
    ],
    'chuyengia': [
        {'field': 'hoTen',                'label': 'Họ tên',                     'aliases': ['Họ tên', 'Họ và tên', 'hoTen']},
        {'field': 'soCCCD',               'label': 'Số CCCD',                     'aliases': ['Số CCCD', 'CCCD', 'soCCCD']},
        {'field': 'ngayCapCCCD',          'label': 'Ngày cấp CCCD',               'aliases': ['Ngày cấp CCCD', 'ngayCapCCCD']},
        {'field': 'noiCapCCCD',           'label': 'Nơi cấp CCCD',                'aliases': ['Nơi cấp CCCD', 'noiCapCCCD']},
        {'field': 'soChungChi',           'label': 'Số chứng chỉ',               'aliases': ['Số chứng chỉ', 'Số chứng chỉ đấu thầu', 'soChungChi']},
        {'field': 'ngayCapChungChi',      'label': 'Ngày cấp chứng chỉ',         'aliases': ['Ngày cấp', 'Ngày cấp chứng chỉ', 'ngayCapChungChi']},
        {'field': 'donViCapChungChi',     'label': 'Đơn vị cấp chứng chỉ',       'aliases': ['Đơn vị cấp', 'Đơn vị cấp chứng chỉ', 'donViCapChungChi']},
        {'field': 'anhChungChi',          'label': 'Ảnh chứng chỉ (Base64)',      'aliases': ['Ảnh chứng chỉ', 'anhChungChi']},
    ],
    'hopdong': [
        {'field': 'soHopDong',            'label': 'Số hợp đồng',                'aliases': ['Số hợp đồng', 'soHopDong']},
        {'field': 'tenHopDong',           'label': 'Tên hợp đồng',               'aliases': ['Tên hợp đồng', 'tenHopDong']},
        {'field': 'ngayKy',               'label': 'Ngày ký',                    'aliases': ['Ngày ký', 'Ngày ký hợp đồng', 'ngayKy']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'chuDauTuId']},
        {'field': 'nhaThauId',            'label': 'Nhà thầu',                    'aliases': ['Nhà thầu', 'nhaThauId']},
        {'field': 'giaTri',               'label': 'Giá trị hợp đồng',          'aliases': ['Giá trị', 'Giá trị hợp đồng', 'giaTri']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'phanLoai',             'label': 'Phân loại',                  'aliases': ['Phân loại', 'phanLoai'],
                                           'options': ['Tư vấn', 'Thẩm định']},
        {'field': 'coQdChiDinh',          'label': 'Có QĐ chỉ định thầu không',   'aliases': ['Có QĐ chỉ định thầu không', 'coQdChiDinh'],
                                           'options': ['Không', 'Có']},
        {'field': 'soQdChiDinh',          'label': 'Số QĐ chỉ định',             'aliases': ['Số QĐ chỉ định', 'soQdChiDinh']},
        {'field': 'ngayQdChiDinh',        'label': 'Ngày QĐ chỉ định',           'aliases': ['Ngày QĐ chỉ định', 'ngayQdChiDinh']},
        {'field': 'soNgayThucHien',       'label': 'Thời gian thực hiện hợp đồng', 'aliases': ['Thời gian thực hiện hợp đồng', 'Thời gian thực hiện', 'Số ngày thực hiện', 'Số ngày', 'soNgayThucHien']},
        {'field': 'goiThauIds',           'label': 'Gói thầu liên kết',            'aliases': ['Gói thầu liên kết', 'Gói thầu', 'goiThauIds']},
    ],
    'phanlo': [
        {'field': 'maPhanLo',             'label': 'Mã phần lô',                 'aliases': ['Mã phần lô', 'maPhanLo']},
        {'field': 'tenPhanLo',            'label': 'Tên phần lô',                'aliases': ['Tên phần lô', 'Tên phân lô', 'tenPhanLo', 'Tên']},
        {'field': 'giaTriPhanLo',         'label': 'Giá trị phần lô',             'aliases': ['Giá trị phần lô (VNĐ)', 'Giá trị phân lô (VNĐ)', 'Giá trị phần lô (VND)', 'Giá trị phân lô (VND)', 'Giá trị phần lô', 'Giá trị phân lô', 'Giá trị', 'giaTriPhanLo']},
        {'field': 'baoDamDuThau',         'label': 'Bảo đảm dự thầu',             'aliases': ['Bảo đảm dự thầu (VNĐ)', 'Bảo đảm dự thầu (VND)', 'Bảo đảm dự thầu', 'baoDamDuThau']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
    ],
    'tuychonmuathem': [
        {'field': 'hangMuc',              'label': 'Hạng mục',                   'aliases': ['Hạng mục', 'Tên hạng mục', 'hangMuc']},
        {'field': 'donVi',                'label': 'Đơn vị',                     'aliases': ['Đơn vị', 'Đơn vị tính', 'ĐVT', 'donVi']},
        {'field': 'soLuong',              'label': 'Khối lượng / Số lượng',       'aliases': ['Khối lượng/ Số lượng', 'Khối lượng', 'Số lượng', 'soLuong', 'khoiLuong']},
        {'field': 'tyLe',                 'label': 'Tỷ lệ phần trăm (%)',         'aliases': ['Tỷ lệ phần trăm (%)', 'Tỷ lệ phần trăm', 'Tỷ lệ (%)', 'Tỷ lệ', 'tyLe', 'phanTram']},
        {'field': 'giaTriUocTinh',        'label': 'Giá trị ước tính',           'aliases': ['Giá trị ước tính', 'Giá trị', 'giaTriUocTinh']},
    ],
    'ketquaqd': [
        {'field': 'maNhaThau',            'label': 'Mã nhà thầu',                 'aliases': ['Mã nhà thầu', 'Mã định danh', 'Mã số thuế', 'Mã', 'maNhaThau']},
        {'field': 'tenNhaThau',           'label': 'Tên nhà thầu',                'aliases': ['Tên nhà thầu', 'Nhà thầu', 'tenNhaThau']},
        {'field': 'trangThai',            'label': 'Trúng thầu/Trượt thầu',       'aliases': ['Trúng thầu/Trượt thầu', 'Trạng thái', 'Kết quả', 'trangThai'],
                                           'options': ['Trúng thầu', 'Trượt thầu']},
        {'field': 'lyDoTruot',            'label': 'Lý do trượt',                 'aliases': ['Lý do trượt', 'Lý do trượt thầu', 'lyDoTruot']},
        {'field': 'giaTrungThau',         'label': 'Giá trúng thầu',               'aliases': ['Giá trúng thầu', 'Giá trúng', 'Giá trúng thầu (VND)', 'giaTrungThau']},
        {'field': 'thoiGianGoiThau',      'label': 'Thời gian thực hiện gói thầu', 'aliases': ['Thời gian thực hiện gói thầu', 'Thời gian gói', 'thoiGianGoiThau']},
        {'field': 'thoiGianHopDong',      'label': 'Thời gian thực hiện hợp đồng', 'aliases': ['Thời gian thực hiện hợp đồng', 'Thời gian hợp đồng', 'thoiGianHopDong']}
    ],
}

def _schema_to_map_cols(entity_type):
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return {entry['field']: entry['aliases'] for entry in schema}

def _schema_to_headers(entity_type):
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return [entry['label'] for entry in schema]

def _schema_to_options(entity_type):
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return {}
    return {entry['label']: entry['options'] for entry in schema if entry.get('options')}

async def import_excel_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)

        import pandas as pd
        from io import BytesIO
        form = await request.form()
        file_obj = form.get('file')
        import_type = form.get('type')
        
        if not file_obj or not import_type:
            return JSONResponse({"error": "Missing file or type parameter"}, status_code=400)

        map_cols = _schema_to_map_cols(import_type)
        if map_cols is None:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)

        file_bytes = await file_obj.read()
        df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
        
        all_possible_headers = []
        for poss in map_cols.values():
            all_possible_headers.extend([x.lower() for x in poss])
            
        first_col = [str(x).strip().lower() for x in df_raw.iloc[:, 0].dropna()]
        vertical_matches = sum(1 for v in first_col if v in all_possible_headers)
        
        first_row = [str(x).strip().lower() for x in df_raw.iloc[0, :].dropna()]
        horizontal_matches = sum(1 for h in first_row if h in all_possible_headers)
        
        is_vertical = vertical_matches > horizontal_matches and vertical_matches >= 2
        
        if is_vertical:
            headers = [str(x).strip() for x in df_raw.iloc[:, 0]]
            records = []
            for col_idx in range(1, df_raw.shape[1]):
                col_vals = df_raw.iloc[:, col_idx]
                if all(str(x).strip() == "" or pd.isna(x) for x in col_vals):
                    continue
                
                row_data = {}
                for r_idx, h in enumerate(headers):
                    val = col_vals.iloc[r_idx] if r_idx < len(col_vals) else ""
                    if pd.isna(val):
                        val = ""
                    row_data[h] = val
                records.append(row_data)
            df = pd.DataFrame(records)
        else:
            df = pd.read_excel(BytesIO(file_bytes))
            
        df.columns = [str(c).strip() for c in df.columns]
        
        rows = []
        
        def find_col(possible_names):
            for name in possible_names:
                for col in df.columns:
                    if col.lower() == name.lower() or col.lower().replace(" ", "") == name.lower().replace(" ", ""):
                        return col
            return None

        def clean_money(val):
            if pd.isna(val):
                return 0
            val_str = str(val).replace("VND", "").replace("đ", "").replace("₫", "").replace(".", "").replace(",", "").strip()
            try:
                return float(val_str)
            except ValueError:
                return 0

        def clean_int(val):
            if pd.isna(val):
                return 0
            try:
                return int(float(str(val).strip()))
            except ValueError:
                return 0

        for idx, row in df.iterrows():
            item = {}
            validation_comments = []
            
            for key, poss in map_cols.items():
                found = find_col(poss)
                val = row[found] if (found is not None) else None
                if pd.isna(val):
                    val = ""
                    
                if key in ['tongMucDauTu', 'giaGoiThau', 'giaTri', 'giaTriPhanLo', 'giaTriUocTinh', 'giaTrungThau', 'baoDamDuThau', 'damBaoDuThau', 'giaDuThau', 'giaSauGiamGia', 'giaTriDamBao']:
                    val = clean_money(val)
                elif key in ['soLuong', 'tyLe']:
                    try:
                        val = float(str(val).strip()) if val != "" else 0.0
                    except ValueError:
                        val = 0.0
                else:
                    if isinstance(val, float) and val.is_integer():
                        val = int(val)
                    val = str(val).strip()
                    
                item[key] = val
            
            if import_type == 'chudautu':
                if not item['tenChuDauTu']:
                    validation_comments.append("Tên chủ đầu tư không được để trống")
                if not item['maChuDauTu']:
                    validation_comments.append("Mã chủ đầu tư không được để trống")
            elif import_type == 'kehoach':
                if not item['tenKeHoach']:
                    validation_comments.append("Tên kế hoạch không được để trống")
                if not item['maKeHoach']:
                    validation_comments.append("Mã kế hoạch không được để trống")
            elif import_type == 'goithau':
                if not item['tenGoiThau']:
                    validation_comments.append("Tên gói thầu không được để trống")
                if not item['maGoiThau']:
                    validation_comments.append("Mã gói thầu không được để trống")
            elif import_type == 'nhathau':
                if not item['tenNhaThau']:
                    validation_comments.append("Tên nhà thầu không được để trống")
                if not item['maNhaThau']:
                    validation_comments.append("Mã nhà thầu không được để trống")
            elif import_type == 'chuyengia':
                if not item['hoTen']:
                    validation_comments.append("Họ và tên không được để trống")
                if not item['soChungChi']:
                    validation_comments.append("Số chứng chỉ không được để trống")
            elif import_type == 'hopdong':
                if not item['tenHopDong']:
                    validation_comments.append("Tên hợp đồng không được để trống")
                if not item['soHopDong']:
                    validation_comments.append("Số hợp đồng không được để trống")
            elif import_type == 'phanlo':
                if not item['tenPhanLo']:
                    validation_comments.append("Tên phần lô không được để trống")
            elif import_type == 'tuychonmuathem':
                if not item['hangMuc']:
                    validation_comments.append("Hạng mục không được để trống")
            
            item['_valid'] = len(validation_comments) == 0
            item['_comment'] = "; ".join(validation_comments) if validation_comments else "Hợp lệ"
            rows.append(item)
            
        return JSONResponse({"success": True, "rows": rows})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_excel_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)

        import_type = request.path_params.get('import_type')
        from io import BytesIO
        cols = _schema_to_headers(import_type)
        if not cols:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)

        options_map = _schema_to_options(import_type)

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Luôn sử dụng định dạng nằm ngang (hàng đại diện cho bản ghi, cột đại diện cho trường thông tin)
        ws.append(cols)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        money_fields = ['tongMucDauTu', 'giaGoiThau', 'giaTri', 'giaTriPhanLo', 'giaTriUocTinh', 'giaTrungThau', 'damBaoDuThau', 'giaDuThau', 'giaSauGiamGia', 'giaTriDamBao', 'baoDamDuThau']
        
        for col_idx in range(1, len(cols) + 1):
            col_name = cols[col_idx - 1]
            field_name = next((entry['field'] for entry in ENTITY_SCHEMA.get(import_type, []) if entry['label'] == col_name), None)
            is_money = field_name in money_fields
            
            for row_idx in range(2, 101):
                ws.row_dimensions[row_idx].height = 24
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if is_money:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            options = options_map.get(col_name)
            if options:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn một giá trị trong danh sách'
                dv.promptTitle = col_name
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}100")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_nhap_lieu_{import_type}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_mothau_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)

        case_type = request.query_params.get('case_type', '1G1T_NO_LOT')
        package_name = request.query_params.get('package_name', 'GoiThau')
        lot_codes_str = request.query_params.get('lot_codes', '')
        
        lot_codes = [c.strip() for c in lot_codes_str.split(',') if c.strip()]
        
        headers = []
        options_map = {
            'Loại nhà thầu': ['Độc lập', 'Liên danh']
        }
        
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for col_idx in range(1, len(headers) + 1):
            col_name = headers[col_idx - 1]
            is_money = any(x in col_name.lower() for x in ['vnd', 'vnđ', 'giá', 'đảm bảo', 'đb'])
            
            for row_idx in range(2, 101):
                ws.row_dimensions[row_idx].height = 24
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if is_money:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            options = options_map.get(col_name)
            if options:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn một giá trị trong danh sách'
                dv.promptTitle = col_name
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}100")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_Mo_Thau_{case_type}_{package_name}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_opening_fin_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)

        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)

        pkg_id_clean = clean_id(package_id)
        if not pkg_id_clean:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT ma_dinh_danh, ten_nha_thau, gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia,
                   hieu_luc_hsdt, thoi_gian_thuc_hien,
                   danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat, danh_gia_ket_luan
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ? AND owner_id = ?
        """, (pkg_id_clean, org_name))
        bids = cursor.fetchall()
        conn.close()
        
        qualified_bids = []
        for b in bids:
            danh_gia_hop_le = b[7]
            danh_gia_nang_luc = b[8]
            danh_gia_ky_thuat = b[9]
            danh_gia_ket_luan = b[10]
            
            is_qualified = False
            if danh_gia_ket_luan:
                is_qualified = (danh_gia_ket_luan == 'Đạt')
            else:
                is_qualified = (danh_gia_hop_le == 'Đạt' and danh_gia_nang_luc == 'Đạt' and danh_gia_ky_thuat != 'Không đạt' and danh_gia_ky_thuat != '')
                
            if is_qualified:
                qualified_bids.append(b)

        headers = ['Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VNĐ)', 'Tỷ lệ %', 'Hiệu lực HSDT', 'Thời gian thực hiện']
        
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "MoHSĐXTC"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in qualified_bids:
            row_data = [
                b[0] or '',
                b[1] or '',
                b[2] or '',
                b[3] or 0.0,
                f"{b[5]} ngày" if b[5] else '',
                b[6] or ''
            ]
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                col_name = headers[col_idx - 1]
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                if col_name == 'Giá dự thầu (VNĐ)':
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name == 'Tỷ lệ %':
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            row_num += 1
            
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mo_Tai_Chinh_{package_name}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_danhgiahsdt_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)

        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)

        pkg_id_clean = clean_id(package_id)
        if not pkg_id_clean:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list FROM goi_thau WHERE id = ? AND owner_id = ?", (pkg_id_clean, org_name))
        gt_row = cursor.fetchone()
        if not gt_row:
            conn.close()
            return JSONResponse({"error": "Package not found or access denied"}, status_code=404)
            
        linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list_str = gt_row
        
        eval_type = request.query_params.get('eval_type', 'technical')

        cursor.execute("""
            SELECT loai_nha_thau, ma_phan_lo, ten_phan_lo, ma_dinh_danh, ten_nha_thau,
                   gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia, hieu_luc_hsdt,
                   gia_tri_dam_bao, hieu_luc_bao_dam_ngay, thoi_gian_thuc_hien,
                   dam_bao_du_thau, hieu_luc_dam_bao, hieu_luc_hsdxt,
                   danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat,
                   lam_ro_hop_le, lam_ro_nang_luc, lam_ro_ky_thuat, lam_ro_tai_chinh,
                   danh_gia_tai_chinh
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ? AND owner_id = ?
        """, (pkg_id_clean, org_name))
        bids = cursor.fetchall()
        conn.close()
        
        is_tu_van = linh_vuc == 'Tư vấn'
        is_1g2t = phuong_thuc_lua_chon == 'Một giai đoạn hai túi hồ sơ'
        is_1g1t = phuong_thuc_lua_chon == 'Một giai đoạn một túi hồ sơ'
        has_phan_lo = phan_lo == 'Có'
        
        case_type = '1G1T_NO_LOT'
        if is_tu_van:
            case_type = 'TU_VAN'
        elif not is_tu_van and is_1g2t:
            if eval_type == 'financial':
                case_type = '1G2T_TC_WITH_LOT' if has_phan_lo else '1G2T_TC_NO_LOT'
            else:
                case_type = '1G2T_WITH_LOT' if has_phan_lo else '1G2T_NO_LOT'
        elif is_1g1t:
            case_type = '1G1T_WITH_LOT' if has_phan_lo else '1G1T_NO_LOT'
            
        headers = []
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G2T_TC_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ %', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSĐXTC (ngày)', 'Thời gian thực hiện (ngày)', 'Làm rõ tài chính', 'Đánh giá tài chính']
        elif case_type == '1G2T_TC_WITH_LOT':
            headers = ['Mã phần lô', 'Tên phần lô', 'Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ %', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSĐXTC (ngày)', 'Thời gian thực hiện (ngày)', 'Làm rõ tài chính', 'Đánh giá tài chính']
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật', 'Làm rõ tài chính']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật', 'Làm rõ tài chính']

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhGiaHSDT"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in bids:
            row_data = []
            if case_type == 'TU_VAN':
                row_data = [b[0], b[3], b[4], b[14], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G2T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G2T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G2T_TC_NO_LOT':
                row_data = [b[0], b[3], b[4], b[5], b[6], b[7], b[8] or '', b[11] or '', b[21] or '', b[22] or '']
            elif case_type == '1G2T_TC_WITH_LOT':
                row_data = [b[1] or '', b[2] or '', b[0], b[3], b[4], b[5], b[6], b[7], b[8] or '', b[11] or '', b[21] or '', b[22] or '']
            elif case_type == '1G1T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '', b[21] or '']
            elif case_type == '1G1T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '', b[21] or '']
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                col_name = headers[col_idx - 1]
                is_money = any(x in col_name.lower() for x in ['vnd', 'vnđ', 'giá', 'đảm bảo', 'đb'])
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                if is_money:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            row_num += 1
            
        for col_idx, h in enumerate(headers, start=1):
            if h in ['Đánh giá hợp lệ', 'Đánh giá năng lực']:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1='"Đạt,Không đạt"', allow_blank=True)
                dv.error = 'Vui lòng chọn Đạt hoặc Không đạt!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn Đạt hoặc Không đạt'
                dv.promptTitle = h
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{row_num + 20}")
                
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        filename = f"Mau_Danh_Gia_HSDT_{package_name}.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except OrgPermissionError as e:
        return JSONResponse({"error": str(e)}, status_code=403)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_ketquaqd_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)

        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)

        pkg_id_clean = clean_id(package_id)
        if not pkg_id_clean:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT nha_thau_trung_thau_id, gia_trung_thau, thoi_gian_goi_thau, thoi_gian_hop_dong FROM goi_thau WHERE id = ? AND owner_id = ?", (pkg_id_clean, org_name))
        gt_row = cursor.fetchone()
        if not gt_row:
            conn.close()
            return JSONResponse({"error": "Package not found or access denied"}, status_code=404)
        
        cursor.execute("""
            SELECT ma_dinh_danh, ten_nha_thau, nha_thau_id, ly_do_truot, loai_nha_thau, ma_phan_lo, ten_phan_lo
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ? AND owner_id = ?
        """, (pkg_id_clean, org_name))
        bids = cursor.fetchall()
        conn.close()
        
        winner_id = gt_row[0] if gt_row else None
        gia_trung = gt_row[1] if gt_row else 0
        tg_goithau = gt_row[2] if gt_row else ""
        tg_hopdong = gt_row[3] if gt_row else ""
        
        headers = ['Mã nhà thầu', 'Tên nhà thầu', 'Trúng thầu/Trượt thầu', 'Lý do trượt', 'Giá trúng thầu (VND)', 'Thời gian thực hiện gói thầu', 'Thời gian thực hiện hợp đồng']
        
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "KetQuaLCNT"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in bids:
            ma_nt, ten_nt, nt_uuid, ly_do, loai_nt, ma_pl, ten_pl = b
            
            is_winner = False
            if winner_id:
                if str(winner_id).startswith("nt-"):
                    is_winner = (str(nt_uuid) == str(winner_id))
                else:
                    try:
                        is_winner = (int(nt_uuid) == int(winner_id))
                    except:
                        is_winner = (str(nt_uuid) == str(winner_id))
            
            if is_winner:
                row_data = [ma_nt, ten_nt, 'Trúng thầu', '', gia_trung, tg_goithau, tg_hopdong]
            else:
                row_data = [ma_nt, ten_nt, 'Trượt thầu', ly_do or 'Nhà thầu xếp hạng 1 trúng thầu', '', '', '']
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                col_name = headers[col_idx - 1]
                is_money = any(x in col_name.lower() for x in ['vnd', 'vnđ', 'giá', 'đảm bảo', 'đb'])
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = thin_border
                if is_money:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            row_num += 1
            
        dv = DataValidation(type="list", formula1='"Trúng thầu,Trượt thầu"', allow_blank=True)
        dv.error = 'Vui lòng chọn Trúng thầu hoặc Trượt thầu!'
        dv.errorTitle = 'Lỗi nhập liệu'
        dv.prompt = 'Chọn trạng thái'
        dv.promptTitle = 'Trúng thầu/Trượt thầu'
        dv.errorStyle = "stop"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"C2:C{row_num + 20}")
        
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        filename = f"Mau_Ket_Qua_LCNT_{package_name}.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except OrgPermissionError as e:
        return JSONResponse({"error": str(e)}, status_code=403)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


async def export_phanlo_excel_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)

        body = await request.json()
        package_name = body.get('package_name', 'GoiThau')
        rows_data = body.get('rows', [])

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "DanhSachPhanLo"

        headers = ["Mã phần lô", "Tên phần lô", "Giá trị phần lô (VNĐ)", "Bảo đảm dự thầu (VNĐ)", "Thời gian thực hiện"]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, item in enumerate(rows_data, start=2):
            ws.row_dimensions[row_idx].height = 24
            
            ma = item.get('maPhanLo') or item.get('ma') or ''
            ten = item.get('tenPhanLo') or item.get('ten') or ''
            
            try:
                gia_tri = float(item.get('giaTriPhanLo') or item.get('gia') or 0)
            except:
                gia_tri = 0
            try:
                bao_dam = float(item.get('baoDamDuThau') or item.get('baodam') or 0)
            except:
                bao_dam = 0
                
            duration = item.get('thoiGianThucHien') or item.get('duration') or ''

            row_values = [ma, ten, gia_tri, bao_dam, duration]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                
                if col_idx in [1, 5]:
                    cell.alignment = center_align
                elif col_idx in [3, 4]:
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align

        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = ""
                if cell.row == 1:
                    val_str = str(cell.value or '')
                else:
                    if cell.column in [3, 4] and isinstance(cell.value, (int, float)):
                        val_str = f"{cell.value:,.0f}"
                    else:
                        val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        safe_pkg_name = "".join(c for c in package_name if c.isalnum() or c in (' ', '_', '-')).strip()
        safe_pkg_name = safe_pkg_name.replace(' ', '_')
        filename = f"Mau_nhap_lieu_phan_lo_{safe_pkg_name}.xlsx"

        import urllib.parse
        quoted_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}"}
        )
    except Exception as e:
        from helpers import log_error
        log_error(e, "export_phanlo_excel_api")
        return JSONResponse({"error": str(e)}, status_code=500)


async def export_tuychonmuathem_excel_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)

        body = await request.json()
        package_name = body.get('package_name', 'GoiThau')
        rows_data = body.get('rows', [])

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "TuyChonMuaThem"

        headers = ["Hạng mục", "Đơn vị", "Khối lượng / Số lượng", "Tỷ lệ phần trăm (%)", "Giá trị ước tính"]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, item in enumerate(rows_data, start=2):
            ws.row_dimensions[row_idx].height = 24
            
            hang_muc = item.get('hangMuc') or ''
            don_vi = item.get('donVi') or ''
            
            try:
                so_luong = float(item.get('soLuong') or 0)
            except:
                so_luong = 0
            try:
                ty_le = float(item.get('tyLe') or 0)
            except:
                ty_le = 0
            try:
                gia_tri = float(item.get('giaTriUocTinh') or 0)
            except:
                gia_tri = 0

            row_values = [hang_muc, don_vi, so_luong, ty_le, gia_tri]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                
                if col_idx == 2:
                    cell.alignment = center_align
                elif col_idx in [3, 4, 5]:
                    if col_idx == 5:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align

        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = ""
                if cell.row == 1:
                    val_str = str(cell.value or '')
                else:
                    if cell.column in [3, 4, 5] and isinstance(cell.value, (int, float)):
                        val_str = f"{cell.value:,.0f}"
                    else:
                        val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        safe_pkg_name = "".join(c for c in package_name if c.isalnum() or c in (' ', '_', '-')).strip()
        safe_pkg_name = safe_pkg_name.replace(' ', '_')
        filename = f"Mau_nhap_lieu_tuy_chon_mua_them_{safe_pkg_name}.xlsx"

        import urllib.parse
        quoted_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}"}
        )
    except Exception as e:
        from helpers import log_error
        log_error(e, "export_tuychonmuathem_excel_api")
        return JSONResponse({"error": str(e)}, status_code=500)



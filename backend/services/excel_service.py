from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from helpers import (
    database,
)
from helpers_py.excel_handler import (
    _schema_to_headers,
    _schema_to_options,
    _schema_to_formats,
)
from helpers_py.sync_mapper import fetch_package_lot_codes


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
BORDER_SIDE = Side(border_style="thin", color="D9D9D9")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def _add_dropdown_sheet(workbook, options_map):
    ranges = {}
    if not options_map:
        return ranges
    sheet = workbook.create_sheet(title="Dropdowns")
    sheet.sheet_state = "hidden"
    for option_index, (header, values) in enumerate(options_map.items(), start=1):
        column_letter = get_column_letter(option_index)
        for value_index, value in enumerate(values, start=1):
            sheet.cell(row=value_index, column=option_index, value=value)
        ranges[header] = f"Dropdowns!${column_letter}$1:${column_letter}${len(values)}"
    return ranges


def _format_cell(cell, field_format):
    if field_format == "currency":
        cell.number_format = "#,##0"
    elif field_format == "date":
        cell.number_format = "dd/mm/yyyy"
    elif field_format == "datetime":
        cell.number_format = 'hh:mm "ngày" dd/mm/yyyy'


def _finalize_widths(sheet):
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = max(max_len + 3, 15)


def _build_configured_workbook(
    title,
    headers,
    rows=None,
    options_map=None,
    formats_map=None,
    empty_rows=0,
    validation_padding=10,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    rows = list(rows or [])
    formats_map = formats_map or {}
    option_ranges = _add_dropdown_sheet(workbook, options_map or {})

    sheet.append(headers)
    sheet.row_dimensions[1].height = 28
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    total_rows = max(len(rows), empty_rows)
    for row_index in range(2, total_rows + 2):
        sheet.row_dimensions[row_index].height = 22
        values = rows[row_index - 2] if row_index - 2 < len(rows) else []
        for column_index, header in enumerate(headers, start=1):
            value = values[column_index - 1] if column_index - 1 < len(values) else None
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = THIN_BORDER
            _format_cell(cell, formats_map.get(header))

    validation_end = max(2 + len(rows) + validation_padding, 1 + empty_rows)
    for column_index, header in enumerate(headers, start=1):
        formula = option_ranges.get(header)
        if not formula:
            continue
        column_letter = get_column_letter(column_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Giá trị nhập không hợp lệ, vui lòng chọn từ danh sách"
        validation.errorTitle = "Lỗi nhập dữ liệu"
        validation.prompt = "Vui lòng chọn giá trị từ danh sách"
        validation.promptTitle = header
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{validation_end}")

    _finalize_widths(sheet)
    return workbook


OPENING_TEMPLATE_HEADERS = {
    "TU_VAN": ["Loại nhà thầu", "Mã nhà thầu", "Tên nhà thầu (Nhập chính xác)", "Hiệu lực E-HSĐXKT (ngày)", "Thời gian thực hiện (ngày)"],
    "1G2T_NO_LOT": ["Loại nhà thầu", "Mã nhà thầu", "Tên nhà thầu (Nhập chính xác)", "Đảm bảo dự thầu (VND)", "Hiệu lực đảm bảo (ngày)", "Hiệu lực E-HSĐXKT (ngày)"],
    "1G2T_WITH_LOT": ["Loại nhà thầu", "Mã phần lô", "Tên phần lô (Tự động điền)", "Mã nhà thầu", "Tên nhà thầu (Nhập chính xác)", "Đảm bảo dự thầu (VND)", "Hiệu lực đảm bảo (ngày)", "Hiệu lực E-HSĐXKT (ngày)"],
    "1G1T_NO_LOT": ["Loại nhà thầu", "Mã nhà thầu", "Tên nhà thầu (Nhập chính xác)", "Giá dự thầu (VND)", "Tỷ lệ giảm giá (%)", "Giá sau giảm giá (nếu có)", "Hiệu lực E-HSDT (ngày)", "Giá trị ĐB DT (VND)", "Hiệu lực ĐB (ngày)", "Thời gian thực hiện (ngày)"],
    "1G1T_WITH_LOT": ["Loại nhà thầu", "Mã phần lô", "Tên phần lô (Tự động điền)", "Mã nhà thầu", "Tên nhà thầu (Nhập chính xác)", "Giá dự thầu (VND)", "Tỷ lệ giảm (%)", "Giá sau giảm giá (nếu có)", "Hiệu lực E-HSDT (ngày)", "Giá trị ĐB (VND)", "Hiệu lực ĐB", "Thời gian thực hiện (ngày)"],
}

def create_excel_template(import_type):
    """Tạo file mẫu Excel nhập liệu cơ bản theo schema thực thể."""
    headers = _schema_to_headers(import_type)
    if not headers:
        raise ValueError(f"Invalid type: {import_type}")
    return _build_configured_workbook(
        "Nhap Lieu",
        headers,
        options_map=_schema_to_options(import_type),
        formats_map=_schema_to_formats(import_type),
        empty_rows=50,
    )

def create_mothau_template(case_type, lot_codes):
    """Tạo template Excel mẫu mở thầu từ cấu hình cột dùng chung."""
    headers = OPENING_TEMPLATE_HEADERS.get(case_type)
    if not headers:
        raise ValueError(f"Invalid opening template type: {case_type}")
    options_map = {"Loại nhà thầu": ["Độc lập", "Liên danh"]}
    if "_WITH_LOT" in case_type and lot_codes:
        options_map["Mã phần lô"] = lot_codes
    currency_headers = {
        header: "currency"
        for header in headers
        if "(VND)" in header or header.startswith("Giá sau giảm")
    }
    return _build_configured_workbook(
        "Mo Thau",
        headers,
        options_map=options_map,
        formats_map=currency_headers,
        empty_rows=50,
    )

def create_opening_fin_template(pkg_id_clean, org_name):
    """Tạo template Excel mở đề xuất tài chính."""
    conn = database.get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT ma_dinh_danh, ten_nha_thau, gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia,
               hieu_luc_hsdt, thoi_gian_thuc_hien,
               danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat, danh_gia_ket_luan
        FROM thong_tin_mo_thau
        WHERE goi_thau_id = ? AND owner_id = ?
    """, (pkg_id_clean, org_name))
    bids = cursor.fetchall()
    conn.close()

    qualified_bids = []
    for b in bids:
        danh_gia_hop_le = b[7]
        danh_gia_nang_luc = b[8]
        danh_gia_ky_thuat = b[9]
        danh_gia_ket_luan = b[10]

        is_qualified = False
        if danh_gia_ket_luan:
            is_qualified = (danh_gia_ket_luan == 'Đạt')
        else:
            is_qualified = (danh_gia_hop_le == 'Đạt' and danh_gia_nang_luc == 'Đạt' and danh_gia_ky_thuat != 'Không đạt' and danh_gia_ky_thuat != '')

        if is_qualified:
            qualified_bids.append(b)

    headers = [
        'Mã định danh',
        'Tên nhà thầu (Nhập chính xác)',
        'Mã phần lô',
        'Tên phần lô (Tự động điền)',
        'Giá dự thầu (VND)',
        'Tỷ lệ giảm giá (%)',
        'Giá sau giảm giá (nếu có)',
        'Hiệu lực E-HSDT (ngày)',
        'Giá trị ĐB DT (VND)',
        'Hiệu lực ĐB (ngày)',
        'Thời gian thực hiện (ngày)'
    ]

    rows = [
        [
            bid[0], bid[1], "", "", bid[2] or "", bid[3] or "",
            bid[4] or "", bid[5] or "", "", "", bid[6] or "",
        ]
        for bid in qualified_bids
    ]
    return _build_configured_workbook(
        "Mo De Xuat Tai Chinh",
        headers,
        rows=rows,
        formats_map={
            "Giá dự thầu (VND)": "currency",
            "Giá sau giảm giá (nếu có)": "currency",
            "Giá trị ĐB DT (VND)": "currency",
        },
    )

def create_danhgiahsdt_template(pkg_id_clean, org_name, eval_type):
    """Tạo file mẫu nhập liệu đánh giá HSDT (kỹ thuật hoặc tài chính)."""
    conn = database.get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT linh_vuc, phuong_thuc_lua_chon, phan_lo FROM goi_thau WHERE id = ? AND owner_id = ?", (pkg_id_clean, org_name))
    gt_row = cursor.fetchone()
    if not gt_row:
        conn.close()
        raise ValueError("Package not found")

    _linh_vuc, _phuong_thuc_lua_chon, phan_lo = gt_row
    lot_codes = fetch_package_lot_codes(cursor, pkg_id_clean, org_name)

    cursor.execute("""
        SELECT loai_nha_thau, ma_phan_lo, ten_phan_lo, ma_dinh_danh, ten_nha_thau,
               gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia, hieu_luc_hsdt,
               gia_tri_dam_bao, hieu_luc_bao_dam_ngay, thoi_gian_thuc_hien,
               dam_bao_du_thau, hieu_luc_dam_bao, hieu_luc_hsdxt,
               danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat,
               lam_ro_hop_le, lam_ro_nang_luc, lam_ro_ky_thuat, lam_ro_tai_chinh,
               danh_gia_tai_chinh,
               nguyen_nhan_khong_dat_hop_le, nguyen_nhan_khong_dat_nang_luc, nguyen_nhan_khong_dat_ky_thuat
        FROM thong_tin_mo_thau
        WHERE goi_thau_id = ? AND owner_id = ?
    """, (pkg_id_clean, org_name))
    bids = cursor.fetchall()
    conn.close()

    has_phan_lo = phan_lo == 'Có'

    headers = []
    options_map = {}

    if eval_type == 'technical':

        if has_phan_lo:
            headers = [
                'Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)',
                'Đánh giá tính hợp lệ', 'Làm rõ tính hợp lệ (nếu có)', 'Nguyên nhân không đạt hợp lệ (nếu có)',
                'Đánh giá năng lực kinh nghiệm', 'Làm rõ năng lực kinh nghiệm (nếu có)', 'Nguyên nhân không đạt năng lực (nếu có)',
                'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật (nếu có)', 'Nguyên nhân không đạt kỹ thuật (nếu có)'
            ]
        else:
            headers = [
                'Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)',
                'Đánh giá tính hợp lệ', 'Làm rõ tính hợp lệ (nếu có)', 'Nguyên nhân không đạt hợp lệ (nếu có)',
                'Đánh giá năng lực kinh nghiệm', 'Làm rõ năng lực kinh nghiệm (nếu có)', 'Nguyên nhân không đạt năng lực (nếu có)',
                'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật (nếu có)', 'Nguyên nhân không đạt kỹ thuật (nếu có)'
            ]

        options_map = {
            'Đánh giá tính hợp lệ': ['Đạt', 'Không đạt'],
            'Đánh giá năng lực kinh nghiệm': ['Đạt', 'Không đạt'],
            'Đánh giá kỹ thuật': ['Đạt', 'Không đạt']
        }

    else:

        if has_phan_lo:
            headers = [
                'Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)',
                'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)',
                'Đánh giá tài chính (Điểm hoặc Xếp hạng)', 'Làm rõ tài chính (nếu có)'
            ]
        else:
            headers = [
                'Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)',
                'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)',
                'Đánh giá tài chính (Điểm hoặc Xếp hạng)', 'Làm rõ tài chính (nếu có)'
            ]

    if has_phan_lo and lot_codes:
        options_map['Mã phần lô'] = lot_codes

    rows = []
    for bid in bids:
        if eval_type == "technical":
            if has_phan_lo:
                row_values = [
                    bid[0], bid[1], bid[2], bid[3], bid[4],
                    bid[15] or "", bid[18] or "", bid[23] or "",
                    bid[16] or "", bid[19] or "", bid[24] or "",
                    bid[17] or "", bid[20] or "", bid[25] or "",
                ]
            else:
                row_values = [
                    bid[0], bid[3], bid[4],
                    bid[15] or "", bid[18] or "", bid[23] or "",
                    bid[16] or "", bid[19] or "", bid[24] or "",
                    bid[17] or "", bid[20] or "", bid[25] or "",
                ]
        elif has_phan_lo:
            row_values = [
                bid[0], bid[1], bid[2], bid[3], bid[4],
                bid[5] or "", bid[6] or "", bid[7] or "",
                bid[22] or "", bid[21] or "",
            ]
        else:
            row_values = [
                bid[0], bid[3], bid[4],
                bid[5] or "", bid[6] or "", bid[7] or "",
                bid[22] or "", bid[21] or "",
            ]
        rows.append(row_values)

    financial_formats = {
        header: "currency"
        for header in headers
        if "(VND)" in header or header.startswith("Giá sau giảm")
    }
    return _build_configured_workbook(
        "Danh gia HSDT",
        headers,
        rows=rows,
        options_map=options_map,
        formats_map=financial_formats,
    )

def create_ketquaqd_template(pkg_id_clean, org_name):
    """Tạo file mẫu nhập kết quả lựa chọn nhà thầu phê duyệt."""
    conn = database.get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT nha_thau_trung_thau_id, gia_trung_thau, thoi_gian_goi_thau, thoi_gian_hop_dong FROM goi_thau WHERE id = ? AND owner_id = ?", (pkg_id_clean, org_name))
    gt_row = cursor.fetchone()
    if not gt_row:
        conn.close()
        raise ValueError("Package not found")

    nha_thau_trung_thau_id, gia_trung_thau, thoi_gian_goi_thau, thoi_gian_hop_dong = gt_row

    cursor.execute("""
        SELECT loai_nha_thau, ma_phan_lo, ten_phan_lo, ma_dinh_danh, ten_nha_thau,
               gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia,
               danh_gia_tai_chinh
        FROM thong_tin_mo_thau
        WHERE goi_thau_id = ? AND owner_id = ?
    """, (pkg_id_clean, org_name))
    bids = cursor.fetchall()
    conn.close()

    headers = [
        'Loại nhà thầu',
        'Mã nhà thầu',
        'Tên nhà thầu (Nhập chính xác)',
        'Kết quả',
        'Lý do trượt thầu (nếu có)',
        'Giá trúng thầu (VND)',
        'Thời gian thực hiện gói thầu (ngày)',
        'Thời gian thực hiện hợp đồng (ngày)'
    ]

    options_map = {
        'Kết quả': ['Trúng thầu', 'Trượt thầu']
    }

    rows = []
    for bid in bids:
        is_winner = bool(nha_thau_trung_thau_id and bid[3] == nha_thau_trung_thau_id)
        rows.append([
            bid[0],
            bid[3],
            bid[4],
            "Trúng thầu" if is_winner else "Trượt thầu",
            "",
            gia_trung_thau if is_winner else "",
            thoi_gian_goi_thau if is_winner else "",
            thoi_gian_hop_dong if is_winner else "",
        ])

    return _build_configured_workbook(
        "Ket Qua LCNT",
        headers,
        rows=rows,
        options_map=options_map,
        formats_map={"Giá trúng thầu (VND)": "currency"},
    )

def create_phanlo_excel(phan_lo_list):
    """Xuất danh sách phân lô bằng builder cấu hình chung."""
    headers = ["Mã phần lô", "Tên phần lô", "Giá trị phần lô (VND)", "Bảo đảm dự thầu (VND)", "Thời gian thực hiện (ngày)"]
    rows = [
        [
            item.get("maPhanLo", ""),
            item.get("tenPhanLo", ""),
            item.get("giaTriPhanLo", 0),
            item.get("baoDamDuThau", 0),
            item.get("thoiGianThucHien", 0),
        ]
        for item in phan_lo_list
    ]
    return _build_configured_workbook(
        "Phan Lo",
        headers,
        rows=rows,
        formats_map={
            "Giá trị phần lô (VND)": "currency",
            "Bảo đảm dự thầu (VND)": "currency",
        },
    )

def create_tuychonmuathem_excel(tuy_chon_list):
    """Xuất tùy chọn mua thêm bằng builder cấu hình chung."""
    headers = ["Hạng mục", "Đơn vị", "Số lượng", "Tỷ lệ phần trăm (%)", "Giá trị ước tính"]
    rows = [
        [
            item.get("hangMuc", ""),
            item.get("donVi", ""),
            item.get("soLuong", 0),
            item.get("tyLe", 0),
            item.get("giaTriUocTinh", 0),
        ]
        for item in tuy_chon_list
    ]
    return _build_configured_workbook(
        "Tuy Chon Mua Them",
        headers,
        rows=rows,
        formats_map={"Giá trị ước tính": "currency"},
    )

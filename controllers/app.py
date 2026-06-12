import sys
import os

# Reconfigure stdout/stderr to use UTF-8 to prevent UnicodeEncodeError on Windows terminals
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
if hasattr(sys.stderr, 'reconfigure'):
    try:
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

import json
import uvicorn
import shutil
from datetime import datetime
import pandas as pd

# ==========================================
# 1. CẤU HÌNH ĐƯỜNG DẪN & TẢI MODULE BIÊN DỊCH
# ==========================================

# Lấy đường dẫn của thư mục controllers/ và thư mục gốc của dự án
current_dir = os.path.dirname(os.path.abspath(__file__)) # controllers/
project_root = os.path.dirname(current_dir) # root
models_dir = os.path.join(project_root, 'models')
controllers_dir = os.path.join(project_root, 'controllers')

# Thêm các thư mục MVC vào sys.path để Python có thể nạp chéo giữa các mô-đun
sys.path.insert(0, project_root)
sys.path.append(models_dir)
sys.path.append(controllers_dir)

# Tự động tải các cấu hình từ file .env nếu có
env_path = os.path.join(project_root, '.env')
if os.path.exists(env_path):
    with open(env_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                k, v = line.split('=', 1)
                k = k.strip()
                v = v.strip().strip("'").strip('"')
                os.environ[k] = v


import importlib.machinery
import importlib.util

def load_and_register(name, filepath):
    """
    Hàm load động các module Python đã được biên dịch sẵn (.pyc) từ các thư mục MVC tương ứng.
    """
    loader = importlib.machinery.SourcelessFileLoader(name, filepath)
    module = importlib.util.module_from_spec(importlib.util.spec_from_loader(name, loader))
    sys.modules[name] = module
    loader.exec_module(module)
    return module

# Tải các module chức năng chính được biên dịch sẵn từ các thư mục MVC mới:
models = load_and_register('models', os.path.join(models_dir, 'models.cpython-314.pyc'))
database = load_and_register('database', os.path.join(models_dir, 'database.cpython-314.pyc'))

# Monkey patch database.get_connection to automatically run optimizations and create indexes
db_indexes_created = False
orig_get_connection = database.get_connection
def optimized_get_connection(*args, **kwargs):
    conn = orig_get_connection(*args, **kwargs)
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA journal_mode = WAL")
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute("PRAGMA cache_size = -65536")   # 64MB cache
        cursor.execute("PRAGMA synchronous = NORMAL")  # An toàn + nhanh hơn FULL
        cursor.execute("PRAGMA temp_store = MEMORY")
        
        global db_indexes_created
        if not db_indexes_created:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_kehoach ON goi_thau(ke_hoach_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_kehoach_chudautu ON ke_hoach_lcnt(chu_dau_tu_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_thongtinmothau_goithau ON thong_tin_mo_thau(goi_thau_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_chudautu_latest ON chu_dau_tu(id_goc, is_latest)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_nhathau_latest ON nha_thau(id_goc, is_latest)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_hopdong_chudautu ON hop_dong(chu_dau_tu_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_hopdong_nhathau ON hop_dong(nha_thau_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_kehoach_latest ON ke_hoach_lcnt(id_goc, is_latest)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_latest ON goi_thau(id_goc, is_latest)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_nhathau ON goi_thau(nha_thau_trung_thau_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_thongtinmothau_nhathau ON thong_tin_mo_thau(nha_thau_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_hopdonggoithau_goithau ON hop_dong_goi_thau(goi_thau_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_deletedrecords_lookup ON deleted_records(owner_id, deleted_at)")
            db_indexes_created = True
    except Exception as e:
        print(f"Error applying SQLite PRAGMAs or indexes: {e}")
    return conn
database.get_connection = optimized_get_connection

exporter = load_and_register('exporter', os.path.join(controllers_dir, 'exporter.cpython-314.pyc'))
import custom_exporter
import uuid
import hashlib
import secrets

def save_base64_image(base64_str: str, subfolder: str, filename_prefix: str) -> str:
    if not base64_str:
        return ""
    if not isinstance(base64_str, str):
        return base64_str
    if not (base64_str.startswith("data:image") or len(base64_str) > 100):
        return base64_str
        
    header = ""
    data_str = base64_str
    if base64_str.startswith("data:image"):
        try:
            parts = base64_str.split(";base64,")
            header = parts[0]
            data_str = parts[1]
        except Exception:
            return base64_str
            
    ext = "png"
    if "jpeg" in header or "jpg" in header:
        ext = "jpg"
    elif "webp" in header:
        ext = "webp"
    elif "gif" in header:
        ext = "gif"
        
    try:
        import base64
        import os
        upload_dir = os.path.join(project_root, "uploads", subfolder)
        os.makedirs(upload_dir, exist_ok=True)
        
        file_data = base64.b64decode(data_str)
        filename = f"{filename_prefix}.{ext}"
        filepath = os.path.join(upload_dir, filename)
        
        # Thử nén và tối ưu hóa bằng Pillow
        try:
            # pyrefly: ignore [missing-import]
            from PIL import Image
            import io
            
            img = Image.open(io.BytesIO(file_data))
            
            # Đặt giới hạn kích thước tối đa tùy theo loại ảnh
            max_size = 1200
            if "sig" in filename_prefix:
                max_size = 600
                
            # Resize nếu vượt quá kích thước tối đa (giữ nguyên tỷ lệ)
            if img.width > max_size or img.height > max_size:
                img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                
            # Lưu ảnh mới, tự động xóa EXIF/Metadata và nén chất lượng
            save_format = "PNG" if ext == "png" else ("JPEG" if ext in ["jpg", "jpeg"] else img.format)
            save_kwargs = {}
            if save_format == "JPEG":
                save_kwargs["quality"] = 100
                save_kwargs["optimize"] = True
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
            elif save_format == "PNG":
                save_kwargs["optimize"] = True
                
            img.save(filepath, format=save_format, **save_kwargs)
        except Exception as pil_err:
            print(f"Pillow optimization failed, falling back to raw save: {pil_err}")
            with open(filepath, "wb") as f:
                f.write(file_data)
                
        return f"uploads/{subfolder}/{filename}"
    except Exception as e:
        print(f"Error saving base64 image: {e}")
        return base64_str

def load_base64_image(db_value: str) -> str:
    if not db_value or not isinstance(db_value, str):
        return ""
    if db_value.startswith("uploads/"):
        try:
            import base64
            import os
            filepath = os.path.join(project_root, db_value)
            if os.path.exists(filepath):
                with open(filepath, "rb") as f:
                    file_data = f.read()
                ext = db_value.split(".")[-1].lower()
                mime = "image/png"
                if ext in ["jpg", "jpeg"]:
                    mime = "image/jpeg"
                elif ext == "webp":
                    mime = "image/webp"
                elif ext == "gif":
                    mime = "image/gif"
                
                base64_data = base64.b64encode(file_data).decode("utf-8")
                return f"data:{mime};base64,{base64_data}"
        except Exception as e:
            print(f"Error loading image path {db_value}: {e}")
    return db_value

def hash_password(password: str, salt: str = None) -> str:
    if salt is None:
        salt = secrets.token_hex(16)
    pwd_hash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    return f"{salt}:{pwd_hash.hex()}"

def verify_password(stored_password: str, provided_password: str) -> bool:
    try:
        if not stored_password:
            return False
        if ":" not in stored_password:
            # Safe plain-text fallback (backward compatibility for old plain DB entries if any)
            return stored_password == provided_password
        salt, stored_hash = stored_password.split(":")
        pwd_hash = hashlib.pbkdf2_hmac('sha256', provided_password.encode('utf-8'), salt.encode('utf-8'), 100000)
        return secrets.compare_digest(stored_hash, pwd_hash.hex())
    except Exception:
        return False

# ==========================================
# ROLE HIERARCHY (thứ bậc phân quyền)
# super_admin kế thừa manager và employee
# manager kế thừa employee
# ==========================================
ROLE_HIERARCHY = {
    'super_admin': ['super_admin', 'manager', 'employee'],
    'manager':     ['manager', 'employee'],
    'employee':    ['employee'],
}

def get_effective_roles(role_str):
    """
    Trả về tập hợp tất cả role hữu hiệu của user, kể cả kế thừa.
    Hỗ trợ nhiều role phân tách bằng dấu phẩy (VD: 'super_admin,manager').
    """
    roles = [r.strip() for r in (role_str or '').split(',') if r.strip()]
    effective = set()
    for r in roles:
        effective.update(ROLE_HIERARCHY.get(r, [r]))
    return effective

class SessionRole(str):
    def __new__(cls, role, user_id):
        instance = super().__new__(cls, role)
        instance.user_id = user_id
        return instance

def verify_session(request, required_role=None):
    token = request.headers.get('X-Session-Token')
    username = request.headers.get('X-Username')
    
    if not token or not username:
        return False, "Thiếu thông tin xác thực phiên làm việc!"
        
    conn = database.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, vai_tro, token_phien, han_su_dung_token FROM tai_khoan WHERE ten_dang_nhap = ? OR (email != '' AND email = ?)", (username, username))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return False, "Tài khoản không tồn tại!"
        
    user = dict(row)
    if user['token_phien'] != token:
        return False, "Phiên làm việc đã hết hạn hoặc không hợp lệ!"

    if user.get('han_su_dung_token'):
        try:
            expiry = datetime.fromisoformat(user['han_su_dung_token'])
            if datetime.utcnow() > expiry:
                return False, "Phiên đăng nhập đã hết hạn! Vui lòng đăng nhập lại."
        except Exception:
            pass
        
    # Kiểm tra quyền theo thứ bậc kế thừa
    if required_role and required_role not in get_effective_roles(user['vai_tro']):
        return False, "Bạn không có quyền thực hiện thao tác này!"
        
    return True, SessionRole(user['vai_tro'], user['id'])

# ==========================================
# CẤU TRÚC DB VÀ CÁC HÀM TRỢ GIÚP ĐỒNG BỘ ĐỘNG
# ==========================================

# 1. Định nghĩa cấu trúc Schema cơ sở dữ liệu (Nguồn sự thật duy nhất)
SCHEMA_DINH_NGHIA = {
    "goi_dich_vu": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "ten_goi": "TEXT",
            "gia_ca": "REAL",
            "han_muc_nhan_su": "INTEGER",
            "mo_ta": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    },
    "tai_khoan": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "ten_dang_nhap": "TEXT UNIQUE",
            "mat_khau": "TEXT",
            "ho_ten": "TEXT",
            "vai_tro": "TEXT",
            "email": "TEXT",
            "token_phien": "TEXT",
            "anh_dai_dien": "TEXT",
            "goi_dich_vu_id": "TEXT DEFAULT 'silver'",
            "ngay_bat_dau_goi": "TEXT",
            "ngay_het_han_goi": "TEXT",
            "han_su_dung_token": "TEXT",
            "thong_tin_thiet_bi_cuoi": "TEXT",
            "da_xac_minh": "INTEGER DEFAULT 0",
            "ma_xac_minh": "TEXT",
            "han_xac_minh": "INTEGER",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": ["FOREIGN KEY (goi_dich_vu_id) REFERENCES goi_dich_vu(id) ON DELETE SET NULL"]
    },
    "chu_dau_tu": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "id_goc": "TEXT",
            "phien_ban": "TEXT NOT NULL DEFAULT '00'",
            "is_latest": "INTEGER NOT NULL DEFAULT 1",
            "ma_chu_dau_tu": "TEXT",
            "ten_chu_dau_tu": "TEXT NOT NULL",
            "ma_so_thue": "TEXT",
            "chuc_vu_nguoi_dung_dau": "TEXT",
            "nguoi_ky_quyet_dinh": "TEXT",
            "chuc_vu_nguoi_ky": "TEXT",
            "danh_xung": "TEXT DEFAULT 'Ông'",
            "dia_chi": "TEXT",
            "so_dien_thoai": "TEXT",
            "so_tai_khoan": "TEXT",
            "noi_mo_tai_khoan": "TEXT",
            "email": "TEXT",
            "ma_qhns": "TEXT",
            "co_quan_chu_quan": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    },
    "ke_hoach_lcnt": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "id_goc": "TEXT",
            "ma_ke_hoach": "TEXT",
            "ma_du_an": "TEXT",
            "phien_ban": "TEXT NOT NULL DEFAULT '00'",
            "is_latest": "INTEGER NOT NULL DEFAULT 1",
            "ten_ke_hoach": "TEXT NOT NULL",
            "ten_du_an_du_toan": "TEXT",
            "loai_hinh_mua_sam": "TEXT",
            "chu_dau_tu_id": "TEXT",
            "tong_muc_dau_tu": "REAL",
            "is_tong_muc_tu_dong": "INTEGER DEFAULT 0",
            "ngay_phe_duyet": "TEXT",
            "quyet_dinh_phe_duyet": "TEXT",
            "thoi_gian_dang_tai": "TEXT",
            "cv_da_thuc_hien": "TEXT",
            "cv_khong_ap_dung": "TEXT",
            "cv_chua_du_dieu_kien": "TEXT",
            "nguon_von": "TEXT",
            "thoi_gian_du_an": "TEXT",
            "dia_diem_quy_mo": "TEXT",
            "thong_tin_khac": "TEXT",
            "so_qd_phe_duyet_du_an": "TEXT",
            "ngay_qd_phe_duyet_du_an": "TEXT",
            "co_quan_phe_duyet_du_an": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": ["FOREIGN KEY (chu_dau_tu_id) REFERENCES chu_dau_tu(id) ON DELETE CASCADE"]
    },
    "nha_thau": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "id_goc": "TEXT",
            "phien_ban": "TEXT NOT NULL DEFAULT '00'",
            "is_latest": "INTEGER NOT NULL DEFAULT 1",
            "ma_nha_thau": "TEXT",
            "ten_nha_thau": "TEXT NOT NULL",
            "loai_nha_thau": "TEXT",
            "thanh_vien_lien_danh": "TEXT",
            "ma_so_thue": "TEXT",
            "nguoi_dai_dien": "TEXT",
            "danh_xung": "TEXT DEFAULT 'Ông'",
            "so_dien_thoai": "TEXT",
            "email": "TEXT",
            "dia_chi": "TEXT",
            "so_tai_khoan": "TEXT",
            "noi_mo_tai_khoan": "TEXT",
            "ma_ngan_hang": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    },
    "goi_thau": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "id_goc": "TEXT",
            "ma_goi_thau": "TEXT",
            "phien_ban": "TEXT NOT NULL DEFAULT '00'",
            "is_latest": "INTEGER NOT NULL DEFAULT 1",
            "ke_hoach_id": "TEXT",
            "ten_goi_thau": "TEXT NOT NULL",
            "gia_goi_thau": "REAL",
            "loai_hop_dong": "TEXT",
            "hinh_thuc_lua_chon": "TEXT",
            "phuong_thuc_lua_chon": "TEXT",
            "thoi_gian_thuc_hien": "TEXT",
            "nguon_von": "TEXT",
            "nha_thau_trung_thau_id": "TEXT",
            "gia_trung_thau": "REAL",
            "linh_vuc": "TEXT",
            "tuy_chon_mua_them": "TEXT DEFAULT 'Không'",
            "thoi_gian_to_chuc": "TEXT",
            "thoi_gian_bat_dau_to_chuc": "TEXT",
            "phan_lo": "TEXT DEFAULT 'Không'",
            "phan_lo_list": "TEXT",
            "tuy_chon_mua_them_list": "TEXT",
            "thoi_gian_dang_tai": "TEXT",
            "thoi_gian_dong_thau": "TEXT",
            "thoi_gian_mo_thau": "TEXT",
            "chuyen_gia_list": "TEXT",
            "tham_dinh_list": "TEXT",
            "so_quyet_dinh": "TEXT",
            "ngay_quyet_dinh": "TEXT",
            "so_quyet_dinh_ket_qua": "TEXT",
            "ngay_quyet_dinh_ket_qua": "TEXT",
            "gia_han_list": "TEXT",
            "yeu_cau_lam_ro_list": "TEXT",
            "tra_loi_lam_ro_list": "TEXT",
            "thoi_gian_goi_thau": "TEXT",
            "thoi_gian_hop_dong": "TEXT",
            "awarded_phan_lo_list": "TEXT",
            "gia_tri_dam_bao_du_thau": "REAL",
            "hieu_luc_hsdt": "INTEGER",
            "hieu_luc_dam_bao_du_thau": "INTEGER",
            "danh_gia_hsdt_metadata": "TEXT",
            "trang_thai": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": [
            "FOREIGN KEY (ke_hoach_id) REFERENCES ke_hoach_lcnt(id) ON DELETE CASCADE",
            "FOREIGN KEY (nha_thau_trung_thau_id) REFERENCES nha_thau(id) ON DELETE SET NULL"
        ]
    },
    "chuyen_gia": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "ho_ten": "TEXT NOT NULL",
            "so_chung_chi": "TEXT",
            "ngay_cap_chung_chi": "TEXT",
            "don_vi_cap_chung_chi": "TEXT",
            "so_cccd": "TEXT",
            "ngay_cap_cccd": "TEXT",
            "noi_cap_cccd": "TEXT",
            "anh_chung_chi": "TEXT",
            "ten_anh_chung_chi": "TEXT",
            "anh_chu_ky": "TEXT",
            "ten_anh_chu_ky": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    },
    "hop_dong": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "ten_hop_dong": "TEXT",
            "so_hop_dong": "TEXT",
            "ngay_ky": "TEXT",
            "chu_dau_tu_id": "TEXT",
            "nha_thau_id": "TEXT",
            "gia_tri": "REAL",
            "loai_hop_dong": "TEXT",
            "thoi_gian_thuc_hien": "TEXT",
            "trang_thai_ho_so": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": [
            "FOREIGN KEY (chu_dau_tu_id) REFERENCES chu_dau_tu(id) ON DELETE CASCADE",
            "FOREIGN KEY (nha_thau_id) REFERENCES nha_thau(id) ON DELETE CASCADE"
        ]
    },
    "hop_dong_goi_thau": {
        "columns": {
            "hop_dong_id": "TEXT",
            "goi_thau_id": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "primary_keys": ["hop_dong_id", "goi_thau_id"],
        "foreign_keys": [
            "FOREIGN KEY (hop_dong_id) REFERENCES hop_dong(id) ON DELETE CASCADE",
            "FOREIGN KEY (goi_thau_id) REFERENCES goi_thau(id) ON DELETE CASCADE"
        ]
    },
    "phan_cong_nhan_su": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "id_nhan_vien": "TEXT NOT NULL",
            "id_muc_tieu": "TEXT NOT NULL",
            "loai_doi_tuong": "TEXT NOT NULL",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "unique_constraints": [
            "UNIQUE(id_nhan_vien, id_muc_tieu, loai_doi_tuong)"
        ]
    },
    "trang_thai_ho_so_giay": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "org_id": "TEXT",
            "name": "TEXT",
            "color": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    },
    "thong_tin_mo_thau": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "owner_id": "TEXT",
            "goi_thau_id": "TEXT",
            "nha_thau_id": "TEXT",
            "ma_phan_lo": "TEXT",
            "ten_phan_lo": "TEXT",
            "ma_dinh_danh": "TEXT",
            "gia_du_thau": "REAL",
            "dam_bao_du_thau": "REAL",
            "hieu_luc_dam_bao": "TEXT",
            "hieu_luc_hsdxt": "TEXT",
            "ty_le_giam_gia": "REAL",
            "gia_sau_giam_gia": "REAL",
            "hieu_luc_hsdt": "INTEGER",
            "gia_tri_dam_bao": "REAL",
            "hieu_luc_bao_dam_ngay": "INTEGER",
            "thoi_gian_thuc_hien": "TEXT",
            "ten_nha_thau": "TEXT",
            "loai_nha_thau": "TEXT",
            "thanh_vien_lien_danh": "TEXT",
            "danh_gia_hop_le": "TEXT",
            "danh_gia_nang_luc": "TEXT",
            "danh_gia_ky_thuat": "TEXT",
            "danh_gia_tai_chinh": "TEXT",
            "danh_gia_ket_luan": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": [
            "FOREIGN KEY (goi_thau_id) REFERENCES goi_thau(id) ON DELETE CASCADE",
            "FOREIGN KEY (nha_thau_id) REFERENCES nha_thau(id) ON DELETE CASCADE"
        ]
    },
    "to_chuc": {
        "columns": {
            "id": "TEXT PRIMARY KEY",
            "ten_to_chuc": "TEXT UNIQUE NOT NULL",
            "quan_ly_id": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "foreign_keys": ["FOREIGN KEY (quan_ly_id) REFERENCES tai_khoan(id) ON DELETE SET NULL"]
    },
    "thanh_vien_to_chuc": {
        "columns": {
            "user_id": "TEXT NOT NULL",
            "to_chuc_id": "TEXT NOT NULL",
            "vai_tro_trong_to_chuc": "TEXT",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))",
            "updated_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "primary_keys": ["user_id", "to_chuc_id"],
        "foreign_keys": [
            "FOREIGN KEY (user_id) REFERENCES tai_khoan(id) ON DELETE CASCADE",
            "FOREIGN KEY (to_chuc_id) REFERENCES to_chuc(id) ON DELETE CASCADE"
        ]
    },
    "goi_thau_chuyen_gia": {
        "columns": {
            "goi_thau_id": "TEXT NOT NULL",
            "chuyen_gia_id": "TEXT NOT NULL",
            "created_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        },
        "primary_keys": ["goi_thau_id", "chuyen_gia_id"],
        "foreign_keys": [
            "FOREIGN KEY (goi_thau_id) REFERENCES goi_thau(id) ON DELETE CASCADE",
            "FOREIGN KEY (chuyen_gia_id) REFERENCES chuyen_gia(id) ON DELETE CASCADE"
        ]
    },
    "deleted_records": {
        "columns": {
            "id": "INTEGER PRIMARY KEY AUTOINCREMENT",
            "table_name": "TEXT NOT NULL",
            "record_id": "TEXT NOT NULL",
            "owner_id": "TEXT NOT NULL",
            "deleted_at": "INTEGER NOT NULL DEFAULT (strftime('%s','now'))"
        }
    }
}

# 2. Ánh xạ đặc biệt giữa JSON camelCase (Frontend) và DB snake_case (Backend)
SPECIAL_FIELD_MAPS = {
    "goi_thau": {
        "chuyen_gia_list": "toChuyenGia",
        "tham_dinh_list": "toThamDinh",
        "nha_thau_trung_thau_id": "nhaThauTrungThauId",
        "thoi_gian_dang_tai": "thoiGianDangTai",
        "thoi_gian_dong_thau": "thoiGianDongThau",
        "thoi_gian_mo_thau": "thoiGianMoThau"
    },
    "chuyen_gia": {
        "so_cccd": "soCCCD",
        "ngay_cap_cccd": "ngayCapCCCD",
        "noi_cap_cccd": "noiCapCCCD"
    },
    "hop_dong": {
        "thoi_gian_thuc_hien": "soNgayThucHien"
    },
    "phan_cong_nhan_su": {
        "id_nhan_vien": "empId",
        "id_muc_tieu": "targetId",
        "loai_doi_tuong": "type"
    }
}

def to_snake_case(name):
    import re
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()

def to_camel_case(snake_str):
    components = snake_str.split('_')
    return components[0] + ''.join(x.title() for x in components[1:])

def clean_id(val):
    if val is None or val == "":
        return None
    val_str = str(val).strip()
    import re
    val_str = re.sub(r'^(cdt-|kh-|gt-|cg-|nt-|hd-|emp-|user-|tm-)', '', val_str)
    return val_str

def khoi_tao_va_di_tru_he_thong():
    try:
        conn = database.get_connection()
        cursor = conn.cursor()
        
        # Kiểm tra và di trú dữ liệu từ cột ten_to_chuc cũ trong tai_khoan trước khi nó bị xóa bởi schema alignment
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='tai_khoan'")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(tai_khoan)")
            cols = [row[1] for row in cursor.fetchall()]
            if 'ten_to_chuc' in cols:
                # Đảm bảo bảng to_chuc và thanh_vien_to_chuc tồn tại
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS to_chuc (
                        id TEXT PRIMARY KEY,
                        ten_to_chuc TEXT UNIQUE NOT NULL,
                        quan_ly_id TEXT,
                        created_at INTEGER NOT NULL DEFAULT (strftime('%s','now')),
                        updated_at INTEGER NOT NULL DEFAULT (strftime('%s','now'))
                    )
                """)
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS thanh_vien_to_chuc (
                        user_id TEXT NOT NULL,
                        to_chuc_id TEXT NOT NULL,
                        vai_tro_trong_to_chuc TEXT,
                        created_at INTEGER NOT NULL DEFAULT (strftime('%s','now')),
                        updated_at INTEGER NOT NULL DEFAULT (strftime('%s','now')),
                        PRIMARY KEY (user_id, to_chuc_id)
                    )
                """)
                
                cursor.execute("SELECT id, ten_to_chuc, vai_tro FROM tai_khoan WHERE ten_to_chuc IS NOT NULL AND ten_to_chuc != ''")
                tk_rows = cursor.fetchall()
                org_managers = {}
                for row in tk_rows:
                    u_id = row['id']
                    vai_tro = row['vai_tro'] or ''
                    orgs = [o.strip() for o in row['ten_to_chuc'].split(',') if o.strip()]
                    for org in orgs:
                        if org not in org_managers:
                            org_managers[org] = u_id
                        else:
                            cursor.execute("SELECT vai_tro FROM tai_khoan WHERE id = ?", (org_managers[org],))
                            mgr_row = cursor.fetchone()
                            current_mgr_role = mgr_row['vai_tro'] or '' if mgr_row else ''
                            def role_weight(role):
                                if 'super_admin' in role: return 3
                                if 'manager' in role: return 2
                                return 1
                            if role_weight(vai_tro) > role_weight(current_mgr_role):
                                org_managers[org] = u_id
                
                import hashlib
                for org_name, mgr_id in org_managers.items():
                    org_hash_id = "org-" + hashlib.md5(org_name.encode('utf-8')).hexdigest()[:16]
                    cursor.execute("""
                        INSERT OR IGNORE INTO to_chuc (id, ten_to_chuc, quan_ly_id)
                        VALUES (?, ?, ?)
                    """, (org_hash_id, org_name, mgr_id))
                
                for row in tk_rows:
                    u_id = row['id']
                    vai_tro = row['vai_tro'] or ''
                    orgs = [o.strip() for o in row['ten_to_chuc'].split(',') if o.strip()]
                    for org in orgs:
                        cursor.execute("SELECT id FROM to_chuc WHERE ten_to_chuc = ?", (org,))
                        org_row = cursor.fetchone()
                        if org_row:
                            org_id = org_row['id']
                            role_in_org = 'employee'
                            if 'super_admin' in vai_tro:
                                role_in_org = 'super_admin'
                            elif 'manager' in vai_tro:
                                role_in_org = 'manager'
                            cursor.execute("""
                                INSERT OR IGNORE INTO thanh_vien_to_chuc (user_id, to_chuc_id, vai_tro_trong_to_chuc)
                                VALUES (?, ?, ?)
                            """, (u_id, org_id, role_in_org))
                print("Đồng bộ: Di trú trước dữ liệu tổ chức từ tai_khoan.ten_to_chuc thành công!")
        
        # ----------------------------------------------------
        # TỰ ĐỘNG KHỞI TẠO BẢNG & SO KHỚP CẤU TRÚC (MIGRATION)
        # ----------------------------------------------------
        for table_name, table_spec in SCHEMA_DINH_NGHIA.items():
            # Kiểm tra xem bảng đã tồn tại trong DB chưa
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'")
            table_exists = cursor.fetchone()
            
            if not table_exists:
                # Nếu bảng chưa tồn tại, tạo bảng động hoàn toàn dựa trên SCHEMA_DINH_NGHIA
                cols_def = []
                primary_keys = table_spec.get("primary_keys", [])
                
                for col_name, col_def in table_spec["columns"].items():
                    if primary_keys and col_name in primary_keys:
                        # Bỏ 'PRIMARY KEY' đơn lẻ nếu sử dụng Khóa chính hỗn hợp (composite key)
                        clean_def = col_def.replace("PRIMARY KEY", "")
                        cols_def.append(f"{col_name} {clean_def}")
                    else:
                        cols_def.append(f"{col_name} {col_def}")
                        
                if primary_keys:
                    cols_def.append(f"PRIMARY KEY ({', '.join(primary_keys)})")
                    
                for constraint in table_spec.get("unique_constraints", []):
                    cols_def.append(constraint)
                    
                for fk in table_spec.get("foreign_keys", []):
                    cols_def.append(fk)
                    
                sql_create = f"CREATE TABLE {table_name} ({', '.join(cols_def)})"
                print(f"Đồng bộ: Tạo bảng mới '{table_name}' theo cấu trúc code định nghĩa.")
                cursor.execute(sql_create)
                continue
            
            # Nếu bảng đã tồn tại, tiến hành đối chiếu từng cột để tự động nâng cấp cấu trúc
            cursor.execute(f"PRAGMA table_info({table_name})")
            current_cols = {row[1]: row for row in cursor.fetchall()}
            expected_cols = table_spec["columns"]
            
            # RÚT GỌN/ĐỔI TÊN ĐẶC BIỆT CỦA CỘT CŨ (NẾU CÓ)
            if table_name == "hop_dong" and "thoi_gian_thuc_hien" not in current_cols and "so_ngay_thuc_hien" in current_cols:
                print("Đồng bộ: Đổi tên cột 'so_ngay_thuc_hien' thành 'thoi_gian_thuc_hien' trong bảng 'hop_dong'")
                try:
                    cursor.execute("ALTER TABLE hop_dong RENAME COLUMN so_ngay_thuc_hien TO thoi_gian_thuc_hien")
                    cursor.execute("PRAGMA table_info(hop_dong)")
                    current_cols = {row[1]: row for row in cursor.fetchall()}
                except Exception as ex:
                    print(f"Lỗi khi đổi tên cột: {ex}")

            # Phát hiện xem có sự lệch kiểu dữ liệu hoặc thiếu cột đặc trưng nào không để xây dựng lại bảng
            rebuild_needed = False
            for col_name, col_def in expected_cols.items():
                if col_name in current_cols:
                    expected_type = col_def.split()[0].upper().replace(",", "")
                    current_type = current_cols[col_name][2].upper()
                    
                    def normalize_type(t):
                        t = t.strip()
                        if not t:
                            return "TEXT"
                        if "INT" in t:
                            return "INTEGER"
                        if "CHAR" in t or "CLOB" in t or "TEXT" in t:
                            return "TEXT"
                        if "BLOB" in t:
                            return "BLOB"
                        if "REAL" in t or "FLOA" in t or "DOUB" in t:
                            return "REAL"
                        return t

                    if normalize_type(expected_type) != normalize_type(current_type):
                        print(f"Đồng bộ: Phát hiện lệch kiểu dữ liệu cột '{col_name}' trong '{table_name}' (Code: {expected_type}, DB: {current_type})")
                        rebuild_needed = True
                        break
                else:
                    # SQLite không cho phép thêm cột mới có DEFAULT động (như strftime) hoặc NOT NULL mà không có default tĩnh qua ALTER TABLE.
                    # Nếu cột thiếu có DEFAULT, NOT NULL, UNIQUE hoặc REFERENCES, ta rebuild lại bảng.
                    col_def_upper = col_def.upper()
                    if "DEFAULT" in col_def_upper or "NOT NULL" in col_def_upper or "UNIQUE" in col_def_upper or "REFERENCES" in col_def_upper:
                        print(f"Đồng bộ: Phát hiện thiếu cột phức tạp '{col_name}' trong '{table_name}', cần xây dựng lại bảng.")
                        rebuild_needed = True
                        break
            
            if rebuild_needed:
                print(f"Đồng bộ: Tiến hành xây dựng lại bảng '{table_name}' để đồng bộ cấu trúc...")
                try:
                    # Tạm thời tắt kiểm tra khóa ngoại để đổi tên / xây dựng lại bảng
                    cursor.execute("PRAGMA foreign_keys = OFF")
                    
                    # 1. Đổi tên bảng hiện tại sang bảng tạm
                    temp_table = f"{table_name}_old_{int(datetime.now().timestamp())}"
                    cursor.execute(f"ALTER TABLE {table_name} RENAME TO {temp_table}")
                    
                    # 2. Tạo bảng mới theo cấu trúc chuẩn từ SCHEMA_DINH_NGHIA
                    cols_def = []
                    primary_keys = table_spec.get("primary_keys", [])
                    for col_name, col_def in table_spec["columns"].items():
                        if primary_keys and col_name in primary_keys:
                            clean_def = col_def.replace("PRIMARY KEY", "")
                            cols_def.append(f"{col_name} {clean_def}")
                        else:
                            cols_def.append(f"{col_name} {col_def}")
                    if primary_keys:
                        cols_def.append(f"PRIMARY KEY ({', '.join(primary_keys)})")
                    for constraint in table_spec.get("unique_constraints", []):
                        cols_def.append(constraint)
                    for fk in table_spec.get("foreign_keys", []):
                        cols_def.append(fk)
                    
                    sql_create = f"CREATE TABLE {table_name} ({', '.join(cols_def)})"
                    cursor.execute(sql_create)
                    
                    # 3. Lấy giao các cột có mặt ở cả hai cấu hình để sao chép dữ liệu
                    cursor.execute(f"PRAGMA table_info({temp_table})")
                    old_cols = [row[1] for row in cursor.fetchall()]
                    common_cols = [c for c in expected_cols.keys() if c in old_cols]
                    
                    if common_cols:
                        cols_str = ", ".join(common_cols)
                        cursor.execute(f"INSERT INTO {table_name} ({cols_str}) SELECT {cols_str} FROM {temp_table}")
                    
                    # 4. Xóa bảng tạm
                    cursor.execute(f"DROP TABLE {temp_table}")
                    cursor.execute("PRAGMA foreign_keys = ON")
                    print(f"Đồng bộ: Xây dựng lại bảng '{table_name}' thành công và bảo toàn dữ liệu!")
                    continue
                except Exception as ex:
                    cursor.execute("PRAGMA foreign_keys = ON")
                    print(f"Lỗi nghiêm trọng khi xây dựng lại bảng '{table_name}': {ex}")

            # 1. Thêm các cột thiếu từ code vào database
            for col_name, col_def in expected_cols.items():
                if col_name not in current_cols:
                    print(f"Đồng bộ: Thêm cột mới '{col_name}' ({col_def}) vào bảng '{table_name}'")
                    alter_def = col_def
                    if "PRIMARY KEY" in col_def.upper():
                        continue
                    if "NOT NULL" in col_def.upper() and "DEFAULT" not in col_def.upper():
                        alter_def = col_def.replace("NOT NULL", "")
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {alter_def}")
            
            # 2. Xóa các cột thừa trong database không còn tồn tại trong code định nghĩa
            for col_name in list(current_cols.keys()):
                if col_name not in expected_cols:
                    print(f"Đồng bộ: Xóa cột thừa '{col_name}' khỏi bảng '{table_name}' để khớp định nghĩa code")
                    try:
                        cursor.execute(f"ALTER TABLE {table_name} DROP COLUMN {col_name}")
                    except Exception as ex:
                        # Một số phiên bản SQLite cũ không hỗ trợ ALTER TABLE DROP COLUMN
                        print(f"Không thể xóa trực tiếp cột '{col_name}' trong SQLite (phiên bản cũ): {ex}")
                        
        # ----------------------------------------------------
        # DI TRÚ DỮ LIỆU TỪ HỆ THỐNG BẢNG TIẾNG ANH CŨ
        # ----------------------------------------------------
        
        # Chuyển đổi system_packages -> goi_dich_vu
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='system_packages'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM system_packages")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO goi_dich_vu (id, ten_goi, gia_ca, han_muc_nhan_su, mo_ta) VALUES (?, ?, ?, ?, ?)",
                               (d['id'], d['name'], d['price'], d['quota'], d['description']))
            print("Đã di trú system_packages -> goi_dich_vu")
  
        # Chuyển đổi users -> tai_khoan
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='users'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM users")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO tai_khoan (id, ten_dang_nhap, mat_khau, ho_ten, vai_tro, email, token_phien, anh_dai_dien, goi_dich_vu_id, ngay_bat_dau_goi, ngay_het_han_goi) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d['username'], d['password'], d['name'], d['role'], d['email'], d['active_session_token'], d.get('avatar'), d.get('package_id', 'silver'), d.get('package_start_date'), d.get('package_end_date')))
                
                # Di trú organization_name của user này sang to_chuc & thanh_vien_to_chuc
                org_name = d.get('organization_name')
                if org_name:
                    import hashlib
                    orgs = [o.strip() for o in org_name.split(',') if o.strip()]
                    for org in orgs:
                        org_hash_id = "org-" + hashlib.md5(org.encode('utf-8')).hexdigest()[:16]
                        cursor.execute("INSERT OR IGNORE INTO to_chuc (id, ten_to_chuc, quan_ly_id) VALUES (?, ?, ?)",
                                       (org_hash_id, org, d['id']))
                        role_in_org = 'employee'
                        if 'super_admin' in d['role']:
                            role_in_org = 'super_admin'
                        elif 'manager' in d['role']:
                            role_in_org = 'manager'
                        cursor.execute("INSERT OR IGNORE INTO thanh_vien_to_chuc (user_id, to_chuc_id, vai_tro_trong_to_chuc) VALUES (?, ?, ?)",
                                       (d['id'], org_hash_id, role_in_org))
            print("Đã di trú users -> tai_khoan")
  
        # Chuyển đổi investors -> chu_dau_tu
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='investors'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM investors")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO chu_dau_tu (id, ten_chu_dau_tu, ma_chu_dau_tu, ma_so_thue, chuc_vu_nguoi_dung_dau, nguoi_ky_quyet_dinh, chuc_vu_nguoi_ky, danh_xung, dia_chi, so_dien_thoai, so_tai_khoan, noi_mo_tai_khoan, email, ma_qhns) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d['name'], d.get('code', ''), d.get('tax_code', ''), d.get('head_position', ''), d.get('signer_name', ''), d.get('signer_position', ''), d.get('honorific', 'Ông'), d.get('address', ''), d.get('phone', ''), d.get('bank_account', ''), d.get('bank_name', ''), d.get('email', ''), d.get('budget_code', '')))
            print("Đã di trú investors -> chu_dau_tu")
  
        # Chuyển đổi plans -> ke_hoach_lcnt
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='plans'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM plans")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO ke_hoach_lcnt (id, id_goc, ma_ke_hoach, phien_ban, ten_ke_hoach, ten_du_an_du_toan, loai_hinh_mua_sam, chu_dau_tu_id, tong_muc_dau_tu, ngay_phe_duyet, quyet_dinh_phe_duyet, thoi_gian_dang_tai, cv_da_thuc_hien, cv_khong_ap_dung, cv_chua_du_dieu_kien, nguon_von, thoi_gian_du_an, dia_diem_quy_mo, thong_tin_khac) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d.get('root_id'), d.get('code'), d.get('version', '00'), d['name'], d.get('project_name'), d.get('loai_hinh'), d.get('investor_id'), d.get('total_investment', 0), d.get('approval_date'), d.get('approval_decision'), d.get('publish_date'), d.get('cv_da_thuc_hien', '[]'), d.get('cv_khong_ap_dung', '[]'), d.get('cv_chua_du_dieu_kien', '[]'), d.get('nguon_von', ''), d.get('thoigian_duan', ''), d.get('diadiem_quymo', ''), d.get('thongtin_khac', '')))
            print("Đã di trú plans -> ke_hoach_lcnt")
  
        # Chuyển đổi contractors -> nha_thau
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='contractors'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM contractors")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO nha_thau (id, ten_nha_thau, loai_nha_thau, thanh_vien_lien_danh, ma_so_thue, nguoi_dai_dien, danh_xung, so_dien_thoai, email, dia_chi, so_tai_khoan, noi_mo_tai_khoan, ma_ngan_hang, ma_nha_thau) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d['name'], d.get('type', 'Độc lập'), d.get('members', '[]'), d.get('tax_code', ''), d.get('representative', ''), d.get('honorific', 'Ông'), d.get('phone', ''), d.get('email', ''), d.get('address', ''), d.get('bank_account', ''), d.get('bank_name', ''), d.get('bank_code', ''), d.get('code', '')))
            print("Đã di trú contractors -> nha_thau")
  
        # Chuyển đổi experts -> chuyen_gia
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='experts'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM experts")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO chuyen_gia (id, ho_ten, so_chung_chi, ngay_cap_chung_chi, so_cccd, ngay_cap_cccd, anh_chung_chi) VALUES (?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d['full_name'], d.get('certificate_code'), d.get('certificate_date'), d.get('cccd'), d.get('cccd_date'), d.get('certificate_image')))
            print("Đã di trú experts -> chuyen_gia")
  
        # Chuyển đổi packages -> goi_thau
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='packages'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM packages")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO goi_thau (id, id_goc, ma_goi_thau, phien_ban, ke_hoach_id, ten_goi_thau, gia_goi_thau, loai_hop_dong, hinh_thuc_lua_chon, phuong_thuc_lua_chon, thoi_gian_thuc_hien, nguon_von, nha_thau_trung_thau_id, linh_vuc, tuy_chon_mua_them, thoi_gian_to_chuc, thoi_gian_bat_dau_to_chuc, phan_lo, phan_lo_list, tuy_chon_mua_them_list, thoi_gian_goi_thau, thoi_gian_hop_dong, awarded_phan_lo_list, trang_thai) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d.get('root_id'), d.get('code'), d.get('version', '00'), d.get('plan_id'), d['name'], d.get('price', 0), d.get('contract_type'), d.get('selection_method'), d.get('phuong_thuc_lua_chon', 'Một giai đoạn một túi hồ sơ'), d.get('execution_time'), d.get('capital_source'), d.get('awarded_contractor_id'), d.get('linh_vuc'), d.get('purchase_option', 'Không'), d.get('org_time'), d.get('org_start_time'), d.get('phan_lo', 'Không'), d.get('phan_lo_list', '[]'), d.get('tuy_chon_mua_them_list', '[]'), d.get('thoi_gian_goi_thau'), d.get('thoi_gian_hop_dong'), d.get('awarded_phan_lo_list', '[]'), d.get('status', 'Chuẩn bị')))
            print("Đã di trú packages -> goi_thau")
  
        # Chuyển đổi contracts -> hop_dong
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='contracts'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM contracts")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO hop_dong (id, ten_hop_dong, so_hop_dong, ngay_ky, chu_dau_tu_id, nha_thau_id, gia_tri, loai_hop_dong, thoi_gian_thuc_hien, trang_thai_ho_so) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                               (d['id'], d.get('name'), d.get('code'), d.get('sign_date'), d.get('investor_id'), d.get('contractor_id'), d.get('value', 0), d.get('type'), str(d.get('execution_time', '')), d.get('paper_status')))
            print("Đã di trú contracts -> hop_dong")
 
        # Chuyển đổi contract_package -> hop_dong_goi_thau
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='contract_package'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM contract_package")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO hop_dong_goi_thau (hop_dong_id, goi_thau_id) VALUES (?, ?)",
                               (d['contract_id'], d['package_id']))
            print("Đã di trú contract_package -> hop_dong_goi_thau")
 
        # Chuyển đổi assignments -> phan_cong_nhan_su
        cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='assignments'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT * FROM assignments")
            for r in cursor.fetchall():
                d = dict(r)
                cursor.execute("INSERT OR IGNORE INTO phan_cong_nhan_su (id, id_nhan_vien, id_muc_tieu, loai_doi_tuong) VALUES (?, ?, ?, ?)",
                               (d['id'], d['emp_id'], d['target_id'], d['type']))
            print("Đã di trú assignments -> phan_cong_nhan_su")
            
        # Dọn dẹp: Xóa sạch các bảng tiếng Anh cũ để tối ưu hóa và sạch sẽ database
        old_tables = ['users', 'system_packages', 'investors', 'plans', 'packages', 'experts', 'contractors', 'contracts', 'contract_package', 'assignments']
        for tbl in old_tables:
            cursor.execute(f"DROP TABLE IF EXISTS {tbl}")
            
        # Gieo dữ liệu bản quyền mặc định nếu trống
        cursor.execute("SELECT COUNT(*) FROM goi_dich_vu")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO goi_dich_vu (id, ten_goi, gia_ca, han_muc_nhan_su, mo_ta) VALUES (?, ?, ?, ?, ?)",
                           ('silver', 'Gói Bạc (Silver)', 15000000.0, 5, 'Phù hợp với đơn vị quy mô nhỏ, quản lý tối đa 5 nhân sự.'))
            cursor.execute("INSERT INTO goi_dich_vu (id, ten_goi, gia_ca, han_muc_nhan_su, mo_ta) VALUES (?, ?, ?, ?, ?)",
                           ('gold', 'Gói Vàng (Gold)', 35000000.0, 15, 'Giải pháp tuyệt vời cho phòng thầu chuyên nghiệp, tối đa 15 nhân sự.'))
            cursor.execute("INSERT INTO goi_dich_vu (id, ten_goi, gia_ca, han_muc_nhan_su, mo_ta) VALUES (?, ?, ?, ?, ?)",
                           ('diamond', 'Gói Kim Cương (Diamond)', 75000000.0, 999, 'Đặc quyền quản trị thầu tối cao, không giới hạn số lượng nhân sự.'))
                           
        # Gieo tài khoản mẫu mặc định nếu trống
        cursor.execute("SELECT COUNT(*) FROM tai_khoan")
        if cursor.fetchone()[0] == 0:
            import uuid
            import hashlib
            admin_uuid = "user-" + str(uuid.uuid4())
            cursor.execute("INSERT INTO tai_khoan (id, ten_dang_nhap, mat_khau, ho_ten, vai_tro, email, goi_dich_vu_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (admin_uuid, 'admin', hash_password('admin123'), 'Vy Tuấn Dương', 'super_admin', 'tuanduong51794@gmail.com', 'diamond'))
            
            # Gieo tổ chức mặc định HTD cho admin
            org_name = 'HTD'
            org_hash_id = "org-" + hashlib.md5(org_name.encode('utf-8')).hexdigest()[:16]
            cursor.execute("""
                INSERT OR IGNORE INTO to_chuc (id, ten_to_chuc, quan_ly_id)
                VALUES (?, ?, ?)
            """, (org_hash_id, org_name, admin_uuid))
            cursor.execute("""
                INSERT OR IGNORE INTO thanh_vien_to_chuc (user_id, to_chuc_id, vai_tro_trong_to_chuc)
                VALUES (?, ?, ?)
            """, (admin_uuid, org_hash_id, 'super_admin'))
                           
        # Cập nhật da_xac_minh = 1 cho các tài khoản hiện tại để tránh bị khóa
        cursor.execute("UPDATE tai_khoan SET da_xac_minh = 1 WHERE da_xac_minh IS NULL OR da_xac_minh = 0")

        # Gán quyền sở hữu mặc định cho các dữ liệu cũ (legacy data)
        cursor.execute("SELECT id FROM tai_khoan WHERE vai_tro = 'super_admin' OR ten_dang_nhap = 'admin' LIMIT 1")
        admin_row = cursor.fetchone()
        admin_id = str(admin_row[0]) if admin_row else "1"
        
        business_tables = [
            "chu_dau_tu", "ke_hoach_lcnt", "goi_thau", "chuyen_gia", 
            "nha_thau", "hop_dong", "phan_cong_nhan_su", 
            "trang_thai_ho_so_giay", "thong_tin_mo_thau"
        ]
        for tbl in business_tables:
            cursor.execute(f"UPDATE {tbl} SET owner_id = ? WHERE owner_id IS NULL OR owner_id = ''", (admin_id,))
            if tbl != "hop_dong_goi_thau":
                cursor.execute(f"CREATE INDEX IF NOT EXISTS idx_{tbl}_owner ON {tbl}(owner_id)")

        # Bổ sung indexes tối ưu hóa hiệu năng
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_kehoach_chudaututu ON ke_hoach_lcnt(chu_dau_tu_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_kehoach ON goi_thau(ke_hoach_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_nhathau ON goi_thau(nha_thau_trung_thau_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_mothau_goithau ON thong_tin_mo_thau(goi_thau_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_mothau_nhathau ON thong_tin_mo_thau(nha_thau_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_hopdong_chudautu ON hop_dong(chu_dau_tu_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_hopdong_nhathau ON hop_dong(nha_thau_id)")
        
        # Index token & email
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_taikhoan_token ON tai_khoan(token_phien) WHERE token_phien IS NOT NULL AND token_phien != ''")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_taikhoan_email ON tai_khoan(email) WHERE email != ''")
        
        # Index versioning
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_chudautu_idgoc ON chu_dau_tu(owner_id, id_goc)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_kehoach_idgoc ON ke_hoach_lcnt(owner_id, id_goc)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_nhathau_idgoc ON nha_thau(owner_id, id_goc)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_idgoc ON goi_thau(owner_id, id_goc)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_goithau_latest ON goi_thau(owner_id, id_goc, is_latest)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_nhathau_latest ON nha_thau(owner_id, id_goc, is_latest)")
        
        # Bổ sung indexes tối ưu hóa hiệu năng versioning còn thiếu
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_chudautu_latest ON chu_dau_tu(owner_id, id_goc, is_latest)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_kehoach_latest ON ke_hoach_lcnt(owner_id, id_goc, is_latest)")

        # Backfill is_latest = 1 cho phiên bản cao nhất
        for tbl in ["chu_dau_tu", "ke_hoach_lcnt", "nha_thau", "goi_thau"]:
            cursor.execute(f"UPDATE {tbl} SET is_latest = 0")
            cursor.execute(f"""
                UPDATE {tbl} SET is_latest = 1 WHERE id IN (
                    SELECT t1.id FROM {tbl} t1
                    INNER JOIN (
                        SELECT COALESCE(id_goc, id) as id_goc_group, MAX(CAST(phien_ban AS INTEGER)) as max_ver
                        FROM {tbl}
                        GROUP BY COALESCE(id_goc, id)
                    ) t2 ON COALESCE(t1.id_goc, t1.id) = t2.id_goc_group AND CAST(t1.phien_ban AS INTEGER) = t2.max_ver
                )
            """)

        # ----------------------------------------------------
        # DỌN DẸP LIÊN KẾT MỒ CÔI TỔ CHỨC
        # ----------------------------------------------------
        try:
            # Dọn dẹp các liên kết mồ côi (không tồn tại trong tai_khoan) để tránh lỗi Foreign Key
            cursor.execute("DELETE FROM thanh_vien_to_chuc WHERE user_id NOT IN (SELECT id FROM tai_khoan)")
            cursor.execute("UPDATE to_chuc SET quan_ly_id = NULL WHERE quan_ly_id NOT IN (SELECT id FROM tai_khoan)")
            print("Đồng bộ: Dọn dẹp các liên kết mồ côi tổ chức thành công!")
        except Exception as migration_ex:
            print("Lỗi khi di trú dữ liệu tổ chức:", migration_ex)

        # ----------------------------------------------------
        # TỰ ĐỘNG ĐỒNG BỘ CHUYÊN GIA GÓI THẦU BẰNG TRIGGERS
        # ----------------------------------------------------
        try:
            # Tạo các trigger tự động đồng bộ khi INSERT / UPDATE goi_thau
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS tg_sync_goithau_chuyengia_insert
                AFTER INSERT ON goi_thau
                FOR EACH ROW
                WHEN NEW.chuyen_gia_list IS NOT NULL AND json_valid(NEW.chuyen_gia_list)
                BEGIN
                    DELETE FROM goi_thau_chuyen_gia WHERE goi_thau_id = NEW.id;
                    INSERT INTO goi_thau_chuyen_gia (goi_thau_id, chuyen_gia_id)
                    SELECT DISTINCT NEW.id, json_extract(value, '$.id')
                    FROM json_each(NEW.chuyen_gia_list)
                    WHERE json_extract(value, '$.id') IS NOT NULL;
                END;
            """)

            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS tg_sync_goithau_chuyengia_update
                AFTER UPDATE OF chuyen_gia_list ON goi_thau
                FOR EACH ROW
                WHEN NEW.chuyen_gia_list IS NOT NULL AND json_valid(NEW.chuyen_gia_list)
                BEGIN
                    DELETE FROM goi_thau_chuyen_gia WHERE goi_thau_id = NEW.id;
                    INSERT INTO goi_thau_chuyen_gia (goi_thau_id, chuyen_gia_id)
                    SELECT DISTINCT NEW.id, json_extract(value, '$.id')
                    FROM json_each(NEW.chuyen_gia_list)
                    WHERE json_extract(value, '$.id') IS NOT NULL;
                END;
            """)

            # Di trú dữ liệu chuyên gia lịch sử (nếu có) vào goi_thau_chuyen_gia
            cursor.execute("SELECT id, chuyen_gia_list FROM goi_thau WHERE chuyen_gia_list IS NOT NULL AND chuyen_gia_list != ''")
            gt_rows = cursor.fetchall()
            for gt_row in gt_rows:
                gt_id = gt_row['id']
                cg_list_str = gt_row['chuyen_gia_list']
                if cg_list_str:
                    try:
                        cg_list = json.loads(cg_list_str)
                        for cg_item in cg_list:
                            cg_id = cg_item.get('id')
                            if cg_id:
                                clean_cg_id = clean_id(cg_id)
                                cursor.execute("""
                                    INSERT OR IGNORE INTO goi_thau_chuyen_gia (goi_thau_id, chuyen_gia_id)
                                    VALUES (?, ?)
                                """, (gt_id, clean_cg_id))
                    except Exception:
                        pass
            print("Đồng bộ: Thiết lập trigger và di trú chuyên gia sang goi_thau_chuyen_gia thành công!")
        except Exception as trigger_ex:
            print("Lỗi khi thiết lập trigger/di trú chuyên gia:", trigger_ex)
                           
        # ----------------------------------------------------
        # DI TRÚ CỘT owner_id TỪ TÊN TỔ CHỨC SANG ID TỔ CHỨC
        # ----------------------------------------------------
        try:
            cursor.execute("SELECT id, ten_to_chuc FROM to_chuc")
            org_rows = cursor.fetchall()
            for org_row in org_rows:
                o_id = org_row['id']
                o_name = org_row['ten_to_chuc']
                for tbl in business_tables:
                    cursor.execute(f"UPDATE {tbl} SET owner_id = ? WHERE owner_id = ?", (o_id, o_name))
            # Also migrate deleted_records owner_id
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='deleted_records'")
            if cursor.fetchone():
                for org_row in org_rows:
                    cursor.execute("UPDATE deleted_records SET owner_id = ? WHERE owner_id = ?", (org_row['id'], org_row['ten_to_chuc']))
            print("Đồng bộ: Di trú cột owner_id từ Tên tổ chức sang ID tổ chức thành công!")
        except Exception as migration_owner_ex:
            print("Lỗi khi di trú owner_id sang ID tổ chức:", migration_owner_ex)
            

                           
        conn.commit()
        conn.close()
        print("Khởi tạo và di trú cơ sở dữ liệu Tiếng Việt thành công!")
    except Exception as e:
        print("Lỗi khởi tạo/di trú database Tiếng Việt:", e)

khoi_tao_va_di_tru_he_thong()

# Import các thành phần của framework Starlette để dựng Web API Server
from starlette.applications import Starlette
from starlette.routing import Route, Mount, WebSocketRoute
from starlette.staticfiles import StaticFiles
from starlette.responses import StreamingResponse, JSONResponse, FileResponse
from starlette.middleware import Middleware
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.background import BackgroundTasks

# ==========================================
# 2. CÁC HÀM TRỢ GIÚP (HELPERS) & LOGGING
# ==========================================

def log_error(e_or_msg, context="System"):
    """
    Tự động ghi log lỗi chi tiết kèm traceback vào file sync_error.log ở thư mục gốc của dự án.
    """
    log_file = os.path.join(project_root, "sync_error.log")
    try:
        import traceback
        now_str = datetime.now().isoformat()
        if isinstance(e_or_msg, Exception):
            tb = traceback.format_exc()
            msg = f"[{now_str}] [{context}] LỖI: {str(e_or_msg)}\n{tb}\n"
        else:
            msg = f"[{now_str}] [{context}] THÔNG BÁO: {str(e_or_msg)}\n"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(msg)
    except Exception:
        pass
    print(f"[{context}] {e_or_msg}")

class ErrorLoggingMiddleware(BaseHTTPMiddleware):
    """
    Middleware giám sát toàn bộ request và tự động ghi log khi có lỗi hệ thống (500) hoặc ngoại lệ chưa xử lý.
    """
    async def dispatch(self, request, call_next):
        try:
            response = await call_next(request)
            if response.status_code >= 500:
                log_error(f"Phản hồi lỗi server {response.status_code}", f"HTTP {request.method} {request.url.path}")
            return response
        except Exception as e:
            log_error(e, f"HTTP {request.method} {request.url.path}")
            raise e

def format_date_str(date_str):
    """
    Chuẩn hóa chuỗi ngày tháng nhận được từ frontend sang định dạng hiển thị Việt Nam (DD/MM/YYYY).
    """
    if not date_str:
        return '--'
    date_str = str(date_str).strip().split(' ')[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(date_str, fmt).strftime('%d/%m/%Y')
        except ValueError:
            pass
    return date_str

class VietnameseFloat(float):
    """
    Lớp kế thừa kiểu số thực (float) tùy chỉnh nhằm định dạng tiền tệ theo phong cách Việt Nam.
    """
    def __format__(self, spec):
        try:
            # Format using standard Python grouping format ',.0f' (thousands separated by comma, 0 decimal places)
            formatted = format(float(self), ",.0f")
            # Replace commas with dots to match Vietnamese standard (e.g. 1.000.000)
            return formatted.replace(",", ".")
        except Exception:
            return super().__format__(spec)

def clean_admin_prefix(name):
    if not name:
        return ""
    import re
    name = name.strip()
    pattern = r"^(thành phố|tỉnh|phường|xã|thị trấn)\s+"
    return re.sub(pattern, "", name, flags=re.IGNORECASE)

# ==========================================
# 2.1 CÁC ENDPOINT ĐĂNG KÝ, ĐĂNG NHẬP, QUÊN MẬT KHẨU
# ==========================================

import smtplib
import random
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_SENDER = os.environ.get("SMTP_SENDER", SMTP_USER)

def gui_email(email_nhan, tieu_de, noi_dung_html):
    """
    Gửi email qua SMTP. Nếu thiếu cấu hình SMTP_USER hoặc SMTP_PASSWORD,
    sẽ ghi thông tin email vào file sync_error.log (Mock Mode).
    """
    if not SMTP_USER or not SMTP_PASSWORD:
        # Mock mode
        msg = f"[MOCK MAIL] Gửi tới: {email_nhan}\nTiêu đề: {tieu_de}\nNội dung:\n{noi_dung_html}\n"
        log_error(msg, context="EmailMock")
        return True
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_SENDER
        msg['To'] = email_nhan
        msg['Subject'] = tieu_de
        msg.attach(MIMEText(noi_dung_html, 'html', 'utf-8'))
        
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SMTP_SENDER, email_nhan, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        log_error(f"Lỗi gửi email tới {email_nhan}: {str(e)}", context="EmailSender")
        return False

async def register_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        role = 'employee'
        
        if not username or not password or not name or not email:
            return JSONResponse({"error": "Vui lòng nhập đầy đủ thông tin bắt buộc!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        
        # Check if username exists
        cursor.execute("SELECT id FROM tai_khoan WHERE ten_dang_nhap = ?", (username,))
        if cursor.fetchone():
            conn.close()
            return JSONResponse({"error": "Tài khoản đăng nhập đã tồn tại!"}, status_code=400)
            
        # Check if email exists
        cursor.execute("SELECT id FROM tai_khoan WHERE email = ?", (email,))
        if cursor.fetchone():
            conn.close()
            return JSONResponse({"error": "Địa chỉ email này đã được sử dụng bởi một tài khoản khác!"}, status_code=400)
            
        import uuid
        user_uuid = "user-" + str(uuid.uuid4())
        
        # Generate verification code
        code = str(random.randint(100000, 999999))
        expiry = int(time.time()) + 600 # 10 minutes from now
        
        cursor.execute(
            "INSERT INTO tai_khoan (id, ten_dang_nhap, mat_khau, ho_ten, vai_tro, email, goi_dich_vu_id, da_xac_minh, ma_xac_minh, han_xac_minh) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (user_uuid, username, hash_password(password), name, role, email, 'silver', 0, code, expiry)
        )
        conn.commit()
        conn.close()
        
        # Send Email
        tieu_de = "[BiddingFlow] Xác thực tài khoản đăng ký mới"
        noi_dung_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 8px;">
                <h2 style="color: #2563eb; text-align: center;">Chào mừng bạn đến với BiddingFlow</h2>
                <p>Xin chào <strong>{name}</strong>,</p>
                <p>Cảm ơn bạn đã đăng ký tài khoản tại BiddingFlow. Mã OTP xác thực email của bạn là:</p>
                <div style="background-color: #f1f5f9; padding: 15px; text-align: center; border-radius: 6px; margin: 20px 0;">
                    <span style="font-size: 24px; font-weight: bold; color: #1e3a8a; letter-spacing: 4px;">{code}</span>
                </div>
                <p style="font-size: 0.9rem; color: #64748b;">Mã OTP này có hiệu lực trong vòng 10 phút. Vui lòng không chia sẻ mã này với bất kỳ ai.</p>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 0.8rem; color: #94a3b8; text-align: center;">Hệ thống Đấu Thầu BiddingFlow</p>
            </div>
        </body>
        </html>
        """
        tasks = BackgroundTasks()
        tasks.add_task(gui_email, email, tieu_de, noi_dung_html)
        
        return JSONResponse(
            {"success": True, "message": "Đăng ký thành công! Vui lòng kiểm tra email để lấy mã xác nhận kích hoạt tài khoản."},
            background=tasks
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def verify_email_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        code = data.get('code', '').strip()
        
        if not username or not code:
            return JSONResponse({"error": "Thiếu thông tin xác thực!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ma_xac_minh, han_xac_minh FROM tai_khoan WHERE ten_dang_nhap = ?", (username,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return JSONResponse({"error": "Tài khoản không tồn tại!"}, status_code=400)
            
        user = dict(row)
        current_time = int(time.time())
        
        if user['ma_xac_minh'] != code:
            conn.close()
            return JSONResponse({"error": "Mã xác nhận không chính xác!"}, status_code=400)
            
        if user['han_xac_minh'] and current_time > user['han_xac_minh']:
            conn.close()
            return JSONResponse({"error": "Mã xác nhận đã hết hạn! Vui lòng yêu cầu mã mới."}, status_code=400)
            
        cursor.execute("UPDATE tai_khoan SET da_xac_minh = 1, ma_xac_minh = NULL, han_xac_minh = NULL WHERE id = ?", (user['id'],))
        conn.commit()
        conn.close()
        
        return JSONResponse({"success": True, "message": "Xác thực email thành công! Bạn có thể đăng nhập ngay bây giờ."})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def resend_code_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        
        if not username:
            return JSONResponse({"error": "Thiếu thông tin người dùng!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ho_ten, email, da_xac_minh FROM tai_khoan WHERE ten_dang_nhap = ?", (username,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return JSONResponse({"error": "Tài khoản không tồn tại!"}, status_code=400)
            
        user = dict(row)
        if user['da_xac_minh']:
            conn.close()
            return JSONResponse({"error": "Tài khoản này đã được xác thực trước đó!"}, status_code=400)
            
        code = str(random.randint(100000, 999999))
        expiry = int(time.time()) + 600
        
        cursor.execute("UPDATE tai_khoan SET ma_xac_minh = ?, han_xac_minh = ? WHERE id = ?", (code, expiry, user['id']))
        conn.commit()
        conn.close()
        
        tieu_de = "[BiddingFlow] Gửi lại mã xác thực tài khoản"
        noi_dung_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 8px;">
                <h2 style="color: #2563eb; text-align: center;">Mã xác thực mới của bạn</h2>
                <p>Xin chào <strong>{user['ho_ten']}</strong>,</p>
                <p>Bạn đã yêu cầu gửi lại mã xác nhận email cho tài khoản BiddingFlow. Mã OTP mới là:</p>
                <div style="background-color: #f1f5f9; padding: 15px; text-align: center; border-radius: 6px; margin: 20px 0;">
                    <span style="font-size: 24px; font-weight: bold; color: #1e3a8a; letter-spacing: 4px;">{code}</span>
                </div>
                <p style="font-size: 0.9rem; color: #64748b;">Mã OTP này có hiệu lực trong vòng 10 phút. Vui lòng không chia sẻ mã này với bất kỳ ai.</p>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 0.8rem; color: #94a3b8; text-align: center;">Hệ thống Đấu Thầu BiddingFlow</p>
            </div>
        </body>
        </html>
        """
        tasks = BackgroundTasks()
        tasks.add_task(gui_email, user['email'], tieu_de, noi_dung_html)
        
        return JSONResponse(
            {"success": True, "message": "Đã gửi lại mã OTP xác nhận vào email của bạn!"},
            background=tasks
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

def get_user_org_names(cursor, user_id):
    cursor.execute("""
        SELECT tc.ten_to_chuc 
        FROM thanh_vien_to_chuc tvtc
        JOIN to_chuc tc ON tvtc.to_chuc_id = tc.id
        WHERE tvtc.user_id = ?
    """, (user_id,))
    rows = cursor.fetchall()
    return ", ".join(row['ten_to_chuc'] for row in rows)

def update_user_organizations(cursor, user_id, organization_name, user_role='employee'):
    import hashlib
    # Parse new organizations
    new_orgs = [o.strip() for o in organization_name.split(',') if o.strip()]
    
    # Get current associations
    cursor.execute("""
        SELECT tc.id, tc.ten_to_chuc 
        FROM thanh_vien_to_chuc tvtc
        JOIN to_chuc tc ON tvtc.to_chuc_id = tc.id
        WHERE tvtc.user_id = ?
    """, (user_id,))
    current_assoc = {row['ten_to_chuc']: row['id'] for row in cursor.fetchall()}
    
    # 1. Add new associations
    for org_name in new_orgs:
        if org_name not in current_assoc:
            # Check if organization already exists in to_chuc by name
            cursor.execute("SELECT id FROM to_chuc WHERE ten_to_chuc = ?", (org_name,))
            org_row = cursor.fetchone()
            if org_row:
                org_id = org_row['id']
            else:
                # Create new organization
                org_id = "org-" + hashlib.md5(org_name.encode('utf-8')).hexdigest()[:16]
                cursor.execute(
                    "INSERT OR IGNORE INTO to_chuc (id, ten_to_chuc, quan_ly_id) VALUES (?, ?, ?)",
                    (org_id, org_name, user_id)
                )
            
            # Create association
            role_in_org = 'employee'
            if 'super_admin' in user_role:
                role_in_org = 'super_admin'
            elif 'manager' in user_role:
                role_in_org = 'manager'
            cursor.execute(
                "INSERT OR IGNORE INTO thanh_vien_to_chuc (user_id, to_chuc_id, vai_tro_trong_to_chuc) VALUES (?, ?, ?)",
                (user_id, org_id, role_in_org)
            )
            
    # 2. Remove associations for organizations no longer specified
    for org_name, org_id in current_assoc.items():
        if org_name not in new_orgs:
            cursor.execute(
                "DELETE FROM thanh_vien_to_chuc WHERE user_id = ? AND to_chuc_id = ?",
                (user_id, org_id)
            )

async def login_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return JSONResponse({"error": "Vui lòng nhập tài khoản và mật khẩu!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM tai_khoan WHERE ten_dang_nhap = ? OR (email != '' AND email = ?)",
            (username, username)
        )
        row = cursor.fetchone()

        
        if not row:
            conn.close()
            return JSONResponse({"error": "Tên đăng nhập hoặc mật khẩu không đúng"}, status_code=400)
            
        user = dict(row)
        if not verify_password(user['mat_khau'], password):
            conn.close()
            return JSONResponse({"error": "Tên đăng nhập hoặc mật khẩu không đúng"}, status_code=400)
            
        if not user.get('da_xac_minh'):
            conn.close()
            return JSONResponse({
                "error": "Tài khoản của bạn chưa được xác thực email. Vui lòng xác thực trước khi đăng nhập!",
                "unverified": True,
                "username": user['ten_dang_nhap']
            }, status_code=400)
            
        # Generate new active session token (uuid) to log out other devices
        session_token = str(uuid.uuid4())
        from datetime import datetime, timedelta
        token_expiry = (datetime.utcnow() + timedelta(hours=24)).isoformat()
        device_info = json.dumps({
            "user_agent": request.headers.get("User-Agent", "")[:200],
            "ip": request.client.host,
            "login_time": datetime.utcnow().isoformat()
        })
        cursor.execute(
            "UPDATE tai_khoan SET token_phien = ?, han_su_dung_token = ?, thong_tin_thiet_bi_cuoi = ? WHERE id = ?",
            (session_token, token_expiry, device_info, user['id'])
        )
        org_names = get_user_org_names(cursor, user['id'])
        conn.commit()
        conn.close()
        
        effective_roles = list(get_effective_roles(user['vai_tro']))
        return JSONResponse({
            "success": True,
            "id": user['id'],
            "session_token": session_token,
            "username": user['ten_dang_nhap'],
            "name": user['ho_ten'],
            "role": user['vai_tro'],
            "effective_roles": effective_roles,
            "email": user['email'],
            "avatar": user.get('anh_dai_dien'),
            "package_id": user.get('goi_dich_vu_id'),
            "organization_name": org_names
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def check_session_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        session_token = data.get('session_token', '').strip()
        
        if not username or not session_token:
            return JSONResponse({"valid": False, "error": "Thiếu thông tin xác thực"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ten_dang_nhap, ho_ten, vai_tro, email, anh_dai_dien, goi_dich_vu_id, token_phien, han_su_dung_token, thong_tin_thiet_bi_cuoi FROM tai_khoan WHERE ten_dang_nhap = ? OR (email != '' AND email = ?)", (username, username))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return JSONResponse({"valid": False, "reason": "user_not_found"})
            
        user = dict(row)
        org_names = get_user_org_names(cursor, user['id'])
        conn.close()
        
        active_token = user.get('token_phien')
        if active_token != session_token:
            return JSONResponse({"valid": False, "reason": "logged_in_elsewhere"})
            
        if user.get('han_su_dung_token'):
            try:
                expiry = datetime.fromisoformat(user['han_su_dung_token'])
                if datetime.utcnow() > expiry:
                    return JSONResponse({"valid": False, "reason": "token_expired"})
            except Exception:
                pass
                
        effective_roles = list(get_effective_roles(user['vai_tro']))
        return JSONResponse({
            "valid": True,
            "device_info": user.get('thong_tin_thiet_bi_cuoi'),
            "user": {
                "id": user['id'],
                "username": user['ten_dang_nhap'],
                "name": user['ho_ten'],
                "role": user['vai_tro'],
                "effective_roles": effective_roles,
                "email": user['email'],
                "avatar": user.get('anh_dai_dien'),
                "package_id": user.get('goi_dich_vu_id'),
                "organization_name": org_names
            }
        })
    except Exception as e:
        return JSONResponse({"valid": False, "error": str(e)}, status_code=500)

async def forgot_password_api(request):
    try:
        data = await request.json()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        
        if not username or not email:
            return JSONResponse({"error": "Vui lòng nhập tài khoản và email đã đăng ký!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ho_ten FROM tai_khoan WHERE ten_dang_nhap = ? AND email = ?", (username, email))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return JSONResponse({"error": "Thông tin tài khoản hoặc email không khớp!"}, status_code=400)
            
        user = dict(row)
        user_id = user['id']
        name = user['ho_ten']
        temp_pwd = secrets.token_hex(4)
        cursor.execute("UPDATE tai_khoan SET mat_khau = ? WHERE id = ?", (hash_password(temp_pwd), user_id))
        conn.commit()
        conn.close()
        
        tieu_de = "[BiddingFlow] Khôi phục mật khẩu tài khoản"
        noi_dung_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 8px;">
                <h2 style="color: #dc2626; text-align: center;">Khôi phục mật khẩu BiddingFlow</h2>
                <p>Xin chào <strong>{name}</strong>,</p>
                <p>Chúng tôi nhận được yêu cầu khôi phục mật khẩu cho tài khoản <strong>{username}</strong>.</p>
                <p>Mật khẩu tạm thời mới của bạn là:</p>
                <div style="background-color: #fef2f2; padding: 15px; text-align: center; border-radius: 6px; margin: 20px 0; border: 1px solid #fca5a5;">
                    <span style="font-size: 22px; font-weight: bold; color: #991b1b; letter-spacing: 2px;">{temp_pwd}</span>
                </div>
                <p>Vui lòng đăng nhập bằng mật khẩu tạm thời này và tiến hành thay đổi mật khẩu ngay lập tức trong phần quản lý tài khoản để đảm bảo bảo mật thông tin.</p>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                <p style="font-size: 0.8rem; color: #94a3b8; text-align: center;">Hệ thống Đấu Thầu BiddingFlow</p>
            </div>
        </body>
        </html>
        """
        tasks = BackgroundTasks()
        tasks.add_task(gui_email, email, tieu_de, noi_dung_html)
        
        return JSONResponse({
            "success": True, 
            "message": "Yêu cầu khôi phục mật khẩu thành công! Mật khẩu mới đã được gửi tới địa chỉ email của bạn. Vui lòng kiểm tra hộp thư (và thư mục Spam nếu không thấy)."
        }, background=tasks)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def update_profile_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        username = data.get('username', '').strip()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        avatar = data.get('avatar', '')
        
        organization_name = data.get('organization_name', '').strip()
        
        if not username or not name or not email:
            return JSONResponse({"error": "Vui lòng điền đầy đủ Họ tên và Email!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        
        # Check if email is in use by another user
        cursor.execute("SELECT ten_dang_nhap FROM tai_khoan WHERE email = ? AND ten_dang_nhap != ?", (email, username))
        if cursor.fetchone():
            conn.close()
            return JSONResponse({"error": "Địa chỉ email này đã được sử dụng bởi một tài khoản khác!"}, status_code=400)
            
        if avatar:
            cursor.execute("UPDATE tai_khoan SET ho_ten = ?, email = ?, anh_dai_dien = ? WHERE ten_dang_nhap = ?", (name, email, avatar, username))
        else:
            cursor.execute("UPDATE tai_khoan SET ho_ten = ?, email = ? WHERE ten_dang_nhap = ?", (name, email, username))
            
        cursor.execute("SELECT id, vai_tro FROM tai_khoan WHERE ten_dang_nhap = ?", (username,))
        u_row = cursor.fetchone()
        if u_row:
            update_user_organizations(cursor, u_row['id'], organization_name, u_row['vai_tro'])
            
        conn.commit()
        conn.close()
        
        return JSONResponse({"success": True, "message": "Cập nhật thông tin tài khoản thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def change_password_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        username = data.get('username', '').strip()
        old_password = data.get('old_password', '').strip()
        new_password = data.get('new_password', '').strip()
        
        if not username or not old_password or not new_password:
            return JSONResponse({"error": "Vui lòng nhập đầy đủ mật khẩu cũ và mới!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT mat_khau, id FROM tai_khoan WHERE ten_dang_nhap = ?", (username,))
        row = cursor.fetchone()
        
        if not row:
            conn.close()
            return JSONResponse({"error": "Người dùng không tồn tại!"}, status_code=400)
            
        user = dict(row)
        if not verify_password(user['mat_khau'], old_password):
            conn.close()
            return JSONResponse({"error": "Mật khẩu cũ không chính xác!"}, status_code=400)
            
        # Update password and invalidate other sessions
        new_token = str(uuid.uuid4())
        cursor.execute(
            "UPDATE tai_khoan SET mat_khau = ?, token_phien = ? WHERE id = ?",
            (hash_password(new_password), new_token, user['id'])
        )
        conn.commit()
        conn.close()
        
        return JSONResponse({
            "success": True, 
            "new_session_token": new_token,
            "message": "Thay đổi mật khẩu thành công! Các phiên đăng nhập trên thiết bị khác đã bị đăng xuất."
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def list_users_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        
        # Lấy thông tin tổ chức và vai trò của người đang yêu cầu
        cursor.execute("SELECT vai_tro FROM tai_khoan WHERE id = ?", (role_or_err.user_id,))
        requester = cursor.fetchone()
        if not requester:
            conn.close()
            return JSONResponse({"error": "Không tìm thấy thông tin tài khoản yêu cầu!"}, status_code=404)
            
        req_role = requester['vai_tro']
        
        # Lấy danh sách ID các tổ chức mà requester thuộc về
        cursor.execute("SELECT to_chuc_id FROM thanh_vien_to_chuc WHERE user_id = ?", (role_or_err.user_id,))
        req_org_ids = [r['to_chuc_id'] for r in cursor.fetchall()]
        
        sql_base = "SELECT id, ten_dang_nhap AS username, ho_ten AS name, vai_tro AS role, email, anh_dai_dien AS avatar, goi_dich_vu_id AS package_id, ngay_bat_dau_goi AS package_start_date, ngay_het_han_goi AS package_end_date FROM tai_khoan"
        
        if 'super_admin' in get_effective_roles(req_role):
            cursor.execute(sql_base)
            users_raw = cursor.fetchall()
        else:
            if not req_org_ids:
                # Nếu người dùng không thuộc tổ chức nào và không phải super_admin, chỉ trả về chính họ
                cursor.execute(sql_base + " WHERE id = ?", (role_or_err.user_id,))
                users_raw = cursor.fetchall()
            else:
                # Lọc những tài khoản có chung tổ chức
                placeholders = ",".join("?" for _ in req_org_ids)
                cursor.execute(f"""
                    SELECT DISTINCT tk.id, tk.ten_dang_nhap AS username, tk.ho_ten AS name, tk.vai_tro AS role, 
                                    tk.email, tk.anh_dai_dien AS avatar, tk.goi_dich_vu_id AS package_id, 
                                    tk.ngay_bat_dau_goi AS package_start_date, tk.ngay_het_han_goi AS package_end_date 
                    FROM tai_khoan tk
                    JOIN thanh_vien_to_chuc tvtc ON tk.id = tvtc.user_id
                    WHERE tvtc.to_chuc_id IN ({placeholders})
                """, req_org_ids)
                users_raw = cursor.fetchall()
                
        users = []
        for row in users_raw:
            u = dict(row)
            u['organization_name'] = get_user_org_names(cursor, u['id'])
            users.append(u)
        conn.close()
        return JSONResponse(users)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def delete_user_api(request):
    try:
        is_valid, role_or_err = verify_session(request, required_role='super_admin')
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        user_id = request.path_params.get('user_id')
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tai_khoan WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": "Xóa người dùng thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def update_user_role_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        effective_roles = get_effective_roles(role_or_err)
        if 'manager' not in effective_roles:
            return JSONResponse({"error": "Bạn không có quyền thực hiện thao tác này!"}, status_code=403)
            
        data = await request.json()
        user_id = data.get('user_id')
        new_role = data.get('role')
        
        if not user_id or not new_role:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc!"}, status_code=400)
            
        # Hỗ trợ nhiều role phân tách bằng dấu phẩy (VD: 'super_admin,manager')
        valid_roles = {'super_admin', 'manager', 'employee'}
        requested_roles = [r.strip() for r in new_role.split(',') if r.strip()]
        if not requested_roles or not all(r in valid_roles for r in requested_roles):
            return JSONResponse({"error": "Vai trò không hợp lệ!"}, status_code=400)
            
        # Nếu người thực hiện không phải là super_admin, không được phép gán vai trò super_admin
        if 'super_admin' not in effective_roles and 'super_admin' in requested_roles:
            return JSONResponse({"error": "Bạn không có quyền gán vai trò Quản trị viên tối cao!"}, status_code=403)
        
        # Chuẩn hóa: nếu có super_admin thì không cần liệt kê lại manager/employee
        normalized_role = ','.join(requested_roles)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE tai_khoan SET vai_tro = ? WHERE id = ?", (normalized_role, user_id))
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": "Cập nhật vai trò người dùng thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def update_user_package_api(request):
    try:
        is_valid, role_or_err = verify_session(request, required_role='super_admin')
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        user_id = data.get('user_id')
        new_package = data.get('package_id')
        
        if not user_id or new_package is None:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc!"}, status_code=400)
            
        pkgs = new_package.split(',')
        for p in pkgs:
            if p not in ['silver', 'gold', 'diamond', 'none', '']:
                return JSONResponse({"error": "Gói đăng ký không hợp lệ!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE tai_khoan SET goi_dich_vu_id = ? WHERE id = ?", (new_package, user_id))
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": "Cập nhật gói đăng ký thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def update_user_metadata_api(request):
    try:
        is_valid, role_or_err = verify_session(request, required_role='super_admin')
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        user_id = data.get('user_id')
        field = data.get('field')
        value = data.get('value')
        
        if not user_id or not field:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc!"}, status_code=400)
            
        field_map = {
            'package_start_date': 'ngay_bat_dau_goi',
            'package_end_date': 'ngay_het_han_goi',
            'name': 'ho_ten',
            'email': 'email'
        }
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        if field == 'organization_name':
            # Get user role
            cursor.execute("SELECT vai_tro FROM tai_khoan WHERE id = ?", (user_id,))
            u_row = cursor.fetchone()
            role = u_row['vai_tro'] if u_row else 'employee'
            update_user_organizations(cursor, user_id, value, role)
        else:
            if field not in field_map:
                conn.close()
                return JSONResponse({"error": "Trường cập nhật không hợp lệ!"}, status_code=400)
            db_field = field_map[field]
            cursor.execute(f"UPDATE tai_khoan SET {db_field} = ? WHERE id = ?", (value, user_id))
            
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": "Cập nhật thông tin thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def list_system_packages_api(request):
    try:
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ten_goi AS name, gia_ca AS price, han_muc_nhan_su AS quota, mo_ta AS description FROM goi_dich_vu")
        packages = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return JSONResponse(packages)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def update_system_package_api(request):
    try:
        is_valid, role_or_err = verify_session(request, required_role='super_admin')
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        pkg_id = data.get('id')
        name = data.get('name')
        price = float(data.get('price', 0))
        quota = int(data.get('quota', 0))
        description = data.get('description', '')
        
        if not pkg_id or not name:
            return JSONResponse({"error": "Thiếu thông tin bắt buộc!"}, status_code=400)
            
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE goi_dich_vu 
            SET ten_goi = ?, gia_ca = ?, han_muc_nhan_su = ?, mo_ta = ?
            WHERE id = ?
        """, (name, price, quota, description, pkg_id))
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": "Cập nhật gói dịch vụ thành công!"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

def safe_int_id(val, prefix=""):
    if not val:
        return None
    val_str = str(val).strip()
    if prefix:
        val_str = val_str.replace(prefix, "")
    import re
    digits = re.findall(r'\d+', val_str)
    if digits:
        return int(digits[0])
    return None

def safe_float(val):
    if not val:
        return 0.0
    try:
        s = str(val).strip()
        if ',' in s and '.' in s:
            if s.find('.') < s.find(','):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            if s.count(',') == 1:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')
        return float(s)
    except Exception:
        return 0.0

def safe_int(val):
    if not val:
        return 0
    try:
        return int(float(val))
    except Exception:
        return 0
def get_active_org(request, user_id):
    active_org = request.headers.get('X-Active-Org')
    if active_org:
        import urllib.parse
        active_org = urllib.parse.unquote(active_org)
    conn = database.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT tc.id, tc.ten_to_chuc 
        FROM thanh_vien_to_chuc tvtc
        JOIN to_chuc tc ON tvtc.to_chuc_id = tc.id
        WHERE tvtc.user_id = ?
    """, (user_id,))
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        return str(user_id)
        
    for row in rows:
        if active_org and (active_org == row['id'] or active_org == row['ten_to_chuc']):
            return row['id']
            
    return rows[0]['id']

active_connections = {}  # owner_id -> set of websocket instances

async def sync_websocket_endpoint(websocket):
    await websocket.accept()
    
    owner_id = None
    try:
        data = await websocket.receive_text()
        msg = json.loads(data)
        if msg.get("action") == "auth":
            token = msg.get("token")
            username = msg.get("username")
            
            conn = database.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, vai_tro, token_phien FROM tai_khoan WHERE ten_dang_nhap = ? OR (email != '' AND email = ?)", (username, username))
            row = cursor.fetchone()
            conn.close()
            
            if row and row['token_phien'] == token:
                user_id = row['id']
                conn = database.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT to_chuc_id 
                    FROM thanh_vien_to_chuc 
                    WHERE user_id = ?
                """, (user_id,))
                user_orgs = [r[0] for r in cursor.fetchall()]
                conn.close()
                if user_orgs:
                    owner_id = user_orgs[0]
                else:
                    owner_id = str(user_id)
                    
        if not owner_id:
            await websocket.close(code=4003)
            return
            
        if owner_id not in active_connections:
            active_connections[owner_id] = set()
        active_connections[owner_id].add(websocket)
        
        while True:
            await websocket.receive_text()
            
    except Exception:
        pass
    finally:
        if owner_id and owner_id in active_connections:
            active_connections[owner_id].discard(websocket)
            if not active_connections[owner_id]:
                del active_connections[owner_id]

def broadcast_websocket_event(owner_id, message):
    if owner_id in active_connections:
        import asyncio
        websockets = list(active_connections[owner_id])
        msg_str = json.dumps(message)
        
        async def broadcast():
            for ws in websockets:
                try:
                    await ws.send_text(msg_str)
                except Exception:
                    pass
        
        loop = asyncio.get_event_loop()
        if loop.is_running():
            loop.create_task(broadcast())

async def sync_api(request):
    """
    [POST] /api/sync
    Đồng bộ dữ liệu thay đổi từ ứng dụng Frontend vào cơ sở dữ liệu SQLite.
    """
    def log_sync_error(msg):
        log_error(msg, "SyncAPI")

    conn = None
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            log_sync_error(f"Xác thực thất bại khi đồng bộ: {role_or_err}")
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        data = await request.json()
        conn = database.get_connection()
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
        cursor = conn.cursor()
        
        org_name = get_active_org(request, role_or_err.user_id)
        current_time = int(datetime.utcnow().timestamp())
        
        # Map of API payload key to DB table name
        TABLE_KEYS = {
            "chudautu": "chu_dau_tu",
            "kehoach": "ke_hoach_lcnt",
            "goithau": "goi_thau",
            "chuyengia": "chuyen_gia",
            "nhathau": "nha_thau",
            "hopdong": "hop_dong",
            "assignments": "phan_cong_nhan_su",
            "custompaperstatuses": "trang_thai_ho_so_giay",
            "thongtinmothau": "thong_tin_mo_thau"
        }
        
        def get_clean_id(tbl, raw_id):
            if raw_id is None:
                return None
            if tbl in ["phan_cong_nhan_su", "trang_thai_ho_so_giay"]:
                return str(raw_id).strip()
            return clean_id(raw_id)
        
        for payload_key, table_name in TABLE_KEYS.items():
            if payload_key not in data:
                continue
            items = data[payload_key]
            if not isinstance(items, list):
                continue
                
            table_spec = SCHEMA_DINH_NGHIA[table_name]
            columns = list(table_spec["columns"].keys())
            
            # 1. Phát hiện và xử lý các bản ghi bị xóa
            incoming_ids = set()
            for item in items:
                raw_id = item.get('id')
                c_id = get_clean_id(table_name, raw_id)
                if c_id:
                    incoming_ids.add(str(c_id))
            
            cursor.execute(f"SELECT id FROM {table_name} WHERE owner_id = ?", (org_name,))
            existing_ids = set(str(row[0]) for row in cursor.fetchall())
            
            deleted_ids = existing_ids - incoming_ids
            if deleted_ids:
                deleted_list = list(deleted_ids)
                # Xóa khỏi bảng chính
                placeholders = ", ".join(["?"] * len(deleted_list))
                cursor.execute(f"DELETE FROM {table_name} WHERE owner_id = ? AND id IN ({placeholders})", (org_name, *deleted_list))
                # Ghi vào bảng deleted_records
                for d_id in deleted_list:
                    cursor.execute(
                        "INSERT INTO deleted_records (table_name, record_id, owner_id, deleted_at) VALUES (?, ?, ?, ?)",
                        (table_name, d_id, org_name, current_time)
                    )
            
            # 2. Thêm hoặc cập nhật (INSERT OR REPLACE) các bản ghi
            for item in items:
                try:
                    row_data = {}
                    for col in columns:
                        if col == "owner_id":
                            val = org_name
                        elif col == "updated_at":
                            val = current_time
                        else:
                            # Rút trích key JSON tương ứng từ trường DB
                            json_key = SPECIAL_FIELD_MAPS.get(table_name, {}).get(col)
                            if not json_key:
                                if col == "id_goc":
                                    json_key = "rootId"
                                else:
                                    json_key = to_camel_case(col)
                                    
                            val = item.get(json_key)
                            
                            # Fallback nếu client gửi key ở dạng raw/snake_case
                            if val is None:
                                val = item.get(col)
                                
                            # Làm sạch tiền tố ID
                            if col == "id" or col.endswith("_id") or col == "id_goc":
                                val = get_clean_id(table_name, val)
                                    
                            # Xử lý các trường kiểu List/Dict sang JSON string
                            if isinstance(val, (list, dict)):
                                val = json.dumps(val)
                            elif col.endswith("_list") or col.startswith("cv_") or col == "thanh_vien_lien_danh":
                                if val is None:
                                    val = "[]"
                                elif not isinstance(val, str):
                                    val = json.dumps(val)
                                    
                            # Chuẩn hóa kiểu dữ liệu số
                            col_type_upper = table_spec["columns"][col].upper()
                            if "REAL" in col_type_upper:
                                val = safe_float(val)
                            elif "INTEGER" in col_type_upper:
                                if val is not None:
                                    val = safe_int(val)
                                    
                            # Gán giá trị mặc định của schema nếu val là None
                            if val is None and "DEFAULT" in col_type_upper:
                                import re
                                default_match = re.search(r"DEFAULT\s+'([^']+)'", col_type_upper)
                                if default_match:
                                    val = default_match.group(1)
                                    
                            # Tối ưu hóa lưu trữ ảnh chuyên gia ra file vật lý
                            if table_name == "chuyen_gia" and col in ["anh_chung_chi", "anh_chu_ky"] and val:
                                ext_suffix = "cert" if col == "anh_chung_chi" else "sig"
                                expert_id = clean_id(item.get('id'))
                                val = save_base64_image(val, "chuyen_gia", f"{expert_id}_{ext_suffix}")
                                    
                        row_data[col] = val
                        
                    # Thực thi INSERT OR REPLACE
                    non_null_row_data = {k: v for k, v in row_data.items() if v is not None}
                    cols_str = ", ".join(non_null_row_data.keys())
                    placeholders = ", ".join(["?"] * len(non_null_row_data))
                    sql = f"INSERT OR REPLACE INTO {table_name} ({cols_str}) VALUES ({placeholders})"
                    cursor.execute(sql, tuple(non_null_row_data.values()))
                    
                    # Ràng buộc thêm: Gắn các gói thầu với hợp đồng (junction table)
                    if table_name == "hop_dong":
                        c_hd_id = get_clean_id("hop_dong", item.get('id'))
                        cursor.execute("DELETE FROM hop_dong_goi_thau WHERE hop_dong_id = ?", (c_hd_id,))
                        for gt_id_str in item.get('goiThauIds', []):
                            if gt_id_str:
                                gt_id = clean_id(gt_id_str)
                                if gt_id is not None:
                                    cursor.execute(
                                        "INSERT OR REPLACE INTO hop_dong_goi_thau (hop_dong_id, goi_thau_id) VALUES (?, ?)",
                                        (c_hd_id, gt_id)
                                    )
                except Exception as item_err:
                    import traceback
                    log_sync_error(f"Lỗi đồng bộ bản ghi trong bảng {table_name} (ID: {item.get('id')}): {item_err}\n{traceback.format_exc()}")
                    
        # Cập nhật cờ is_latest cho các bảng versioning sau khi lưu xong dữ liệu
        for tbl in ["chu_dau_tu", "ke_hoach_lcnt", "nha_thau", "goi_thau"]:
            cursor.execute(f"UPDATE {tbl} SET is_latest = 0 WHERE owner_id = ?", (org_name,))
            cursor.execute(f"""
                UPDATE {tbl} SET is_latest = 1 WHERE owner_id = ? AND id IN (
                    SELECT t1.id FROM {tbl} t1
                    INNER JOIN (
                        SELECT COALESCE(id_goc, id) as id_goc_group, MAX(CAST(phien_ban AS INTEGER)) as max_ver
                        FROM {tbl}
                        WHERE owner_id = ?
                        GROUP BY COALESCE(id_goc, id)
                    ) t2 ON COALESCE(t1.id_goc, t1.id) = t2.id_goc_group AND CAST(t1.phien_ban AS INTEGER) = t2.max_ver
                    WHERE t1.owner_id = ?
                )
            """, (org_name, org_name, org_name))
                    
        conn.commit()
        
        # Broadcast WebSocket update
        broadcast_websocket_event(org_name, {"event": "db_changed", "sender_session": request.headers.get('X-Session-Token')})
        
        return JSONResponse({"status": "success", "timestamp": current_time})
    except Exception as e:
        import traceback
        log_sync_error(f"Lỗi tổng quát sync_api: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

async def get_all_data_api(request):
    """
    [GET] /api/get-all-data
    Trả về dữ liệu thay đổi từ lần đồng bộ trước (nếu truyền since) hoặc toàn bộ dữ liệu.
    """
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        since = 0
        try:
            since = int(request.query_params.get('since', 0))
        except Exception:
            pass
            
        conn = database.get_connection()
        cursor = conn.cursor()
        current_time = int(datetime.utcnow().timestamp())
        
        # Hàm ánh xạ động DB snake_case sang JSON camelCase
        def map_db_to_json(table_name, row_dict):
            item = {}
            table_spec = SCHEMA_DINH_NGHIA[table_name]
            
            for col in table_spec["columns"].keys():
                json_key = SPECIAL_FIELD_MAPS.get(table_name, {}).get(col)
                if not json_key:
                    if col == "id_goc":
                        json_key = "rootId"
                    else:
                        json_key = to_camel_case(col)
                        
                val = row_dict.get(col)
                
                # 1. Thêm tiền tố ID cho client
                if col == "id" or col.endswith("_id") or col == "id_goc":
                    if table_name != "phan_cong_nhan_su" and val is not None:
                        prefix = ""
                        if col == "id":
                            prefix_map = {
                                "chu_dau_tu": "cdt-",
                                "ke_hoach_lcnt": "kh-",
                                "goi_thau": "gt-",
                                "chuyen_gia": "cg-",
                                "nha_thau": "nt-",
                                "hop_dong": "hd-",
                                "thong_tin_mo_thau": "tm-"
                            }
                            prefix = prefix_map.get(table_name, "")
                        elif col == "chu_dau_tu_id":
                            prefix = "cdt-"
                        elif col == "ke_hoach_id":
                            prefix = "kh-"
                        elif col == "goi_thau_id":
                            prefix = "gt-"
                        elif col == "nha_thau_trung_thau_id" or col == "nha_thau_id":
                            prefix = "nt-"
                        elif col == "id_goc":
                            prefix_map = {
                                "ke_hoach_lcnt": "kh-",
                                "goi_thau": "gt-",
                                "chu_dau_tu": "cdt-",
                                "nha_thau": "nt-"
                            }
                            prefix = prefix_map.get(table_name, "")
                            
                        val = f"{prefix}{val}"
                        
                # 2. Xử lý các trường dạng List/Dict đã lưu chuỗi JSON
                is_json_field = (
                    col.endswith("_list") or 
                    col.startswith("cv_") or 
                    col == "thanh_vien_lien_danh"
                )
                if is_json_field:
                    if val:
                        try:
                            val = json.loads(val)
                        except Exception:
                            val = []
                    else:
                        val = []
                        
                item[json_key] = val
            return item

        org_name = get_active_org(request, role_or_err.user_id)
        
        # Check scale for Server-side Pagination flag
        # Calculate total records across versionable/heavy tables
        heavy_tables = ["chu_dau_tu", "ke_hoach_lcnt", "goi_thau", "nha_thau", "chuyen_gia", "hop_dong"]
        total_records = 0
        for tbl in heavy_tables:
            cursor.execute(f"SELECT COUNT(*) FROM {tbl} WHERE owner_id = ?", (org_name,))
            total_records += cursor.fetchone()[0]
            
        use_server_pagination = total_records > 10000
        
        # Helper query function
        def query_table(tbl):
            if use_server_pagination:
                # If using server pagination, do not fetch all data, client will fetch paginated
                return []
            if since > 0:
                cursor.execute(f"SELECT * FROM {tbl} WHERE owner_id = ? AND updated_at > ?", (org_name, since))
            else:
                cursor.execute(f"SELECT * FROM {tbl} WHERE owner_id = ?", (org_name,))
            return cursor.fetchall()

        # 1. Chudautu
        chudautu = []
        for row in query_table("chu_dau_tu"):
            chudautu.append(map_db_to_json("chu_dau_tu", dict(row)))
            
        # 2. Kehoach
        kehoach = []
        for row in query_table("ke_hoach_lcnt"):
            item = map_db_to_json("ke_hoach_lcnt", dict(row))
            for list_key in ["cvDaThucHienList", "cvKhongApDungList", "cvChuaDuDieuKienList"]:
                if item.get(list_key) is None:
                    item[list_key] = []
            kehoach.append(item)
            
        # 3. Chuyengia
        chuyengia = []
        for row in query_table("chuyen_gia"):
            row_dict = dict(row)
            img = load_base64_image(row_dict.get("anh_chung_chi", ""))
            sig = load_base64_image(row_dict.get("anh_chu_ky", ""))
            item = map_db_to_json("chuyen_gia", row_dict)
            item["anhChungChi"] = img
            item["anhChuKy"] = sig
            chuyengia.append(item)
            
        # 4. Nhathau
        nhathau = []
        for row in query_table("nha_thau"):
            nhathau.append(map_db_to_json("nha_thau", dict(row)))
            
        # 5. Goithau
        goithau = []
        for row in query_table("goi_thau"):
            row_dict = dict(row)
            item = map_db_to_json("goi_thau", row_dict)
            cg_ids = []
            if item.get("toChuyenGia"):
                for x in item.get("toChuyenGia", []):
                    if isinstance(x, dict) and 'id' in x:
                        cg_ids.append(f"cg-{x['id']}")
            item["chuyenGiaIds"] = cg_ids
            for list_key in ["phanLoList", "tuyChonMuaThemList", "awardedPhanLoList", "toChuyenGia", "toThamDinh", "giaHanList", "yeuCauLamRoList", "traLoiLamRoList"]:
                if item.get(list_key) is None:
                    item[list_key] = []
            goithau.append(item)
            
        # 6. Hopdong
        hopdong = []
        for row in query_table("hop_dong"):
            row_dict = dict(row)
            item = map_db_to_json("hop_dong", row_dict)
            # Lấy danh sách gói thầu thuộc hợp đồng này (junction table)
            goithau_ids = []
            cursor.execute("SELECT goi_thau_id FROM hop_dong_goi_thau WHERE hop_dong_id = ?", (row_dict["id"],))
            for subrow in cursor.fetchall():
                goithau_ids.append(f"gt-{subrow[0]}")
            item["goiThauIds"] = goithau_ids
            hopdong.append(item)
            
        # 7. Assignments
        assignments = []
        if since > 0:
            cursor.execute("SELECT * FROM phan_cong_nhan_su WHERE owner_id = ? AND updated_at > ?", (org_name, since))
        else:
            cursor.execute("SELECT * FROM phan_cong_nhan_su WHERE owner_id = ?", (org_name,))
        for row in cursor.fetchall():
            assignments.append(map_db_to_json("phan_cong_nhan_su", dict(row)))
            
        # 8. Custom Paper Statuses
        custompaperstatuses = []
        if since > 0:
            cursor.execute("SELECT * FROM trang_thai_ho_so_giay WHERE owner_id = ? AND updated_at > ?", (org_name, since))
        else:
            cursor.execute("SELECT * FROM trang_thai_ho_so_giay WHERE owner_id = ?", (org_name,))
        for row in cursor.fetchall():
            custompaperstatuses.append(map_db_to_json("trang_thai_ho_so_giay", dict(row)))
            
        # 9. Thong Tin Mo Thau
        thongtinmothau = []
        for row in query_table("thong_tin_mo_thau"):
            thongtinmothau.append(map_db_to_json("thong_tin_mo_thau", dict(row)))
            
        # 10. Deletions
        deletions = []
        if since > 0:
            cursor.execute("SELECT table_name, record_id FROM deleted_records WHERE owner_id = ? AND deleted_at > ?", (org_name, since))
            TABLE_KEYS = {
                "chudautu": "chu_dau_tu",
                "kehoach": "ke_hoach_lcnt",
                "goithau": "goi_thau",
                "chuyengia": "chuyen_gia",
                "nhathau": "nha_thau",
                "hopdong": "hop_dong",
                "assignments": "phan_cong_nhan_su",
                "custompaperstatuses": "trang_thai_ho_so_giay",
                "thongtinmothau": "thong_tin_mo_thau"
            }
            TABLE_KEYS_INV = {v: k for k, v in TABLE_KEYS.items()}
            for row in cursor.fetchall():
                tbl_key = TABLE_KEYS_INV.get(row[0])
                if tbl_key:
                    prefix_map = {
                        "chu_dau_tu": "cdt-",
                        "ke_hoach_lcnt": "kh-",
                        "goi_thau": "gt-",
                        "chuyen_gia": "cg-",
                        "nha_thau": "nt-",
                        "hop_dong": "hd-",
                        "thong_tin_mo_thau": "tm-"
                    }
                    pfx = prefix_map.get(row[0], "")
                    deletions.append({"table": tbl_key, "id": f"{pfx}{row[1]}"})
                    
        conn.close()
        
        return JSONResponse({
            "chudautu": chudautu,
            "kehoach": kehoach,
            "chuyengia": chuyengia,
            "nhathau": nhathau,
            "goithau": goithau,
            "hopdong": hopdong,
            "assignments": assignments,
            "custompaperstatuses": custompaperstatuses,
            "thongtinmothau": thongtinmothau,
            "deletions": deletions,
            "useServerSidePagination": use_server_pagination,
            "timestamp": current_time
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


async def export_report_api(request):
    """
    [GET] /api/export-report/{package_id}
    Xuất báo cáo đánh giá hồ sơ mời thầu ra file Word (.docx).
    """
    package_id = request.path_params.get('package_id')
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        conn = database.get_connection()
        cursor = conn.cursor()
        
        org_name = get_active_org(request, user_id)

        cursor.execute("SELECT * FROM goi_thau WHERE id = ? AND owner_id = ?", (package_id, org_name))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return JSONResponse({"error": f"Package with id {package_id} not found"}, status_code=404)
        pkg = dict(row)
        
        cursor.execute("SELECT * FROM ke_hoach_lcnt WHERE id = ? AND owner_id = ?", (pkg['ke_hoach_id'], org_name))
        row_plan = cursor.fetchone()
        if not row_plan:
            conn.close()
            return JSONResponse({"error": f"Plan linked to package not found"}, status_code=404)
        plan = dict(row_plan)
        
        investor_name = '--'
        investor_address = ''
        if plan.get('chu_dau_tu_id'):
            cursor.execute("SELECT * FROM chu_dau_tu WHERE id = ?", (plan['chu_dau_tu_id'],))
            row_inv = cursor.fetchone()
            if row_inv:
                inv_data = dict(row_inv)
                investor_name = inv_data.get('ten_chu_dau_tu', '--')
                investor_address = inv_data.get('dia_chi', '')
                
        expert_ids = []
        if pkg.get('chuyen_gia_list'):
            try:
                cg_list = json.loads(pkg['chuyen_gia_list'])
                expert_ids = [int(x['id']) for x in cg_list if 'id' in x]
            except Exception:
                pass
                
        chuyen_gia_list = []
        if expert_ids:
            placeholders = ",".join(["?"] * len(expert_ids))
            cursor.execute(f"SELECT * FROM chuyen_gia WHERE id IN ({placeholders})", expert_ids)
            for row_cg in cursor.fetchall():
                cg = dict(row_cg)
                chuyen_gia_list.append({
                    'full_name': cg['ho_ten'],
                    'title': cg.get('chuc_danh', '--'),
                    'certificate_code': cg.get('so_chung_chi', '--'),
                    'certificate_date': format_date_str(cg.get('ngay_cap_chung_chi')),
                    'cccd': cg.get('so_cccd', '--'),
                    'cccd_date': format_date_str(cg.get('ngay_cap_cccd')),
                    'certificate_image': load_base64_image(cg.get('anh_chung_chi', ''))
                })
                
        awarded_id = pkg.get('nha_thau_trung_thau_id')
        nha_thau_list = []
        if awarded_id:
            cursor.execute("SELECT * FROM nha_thau WHERE id = ?", (int(awarded_id),))
            row_nt = cursor.fetchone()
            if row_nt:
                nt = dict(row_nt)
                
                # Cố gắng lấy thông tin liên danh thực tế được ghi nhận trong Biên bản mở thầu
                cursor.execute(
                    "SELECT ten_nha_thau, loai_nha_thau, thanh_vien_lien_danh FROM thong_tin_mo_thau WHERE goi_thau_id = ? AND nha_thau_id = ?",
                    (int(pkg['id']), int(awarded_id))
                )
                row_mt = cursor.fetchone()
                
                nt_name = nt['ten_nha_thau']
                nt_type = nt.get('loai_nha_thau', 'Độc lập')
                members_parsed = []
                
                if row_mt:
                    mt_data = dict(row_mt)
                    if mt_data.get('ten_nha_thau'):
                        nt_name = mt_data['ten_nha_thau']
                    if mt_data.get('loai_nha_thau'):
                        nt_type = mt_data['loai_nha_thau']
                    if mt_data.get('thanh_vien_lien_danh'):
                        try:
                            members_parsed = json.loads(mt_data['thanh_vien_lien_danh'])
                        except Exception:
                            members_parsed = []
                else:
                    # Hỗ trợ tương thích ngược nếu không tìm thấy dòng mở thầu
                    if nt.get('loai_nha_thau') == "Liên danh" and nt.get('thanh_vien_lien_danh'):
                        try:
                            members_parsed = json.loads(nt['thanh_vien_lien_danh'])
                        except Exception:
                            members_parsed = [nt['thanh_vien_lien_danh']]
                
                nha_thau_list.append({
                    'name': nt_name,
                    'type': nt_type,
                    'members': members_parsed,
                    'tax_code': nt.get('ma_so_thue', ''),
                    'representative': nt.get('nguoi_dai_dien', ''),
                    'phone': nt.get('so_dien_thoai', ''),
                    'awarded_price': pkg.get('gia_goi_thau', 0)
                })
        
        conn.close()
            
        context = {
            'ten_chu_dau_tu': investor_name,
            'so_quyet_dinh': plan.get('quyet_dinh_phe_duyet', '--'),
            'ngay_phe_duyet': format_date_str(plan.get('ngay_phe_duyet')),
            'ten_ke_hoach': plan.get('ten_ke_hoach', '--'),
            'ma_ke_hoach': plan.get('ma_ke_hoach', '--'),
            'tong_muc_dau_tu': VietnameseFloat(plan.get('tong_muc_dau_tu', 0)),
            'ten_goi_thau': pkg.get('ten_goi_thau', '--'),
            'ma_goi_thau': pkg.get('ma_goi_thau', '--'),
            'gia_goi_thau': VietnameseFloat(pkg.get('gia_goi_thau', 0)),
            'phuong_thuc_lua_chon': pkg.get('phuong_thuc_lua_chon') or pkg.get('hinh_thuc_lua_chon', '--'),
            'loai_hop_dong': pkg.get('loai_hop_dong', '--'),
            'thoi_gian_thuc_hien': pkg.get('thoi_gian_thuc_hien', '--'),
            'nguon_von': pkg.get('nguon_von', '--'),
            'linh_vuc': pkg.get('linh_vuc', '--'),
            'tuy_chon_mua_them': pkg.get('tuy_chon_mua_them', 'Không'),
            'thoi_gian_to_chuc': pkg.get('thoi_gian_to_chuc', '--'),
            'thoi_gian_bat_dau_to_chuc': format_date_str(pkg.get('thoi_gian_bat_dau_to_chuc')),
            'chuyen_gia': chuyen_gia_list,
            'nha_thau': nha_thau_list
        }
        
        active_tpl = custom_exporter.get_active_template(user_id)
        if active_tpl == 'mau_bao_cao_dau_thau.docx':
            docx_stream = exporter.export_docx_report(context)
        else:
            user_dir = custom_exporter.get_user_template_dir(user_id)
            tpl_path = os.path.join(user_dir, active_tpl)
            
            # Map custom dictionary variables
            custom_context = {}
            custom_context['Ten_Chu_Dau_Tu'] = investor_name
            
            # Parse address
            inv_parts = [p.strip() for p in (investor_address or '').split('|') if p.strip()]
            dia_chi_day_du = ", ".join(inv_parts) if inv_parts else "--"
            raw_xa = inv_parts[1] if len(inv_parts) > 1 else ""
            raw_tinh = inv_parts[2] if len(inv_parts) > 2 else ""
            
            clean_tinh = clean_admin_prefix(raw_tinh) or "--"
            clean_xa = clean_admin_prefix(raw_xa) or "--"
            
            custom_context['Dia_Chi_Day_Du_CDT'] = dia_chi_day_du
            custom_context['Tinh_Rieng_CDT'] = clean_tinh
            custom_context['Xa_Rieng_CDT'] = clean_xa
            custom_context['Dia_Chi_Rut_Gon_CDT'] = clean_tinh
            custom_context['So_Quyet_Dinh'] = plan.get('quyet_dinh_phe_duyet', '--')
            custom_context['Ngay_Phe_Duyet'] = format_date_str(plan.get('ngay_phe_duyet'))
            custom_context['Ten_Ke_Hoach'] = plan.get('ten_ke_hoach', '--')
            custom_context['Ma_Ke_Hoach'] = plan.get('ma_ke_hoach', '--')
            custom_context['Ma_Du_An'] = plan.get('ma_du_an', '--')
            custom_context['Tong_Muc_Dau_Tu'] = f"{VietnameseFloat(plan.get('tong_muc_dau_tu', 0))} VND"
            custom_context['Ten_Goi_Thau'] = pkg.get('ten_goi_thau', '--')
            custom_context['Ma_Goi_Thau'] = pkg.get('ma_goi_thau', '--')
            custom_context['Gia_Goi_Thau'] = f"{VietnameseFloat(pkg.get('gia_goi_thau', 0))} VND"
            custom_context['Phuong_Thuc_Lua_Chon'] = pkg.get('phuong_thuc_lua_chon') or pkg.get('hinh_thuc_lua_chon', '--')
            custom_context['Loai_Hop_Dong'] = pkg.get('loai_hop_dong', '--')
            custom_context['Thoi_Gian_Thuc_Hien'] = pkg.get('thoi_gian_thuc_hien', '--')
            custom_context['Nguon_Von'] = pkg.get('nguon_von', '--')
            custom_context['Linh_Vuc'] = pkg.get('linh_vuc', '--')
            custom_context['Tuy_Chon_Mua_Them'] = pkg.get('tuy_chon_mua_them', 'Không')
            custom_context['Thoi_Gian_To_Chuc'] = pkg.get('thoi_gian_to_chuc', '--')
            custom_context['Thoi_Gian_Bat_Dau_To_Chuc'] = format_date_str(pkg.get('thoi_gian_bat_dau_to_chuc'))
            
            custom_context['Danh_Sach_Chuyen_Gia'] = []
            for cg in chuyen_gia_list:
                custom_context['Danh_Sach_Chuyen_Gia'].append({
                    'Ho_Ten': cg.get('full_name', '--'),
                    'So_CCCD': cg.get('cccd', '--'),
                    'So_Chung_Chi': cg.get('certificate_code', '--'),
                    'Ngay_Cap_Chung_Chi': format_date_str(cg.get('certificate_date', '')),
                    'Don_Vi_Cap_Chung_Chi': cg.get('don_vi_cap', '--') or '--',
                    'Chuc_Vu': cg.get('title', '--')
                })
                
            custom_context['Danh_Sach_Nha_Thau'] = []
            for nt in nha_thau_list:
                custom_context['Danh_Sach_Nha_Thau'].append({
                    'Ten_Nha_Thau': nt.get('name', '--'),
                    'Loai_Nha_Thau': nt.get('type', '--'),
                    'Ma_So_Thue': nt.get('tax_code', '--') or '--',
                    'Nguoi_Dai_Dien': nt.get('representative', '--') or '--',
                    'So_Dien_Thoai': nt.get('phone', '--') or '--',
                    'Gia_Trung_Thau': f"{VietnameseFloat(nt.get('awarded_price', 0))} VND" if nt.get('awarded_price') else '--'
                })
                
            # Set global winner variables for conditional templates
            winning_nt = nha_thau_list[0] if nha_thau_list else {}
            custom_context['Loai_Nha_Thau'] = winning_nt.get('type', '--')
            custom_context['Thanh_Vien_Lien_Danh'] = winning_nt.get('members', [])
                
            docx_stream = custom_exporter.generate_report_from_custom_template(tpl_path, custom_context)
        
        return StreamingResponse(
            docx_stream,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename=Bao_cao_danh_gia_goi_thau_{pkg['ma_goi_thau']}.docx"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

async def list_templates_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        templates = custom_exporter.list_templates(user_id)
        return JSONResponse(templates)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def set_active_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        
        data = await request.json()
        filename = data.get('filename')
        if not filename:
            return JSONResponse({"error": "Filename is required"}, status_code=400)
        custom_exporter.set_active_template(filename, user_id)
        return JSONResponse({"status": "success"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def upload_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id

        form = await request.form()
        file_obj = form.get('file')
        if not file_obj:
            return JSONResponse({"error": "No file uploaded"}, status_code=400)
        
        file_bytes = await file_obj.read()
        valid, msg = custom_exporter.validate_template_syntax(file_bytes)
        
        if not valid:
            return JSONResponse({"success": False, "error": msg}, status_code=200)
        
        user_dir = custom_exporter.get_user_template_dir(user_id)
        save_path = os.path.join(user_dir, file_obj.filename)
        with open(save_path, 'wb') as f:
            f.write(file_bytes)
            
        custom_exporter.set_active_template(file_obj.filename, user_id)
        return JSONResponse({"success": True, "message": "Tải biểu mẫu lên thành công và đã được kích hoạt!"})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# =============================================================================
# ENTITY_SCHEMA — Nguồn dữ liệu duy nhất (single source of truth)
# Định nghĩa tất cả trường dữ liệu cho mỗi loại thực thể.
# Mỗi trường gồm:
#   - label   : Tiêu đề hiển thị trong file Excel xuất ra
#   - aliases : Danh sách tên cột có thể chấp nhận khi nhập vào (import)
#   - options : (Tuỳ chọn) Danh sách giá trị cho phép — tạo dropdown trong Excel
# Để thêm trường mới: chỉ cần thêm entry vào đây, cả export và import đều tự động cập nhật.
# =============================================================================
ENTITY_SCHEMA = {
    'chudautu': [
        {'field': 'maChuDauTu',           'label': 'Mã chủ đầu tư',              'aliases': ['Mã chủ đầu tư', 'Mã CĐT', 'maChuDauTu']},
        {'field': 'tenChuDauTu',           'label': 'Tên chủ đầu tư',             'aliases': ['Tên chủ đầu tư', 'Tên CĐT', 'tenChuDauTu']},
        {'field': 'maSoThue',              'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'chucVuNguoiDungDau',   'label': 'Chức vụ người đứng đầu',      'aliases': ['Chức vụ người đứng đầu', 'Chức vụ', 'chucVuNguoiDungDau']},
        {'field': 'nguoiKyQuyetDinh',     'label': 'Người ký quyết định',          'aliases': ['Người ký quyết định', 'Người ký', 'nguoiKyQuyetDinh']},
        {'field': 'chucVuNguoiKy',        'label': 'Chức vụ người ký',             'aliases': ['Chức vụ người ký', 'chucVuNguoiKy']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'Ông/Bà', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'diaChi',               'label': 'Địa chỉ trụ sở',              'aliases': ['Địa chỉ', 'Địa chỉ trụ sở', 'diaChi']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'STK', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'Ngân hàng', 'noiMoTaiKhoan']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'Địa chỉ email', 'email']},
        {'field': 'maQHNS',               'label': 'Mã QHNS',                    'aliases': ['Mã QHNS', 'maQHNS']},
    ],
    'kehoach': [
        {'field': 'maKeHoach',            'label': 'Mã kế hoạch',                'aliases': ['Mã kế hoạch', 'maKeHoach']},
        {'field': 'maDuan',               'label': 'Mã dự án',                   'aliases': ['Mã dự án', 'maDuan']},
        {'field': 'tenKeHoach',           'label': 'Tên kế hoạch',               'aliases': ['Tên kế hoạch', 'tenKeHoach']},
        {'field': 'loaiHinhMuaSam',       'label': 'Loại hình',                   'aliases': ['Loại hình', 'Loại hình mua sắm', 'loaiHinhMuaSam'],
                                           'options': ['Dự án', 'Dự toán mua sắm']},
        {'field': 'tenDuAnDuToan',        'label': 'Tên dự án',                  'aliases': ['Tên dự án', 'Dự án', 'Dự án / Dự toán', 'tenDuAnDuToan']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'Mã chủ đầu tư', 'chuDauTuId']},
        {'field': 'tongMucDauTu',         'label': 'Tổng mức đầu tư',             'aliases': ['Tổng giá trị', 'Tổng mức đầu tư', 'tongMucDauTu']},
        {'field': 'ngayPheDuyet',         'label': 'Ngày phê duyệt',               'aliases': ['Ngày phê duyệt', 'ngayPheDuyet']},
        {'field': 'quyetDinhPheDuyet',    'label': 'Số quyết định',               'aliases': ['Số quyết định', 'Quyết định phê duyệt', 'quyetDinhPheDuyet']},
        {'field': 'thoiGianDangMa',       'label': 'Thời gian đăng',               'aliases': ['Thời gian đăng', 'Thời gian đăng mã', 'thoiGianDangMa']},
        {'field': 'soQdPheDuyetDuAn',     'label': 'Số QĐ phê duyệt dự án',       'aliases': ['Số QĐ phê duyệt dự án', 'soQdPheDuyetDuAn']},
        {'field': 'ngayQdPheDuyetDuAn',   'label': 'Ngày QĐ phê duyệt dự án',     'aliases': ['Ngày QĐ phê duyệt dự án', 'ngayQdPheDuyetDuAn']},
        {'field': 'coQuanPheDuyetDuAn',   'label': 'Cơ quan phê duyệt dự án',     'aliases': ['Cơ quan phê duyệt dự án', 'coQuanPheDuyetDuAn']},
    ],
    'goithau': [
        {'field': 'maGoiThau',            'label': 'Mã gói thầu',                 'aliases': ['Mã gói thầu', 'maGoiThau']},
        {'field': 'tenGoiThau',           'label': 'Tên gói thầu',                'aliases': ['Tên gói thầu', 'tenGoiThau']},
        {'field': 'keHoachId',            'label': 'Kế hoạch',                    'aliases': ['Kế hoạch', 'Kế hoạch liên kết', 'keHoachId']},
        {'field': 'giaGoiThau',           'label': 'Giá gói thầu',               'aliases': ['Giá gói thầu', 'Giá gói', 'giaGoiThau']},
        {'field': 'hinhThucLuaChon',      'label': 'Hình thức',                   'aliases': ['Hình thức', 'Hình thức lựa chọn', 'hinhThucLuaChon'],
                                           'options': ['Đấu thầu rộng rãi', 'Đấu thầu hạn chế', 'Chỉ định thầu', 'Chỉ định thầu rút gọn', 'Chào hàng cạnh tranh', 'Lựa chọn nhà thầu trong trường hợp đặc biệt']},
        {'field': 'phuongThucLuaChon',    'label': 'Phương thức',                 'aliases': ['Phương thức', 'Phương thức lựa chọn', 'phuongThucLuaChon'],
                                           'options': ['Một giai đoạn một túi hồ sơ', 'Một giai đoạn hai túi hồ sơ', 'Hai giai đoạn một túi hồ sơ', 'Hai giai đoạn hai túi hồ sơ', 'Không có']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
        {'field': 'trangThai',            'label': 'Trạng thái',                   'aliases': ['Trạng thái', 'trangThai'],
                                           'options': ['Chuẩn bị', 'Đang mời thầu', 'Đã mở thầu', 'Đang chấm thầu', 'Đã có kết quả', 'Hủy thầu']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'nguonVon',             'label': 'Nguồn vốn',                   'aliases': ['Nguồn vốn', 'nguonVon']},
        {'field': 'linhVuc',              'label': 'Lĩnh vực',                    'aliases': ['Lĩnh vực', 'linhVuc'],
                                           'options': ['Tư vấn', 'Phi tư vấn', 'Xây lắp', 'Hỗn hợp', 'Hàng hóa']},
        {'field': 'tuyChonMuaThem',       'label': 'Tùy chọn mua thêm',           'aliases': ['Tùy chọn mua thêm', 'tuyChonMuaThem'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianToChuc',       'label': 'Thời gian tổ chức',            'aliases': ['Thời gian tổ chức', 'thoiGianToChuc']},
        {'field': 'thoiGianBatDauToChuc', 'label': 'Thời gian bắt đầu tổ chức',  'aliases': ['Thời gian bắt đầu tổ chức', 'thoiGianBatDauToChuc']},
        {'field': 'quaMang',              'label': 'Qua mạng / Trực tiếp',        'aliases': ['Qua mạng / Trực tiếp', 'Qua mạng', 'quaMang'],
                                           'options': ['Qua mạng', 'Trực tiếp']},
        {'field': 'trongNuocQuocTe',      'label': 'Trong nước / Quốc tế',       'aliases': ['Trong nước / Quốc tế', 'Trong nước', 'trongNuocQuocTe'],
                                           'options': ['Trong nước', 'Quốc tế']},
        {'field': 'phanLo',               'label': 'Phân lô / Không phân lô',     'aliases': ['Phân lô / Không phân lô', 'Phân lô', 'phanLo'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianDangTai',      'label': 'Thời gian đăng tải',          'aliases': ['Thời gian đăng tải', 'thoiGianDangTai']},
        {'field': 'thoiGianDongThau',     'label': 'Thời gian đóng thầu',         'aliases': ['Thời gian đóng thầu', 'thoiGianDongThau']},
        {'field': 'thoiGianMoThau',       'label': 'Thời gian mở thầu',           'aliases': ['Thời gian mở thầu', 'thoiGianMoThau']},
        {'field': 'soQuyetDinh',          'label': 'Số quyết định phê duyệt',     'aliases': ['Số quyết định phê duyệt', 'Số quyết định', 'soQuyetDinh']},
        {'field': 'ngayQuyetDinh',        'label': 'Ngày quyết định phê duyệt',   'aliases': ['Ngày quyết định phê duyệt', 'Ngày quyết định', 'ngayQuyetDinh']},
        {'field': 'soQuyetDinhKetQua',    'label': 'Số quyết định phê duyệt kết quả LCNT', 'aliases': ['Số quyết định phê duyệt kết quả LCNT', 'Số quyết định kết quả', 'soQuyetDinhKetQua']},
        {'field': 'ngayQuyetDinhKetQua',  'label': 'Ngày ký quyết định kết quả LCNT',     'aliases': ['Ngày ký quyết định kết quả LCNT', 'Ngày quyết định kết quả', 'ngayQuyetDinhKetQua']},
        {'field': 'nhaThauTrungThauId',   'label': 'Nhà thầu trúng thầu',          'aliases': ['Nhà thầu trúng thầu', 'Nhà thầu', 'nhaThauTrungThauId']},
        {'field': 'giaTrungThau',         'label': 'Giá trúng thầu',               'aliases': ['Giá trúng thầu', 'giaTrungThau']},
        {'field': 'thoiGianGoiThau',      'label': 'Thời gian gói thầu',          'aliases': ['Thời gian gói thầu', 'thoiGianGoiThau']},
        {'field': 'thoiGianHopDong',      'label': 'Thời gian hợp đồng',          'aliases': ['Thời gian hợp đồng', 'thoiGianHopDong']},
    ],
    'nhathau': [
        {'field': 'loaiNhaThau',          'label': 'Loại nhà thầu',               'aliases': ['Loại nhà thầu', 'loaiNhaThau'],
                                           'options': ['Độc lập', 'Liên danh']},
        {'field': 'maNhaThau',            'label': 'Mã nhà thầu',                 'aliases': ['Mã nhà thầu', 'Mã định danh', 'Mã nhà thầu', 'maNhaThau']},
        {'field': 'tenNhaThau',           'label': 'Tên nhà thầu',                'aliases': ['Tên nhà thầu', 'tenNhaThau']},
        {'field': 'maSoThue',             'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'nguoiDaiDien',         'label': 'Người đại diện',              'aliases': ['Người đại diện', 'nguoiDaiDien']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'email']},
        {'field': 'diaChi',               'label': 'Địa chỉ',                    'aliases': ['Địa chỉ', 'diaChi']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'noiMoTaiKhoan']},
    ],
    'chuyengia': [
        {'field': 'hoTen',                'label': 'Họ tên',                     'aliases': ['Họ tên', 'Họ và tên', 'hoTen']},
        {'field': 'soCCCD',               'label': 'Số CCCD',                     'aliases': ['Số CCCD', 'CCCD', 'soCCCD']},
        {'field': 'ngayCapCCCD',          'label': 'Ngày cấp CCCD',               'aliases': ['Ngày cấp CCCD', 'ngayCapCCCD']},
        {'field': 'noiCapCCCD',           'label': 'Nơi cấp CCCD',                'aliases': ['Nơi cấp CCCD', 'noiCapCCCD']},
        {'field': 'soChungChi',           'label': 'Số chứng chỉ',               'aliases': ['Số chứng chỉ', 'Số chứng chỉ đấu thầu', 'soChungChi']},
        {'field': 'ngayCapChungChi',      'label': 'Ngày cấp chứng chỉ',         'aliases': ['Ngày cấp', 'Ngày cấp chứng chỉ', 'ngayCapChungChi']},
        {'field': 'donViCapChungChi',     'label': 'Đơn vị cấp chứng chỉ',       'aliases': ['Đơn vị cấp', 'Đơn vị cấp chứng chỉ', 'donViCapChungChi']},
    ],
    'hopdong': [
        {'field': 'soHopDong',            'label': 'Số hợp đồng',                'aliases': ['Số hợp đồng', 'soHopDong']},
        {'field': 'tenHopDong',           'label': 'Tên hợp đồng',               'aliases': ['Tên hợp đồng', 'tenHopDong']},
        {'field': 'ngayKy',               'label': 'Ngày ký',                    'aliases': ['Ngày ký', 'Ngày ký hợp đồng', 'ngayKy']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'chuDauTuId']},
        {'field': 'nhaThauId',            'label': 'Nhà thầu',                    'aliases': ['Nhà thầu', 'nhaThauId']},
        {'field': 'giaTri',               'label': 'Giá trị hợp đồng',          'aliases': ['Giá trị', 'Giá trị hợp đồng', 'giaTri']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'soNgayThucHien',       'label': 'Thời gian thực hiện hợp đồng', 'aliases': ['Thời gian thực hiện hợp đồng', 'Thời gian thực hiện', 'Số ngày thực hiện', 'Số ngày', 'soNgayThucHien']},
        {'field': 'goiThauIds',           'label': 'Gói thầu liên kết',            'aliases': ['Gói thầu liên kết', 'Gói thầu', 'goiThauIds']},
    ],
    'phanlo': [
        {'field': 'tenPhanLo',            'label': 'Tên phần lô',                'aliases': ['Tên phần lô', 'Tên phân lô', 'tenPhanLo', 'Tên']},
        {'field': 'giaTriPhanLo',         'label': 'Giá trị phần lô',             'aliases': ['Giá trị phần lô', 'Giá trị phân lô', 'Giá trị', 'giaTriPhanLo']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
    ],
    'tuychonmuathem': [
        {'field': 'hangMuc',              'label': 'Hạng mục',                   'aliases': ['Hạng mục', 'Tên hạng mục', 'hangMuc']},
        {'field': 'donVi',                'label': 'Đơn vị',                    'aliases': ['Đơn vị', 'Đơn vị tính', 'ĐVT', 'donVi']},
        {'field': 'soLuong',              'label': 'Khối lượng / Số lượng',       'aliases': ['Khối lượng/ Số lượng', 'Khối lượng', 'Số lượng', 'soLuong', 'khoiLuong']},
        {'field': 'tyLe',                 'label': 'Tỷ lệ phần trăm (%)',         'aliases': ['Tỷ lệ phần trăm (%)', 'Tỷ lệ phần trăm', 'Tỷ lệ (%)', 'Tỷ lệ', 'tyLe', 'phanTram']},
        {'field': 'giaTriUocTinh',        'label': 'Giá trị ước tính',           'aliases': ['Giá trị ước tính', 'Giá trị', 'giaTriUocTinh']},
    ],
}


def _schema_to_map_cols(entity_type):
    """Chuyển ENTITY_SCHEMA thành map_cols dùng cho import (field → aliases)."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return {entry['field']: entry['aliases'] for entry in schema}


def _schema_to_headers(entity_type):
    """Lấy danh sách label (tiêu đề cột) cho template Excel."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return [entry['label'] for entry in schema]


def _schema_to_options(entity_type):
    """Lấy options_map (label → danh sách giá trị dropdown) cho validation Excel."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return {}
    return {entry['label']: entry['options'] for entry in schema if entry.get('options')}


async def import_excel_api(request):
    try:
        from io import BytesIO
        form = await request.form()
        file_obj = form.get('file')
        import_type = form.get('type')
        
        if not file_obj or not import_type:
            return JSONResponse({"error": "Missing file or type parameter"}, status_code=400)

        # Tự động lấy map_cols từ ENTITY_SCHEMA — không còn hardcoded
        map_cols = _schema_to_map_cols(import_type)
        if map_cols is None:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)


        file_bytes = await file_obj.read()
        df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
        
        all_possible_headers = []
        for poss in map_cols.values():
            all_possible_headers.extend([x.lower() for x in poss])
            
        first_col = [str(x).strip().lower() for x in df_raw.iloc[:, 0].dropna()]
        vertical_matches = sum(1 for v in first_col if v in all_possible_headers)
        
        if vertical_matches >= 3 or (df_raw.shape[1] >= 2 and vertical_matches >= 1):
            headers = [str(x).strip() for x in df_raw.iloc[:, 0]]
            records = []
            for col_idx in range(1, df_raw.shape[1]):
                col_vals = df_raw.iloc[:, col_idx]
                # Bỏ qua nếu cả cột trống rỗng
                if all(str(x).strip() == "" or pd.isna(x) for x in col_vals):
                    continue
                
                row_data = {}
                for r_idx, h in enumerate(headers):
                    val = col_vals.iloc[r_idx] if r_idx < len(col_vals) else ""
                    if pd.isna(val):
                        val = ""
                    row_data[h] = val
                records.append(row_data)
            df = pd.DataFrame(records)
        else:
            df = pd.read_excel(BytesIO(file_bytes))
            
        df.columns = [str(c).strip() for c in df.columns]
        
        rows = []
        
        def find_col(possible_names):
            for name in possible_names:
                for col in df.columns:
                    if col.lower() == name.lower() or col.lower().replace(" ", "") == name.lower().replace(" ", ""):
                        return col
            return None

        def clean_money(val):
            if pd.isna(val):
                return 0
            val_str = str(val).replace("VND", "").replace("đ", "").replace("₫", "").replace(".", "").replace(",", "").strip()
            try:
                return float(val_str)
            except ValueError:
                return 0

        def clean_int(val):
            if pd.isna(val):
                return 0
            try:
                return int(float(str(val).strip()))
            except ValueError:
                return 0

        for idx, row in df.iterrows():
            item = {}
            validation_comments = []
            
            for key, poss in map_cols.items():
                found = find_col(poss)
                val = row[found] if (found is not None) else None
                if pd.isna(val):
                    val = ""
                    
                if key in ['tongMucDauTu', 'giaGoiThau', 'giaTri', 'giaTriPhanLo', 'giaTriUocTinh', 'giaTrungThau']:
                    val = clean_money(val)
                elif key in ['thoiGianThucHien']:
                    val = clean_int(val)
                elif key in ['soLuong', 'tyLe']:
                    try:
                        val = float(str(val).strip()) if val != "" else 0.0
                    except ValueError:
                        val = 0.0
                else:
                    val = str(val).strip()
                    
                item[key] = val
            
            if import_type == 'chudautu':
                if not item['tenChuDauTu']:
                    validation_comments.append("Tên chủ đầu tư không được để trống")
                if not item['maChuDauTu']:
                    validation_comments.append("Mã chủ đầu tư không được để trống")
            elif import_type == 'kehoach':
                if not item['tenKeHoach']:
                    validation_comments.append("Tên kế hoạch không được để trống")
                if not item['maKeHoach']:
                    validation_comments.append("Mã kế hoạch không được để trống")
            elif import_type == 'goithau':
                if not item['tenGoiThau']:
                    validation_comments.append("Tên gói thầu không được để trống")
                if not item['maGoiThau']:
                    validation_comments.append("Mã gói thầu không được để trống")
            elif import_type == 'nhathau':
                if not item['tenNhaThau']:
                    validation_comments.append("Tên nhà thầu không được để trống")
                if not item['maNhaThau']:
                    validation_comments.append("Mã nhà thầu không được để trống")
            elif import_type == 'chuyengia':
                if not item['hoTen']:
                    validation_comments.append("Họ và tên không được để trống")
                if not item['soChungChi']:
                    validation_comments.append("Số chứng chỉ không được để trống")
            elif import_type == 'hopdong':
                if not item['tenHopDong']:
                    validation_comments.append("Tên hợp đồng không được để trống")
                if not item['soHopDong']:
                    validation_comments.append("Số hợp đồng không được để trống")
            elif import_type == 'phanlo':
                if not item['tenPhanLo']:
                    validation_comments.append("Tên phần lô không được để trống")
            elif import_type == 'tuychonmuathem':
                if not item['hangMuc']:
                    validation_comments.append("Hạng mục không được để trống")
            
            item['_valid'] = len(validation_comments) == 0
            item['_comment'] = "; ".join(validation_comments) if validation_comments else "Hợp lệ"
            rows.append(item)
            
        return JSONResponse({"success": True, "rows": rows})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_excel_template_api(request):
    try:
        # Tự động lấy danh sách cột và dropdown từ ENTITY_SCHEMA — không còn hardcoded
        import_type = request.path_params.get('import_type')
        from io import BytesIO
        cols = _schema_to_headers(import_type)
        if not cols:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)

        options_map = _schema_to_options(import_type)


        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        # Store dropdown options in a hidden 'Dropdowns' sheet to ensure robust cross-locale compatibility
        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                # Excel formula reference to the hidden dropdown sheet range
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        is_vertical = len(cols) > 5  # Tự động xác định: Nếu quá 5 trường thì để dạng dọc cho dễ theo dõi, nhập liệu

        if is_vertical:
            for idx, col_name in enumerate(cols, start=1):
                ws.row_dimensions[idx].height = 24
                cell_a = ws.cell(row=idx, column=1, value=col_name)
                
                cell_a.font = Font(name="Calibri", size=11, bold=True)
                cell_a.alignment = left_align
                cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_a.border = thin_border
                
                for col_idx in range(2, 12):
                    cell_data = ws.cell(row=idx, column=col_idx)
                    cell_data.border = thin_border

                options = options_map.get(col_name)
                if options:
                    dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                    dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                    dv.errorTitle = 'Lỗi nhập liệu'
                    dv.prompt = 'Chọn một giá trị trong danh sách'
                    dv.promptTitle = col_name
                    dv.errorStyle = "stop"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    ws.add_data_validation(dv)
                    for col_idx in range(2, 12):
                        dv.add(ws.cell(row=idx, column=col_idx))

            # Tự động căn chỉnh độ rộng cột dạng dọc
            max_len_a = max(len(str(cell.value or '')) for cell in ws['A'])
            ws.column_dimensions['A'].width = max(max_len_a + 5, 25)
            for col_idx in range(2, 12):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 18
        else:
            ws.append(cols)
            ws.row_dimensions[1].height = 28
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            ws.row_dimensions[2].height = 24
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.border = thin_border
                col_name = cols[col_idx - 1]
                options = options_map.get(col_name)
                if options:
                    col_letter = get_column_letter(col_idx)
                    dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                    dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                    dv.errorTitle = 'Lỗi nhập liệu'
                    dv.prompt = 'Chọn một giá trị trong danh sách'
                    dv.promptTitle = col_name
                    dv.errorStyle = "stop"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}100")

            # Tự động căn chỉnh độ rộng cột dạng ngang
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_nhap_lieu_{import_type}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_mothau_template_api(request):
    try:
        case_type = request.query_params.get('case_type', '1G1T_NO_LOT')
        package_name = request.query_params.get('package_name', 'GoiThau')
        lot_codes_str = request.query_params.get('lot_codes', '')
        
        lot_codes = [c.strip() for c in lot_codes_str.split(',') if c.strip()]
        
        headers = []
        options_map = {
            'Loại nhà thầu': ['Độc lập', 'Liên danh']
        }
        
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        # Store dropdown options in a hidden 'Dropdowns' sheet
        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[2].height = 24
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.border = thin_border
            col_name = headers[col_idx - 1]
            options = options_map.get(col_name)
            if options:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn một giá trị trong danh sách'
                dv.promptTitle = col_name
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}100")

        # Column width adjustment
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_Mo_Thau_{case_type}_{package_name}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_danhgiahsdt_template_api(request):
    try:
        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)
            
        pkg_id_int = safe_int_id(package_id, "gt-")
        if pkg_id_int is None:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list FROM goi_thau WHERE id = ?", (pkg_id_int,))
        gt_row = cursor.fetchone()
        if not gt_row:
            conn.close()
            return JSONResponse({"error": "Package not found"}, status_code=404)
            
        linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list_str = gt_row
        
        cursor.execute("""
            SELECT loai_nha_thau, ma_phan_lo, ten_phan_lo, ma_dinh_danh, ten_nha_thau,
                   gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia, hieu_luc_hsdt,
                   gia_tri_dam_bao, hieu_luc_bao_dam_ngay, thoi_gian_thuc_hien,
                   dam_bao_du_thau, hieu_luc_dam_bao, hieu_luc_hsdxt,
                   danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ?
        """, (pkg_id_int,))
        bids = cursor.fetchall()
        conn.close()
        
        is_tu_van = linh_vuc == 'Tư vấn'
        is_1g2t = phuong_thuc_lua_chon == 'Một giai đoạn hai túi hồ sơ'
        is_1g1t = phuong_thuc_lua_chon == 'Một giai đoạn một túi hồ sơ'
        has_phan_lo = phan_lo == 'Có'
        
        case_type = '1G1T_NO_LOT'
        if is_tu_van:
            case_type = 'TU_VAN'
        elif not is_tu_van and is_1g2t:
            case_type = '1G2T_WITH_LOT' if has_phan_lo else '1G2T_NO_LOT'
        elif is_1g1t:
            case_type = '1G1T_WITH_LOT' if has_phan_lo else '1G1T_NO_LOT'
            
        headers = []
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Đánh giá năng lực', 'Đánh giá kỹ thuật']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Đánh giá năng lực', 'Đánh giá kỹ thuật']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Đánh giá năng lực', 'Đánh giá kỹ thuật']
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Đánh giá năng lực', 'Đánh giá kỹ thuật']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Đánh giá năng lực', 'Đánh giá kỹ thuật']

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhGiaHSDT"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in bids:
            row_data = []
            if case_type == 'TU_VAN':
                row_data = [b[0], b[3], b[4], b[14], b[11], b[15] or '', b[16] or '', b[17] or '']
            elif case_type == '1G2T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[16] or '', b[17] or '']
            elif case_type == '1G2T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[16] or '', b[17] or '']
            elif case_type == '1G1T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[16] or '', b[17] or '']
            elif case_type == '1G1T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[16] or '', b[17] or '']
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border
            row_num += 1
            
        for col_idx, h in enumerate(headers, start=1):
            if h in ['Đánh giá hợp lệ', 'Đánh giá năng lực']:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1='"Đạt,Không đạt"', allow_blank=True)
                dv.error = 'Vui lòng chọn Đạt hoặc Không đạt!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn Đạt hoặc Không đạt'
                dv.promptTitle = h
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{row_num + 20}")
                
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        filename = f"Mau_Danh_Gia_HSDT_{package_name}.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def paginate_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
            
        params = request.query_params
        table_key = params.get("table")
        page = int(params.get("page", 1))
        page_size = int(params.get("pageSize", 10))
        search = params.get("search", "").strip().lower()
        
        TABLE_KEYS = {
            "chudautu": "chu_dau_tu",
            "kehoach": "ke_hoach_lcnt",
            "goithau": "goi_thau",
            "chuyengia": "chuyen_gia",
            "nhathau": "nha_thau",
            "hopdong": "hop_dong"
        }
        
        if table_key not in TABLE_KEYS:
            return JSONResponse({"error": "Invalid table key"}, status_code=400)
            
        table_name = TABLE_KEYS[table_key]
        org_name = get_active_org(request, role_or_err.user_id)
        
        # Build query
        query_parts = ["owner_id = ?"]
        query_params = [org_name]
        
        # Apply versioning filter for tables that support it
        versioned_tables = ["chu_dau_tu", "ke_hoach_lcnt", "goi_thau", "nha_thau"]
        if table_name in versioned_tables:
            query_parts.append("is_latest = 1")
            
        # Apply search filter
        if search:
            search_like = f"%{search}%"
            if table_name == "ke_hoach_lcnt":
                query_parts.append("(ma_ke_hoach LIKE ? OR ten_ke_hoach LIKE ? OR ten_du_an_du_toan LIKE ?)")
                query_params.extend([search_like, search_like, search_like])
            elif table_name == "goi_thau":
                query_parts.append("(ma_goi_thau LIKE ? OR ten_goi_thau LIKE ?)")
                query_params.extend([search_like, search_like])
            elif table_name == "chu_dau_tu":
                query_parts.append("(ma_chu_dau_tu LIKE ? OR ten_chu_dau_tu LIKE ? OR ma_so_thue LIKE ?)")
                query_params.extend([search_like, search_like, search_like])
            elif table_name == "nha_thau":
                query_parts.append("(ma_nha_thau LIKE ? OR ten_nha_thau LIKE ? OR ma_so_thue LIKE ?)")
                query_params.extend([search_like, search_like, search_like])
            elif table_name == "chuyen_gia":
                query_parts.append("(ho_ten LIKE ? OR so_cccd LIKE ? OR so_chung_chi LIKE ?)")
                query_params.extend([search_like, search_like, search_like])
            elif table_name == "hop_dong":
                query_parts.append("(so_hop_dong LIKE ? OR ten_hop_dong LIKE ?)")
                query_params.extend([search_like, search_like])
                
        # Apply specific filters (e.g. trangThai and hinhThuc for goi_thau)
        if table_name == "goi_thau":
            trang_thai = params.get("trangThai", "")
            hinh_thuc = params.get("hinhThuc", "")
            if trang_thai:
                query_parts.append("trang_thai = ?")
                query_params.append(trang_thai)
            if hinh_thuc:
                query_parts.append("hinh_thuc_lua_chon = ?")
                query_params.append(hinh_thuc)
                
        # Get count
        where_clause = " AND ".join(query_parts)
        conn = database.get_connection()
        cursor = conn.cursor()
        
        count_sql = f"SELECT COUNT(*) FROM {table_name} WHERE {where_clause}"
        cursor.execute(count_sql, tuple(query_params))
        total_items = cursor.fetchone()[0]
        
        # Get paginated items
        offset = (page - 1) * page_size
        items_sql = f"SELECT * FROM {table_name} WHERE {where_clause} LIMIT ? OFFSET ?"
        cursor.execute(items_sql, tuple(query_params + [page_size, offset]))
        rows = cursor.fetchall()
        
        # Map DB snake_case to JSON camelCase
        def map_db_to_json(tbl, row_dict):
            item = {}
            table_spec = SCHEMA_DINH_NGHIA[tbl]
            for col in table_spec["columns"].keys():
                json_key = SPECIAL_FIELD_MAPS.get(tbl, {}).get(col)
                if not json_key:
                    if col == "id_goc":
                        json_key = "rootId"
                    else:
                        json_key = to_camel_case(col)
                val = row_dict.get(col)
                
                # Prepend prefix
                if col == "id" or col.endswith("_id") or col == "id_goc":
                    if tbl != "phan_cong_nhan_su" and val is not None:
                        prefix = ""
                        if col == "id":
                            prefix_map = {
                                "chu_dau_tu": "cdt-",
                                "ke_hoach_lcnt": "kh-",
                                "goi_thau": "gt-",
                                "chuyen_gia": "cg-",
                                "nha_thau": "nt-",
                                "hop_dong": "hd-",
                                "thong_tin_mo_thau": "tm-"
                            }
                            prefix = prefix_map.get(tbl, "")
                        elif col == "chu_dau_tu_id":
                            prefix = "cdt-"
                        elif col == "ke_hoach_id":
                            prefix = "kh-"
                        elif col == "goi_thau_id":
                            prefix = "gt-"
                        elif col == "nha_thau_trung_thau_id" or col == "nha_thau_id":
                            prefix = "nt-"
                        elif col == "id_goc":
                            prefix_map = {
                                "ke_hoach_lcnt": "kh-",
                                "goi_thau": "gt-",
                                "chu_dau_tu": "cdt-",
                                "nha_thau": "nt-"
                            }
                            prefix = prefix_map.get(tbl, "")
                        val = f"{prefix}{val}"
                        
                is_json_field = col.endswith("_list") or col.startswith("cv_") or col == "thanh_vien_lien_danh"
                if is_json_field:
                    if val:
                        try:
                            val = json.loads(val)
                        except Exception:
                            val = []
                    else:
                        val = []
                item[json_key] = val
            return item
            
        items = []
        for row in rows:
            row_dict = dict(row)
            # handle base64 images for chuyengia
            if table_name == "chuyen_gia":
                row_dict["anh_chung_chi"] = load_base64_image(row_dict.get("anh_chung_chi", ""))
                row_dict["anh_chu_ky"] = load_base64_image(row_dict.get("anh_chu_ky", ""))
                
            item = map_db_to_json(table_name, row_dict)
            
            # Additional relationships for goithau/hopdong
            if table_name == "goi_thau":
                cg_ids = []
                if item.get("toChuyenGia"):
                    for x in item.get("toChuyenGia", []):
                        if isinstance(x, dict) and 'id' in x:
                            cg_ids.append(f"cg-{x['id']}")
                item["chuyenGiaIds"] = cg_ids
                for list_key in ["phanLoList", "tuyChonMuaThemList", "awardedPhanLoList", "toChuyenGia", "toThamDinh", "giaHanList", "yeuCauLamRoList", "traLoiLamRoList"]:
                    if item.get(list_key) is None:
                        item[list_key] = []
            elif table_name == "hop_dong":
                goithau_ids = []
                cursor.execute("SELECT goi_thau_id FROM hop_dong_goi_thau WHERE hop_dong_id = ?", (row_dict["id"],))
                for subrow in cursor.fetchall():
                    goithau_ids.append(f"gt-{subrow[0]}")
                item["goiThauIds"] = goithau_ids
                
            # If versioned table, query all versions for the dropdown
            if table_name in versioned_tables:
                root_col = "id_goc" if "id_goc" in row_dict and row_dict["id_goc"] else "id"
                root_val = row_dict.get("id_goc") or row_dict.get("id")
                cursor.execute(f"SELECT id, phien_ban FROM {table_name} WHERE owner_id = ? AND (id_goc = ? OR id = ?) ORDER BY CAST(phien_ban AS INTEGER) DESC", (org_name, root_val, root_val))
                versions = []
                for v_row in cursor.fetchall():
                    prefix_map = {
                        "chu_dau_tu": "cdt-",
                        "ke_hoach_lcnt": "kh-",
                        "goi_thau": "gt-",
                        "nha_thau": "nt-"
                    }
                    pfx = prefix_map.get(table_name, "")
                    versions.append({
                        "id": f"{pfx}{v_row[0]}",
                        "phienBan": v_row[1]
                    })
                item["allVersions"] = versions
                
            items.append(item)
            
        conn.close()
        return JSONResponse({
            "items": items,
            "totalItems": total_items
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

async def index(request):
    """
    [GET] /
    Trả về tệp index.html từ thư mục views.
    """
    return FileResponse(os.path.join(project_root, 'views', 'index.html'))

# ==========================================
# 4. KHAI BÁO PATH ROUTING & STATIC FILES
# ==========================================
routes = [
    Route("/", index, methods=["GET"]),
    Route("/api/sync", sync_api, methods=["POST"]),
    Route("/api/paginate", paginate_api, methods=["GET"]),
    Route("/api/get-all-data", get_all_data_api, methods=["GET"]),
    WebSocketRoute("/ws/sync", sync_websocket_endpoint),
    Route("/api/export-report/{package_id}", export_report_api, methods=["GET"]),
    Route("/api/templates", list_templates_api, methods=["GET"]),
    Route("/api/templates/active", set_active_template_api, methods=["POST"]),
    Route("/api/templates/upload", upload_template_api, methods=["POST"]),
    Route("/api/import-excel", import_excel_api, methods=["POST"]),
    Route("/api/export-excel-template/{import_type}", export_excel_template_api, methods=["GET"]),
    Route("/api/export-mothau-template", export_mothau_template_api, methods=["GET"]),
    Route("/api/export-danhgiahsdt-template", export_danhgiahsdt_template_api, methods=["GET"]),
    Route("/api/system-packages", list_system_packages_api, methods=["GET"]),
    Route("/api/system-packages/update", update_system_package_api, methods=["POST"]),
    
    # Auth Routes
    Route("/api/auth/register", register_api, methods=["POST"]),
    Route("/api/auth/verify", verify_email_api, methods=["POST"]),
    Route("/api/auth/resend-code", resend_code_api, methods=["POST"]),
    Route("/api/auth/login", login_api, methods=["POST"]),
    Route("/api/auth/check-session", check_session_api, methods=["POST"]),
    Route("/api/auth/forgot-password", forgot_password_api, methods=["POST"]),
    Route("/api/auth/update-profile", update_profile_api, methods=["POST"]),
    Route("/api/auth/change-password", change_password_api, methods=["POST"]),
    Route("/api/auth/users", list_users_api, methods=["GET"]),
    Route("/api/auth/users/{user_id}", delete_user_api, methods=["DELETE"]),
    Route("/api/auth/users/update-role", update_user_role_api, methods=["POST"]),
    Route("/api/auth/users/update-package", update_user_package_api, methods=["POST"]),
    Route("/api/auth/users/update-metadata", update_user_metadata_api, methods=["POST"]),
    
    # SPA Clean Paths Fallback to serve index.html for browser routes (Kebab-Case Standardized)
    Route("/tong-quan", index, methods=["GET"]),
    Route("/ke-hoach", index, methods=["GET"]),
    Route("/ke-hoach/{action}", index, methods=["GET"]),
    Route("/goi-thau", index, methods=["GET"]),
    Route("/goi-thau/{action}", index, methods=["GET"]),
    Route("/mothau", index, methods=["GET"]),
    Route("/mothau/{action}", index, methods=["GET"]),
    Route("/danh-gia-hsdt", index, methods=["GET"]),
    Route("/danh-gia-hsdt/{action}", index, methods=["GET"]),
    Route("/chu-dau-tu", index, methods=["GET"]),
    Route("/chu-dau-tu/{action}", index, methods=["GET"]),
    Route("/nha-thau", index, methods=["GET"]),
    Route("/nha-thau/{action}", index, methods=["GET"]),
    Route("/chuyen-gia", index, methods=["GET"]),
    Route("/chuyen-gia/{action}", index, methods=["GET"]),
    Route("/hop-dong", index, methods=["GET"]),
    Route("/hop-dong/{action}", index, methods=["GET"]),
    Route("/bieu-mau", index, methods=["GET"]),
    Route("/tong-quan-admin", index, methods=["GET"]),
    Route("/quan-ly-tai-khoan", index, methods=["GET"]),
    Route("/nhan-su", index, methods=["GET"]),
    Route("/trang-thai-ho-so", index, methods=["GET"]),
    Route("/trang-ca-nhan", index, methods=["GET"]),
    Route("/goi-thau-chi-tiet", index, methods=["GET"]),
    Route("/goi-thau-chi-tiet/{action}", index, methods=["GET"]),

    # Mount các thư mục MVC tĩnh để client có thể tải ES Modules
    Mount("/models", app=StaticFiles(directory=os.path.join(project_root, 'models')), name="models"),
    Mount("/views", app=StaticFiles(directory=os.path.join(project_root, 'views')), name="views"),
    Mount("/controllers", app=StaticFiles(directory=os.path.join(project_root, 'controllers')), name="controllers"),
    
    # Mount gốc views cho tệp index.html và style.css
    Mount("/", app=StaticFiles(directory=os.path.join(project_root, 'views'), html=True), name="static")
]

middleware = [
    Middleware(CORSMiddleware, allow_origins=['*'], allow_methods=['*'], allow_headers=['*']),
    Middleware(ErrorLoggingMiddleware)
]

# Cấu hình Host/Port và chế độ Debug qua file .env / biến môi trường
APP_HOST = os.environ.get("APP_HOST", "127.0.0.1")
APP_PORT = int(os.environ.get("APP_PORT", "8000"))
APP_DEBUG = os.environ.get("APP_DEBUG", "True").lower() == "true"

app = Starlette(debug=APP_DEBUG, routes=routes, middleware=middleware)

# ==========================================
# 5. KHỞI CHẠY MÁY CHỦ UVICORN
# ==========================================
if __name__ == "__main__":
    # Khởi chạy server sử dụng đường dẫn import dạng module chính xác 'controllers.app:app'
    uvicorn.run("controllers.app:app", host=APP_HOST, port=APP_PORT, reload=APP_DEBUG)

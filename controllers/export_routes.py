import os
import json
import secrets
from datetime import datetime
from io import BytesIO
import pandas as pd

from starlette.responses import StreamingResponse, JSONResponse

# Import helpers from helpers.py
from helpers import (
    database,
    verify_session,
    clean_id,
    VietnameseFloat,
    SCHEMA_DINH_NGHIA,
    SPECIAL_FIELD_MAPS,
    to_camel_case,
    get_active_org,
    load_base64_image
)

import custom_exporter

# ==========================================
# BIỂU MẪU CHUẨN HOÁ & NHẬP XUẤT EXCEL / WORD
# ==========================================

# =============================================================================
# ENTITY_SCHEMA — Nguồn dữ liệu duy nhất (single source of truth)
# Định nghĩa tất cả trường dữ liệu cho mỗi loại thực thể.
# Mỗi trường gồm:
#   - label   : Tiêu đề hiển thị trong file Excel xuất ra
#   - aliases : Danh sách tên cột có thể chấp nhận khi nhập vào (import)
#   - options : (Tuỳ chọn) Danh sách giá trị cho phép — tạo dropdown trong Excel
# Để thêm trường mới: chỉ cần thêm entry vào đây, cả export và import đều tự động cập nhật.
# =============================================================================
ENTITY_SCHEMA = {
    'chudautu': [
        {'field': 'maChuDauTu',           'label': 'Mã chủ đầu tư',              'aliases': ['Mã chủ đầu tư', 'Mã CĐT', 'maChuDauTu']},
        {'field': 'tenChuDauTu',           'label': 'Tên chủ đầu tư',             'aliases': ['Tên chủ đầu tư', 'Tên CĐT', 'tenChuDauTu']},
        {'field': 'maSoThue',              'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'chucVuNguoiDungDau',   'label': 'Chức vụ người đứng đầu',      'aliases': ['Chức vụ người đứng đầu', 'Chức vụ', 'chucVuNguoiDungDau']},
        {'field': 'nguoiKyQuyetDinh',     'label': 'Người ký quyết định',          'aliases': ['Người ký quyết định', 'Người ký', 'nguoiKyQuyetDinh']},
        {'field': 'chucVuNguoiKy',        'label': 'Chức vụ người ký',             'aliases': ['Chức vụ người ký', 'chucVuNguoiKy']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'Ông/Bà', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'diaChi',               'label': 'Địa chỉ trụ sở',              'aliases': ['Địa chỉ', 'Địa chỉ trụ sở', 'diaChi']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'STK', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'Ngân hàng', 'noiMoTaiKhoan']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'Địa chỉ email', 'email']},
        {'field': 'maQHNS',               'label': 'Mã QHNS',                    'aliases': ['Mã QHNS', 'maQHNS']},
    ],
    'kehoach': [
        {'field': 'maKeHoach',            'label': 'Mã kế hoạch',                'aliases': ['Mã kế hoạch', 'maKeHoach']},
        {'field': 'maDuan',               'label': 'Mã dự án',                   'aliases': ['Mã dự án', 'maDuan']},
        {'field': 'tenKeHoach',           'label': 'Tên kế hoạch',               'aliases': ['Tên kế hoạch', 'tenKeHoach']},
        {'field': 'loaiHinhMuaSam',       'label': 'Loại hình',                   'aliases': ['Loại hình', 'Loại hình mua sắm', 'loaiHinhMuaSam'],
                                           'options': ['Dự án', 'Dự toán mua sắm']},
        {'field': 'tenDuAnDuToan',        'label': 'Tên dự án',                  'aliases': ['Tên dự án', 'Dự án', 'Dự án / Dự toán', 'tenDuAnDuToan']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'Mã chủ đầu tư', 'chuDauTuId']},
        {'field': 'tongMucDauTu',         'label': 'Tổng mức đầu tư',             'aliases': ['Tổng giá trị', 'Tổng mức đầu tư', 'tongMucDauTu']},
        {'field': 'ngayPheDuyet',         'label': 'Ngày phê duyệt',               'aliases': ['Ngày phê duyệt', 'ngayPheDuyet']},
        {'field': 'quyetDinhPheDuyet',    'label': 'Số quyết định',               'aliases': ['Số quyết định', 'Quyết định phê duyệt', 'quyetDinhPheDuyet']},
        {'field': 'thoiGianDangMa',       'label': 'Thời gian đăng',               'aliases': ['Thời gian đăng', 'Thời gian đăng mã', 'thoiGianDangMa']},
        {'field': 'soQdPheDuyetDuAn',     'label': 'Số QĐ phê duyệt dự án',       'aliases': ['Số QĐ phê duyệt dự án', 'soQdPheDuyetDuAn']},
        {'field': 'ngayQdPheDuyetDuAn',   'label': 'Ngày QĐ phê duyệt dự án',     'aliases': ['Ngày QĐ phê duyệt dự án', 'ngayQdPheDuyetDuAn']},
        {'field': 'coQuanPheDuyetDuAn',   'label': 'Cơ quan phê duyệt dự án',     'aliases': ['Cơ quan phê duyệt dự án', 'coQuanPheDuyetDuAn']},
    ],
    'goithau': [
        {'field': 'maGoiThau',            'label': 'Mã gói thầu',                 'aliases': ['Mã gói thầu', 'maGoiThau']},
        {'field': 'tenGoiThau',           'label': 'Tên gói thầu',                'aliases': ['Tên gói thầu', 'tenGoiThau']},
        {'field': 'keHoachId',            'label': 'Kế hoạch',                    'aliases': ['Kế hoạch', 'Kế hoạch liên kết', 'keHoachId']},
        {'field': 'giaGoiThau',           'label': 'Giá gói thầu',               'aliases': ['Giá gói thầu', 'Giá gói', 'giaGoiThau']},
        {'field': 'hinhThucLuaChon',      'label': 'Hình thức',                   'aliases': ['Hình thức', 'Hình thức lựa chọn', 'hinhThucLuaChon'],
                                           'options': ['Đấu thầu rộng rãi', 'Đấu thầu hạn chế', 'Chỉ định thầu', 'Chỉ định thầu rút gọn', 'Chào hàng cạnh tranh', 'Lựa chọn nhà thầu trong trường hợp đặc biệt']},
        {'field': 'phuongThucLuaChon',    'label': 'Phương thức',                 'aliases': ['Phương thức', 'Phương thức lựa chọn', 'phuongThucLuaChon'],
                                           'options': ['Một giai đoạn một túi hồ sơ', 'Một giai đoạn hai túi hồ sơ', 'Hai giai đoạn một túi hồ sơ', 'Hai giai đoạn hai túi hồ sơ', 'Không có']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
        {'field': 'trangThai',            'label': 'Trạng thái',                   'aliases': ['Trạng thái', 'trangThai'],
                                           'options': ['Chuẩn bị', 'Đang mời thầu', 'Đã mở thầu', 'Đang chấm thầu', 'Đã có kết quả', 'Hủy thầu']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'nguonVon',             'label': 'Nguồn vốn',                   'aliases': ['Nguồn vốn', 'nguonVon']},
        {'field': 'linhVuc',              'label': 'Lĩnh vực',                    'aliases': ['Lĩnh vực', 'linhVuc'],
                                           'options': ['Tư vấn', 'Phi tư vấn', 'Xây lắp', 'Hỗn hợp', 'Hàng hóa']},
        {'field': 'tuyChonMuaThem',       'label': 'Tùy chọn mua thêm',           'aliases': ['Tùy chọn mua thêm', 'tuyChonMuaThem'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianToChuc',       'label': 'Thời gian tổ chức',            'aliases': ['Thời gian tổ chức', 'thoiGianToChuc']},
        {'field': 'thoiGianBatDauToChuc', 'label': 'Thời gian bắt đầu tổ chức',  'aliases': ['Thời gian bắt đầu tổ chức', 'thoiGianBatDauToChuc']},
        {'field': 'quaMang',              'label': 'Qua mạng / Trực tiếp',        'aliases': ['Qua mạng / Trực tiếp', 'Qua mạng', 'quaMang'],
                                           'options': ['Qua mạng', 'Trực tiếp']},
        {'field': 'trongNuocQuocTe',      'label': 'Trong nước / Quốc tế',       'aliases': ['Trong nước / Quốc tế', 'Trong nước', 'trongNuocQuocTe'],
                                           'options': ['Trong nước', 'Quốc tế']},
        {'field': 'phanLo',               'label': 'Phân lô / Không phân lô',     'aliases': ['Phân lô / Không phân lô', 'Phân lô', 'phanLo'],
                                           'options': ['Có', 'Không']},
        {'field': 'thoiGianDangTai',      'label': 'Thời gian đăng tải',          'aliases': ['Thời gian đăng tải', 'thoiGianDangTai']},
        {'field': 'thoiGianDongThau',     'label': 'Thời gian đóng thầu',         'aliases': ['Thời gian đóng thầu', 'thoiGianDongThau']},
        {'field': 'thoiGianMoThau',       'label': 'Thời gian mở thầu',           'aliases': ['Thời gian mở thầu', 'thoiGianMoThau']},
        {'field': 'soQuyetDinh',          'label': 'Số quyết định phê duyệt',     'aliases': ['Số quyết định phê duyệt', 'Số quyết định', 'soQuyetDinh']},
        {'field': 'ngayQuyetDinh',        'label': 'Ngày quyết định phê duyệt',   'aliases': ['Ngày quyết định phê duyệt', 'Ngày quyết định', 'ngayQuyetDinh']},
        {'field': 'soQuyetDinhKetQua',    'label': 'Số quyết định phê duyệt kết quả LCNT', 'aliases': ['Số quyết định phê duyệt kết quả LCNT', 'Số quyết định kết quả', 'soQuyetDinhKetQua']},
        {'field': 'ngayQuyetDinhKetQua',  'label': 'Ngày ký quyết định kết quả LCNT',     'aliases': ['Ngày ký quyết định kết quả LCNT', 'Ngày quyết định kết quả', 'ngayQuyetDinhKetQua']},
        {'field': 'nhaThauTrungThauId',   'label': 'Nhà thầu trúng thầu',          'aliases': ['Nhà thầu trúng thầu', 'Nhà thầu', 'nhaThauTrungThauId']},
        {'field': 'giaTrungThau',         'label': 'Giá trúng thầu',               'aliases': ['Giá trúng thầu', 'giaTrungThau']},
        {'field': 'thoiGianGoiThau',      'label': 'Thời gian gói thầu',          'aliases': ['Thời gian gói thầu', 'thoiGianGoiThau']},
        {'field': 'thoiGianHopDong',      'label': 'Thời gian hợp đồng',          'aliases': ['Thời gian hợp đồng', 'thoiGianHopDong']},
        {'field': 'giaHanList',           'label': 'Gia hạn thời điểm đóng thầu',  'aliases': ['Gia hạn thời điểm đóng thầu', 'giaHanList']},
        {'field': 'yeuCauLamRoList',      'label': 'Yêu cầu làm rõ HSMT',          'aliases': ['Yêu cầu làm rõ HSMT', 'yeuCauLamRoList']},
        {'field': 'traLoiLamRoList',      'label': 'Trả lời làm rõ HSMT',          'aliases': ['Trả lời làm rõ HSMT', 'traLoiLamRoList']},
        {'field': 'giaToDamBaoDuThau',    'label': 'Giá trị bảo đảm dự thầu',      'aliases': ['Giá trị bảo đảm dự thầu', 'giaToDamBaoDuThau']},
        {'field': 'hieuLucHsdtGoiThau',   'label': 'Hiệu lực HSDT',                'aliases': ['Hiệu lực HSDT', 'hieuLucHsdtGoiThau']},
        {'field': 'hieuLucDamBaoDuThau',  'label': 'Hiệu lực bảo đảm dự thầu',     'aliases': ['Hiệu lực bảo đảm dự thầu', 'hieuLucDamBaoDuThau']},
        {'field': 'awardedPhanLoList',    'label': 'Danh sách phân lô trúng thầu', 'aliases': ['Danh sách phân lô trúng thầu', 'awardedPhanLoList']},
    ],
    'nhathau': [
        {'field': 'loaiNhaThau',          'label': 'Loại nhà thầu',               'aliases': ['Loại nhà thầu', 'loaiNhaThau'],
                                           'options': ['Độc lập', 'Liên danh']},
        {'field': 'maNhaThau',            'label': 'Mã nhà thầu',                 'aliases': ['Mã nhà thầu', 'Mã định danh', 'Mã nhà thầu', 'maNhaThau']},
        {'field': 'tenNhaThau',           'label': 'Tên nhà thầu',                'aliases': ['Tên nhà thầu', 'tenNhaThau']},
        {'field': 'maSoThue',             'label': 'Mã số thuế',                  'aliases': ['Mã số thuế', 'MST', 'maSoThue']},
        {'field': 'nguoiDaiDien',         'label': 'Người đại diện',              'aliases': ['Người đại diện', 'nguoiDaiDien']},
        {'field': 'danhXung',             'label': 'Danh xưng',                   'aliases': ['Danh xưng', 'danhXung'],
                                           'options': ['Ông', 'Bà']},
        {'field': 'soDienThoai',          'label': 'Số điện thoại',               'aliases': ['Số điện thoại', 'SĐT', 'soDienThoai']},
        {'field': 'email',                'label': 'Email',                       'aliases': ['Email', 'email']},
        {'field': 'diaChi',               'label': 'Địa chỉ',                    'aliases': ['Địa chỉ', 'diaChi']},
        {'field': 'soTaiKhoan',           'label': 'Số tài khoản',               'aliases': ['Số tài khoản', 'soTaiKhoan']},
        {'field': 'noiMoTaiKhoan',        'label': 'Nơi mở tài khoản',            'aliases': ['Nơi mở tài khoản', 'noiMoTaiKhoan']},
    ],
    'chuyengia': [
        {'field': 'hoTen',                'label': 'Họ tên',                     'aliases': ['Họ tên', 'Họ và tên', 'hoTen']},
        {'field': 'soCCCD',               'label': 'Số CCCD',                     'aliases': ['Số CCCD', 'CCCD', 'soCCCD']},
        {'field': 'ngayCapCCCD',          'label': 'Ngày cấp CCCD',               'aliases': ['Ngày cấp CCCD', 'ngayCapCCCD']},
        {'field': 'noiCapCCCD',           'label': 'Nơi cấp CCCD',                'aliases': ['Nơi cấp CCCD', 'noiCapCCCD']},
        {'field': 'soChungChi',           'label': 'Số chứng chỉ',               'aliases': ['Số chứng chỉ', 'Số chứng chỉ đấu thầu', 'soChungChi']},
        {'field': 'ngayCapChungChi',      'label': 'Ngày cấp chứng chỉ',         'aliases': ['Ngày cấp', 'Ngày cấp chứng chỉ', 'ngayCapChungChi']},
        {'field': 'donViCapChungChi',     'label': 'Đơn vị cấp chứng chỉ',       'aliases': ['Đơn vị cấp', 'Đơn vị cấp chứng chỉ', 'donViCapChungChi']},
    ],
    'hopdong': [
        {'field': 'soHopDong',            'label': 'Số hợp đồng',                'aliases': ['Số hợp đồng', 'soHopDong']},
        {'field': 'tenHopDong',           'label': 'Tên hợp đồng',               'aliases': ['Tên hợp đồng', 'tenHopDong']},
        {'field': 'ngayKy',               'label': 'Ngày ký',                    'aliases': ['Ngày ký', 'Ngày ký hợp đồng', 'ngayKy']},
        {'field': 'chuDauTuId',           'label': 'Chủ đầu tư',                  'aliases': ['Chủ đầu tư', 'chuDauTuId']},
        {'field': 'nhaThauId',            'label': 'Nhà thầu',                    'aliases': ['Nhà thầu', 'nhaThauId']},
        {'field': 'giaTri',               'label': 'Giá trị hợp đồng',          'aliases': ['Giá trị', 'Giá trị hợp đồng', 'giaTri']},
        {'field': 'loaiHopDong',          'label': 'Loại hợp đồng',               'aliases': ['Loại hợp đồng', 'loaiHopDong'],
                                           'options': ['Trọn gói', 'Theo đơn giá cố định', 'Theo đơn giá điều chỉnh', 'Theo thời gian', 'Hợp đồng theo tỷ lệ phần trăm', 'Hợp đồng hỗn hợp']},
        {'field': 'soNgayThucHien',       'label': 'Thời gian thực hiện hợp đồng', 'aliases': ['Thời gian thực hiện hợp đồng', 'Thời gian thực hiện', 'Số ngày thực hiện', 'Số ngày', 'soNgayThucHien']},
        {'field': 'goiThauIds',           'label': 'Gói thầu liên kết',            'aliases': ['Gói thầu liên kết', 'Gói thầu', 'goiThauIds']},
    ],
    'phanlo': [
        {'field': 'tenPhanLo',            'label': 'Tên phần lô',                'aliases': ['Tên phần lô', 'Tên phân lô', 'tenPhanLo', 'Tên']},
        {'field': 'giaTriPhanLo',         'label': 'Giá trị phần lô',             'aliases': ['Giá trị phần lô', 'Giá trị phân lô', 'Giá trị', 'giaTriPhanLo']},
        {'field': 'thoiGianThucHien',     'label': 'Thời gian thực hiện',          'aliases': ['Thời gian thực hiện', 'Thời gian', 'thoiGianThucHien']},
    ],
    'tuychonmuathem': [
        {'field': 'hangMuc',              'label': 'Hạng mục',                   'aliases': ['Hạng mục', 'Tên hạng mục', 'hangMuc']},
        {'field': 'donVi',                'label': 'Đơn vị',                    'aliases': ['Đơn vị', 'Đơn vị tính', 'ĐVT', 'donVi']},
        {'field': 'soLuong',              'label': 'Khối lượng / Số lượng',       'aliases': ['Khối lượng/ Số lượng', 'Khối lượng', 'Số lượng', 'soLuong', 'khoiLuong']},
        {'field': 'tyLe',                 'label': 'Tỷ lệ phần trăm (%)',         'aliases': ['Tỷ lệ phần trăm (%)', 'Tỷ lệ phần trăm', 'Tỷ lệ (%)', 'Tỷ lệ', 'tyLe', 'phanTram']},
        {'field': 'giaTriUocTinh',        'label': 'Giá trị ước tính',           'aliases': ['Giá trị ước tính', 'Giá trị', 'giaTriUocTinh']},
    ],
    'ketquaqd': [
        {'field': 'maNhaThau',            'label': 'Mã nhà thầu',                 'aliases': ['Mã nhà thầu', 'Mã định danh', 'Mã số thuế', 'Mã', 'maNhaThau']},
        {'field': 'tenNhaThau',           'label': 'Tên nhà thầu',                'aliases': ['Tên nhà thầu', 'Nhà thầu', 'tenNhaThau']},
        {'field': 'trangThai',            'label': 'Trúng thầu/Trượt thầu',       'aliases': ['Trúng thầu/Trượt thầu', 'Trạng thái', 'Kết quả', 'trangThai'],
                                           'options': ['Trúng thầu', 'Trượt thầu']},
        {'field': 'lyDoTruot',            'label': 'Lý do trượt',                 'aliases': ['Lý do trượt', 'Lý do trượt thầu', 'lyDoTruot']},
        {'field': 'giaTrungThau',         'label': 'Giá trúng thầu',               'aliases': ['Giá trúng thầu', 'Giá trúng', 'Giá trúng thầu (VND)', 'giaTrungThau']},
        {'field': 'thoiGianGoiThau',      'label': 'Thời gian thực hiện gói thầu', 'aliases': ['Thời gian thực hiện gói thầu', 'Thời gian gói', 'thoiGianGoiThau']},
        {'field': 'thoiGianHopDong',      'label': 'Thời gian thực hiện hợp đồng', 'aliases': ['Thời gian thực hiện hợp đồng', 'Thời gian hợp đồng', 'thoiGianHopDong']}
    ],
}

def _schema_to_map_cols(entity_type):
    """Chuyển ENTITY_SCHEMA thành map_cols dùng cho import (field → aliases)."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return {entry['field']: entry['aliases'] for entry in schema}

def _schema_to_headers(entity_type):
    """Lấy danh sách label (tiêu đề cột) cho template Excel."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return None
    return [entry['label'] for entry in schema]

def _schema_to_options(entity_type):
    """Lấy options_map (label → danh sách giá trị dropdown) cho validation Excel."""
    schema = ENTITY_SCHEMA.get(entity_type)
    if not schema:
        return {}
    return {entry['label']: entry['options'] for entry in schema if entry.get('options')}

async def export_report_api(request):
    """
    [GET] /api/export-report/{package_id}
    Xuất báo cáo đánh giá hồ sơ mời thầu ra file Word (.docx).
    """
    package_id = clean_id(request.path_params.get('package_id'))
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        conn = database.get_connection()
        cursor = conn.cursor()
        
        org_name = get_active_org(request, user_id)

        cursor.execute("SELECT * FROM goi_thau WHERE id = ? AND owner_id = ?", (package_id, org_name))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return JSONResponse({"error": f"Package with id {package_id} not found"}, status_code=404)
        pkg = dict(row)
        
        cursor.execute("SELECT * FROM ke_hoach_lcnt WHERE id = ? AND owner_id = ?", (pkg['ke_hoach_id'], org_name))
        row_plan = cursor.fetchone()
        if not row_plan:
            conn.close()
            return JSONResponse({"error": f"Plan linked to package not found"}, status_code=404)
        plan = dict(row_plan)
        
        investor_name = '--'
        investor_address = ''
        if plan.get('chu_dau_tu_id'):
            cursor.execute("SELECT * FROM chu_dau_tu WHERE id = ?", (plan['chu_dau_tu_id'],))
            row_inv = cursor.fetchone()
            if row_inv:
                inv_data = dict(row_inv)
                investor_name = inv_data.get('ten_chu_dau_tu', '--')
                investor_address = inv_data.get('dia_chi', '')
                
        expert_ids = []
        cg_meta = {}
        if pkg.get('chuyen_gia_list'):
            try:
                cg_list = json.loads(pkg['chuyen_gia_list'])
                for x in cg_list:
                    cg_id = x.get('chuyenGiaId') or x.get('id')
                    if cg_id:
                        norm_id = str(cg_id).replace("cg-", "")
                        expert_ids.append(norm_id)
                        cg_meta[norm_id] = {
                            'chuc_vu': x.get('chucVu') or x.get('chuc_vu') or 'Tổ viên',
                            'cong_viec': x.get('congViec') or x.get('cong_viec') or ''
                        }
            except Exception:
                pass
                
        chuyen_gia_list = []
        if expert_ids:
            placeholders = ",".join(["?"] * len(expert_ids))
            cursor.execute(f"SELECT * FROM chuyen_gia WHERE id IN ({placeholders})", expert_ids)
            for row_cg in cursor.fetchall():
                cg = dict(row_cg)
                cg_id_str = str(cg.get('id', ''))
                meta = cg_meta.get(cg_id_str, {})
                cg['chuc_vu'] = meta.get('chuc_vu', 'Tổ viên')
                cg['cong_viec'] = meta.get('cong_viec', '')
                chuyen_gia_list.append(cg)
                
        # Resolve Tổ thẩm định (Appraisal Team)
        appraisal_ids = []
        td_meta = {}
        if pkg.get('tham_dinh_list'):
            try:
                td_list = json.loads(pkg['tham_dinh_list'])
                for x in td_list:
                    td_id = x.get('chuyenGiaId') or x.get('id')
                    if td_id:
                        norm_id = str(td_id).replace("cg-", "")
                        appraisal_ids.append(norm_id)
                        td_meta[norm_id] = {
                            'chuc_vu': x.get('chucVu') or x.get('chuc_vu') or 'Tổ viên',
                            'cong_viec': x.get('congViec') or x.get('cong_viec') or ''
                        }
            except Exception:
                pass
                
        tham_dinh_list = []
        if appraisal_ids:
            placeholders = ",".join(["?"] * len(appraisal_ids))
            cursor.execute(f"SELECT * FROM chuyen_gia WHERE id IN ({placeholders})", appraisal_ids)
            for row_td in cursor.fetchall():
                td = dict(row_td)
                td_id_str = str(td.get('id', ''))
                meta = td_meta.get(td_id_str, {})
                td['chuc_vu'] = meta.get('chuc_vu', 'Tổ viên')
                td['cong_viec'] = meta.get('cong_viec', '')
                tham_dinh_list.append(td)
                
        awarded_id = pkg.get('nha_thau_trung_thau_id')
        if not awarded_id and pkg.get('trang_thai') == 'Đã có kết quả':
            cursor.execute("""
                SELECT nha_thau_id FROM thong_tin_mo_thau 
                WHERE goi_thau_id = ? 
                  AND (danh_gia_ket_luan = 'Đạt' 
                       OR (danh_gia_hop_le = 'Đạt' AND danh_gia_nang_luc = 'Đạt' AND (danh_gia_ky_thuat = 'Đạt' OR danh_gia_ky_thuat = '')))
                LIMIT 1
            """, (package_id,))
            healed_row = cursor.fetchone()
            if healed_row:
                awarded_id = healed_row[0]
            else:
                cursor.execute("SELECT nha_thau_id FROM thong_tin_mo_thau WHERE goi_thau_id = ? LIMIT 1", (package_id,))
                healed_row = cursor.fetchone()
                if healed_row:
                    awarded_id = healed_row[0]

        nha_thau_list = []
        if awarded_id:
            cursor.execute("SELECT * FROM nha_thau WHERE id = ?", (awarded_id,))
            row_nt = cursor.fetchone()
            if row_nt:
                nt = dict(row_nt)
                
                # Cố gắng lấy thông tin liên danh thực tế được ghi nhận trong Biên bản mở thầu
                cursor.execute(
                    "SELECT ten_nha_thau, loai_nha_thau, thanh_vien_lien_danh FROM thong_tin_mo_thau WHERE goi_thau_id = ? AND nha_thau_id = ?",
                    (pkg['id'], awarded_id)
                )
                row_mt = cursor.fetchone()
                
                nt_name = nt['ten_nha_thau']
                nt_type = nt.get('loai_nha_thau', 'Độc lập')
                members_parsed = []
                
                if row_mt:
                    mt_data = dict(row_mt)
                    if mt_data.get('ten_nha_thau'):
                        nt_name = mt_data['ten_nha_thau']
                    if mt_data.get('loai_nha_thau'):
                        nt_type = mt_data['loai_nha_thau']
                    if mt_data.get('thanh_vien_lien_danh'):
                        try:
                            members_parsed = json.loads(mt_data['thanh_vien_lien_danh'])
                        except Exception:
                            members_parsed = []
                else:
                    # Hỗ trợ tương thích ngược nếu không tìm thấy dòng mở thầu
                    if nt.get('loai_nha_thau') == "Liên danh" and nt.get('thanh_vien_lien_danh'):
                        try:
                            members_parsed = json.loads(nt['thanh_vien_lien_danh'])
                        except Exception:
                            members_parsed = [nt['thanh_vien_lien_danh']]
                
                nt['ten_nha_thau'] = nt_name
                nt['loai_nha_thau'] = nt_type
                nt['members'] = members_parsed
                nt['awarded_price'] = pkg.get('gia_goi_thau', 0)
                nha_thau_list.append(nt)
        
        # Fetch contract details if any
        cursor.execute("""
            SELECT hd.* FROM hop_dong hd
            INNER JOIN hop_dong_goi_thau hdgt ON hd.id = hdgt.hop_dong_id
            WHERE hdgt.goi_thau_id = ? AND hd.owner_id = ?
        """, (package_id, org_name))
        row_hd = cursor.fetchone()
        contract_data = dict(row_hd) if row_hd else {}

        # Fetch first expert details
        first_expert_data = {}
        if expert_ids:
            cursor.execute("SELECT * FROM chuyen_gia WHERE id = ?", (expert_ids[0],))
            row_cg = cursor.fetchone()
            if row_cg:
                first_expert_data = dict(row_cg)
                
        # Fetch thong_tin_mo_thau for the winning contractor
        mt_data = {}
        if awarded_id:
            cursor.execute(
                "SELECT * FROM thong_tin_mo_thau WHERE goi_thau_id = ? AND nha_thau_id = ?",
                (package_id, awarded_id)
            )
            row_mt = cursor.fetchone()
            if row_mt:
                mt_data = dict(row_mt)

        # Fetch current user data
        cursor.execute("SELECT * FROM tai_khoan WHERE id = ?", (user_id,))
        row_user = cursor.fetchone()
        user_data = dict(row_user) if row_user else {}
        
        # Fetch organization data
        cursor.execute("SELECT * FROM to_chuc WHERE ten_to_chuc = ?", (org_name,))
        row_org = cursor.fetchone()
        org_data = dict(row_org) if row_org else {}
        
        # Fetch service package data
        gdv_data = {}
        if user_data.get('goi_dich_vu_id'):
            cursor.execute("SELECT * FROM goi_dich_vu WHERE id = ?", (user_data['goi_dich_vu_id'],))
            row_gdv = cursor.fetchone()
            if row_gdv:
                gdv_data = dict(row_gdv)

        # Fetch custom Word mapping configurations
        cursor.execute("SELECT ten_bien, source_table, source_column FROM cau_hinh_bien_word WHERE owner_id = ?", (org_name,))
        mappings_rows = cursor.fetchall()
        
        row_by_table = {
            'chu_dau_tu': inv_data if 'inv_data' in locals() else {},
            'ke_hoach_lcnt': plan,
            'goi_thau': pkg,
            'nha_thau': nt if (awarded_id and 'nt' in locals()) else {},
            'hop_dong': contract_data,
            'chuyen_gia': first_expert_data,
            'thong_tin_mo_thau': mt_data,
            'tai_khoan': user_data,
            'to_chuc': org_data,
            'goi_dich_vu': gdv_data
        }
        
        custom_vars_list = []
        custom_evaluated_values = {}
        for m_row in mappings_rows:
            ten_bien = m_row[0].lower()
            src_table = m_row[1]
            src_column = m_row[2]
            
            custom_vars_list.append(ten_bien)
            
            tbl_data = row_by_table.get(src_table, {})
            val = tbl_data.get(src_column)
            if src_table == 'goi_thau' and src_column in ['gia_han_list', 'yeu_cau_lam_ro_list', 'tra_loi_lam_ro_list'] and val:
                try:
                    parsed_list = json.loads(val) if isinstance(val, str) else val
                    if isinstance(parsed_list, list):
                        formatted_items = []
                        for idx, item in enumerate(parsed_list):
                            if src_column == 'gia_han_list':
                                tg = item.get('thoiGianDongThau', '')
                                ld = item.get('lyDoGiaHan', '')
                                formatted_items.append(f'Lần {idx+1}: Gia hạn đến {tg} (Lý do: {ld})')
                            elif src_column == 'yeu_cau_lam_ro_list':
                                tg = item.get('thoiGianYeuCau', '')
                                nd = item.get('noiDungYeuCau', '')
                                formatted_items.append(f'Lần {idx+1}: Yêu cầu làm rõ lúc {tg} (Nội dung: {nd})')
                            elif src_column == 'tra_loi_lam_ro_list':
                                tg = item.get('thoiGianTraLoi', '')
                                nd = item.get('noiDungTraLoi', '')
                                formatted_items.append(f'Lần {idx+1}: Trả lời làm rõ lúc {tg} (Nội dung: {nd})')
                        val = '\n'.join(formatted_items) if formatted_items else '--'
                except Exception:
                    pass

            if val is None:
                val = '--'
            elif isinstance(val, (int, float)) and ('gia' in src_column or 'tong_muc' in src_column or 'gia_tri' in src_column):
                val = f'{VietnameseFloat(val)} VND'
            elif isinstance(val, (int, float)):
                val = str(val)
            else:
                val = str(val)

            custom_evaluated_values[ten_bien] = val

        conn.close()

        context = {
            'chuyen_gia': chuyen_gia_list,
            'tham_dinh': tham_dinh_list,
            'nha_thau': nha_thau_list
        }

        # Parse lists for jinja2 templates
        for key_name in ['gia_han_list', 'yeu_cau_lam_ro_list', 'tra_loi_lam_ro_list']:
            try:
                raw_val = pkg.get(key_name)
                context[key_name] = json.loads(raw_val) if isinstance(raw_val, str) else (raw_val or [])
            except:
                context[key_name] = []
        active_tpl = custom_exporter.get_active_template(user_id)
        if request.query_params.get('type') == 'contract':
            active_tpl = 'mau_hop_dong_lcnt.docx'
            
        # Map custom dictionary variables
        custom_context = {}
        
        custom_context['Danh_Sach_Chuyen_Gia'] = []
        for idx, cg in enumerate(chuyen_gia_list):
            item = {'STT': idx + 1}
            for k, v in cg.items():
                item[k.lower()] = v
            for m_row in mappings_rows:
                ten_bien = m_row[0].lower()
                src_table = m_row[1]
                src_column = m_row[2]
                if src_table == 'chuyen_gia':
                    val = cg.get(src_column)
                    if val is None:
                        val = '--'
                    elif isinstance(val, (int, float)):
                        val = str(val)
                    else:
                        val = str(val)
                    item[ten_bien] = val
            custom_context['Danh_Sach_Chuyen_Gia'].append(item)
            
        custom_context['Danh_Sach_Tham_Dinh'] = []
        for idx, td in enumerate(tham_dinh_list):
            item = {'STT': idx + 1}
            for k, v in td.items():
                item[k.lower()] = v
            for m_row in mappings_rows:
                ten_bien = m_row[0].lower()
                src_table = m_row[1]
                src_column = m_row[2]
                if src_table == 'chuyen_gia':
                    val = td.get(src_column)
                    if val is None:
                        val = '--'
                    elif isinstance(val, (int, float)):
                        val = str(val)
                    else:
                        val = str(val)
                    item[ten_bien] = val
            custom_context['Danh_Sach_Tham_Dinh'].append(item)
            
        custom_context['Danh_Sach_Nha_Thau'] = []
        for idx, nt in enumerate(nha_thau_list):
            item = {'STT': idx + 1}
            for k, v in nt.items():
                item[k.lower()] = v
            for m_row in mappings_rows:
                ten_bien = m_row[0].lower()
                src_table = m_row[1]
                src_column = m_row[2]
                if src_table == 'nha_thau':
                    val = nt.get(src_column)
                    if val is None:
                        val = '--'
                    elif isinstance(val, (int, float)):
                        val = str(val)
                    else:
                        val = str(val)
                    item[ten_bien] = val
            custom_context['Danh_Sach_Nha_Thau'].append(item)

        # Phan lo list
        custom_context['Danh_Sach_Phan_Lo'] = []
        try:
            pl_list = json.loads(pkg.get('phan_lo_list') or '[]')
            # Create a contractor ID -> Name lookup
            nt_name_map = {str(nt.get('id', '')): nt.get('ten_nha_thau', '--') for nt in nha_thau_list}
            for idx, pl in enumerate(pl_list):
                # Resolve contractor name
                nt_id = str(pl.get('nhaThauTrungThauId', ''))
                nt_name = nt_name_map.get(nt_id, pl.get('nhaThauTrungThauName', '--'))
                if nt_name == '--' and pl.get('tenNhaThau'):
                    nt_name = pl.get('tenNhaThau')
                
                # Resolve value
                raw_val = pl.get('giaTrungThau') or pl.get('giaTriPhanLo') or 0
                
                custom_context['Danh_Sach_Phan_Lo'].append({
                    'STT': idx + 1,
                    'Ten_Phan_Lo': pl.get('tenPhanLo', '--'),
                    'Gia_Tri_Phan_Lo': f"{VietnameseFloat(raw_val)} VND" if raw_val else '0 VND',
                    'Gia_Tri_So': float(raw_val),
                    'Nha_Thau_Trung': nt_name,
                    'Thoi_Gian_Thuc_Hien': pl.get('thoiGianHopDong') or pl.get('thoiGianThucHien') or '--'
                })
        except Exception:
            pass

        # Tuy chon mua them list
        custom_context['Danh_Sach_Tuy_Chon_Mua_Them'] = []
        try:
            tc_list = json.loads(pkg.get('tuy_chon_mua_them_list') or '[]')
            for idx, tc in enumerate(tc_list):
                custom_context['Danh_Sach_Tuy_Chon_Mua_Them'].append({
                    'STT': idx + 1,
                    'Hang_Muc': tc.get('hangMuc', '--'),
                    'Don_Vi': tc.get('donVi', '--'),
                    'So_Luong': str(tc.get('soLuong', 0)),
                    'Ty_Le': str(tc.get('tyLe', 0)),
                    'Gia_Tri_Uoc_Tinh': f"{VietnameseFloat(tc.get('giaTriUocTinh', 0))} VND" if tc.get('giaTriUocTinh') else '0 VND'
                })
        except Exception:
            pass
            
        # Danh sach nha thau truot thau
        custom_context['Danh_Sach_Nha_Thau_Truot'] = []
        try:
            # 1. Thu thap tat ca ID nha thau trung thau cua goi thau (ke ca phan lo)
            winning_ids = set()
            if awarded_id:
                winning_ids.add(str(awarded_id))
            
            if pkg.get('phan_lo') == 'Có':
                try:
                    pl_list = json.loads(pkg.get('phan_lo_list') or '[]')
                    for pl in pl_list:
                        nt_id = pl.get('nhaThauTrungThauId')
                        if nt_id:
                            winning_ids.add(str(nt_id))
                except Exception:
                    pass

            # 2. Lay toan bo ho so du thau va nhom theo tung nha thau
            cursor.execute("SELECT * FROM thong_tin_mo_thau WHERE goi_thau_id = ?", (package_id,))
            all_bids = [dict(r) for r in cursor.fetchall()]
            
            bids_by_nt = {}
            for bid in all_bids:
                nt_id = str(bid.get('nha_thau_id') or '')
                if not nt_id:
                    continue
                if nt_id not in bids_by_nt:
                    bids_by_nt[nt_id] = []
                bids_by_nt[nt_id].append(bid)

            idx_truot = 1
            for nt_id, nt_bids in bids_by_nt.items():
                # Neu nha thau da trung bat ky lo nao, ho khong nam trong danh sach truot thau tong hop
                if nt_id in winning_ids:
                    continue
                
                # Nha thau thuc su truot tat ca cac phan lo tham gia
                lot_reasons = []
                for bid in nt_bids:
                    ly_do = bid.get('ly_do_truot') or ''
                    if not ly_do:
                        ket_luan = bid.get('danh_gia_ket_luan')
                        if ket_luan == 'Không đạt':
                            failed_steps = []
                            if bid.get('danh_gia_hop_le') == 'Không đạt':
                                failed_steps.append("Đánh giá hợp lệ")
                            if bid.get('danh_gia_nang_luc') == 'Không đạt':
                                failed_steps.append("Đánh giá năng lực")
                            if bid.get('danh_gia_ky_thuat') == 'Không đạt' or (bid.get('danh_gia_ky_thuat') and 'không đạt' in str(bid.get('danh_gia_ky_thuat')).lower()):
                                failed_steps.append("Đánh giá kỹ thuật")
                            if bid.get('danh_gia_tai_chinh') == 'Không đạt' or (bid.get('danh_gia_tai_chinh') and 'không đạt' in str(bid.get('danh_gia_tai_chinh')).lower()):
                                failed_steps.append("Đánh giá tài chính")
                            
                            if failed_steps:
                                ly_do = f"Không đạt ở bước: {', '.join(failed_steps)}"
                            else:
                                ly_do = "Không đạt đánh giá chi tiết"
                        else:
                            ly_do = "Đạt yêu cầu kỹ thuật nhưng giá dự thầu xếp sau"
                    
                    if pkg.get('phan_lo') == 'Có' and (bid.get('ten_phan_lo') or bid.get('ma_phan_lo')):
                        ten_lo = bid.get('ten_phan_lo') or bid.get('ma_phan_lo')
                        lot_reasons.append(f"{ten_lo}: {ly_do}")
                    else:
                        lot_reasons.append(ly_do)
                
                tong_hop_ly_do = "; ".join(lot_reasons)
                first_bid = nt_bids[0]
                custom_context['Danh_Sach_Nha_Thau_Truot'].append({
                    'STT': idx_truot,
                    'Ten_Nha_Thau': first_bid.get('ten_nha_thau') or '--',
                    'Ma_Nha_Thau': first_bid.get('ma_dinh_danh') or first_bid.get('ma_phan_lo') or '--',
                    'Ly_Do_Truot': tong_hop_ly_do
                })
                idx_truot += 1
        except Exception as e:
            print("Error preparing Danh_Sach_Nha_Thau_Truot:", e)
            
        # Set global winner variables for conditional templates
        winning_nt = nha_thau_list[0] if nha_thau_list else {}
        custom_context['Thanh_Vien_Lien_Danh'] = []
        members = winning_nt.get('members', []) or []
        for idx, m in enumerate(members):
            role = "Liên danh phụ (Thành viên)"
            if idx == 0:
                role = "Liên danh chính (Đứng đầu liên danh)"
            custom_context['Thanh_Vien_Lien_Danh'].append({
                'STT': idx + 1,
                'Ten_TV': m.get('tenNhaThau') or m.get('ten_tv') or m.get('name') or '--',
                'MST_TV': m.get('maSoThue') or m.get('mst_tv') or m.get('tax_code') or '--',
                'Vai_Tro_TV': m.get('vaiTro') or m.get('vai_tro_tv') or role,
                'Nguoi_Dai_Dien_TV': m.get('nguoiDaiDien') or m.get('representative') or '--',
                'Dia_Chi_TV': m.get('diaChi') or m.get('address') or '--',
                'So_Tai_Khoan_TV': m.get('soTaiKhoan') or m.get('so_tai_khoan') or '--',
                'Noi_Mo_Tai_Khoan_TV': m.get('noiMoTaiKhoan') or m.get('noi_mo_tai_khoan') or '--',
            })
        
        # Create a unified context
        unified_context = {}
        unified_context.update(context)
        unified_context.update(custom_context)
        for k, v in custom_evaluated_values.items():
            unified_context[k] = v

        if active_tpl == 'mau_bao_cao_dau_thau.docx':
            tpl_path = os.path.join(custom_exporter.TEMPLATE_DIR, active_tpl)
        elif active_tpl == 'mau_hop_dong_lcnt.docx':
            tpl_path = os.path.join(custom_exporter.TEMPLATE_DIR, active_tpl)
        else:
            user_dir = custom_exporter.get_user_template_dir(user_id)
            tpl_path = os.path.join(user_dir, active_tpl)
            
        docx_stream = custom_exporter.generate_report_from_custom_template(tpl_path, unified_context, custom_vars_list)
        
        return StreamingResponse(
            docx_stream,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename=Bao_cao_danh_gia_goi_thau_{pkg['ma_goi_thau']}.docx"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)

async def list_templates_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        templates = custom_exporter.list_templates(user_id)
        return JSONResponse(templates)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def set_active_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id
        
        data = await request.json()
        filename = data.get('filename')
        if not filename:
            return JSONResponse({"error": "Filename is required"}, status_code=400)
        custom_exporter.set_active_template(filename, user_id)
        return JSONResponse({"status": "success"})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def upload_template_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        user_id = role_or_err.user_id

        form = await request.form()
        file_obj = form.get('file')
        if not file_obj:
            return JSONResponse({"error": "No file uploaded"}, status_code=400)
        
        file_bytes = await file_obj.read()
        valid, msg = custom_exporter.validate_template_syntax(file_bytes)
        
        if not valid:
            return JSONResponse({"success": False, "error": msg}, status_code=200)
        
        user_dir = custom_exporter.get_user_template_dir(user_id)
        save_path = os.path.join(user_dir, file_obj.filename)
        with open(save_path, 'wb') as f:
            f.write(file_bytes)
            
        custom_exporter.set_active_template(file_obj.filename, user_id)
        return JSONResponse({"success": True, "message": "Tải biểu mẫu lên thành công và đã được kích hoạt!"})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

async def list_word_mappings_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)
        
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ten_bien, source_table, source_column FROM cau_hinh_bien_word WHERE owner_id = ?", (org_name,))
        rows = cursor.fetchall()
        conn.close()
        
        mappings = []
        for r in rows:
            mappings.append({
                "id": r[0],
                "tenBien": r[1],
                "sourceTable": r[2],
                "sourceColumn": r[3]
            })
        return JSONResponse(mappings)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def save_word_mapping_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)
        
        data = await request.json()
        m_id = data.get('id')
        ten_bien = data.get('tenBien', '').strip().lower()
        source_table = data.get('sourceTable', '').strip()
        source_column = data.get('sourceColumn', '').strip()
        
        if not ten_bien or not source_table or not source_column:
            return JSONResponse({"error": "Vui lòng nhập đầy đủ thông tin!"}, status_code=400)
            
        import re
        if not re.match(r'^[A-Za-z0-9_]+$', ten_bien):
            return JSONResponse({"error": "Tên biến chỉ được chứa chữ cái, chữ số và dấu gạch dưới!"}, status_code=400)
            
        if not m_id:
            m_id = "wm-" + secrets.token_hex(8)
            
        current_time = int(datetime.utcnow().timestamp())
        
        conn = database.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO cau_hinh_bien_word (id, owner_id, ten_bien, source_table, source_column, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (m_id, org_name, ten_bien, source_table, source_column, current_time))
            conn.commit()
        except Exception as db_err:
            conn.close()
            if "UNIQUE" in str(db_err) or "constraint failed" in str(db_err):
                return JSONResponse({"error": f"Tên biến '{ten_bien}' đã tồn tại trong hệ thống của bạn!"}, status_code=400)
            return JSONResponse({"error": str(db_err)}, status_code=500)
        conn.close()
        
        return JSONResponse({"success": True, "id": m_id})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def delete_word_mapping_api(request):
    try:
        is_valid, role_or_err = verify_session(request)
        if not is_valid:
            return JSONResponse({"error": role_or_err}, status_code=403)
        org_name = get_active_org(request, role_or_err.user_id)
        
        mapping_id = request.path_params.get('mapping_id')
        
        conn = database.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cau_hinh_bien_word WHERE id = ? AND owner_id = ?", (mapping_id, org_name))
        conn.commit()
        conn.close()
        
        return JSONResponse({"success": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def import_excel_api(request):
    try:
        from io import BytesIO
        form = await request.form()
        file_obj = form.get('file')
        import_type = form.get('type')
        
        if not file_obj or not import_type:
            return JSONResponse({"error": "Missing file or type parameter"}, status_code=400)

        # Tự động lấy map_cols từ ENTITY_SCHEMA — không còn hardcoded
        map_cols = _schema_to_map_cols(import_type)
        if map_cols is None:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)


        file_bytes = await file_obj.read()
        df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
        
        all_possible_headers = []
        for poss in map_cols.values():
            all_possible_headers.extend([x.lower() for x in poss])
            
        first_col = [str(x).strip().lower() for x in df_raw.iloc[:, 0].dropna()]
        vertical_matches = sum(1 for v in first_col if v in all_possible_headers)
        
        if vertical_matches >= 3 or (df_raw.shape[1] >= 2 and vertical_matches >= 1):
            headers = [str(x).strip() for x in df_raw.iloc[:, 0]]
            records = []
            for col_idx in range(1, df_raw.shape[1]):
                col_vals = df_raw.iloc[:, col_idx]
                if all(str(x).strip() == "" or pd.isna(x) for x in col_vals):
                    continue
                
                row_data = {}
                for r_idx, h in enumerate(headers):
                    val = col_vals.iloc[r_idx] if r_idx < len(col_vals) else ""
                    if pd.isna(val):
                        val = ""
                    row_data[h] = val
                records.append(row_data)
            df = pd.DataFrame(records)
        else:
            df = pd.read_excel(BytesIO(file_bytes))
            
        df.columns = [str(c).strip() for c in df.columns]
        
        rows = []
        
        def find_col(possible_names):
            for name in possible_names:
                for col in df.columns:
                    if col.lower() == name.lower() or col.lower().replace(" ", "") == name.lower().replace(" ", ""):
                        return col
            return None

        def clean_money(val):
            if pd.isna(val):
                return 0
            val_str = str(val).replace("VND", "").replace("đ", "").replace("₫", "").replace(".", "").replace(",", "").strip()
            try:
                return float(val_str)
            except ValueError:
                return 0

        def clean_int(val):
            if pd.isna(val):
                return 0
            try:
                return int(float(str(val).strip()))
            except ValueError:
                return 0

        for idx, row in df.iterrows():
            item = {}
            validation_comments = []
            
            for key, poss in map_cols.items():
                found = find_col(poss)
                val = row[found] if (found is not None) else None
                if pd.isna(val):
                    val = ""
                    
                if key in ['tongMucDauTu', 'giaGoiThau', 'giaTri', 'giaTriPhanLo', 'giaTriUocTinh', 'giaTrungThau']:
                    val = clean_money(val)
                elif key in ['thoiGianThucHien']:
                    val = clean_int(val)
                elif key in ['soLuong', 'tyLe']:
                    try:
                        val = float(str(val).strip()) if val != "" else 0.0
                    except ValueError:
                        val = 0.0
                else:
                    val = str(val).strip()
                    
                item[key] = val
            
            if import_type == 'chudautu':
                if not item['tenChuDauTu']:
                    validation_comments.append("Tên chủ đầu tư không được để trống")
                if not item['maChuDauTu']:
                    validation_comments.append("Mã chủ đầu tư không được để trống")
            elif import_type == 'kehoach':
                if not item['tenKeHoach']:
                    validation_comments.append("Tên kế hoạch không được để trống")
                if not item['maKeHoach']:
                    validation_comments.append("Mã kế hoạch không được để trống")
            elif import_type == 'goithau':
                if not item['tenGoiThau']:
                    validation_comments.append("Tên gói thầu không được để trống")
                if not item['maGoiThau']:
                    validation_comments.append("Mã gói thầu không được để trống")
            elif import_type == 'nhathau':
                if not item['tenNhaThau']:
                    validation_comments.append("Tên nhà thầu không được để trống")
                if not item['maNhaThau']:
                    validation_comments.append("Mã nhà thầu không được để trống")
            elif import_type == 'chuyengia':
                if not item['hoTen']:
                    validation_comments.append("Họ và tên không được để trống")
                if not item['soChungChi']:
                    validation_comments.append("Số chứng chỉ không được để trống")
            elif import_type == 'hopdong':
                if not item['tenHopDong']:
                    validation_comments.append("Tên hợp đồng không được để trống")
                if not item['soHopDong']:
                    validation_comments.append("Số hợp đồng không được để trống")
            elif import_type == 'phanlo':
                if not item['tenPhanLo']:
                    validation_comments.append("Tên phần lô không được để trống")
            elif import_type == 'tuychonmuathem':
                if not item['hangMuc']:
                    validation_comments.append("Hạng mục không được để trống")
            
            item['_valid'] = len(validation_comments) == 0
            item['_comment'] = "; ".join(validation_comments) if validation_comments else "Hợp lệ"
            rows.append(item)
            
        return JSONResponse({"success": True, "rows": rows})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_excel_template_api(request):
    try:
        import_type = request.path_params.get('import_type')
        from io import BytesIO
        cols = _schema_to_headers(import_type)
        if not cols:
            return JSONResponse({"error": f"Invalid type: {import_type}"}, status_code=400)

        options_map = _schema_to_options(import_type)

        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        # Store dropdown options in a hidden 'Dropdowns' sheet to ensure robust cross-locale compatibility
        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        is_vertical = len(cols) > 5

        if is_vertical:
            for idx, col_name in enumerate(cols, start=1):
                ws.row_dimensions[idx].height = 24
                cell_a = ws.cell(row=idx, column=1, value=col_name)
                
                cell_a.font = Font(name="Calibri", size=11, bold=True)
                cell_a.alignment = left_align
                cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell_a.border = thin_border
                
                for col_idx in range(2, 12):
                    cell_data = ws.cell(row=idx, column=col_idx)
                    cell_data.border = thin_border

                options = options_map.get(col_name)
                if options:
                    dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                    dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                    dv.errorTitle = 'Lỗi nhập liệu'
                    dv.prompt = 'Chọn một giá trị trong danh sách'
                    dv.promptTitle = col_name
                    dv.errorStyle = "stop"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    ws.add_data_validation(dv)
                    for col_idx in range(2, 12):
                        dv.add(ws.cell(row=idx, column=col_idx))

            max_len_a = max(len(str(cell.value or '')) for cell in ws['A'])
            ws.column_dimensions['A'].width = max(max_len_a + 5, 25)
            for col_idx in range(2, 12):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 18
        else:
            ws.append(cols)
            ws.row_dimensions[1].height = 28
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            ws.row_dimensions[2].height = 24
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.border = thin_border
                col_name = cols[col_idx - 1]
                options = options_map.get(col_name)
                if options:
                    col_letter = get_column_letter(col_idx)
                    dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                    dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                    dv.errorTitle = 'Lỗi nhập liệu'
                    dv.prompt = 'Chọn một giá trị trong danh sách'
                    dv.promptTitle = col_name
                    dv.errorStyle = "stop"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}100")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_nhap_lieu_{import_type}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_mothau_template_api(request):
    try:
        case_type = request.query_params.get('case_type', '1G1T_NO_LOT')
        package_name = request.query_params.get('package_name', 'GoiThau')
        lot_codes_str = request.query_params.get('lot_codes', '')
        
        lot_codes = [c.strip() for c in lot_codes_str.split(',') if c.strip()]
        
        headers = []
        options_map = {
            'Loại nhà thầu': ['Độc lập', 'Liên danh']
        }
        
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô (Tự động điền)', 'Mã nhà thầu', 'Tên nhà thầu (Nhập chính xác)', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)']
            if lot_codes:
                options_map['Mã phần lô'] = lot_codes

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Nhap Lieu"

        options_ranges = {}
        if options_map:
            ws_options = wb.create_sheet(title="Dropdowns")
            ws_options.sheet_state = 'hidden'
            for opt_idx, (opt_col_name, opt_values) in enumerate(options_map.items(), start=1):
                opt_col_letter = get_column_letter(opt_idx)
                for val_idx, val in enumerate(opt_values, start=1):
                    ws_options.cell(row=val_idx, column=opt_idx, value=val)
                options_ranges[opt_col_name] = f"Dropdowns!${opt_col_letter}$1:${opt_col_letter}${len(opt_values)}"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[2].height = 24
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.border = thin_border
            col_name = headers[col_idx - 1]
            options = options_map.get(col_name)
            if options:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1=options_ranges[col_name], allow_blank=True)
                dv.error = 'Dữ liệu chọn không hợp lệ. Vui lòng chọn giá trị từ danh sách!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn một giá trị trong danh sách'
                dv.promptTitle = col_name
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}100")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        filename = f"Mau_Mo_Thau_{case_type}_{package_name}.xlsx"
        
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_danhgiahsdt_template_api(request):
    try:
        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)

        pkg_id_clean = clean_id(package_id)
        if not pkg_id_clean:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list FROM goi_thau WHERE id = ?", (pkg_id_clean,))
        gt_row = cursor.fetchone()
        if not gt_row:
            conn.close()
            return JSONResponse({"error": "Package not found"}, status_code=404)
            
        linh_vuc, phuong_thuc_lua_chon, phan_lo, phan_lo_list_str = gt_row
        
        cursor.execute("""
            SELECT loai_nha_thau, ma_phan_lo, ten_phan_lo, ma_dinh_danh, ten_nha_thau,
                   gia_du_thau, ty_le_giam_gia, gia_sau_giam_gia, hieu_luc_hsdt,
                   gia_tri_dam_bao, hieu_luc_bao_dam_ngay, thoi_gian_thuc_hien,
                   dam_bao_du_thau, hieu_luc_dam_bao, hieu_luc_hsdxt,
                   danh_gia_hop_le, danh_gia_nang_luc, danh_gia_ky_thuat,
                   lam_ro_hop_le, lam_ro_nang_luc, lam_ro_ky_thuat, lam_ro_tai_chinh
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ?
        """, (pkg_id_clean,))
        bids = cursor.fetchall()
        conn.close()
        
        is_tu_van = linh_vuc == 'Tư vấn'
        is_1g2t = phuong_thuc_lua_chon == 'Một giai đoạn hai túi hồ sơ'
        is_1g1t = phuong_thuc_lua_chon == 'Một giai đoạn một túi hồ sơ'
        has_phan_lo = phan_lo == 'Có'
        
        case_type = '1G1T_NO_LOT'
        if is_tu_van:
            case_type = 'TU_VAN'
        elif not is_tu_van and is_1g2t:
            case_type = '1G2T_WITH_LOT' if has_phan_lo else '1G2T_NO_LOT'
        elif is_1g1t:
            case_type = '1G1T_WITH_LOT' if has_phan_lo else '1G1T_NO_LOT'
            
        headers = []
        if case_type == 'TU_VAN':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Hiệu lực E-HSĐXKT (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G2T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G2T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Đảm bảo dự thầu (VND)', 'Hiệu lực đảm bảo (ngày)', 'Hiệu lực E-HSĐXKT (ngày)', 'Đánh giá hợp lệ', 'Làm rõ tính hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực kinh nghiệm', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật']
        elif case_type == '1G1T_NO_LOT':
            headers = ['Loại nhà thầu', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm giá (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB DT (VND)', 'Hiệu lực ĐB (ngày)', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật', 'Làm rõ tài chính']
        elif case_type == '1G1T_WITH_LOT':
            headers = ['Loại nhà thầu', 'Mã phần lô', 'Tên phần lô', 'Mã nhà thầu', 'Tên nhà thầu', 'Giá dự thầu (VND)', 'Tỷ lệ giảm (%)', 'Giá sau giảm giá (nếu có)', 'Hiệu lực E-HSDT (ngày)', 'Giá trị ĐB (VND)', 'Hiệu lực ĐB', 'Thời gian thực hiện (ngày)', 'Đánh giá hợp lệ', 'Làm rõ hợp lệ', 'Đánh giá năng lực', 'Làm rõ năng lực', 'Đánh giá kỹ thuật', 'Làm rõ kỹ thuật', 'Làm rõ tài chính']

        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhGiaHSDT"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in bids:
            row_data = []
            if case_type == 'TU_VAN':
                row_data = [b[0], b[3], b[4], b[14], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G2T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G2T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[12], b[13], b[14], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '']
            elif case_type == '1G1T_NO_LOT':
                row_data = [b[0], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '', b[21] or '']
            elif case_type == '1G1T_WITH_LOT':
                row_data = [b[0], b[1], b[2], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10], b[11], b[15] or '', b[18] or '', b[16] or '', b[19] or '', b[17] or '', b[20] or '', b[21] or '']
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border
            row_num += 1
            
        for col_idx, h in enumerate(headers, start=1):
            if h in ['Đánh giá hợp lệ', 'Đánh giá năng lực']:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(type="list", formula1='"Đạt,Không đạt"', allow_blank=True)
                dv.error = 'Vui lòng chọn Đạt hoặc Không đạt!'
                dv.errorTitle = 'Lỗi nhập liệu'
                dv.prompt = 'Chọn Đạt hoặc Không đạt'
                dv.promptTitle = h
                dv.errorStyle = "stop"
                dv.showErrorMessage = True
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{row_num + 20}")
                
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        filename = f"Mau_Danh_Gia_HSDT_{package_name}.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

async def export_ketquaqd_template_api(request):
    try:
        package_id = request.query_params.get('package_id', '')
        package_name = request.query_params.get('package_name', 'GoiThau')
        
        if not package_id:
            return JSONResponse({"error": "Missing package_id parameter"}, status_code=400)

        pkg_id_clean = clean_id(package_id)
        if not pkg_id_clean:
            return JSONResponse({"error": "Invalid package_id format"}, status_code=400)

        conn = database.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT nha_thau_trung_thau_id, gia_trung_thau, thoi_gian_goi_thau, thoi_gian_hop_dong FROM goi_thau WHERE id = ?", (pkg_id_clean,))
        gt_row = cursor.fetchone()
        
        cursor.execute("""
            SELECT ma_dinh_danh, ten_nha_thau, nha_thau_id, ly_do_truot, loai_nha_thau, ma_phan_lo, ten_phan_lo
            FROM thong_tin_mo_thau
            WHERE goi_thau_id = ?
        """, (pkg_id_clean,))
        bids = cursor.fetchall()
        conn.close()
        
        winner_id = gt_row[0] if gt_row else None
        gia_trung = gt_row[1] if gt_row else 0
        tg_goithau = gt_row[2] if gt_row else ""
        tg_hopdong = gt_row[3] if gt_row else ""
        
        headers = ['Mã nhà thầu', 'Tên nhà thầu', 'Trúng thầu/Trượt thầu', 'Lý do trượt', 'Giá trúng thầu (VND)', 'Thời gian thực hiện gói thầu', 'Thời gian thực hiện hợp đồng']
        
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "KetQuaLCNT"
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        row_num = 2
        for b in bids:
            ma_nt, ten_nt, nt_uuid, ly_do, loai_nt, ma_pl, ten_pl = b
            
            # check if winner
            is_winner = False
            if winner_id:
                if str(winner_id).startswith("nt-"):
                    is_winner = (str(nt_uuid) == str(winner_id))
                else:
                    try:
                        is_winner = (int(nt_uuid) == int(winner_id))
                    except:
                        is_winner = (str(nt_uuid) == str(winner_id))
            
            if is_winner:
                row_data = [ma_nt, ten_nt, 'Trúng thầu', '', gia_trung, tg_goithau, tg_hopdong]
            else:
                row_data = [ma_nt, ten_nt, 'Trượt thầu', ly_do or 'Đạt yêu cầu kỹ thuật nhưng giá dự thầu xếp sau', '', '', '']
                
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 22
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border
            row_num += 1
            
        dv = DataValidation(type="list", formula1='"Trúng thầu,Trượt thầu"', allow_blank=True)
        dv.error = 'Vui lòng chọn Trúng thầu hoặc Trượt thầu!'
        dv.errorTitle = 'Lỗi nhập liệu'
        dv.prompt = 'Chọn trạng thái'
        dv.promptTitle = 'Trúng thầu/Trượt thầu'
        dv.errorStyle = "stop"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"C2:C{row_num + 20}")
        
        for col in ws.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        filename = f"Mau_Ket_Qua_LCNT_{package_name}.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from backend.documents import excel_handler
from backend.documents import excel_service


def _workbook_bytes(headers, rows, *, vertical=False):
    workbook = Workbook()
    sheet = workbook.active
    if vertical:
        for index, header in enumerate(headers, 1):
            sheet.cell(index, 1, header)
            sheet.cell(index, 2, rows[0][index - 1] if rows and index - 1 < len(rows[0]) else "")
    else:
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class _Cursor:
    def __init__(self, *, one=None, rows=None):
        self.one = one
        self.rows = rows or []
        self.calls = []

    def execute(self, sql, params=()):
        self.calls.append((" ".join(sql.split()), params))
        return self

    def fetchone(self):
        return self.one

    def fetchall(self):
        return self.rows


class _Connection:
    def __init__(self, cursor):
        self._cursor = cursor
        self.closed = 0

    def cursor(self):
        return self._cursor

    def close(self):
        self.closed += 1


class _Database:
    def __init__(self, *connections):
        self.connections = list(connections)

    def get_connection(self):
        return self.connections.pop(0)


def test_schema_metadata_cleaners_and_invalid_types():
    assert excel_handler._schema_to_map_cols("nhathau")
    assert excel_handler._schema_to_headers("nhathau")
    assert excel_handler._schema_to_options("nhathau")
    assert excel_handler._schema_to_formats("nhathau")
    assert excel_handler._schema_to_map_cols("missing") is None
    assert excel_handler._schema_to_headers("missing") is None
    assert excel_handler._schema_to_options("missing") == {}
    assert excel_handler._schema_to_formats("missing") == {}
    assert excel_handler.clean_money(None) == 0
    assert excel_handler.clean_money("1.234 VND") == 1234
    assert excel_handler.clean_money("invalid") == 0
    assert excel_handler.clean_int(None) == 0
    assert excel_handler.clean_int("2.0") == 2
    assert excel_handler.clean_int("invalid") == 0
    with pytest.raises(ValueError):
        excel_handler.parse_excel(b"", "missing")


def test_plan_excel_schema_covers_submission_approval_workflow():
    schema = excel_handler.ENTITY_SCHEMA["kehoach"]
    fields = {entry["field"]: entry for entry in schema}

    expected_fields = {
        "pheDuyet",
        "soToTrinhDuToan",
        "ngayTrinhDuToan",
        "ngayPheDuyetDuToan",
        "soQdPheDuyetDuToan",
        "soToTrinhKeHoach",
        "soToTrinhDuToanKeHoach",
        "ngayTrinhKeHoach",
    }
    assert expected_fields <= fields.keys()
    assert fields["pheDuyet"]["options"] == ["Kế hoạch", "Dự toán và kế hoạch"]
    assert "Số TTr dự toán" in fields["soToTrinhDuToan"]["aliases"]
    assert "Số TTr kế hoạch" in fields["soToTrinhKeHoach"]["aliases"]
    assert "Số TTr dự toán và kế hoạch" in fields["soToTrinhDuToanKeHoach"]["aliases"]

    workbook = excel_service.create_excel_template("kehoach")
    headers = [cell.value for cell in workbook["Nhap Lieu"][1]]
    assert "Số tờ trình dự toán" in headers
    assert "Số tờ trình kế hoạch" in headers
    assert "Số tờ trình dự toán và kế hoạch" in headers
    formats = excel_handler._schema_to_formats("kehoach")
    assert formats["Số tờ trình dự toán"] == "text"
    assert formats["Ngày trình dự toán"] == "date"
    assert formats["Ngày trình kế hoạch"] == "date"


def test_plan_excel_import_maps_submission_number_aliases():
    headers = [
        "Mã kế hoạch",
        "Tên kế hoạch",
        "Hình thức phê duyệt",
        "Số TTr dự toán",
        "Số TTr kế hoạch",
        "Số TTr dự toán và kế hoạch",
    ]
    values = [
        "KH-01",
        "Kế hoạch thử nghiệm",
        "Kế hoạch",
        "01/TTr-CĐT",
        "02/TTr-CĐT",
        "03/TTr-CĐT",
    ]

    result = excel_handler.parse_excel(_workbook_bytes(headers, [values]), "kehoach")

    assert result[0]["isValid"]
    assert result[0]["data"]["pheDuyet"] == "Kế hoạch"
    assert result[0]["data"]["soToTrinhDuToan"] == "01/TTr-CĐT"
    assert result[0]["data"]["soToTrinhKeHoach"] == "02/TTr-CĐT"
    assert result[0]["data"]["soToTrinhDuToanKeHoach"] == "03/TTr-CĐT"


@pytest.mark.parametrize(
    ("import_type", "overrides", "expected_comment"),
    [
        ("chudautu", {"maSoThue": "bad", "email": "bad", "soDienThoai": "x"}, "Tên chủ đầu tư"),
        ("kehoach", {"tongMucDauTu": -1}, "Tên kế hoạch"),
        ("goithau", {"giaGoiThau": -1, "thoiGianThucHien": 0}, "Tên gói thầu"),
        ("nhathau", {"maSoThue": "bad", "email": "bad", "soDienThoai": "x"}, "Tên nhà thầu"),
        ("chuyengia", {"soCCCD": "123", "email": "bad"}, "Họ và tên"),
        ("hopdong", {"giaTri": -1}, "Tên hợp đồng"),
        ("phanlo", {"maPhanLo": "PL-01"}, "Tên phần lô"),
        ("tuychonmuathem", {"donVi": "cái"}, "Hạng mục"),
    ],
)
def test_parse_horizontal_excel_validates_each_import_contract(import_type, overrides, expected_comment):
    schema = excel_handler.ENTITY_SCHEMA[import_type]
    headers = [entry["label"] for entry in schema]
    values = []
    for entry in schema:
        value = overrides.get(entry["field"], "")
        values.append(value)
    result = excel_handler.parse_excel(_workbook_bytes(headers, [values]), import_type)
    assert len(result) == 1
    assert not result[0]["isValid"]
    assert expected_comment in result[0]["comments"]


def test_parse_vertical_excel_and_numeric_normalization():
    import_type = "phanlo"
    schema = excel_handler.ENTITY_SCHEMA[import_type]
    headers = [entry["label"] for entry in schema]
    values_by_field = {
        "maPhanLo": 1.0,
        "tenPhanLo": "Lô 1",
        "giaTriPhanLo": "1.234 VND",
        "baoDamDuThau": "200",
        "thoiGianThucHien": "30",
    }
    values = [values_by_field.get(entry["field"], "") for entry in schema]
    result = excel_handler.parse_excel(_workbook_bytes(headers, [values], vertical=True), import_type)
    assert result[0]["isValid"]
    assert result[0]["rowIdx"] == 1
    assert result[0]["data"]["maPhanLo"] == "1"
    assert result[0]["data"]["giaTriPhanLo"] == 1234


def test_configured_workbook_styles_formats_dropdowns_and_widths():
    workbook = excel_service._build_configured_workbook(
        "Sheet",
        ["Money", "Date", "DateTime", "Choice"],
        rows=[[1000, "2026-01-01", "2026-01-01 10:00", "A"]],
        options_map={"Choice": ["A", "B"]},
        formats_map={"Money": "currency", "Date": "date", "DateTime": "datetime"},
        empty_rows=3,
    )
    sheet = workbook["Sheet"]
    assert sheet["A2"].number_format == "#,##0"
    assert sheet["B2"].number_format == "dd/mm/yyyy"
    assert "ngày" in sheet["C2"].number_format
    assert sheet.column_dimensions["A"].width >= 15
    assert workbook["Dropdowns"].sheet_state == "hidden"
    assert sheet.data_validations.count == 1
    assert excel_service._add_dropdown_sheet(Workbook(), {}) == {}


@pytest.mark.parametrize("case_type", list(excel_service.OPENING_TEMPLATE_HEADERS))
def test_basic_and_opening_templates(case_type):
    workbook = excel_service.create_excel_template("nhathau")
    assert workbook["Nhap Lieu"].max_row == 51
    with pytest.raises(ValueError):
        excel_service.create_excel_template("missing")
    opening = excel_service.create_mothau_template(case_type, ["L1"])
    assert opening["Mo Thau"].max_row == 51
    with pytest.raises(ValueError):
        excel_service.create_mothau_template("missing", [])


def test_opening_financial_template_filters_qualified_bids(monkeypatch):
    rows = [
        ("m1", "A", 100, 0, 100, 30, 20, "", "", "", "Đạt"),
        ("m2", "B", 200, 0, 200, 30, 20, "Đạt", "Đạt", "Đạt", ""),
        ("m3", "C", 300, 0, 300, 30, 20, "Đạt", "Đạt", "Không đạt", ""),
    ]
    cursor = _Cursor(rows=rows)
    conn = _Connection(cursor)
    monkeypatch.setattr(excel_service, "database", _Database(conn))
    workbook = excel_service.create_opening_fin_template("pkg", "org")
    assert workbook["Mo De Xuat Tai Chinh"].max_row == 3
    assert conn.closed == 1


def _evaluation_bid():
    return tuple(
        [
            "Độc lập", "L1", "Lô 1", "m1", "A", 100, 5, 95, 30, 10, 20, 40,
            "Đạt", "Đạt", "Đạt", "clear1", "clear2", "clear3", "clear4", "Xếp hạng 1",
            "", "", "",
        ]
    )


@pytest.mark.parametrize(
    ("eval_type", "has_lots", "expected_columns"),
    [
        ("technical", True, 14),
        ("technical", False, 12),
        ("financial", True, 10),
        ("financial", False, 8),
    ],
)
def test_evaluation_templates_cover_modes(monkeypatch, eval_type, has_lots, expected_columns):
    cursor = _Cursor(one=("goods", "one-stage", "Có" if has_lots else "Không"), rows=[_evaluation_bid()])
    conn = _Connection(cursor)
    monkeypatch.setattr(excel_service, "database", _Database(conn))
    monkeypatch.setattr(excel_service, "fetch_package_lot_codes", lambda *_args: ["L1"])
    workbook = excel_service.create_danhgiahsdt_template("pkg", "org", eval_type)
    assert workbook["Danh gia HSDT"].max_column == expected_columns
    assert workbook["Danh gia HSDT"].max_row == 2


def test_evaluation_and_result_templates_reject_missing_package(monkeypatch):
    for function in (excel_service.create_danhgiahsdt_template, excel_service.create_ketquaqd_template):
        conn = _Connection(_Cursor(one=None))
        monkeypatch.setattr(excel_service, "database", _Database(conn))
        args = ("pkg", "org", "technical") if function is excel_service.create_danhgiahsdt_template else ("pkg", "org")
        with pytest.raises(ValueError):
            function(*args)
        assert conn.closed == 1


def test_result_and_small_list_exports(monkeypatch):
    cursor = _Cursor(one=("m1", 90, 20, 30), rows=[
        ("Độc lập", "", "", "m1", "A", 100, 10, 90, "Xếp hạng 1"),
        ("Độc lập", "", "", "m2", "B", 120, 0, 120, "Xếp hạng 2"),
    ])
    monkeypatch.setattr(excel_service, "database", _Database(_Connection(cursor)))
    result = excel_service.create_ketquaqd_template("pkg", "org")
    assert result["Ket Qua LCNT"]["D2"].value == "Trúng thầu"
    assert result["Ket Qua LCNT"]["D3"].value == "Trượt thầu"

    lots = excel_service.create_phanlo_excel([
        {"maPhanLo": "L1", "tenPhanLo": "Lô", "giaTriPhanLo": 100, "baoDamDuThau": 10, "thoiGianThucHien": 20}
    ])
    options = excel_service.create_tuychonmuathem_excel([
        {"hangMuc": "X", "donVi": "cái", "soLuong": 1, "tyLe": 5, "giaTriUocTinh": 50}
    ])
    assert lots["Phan Lo"].max_row == 2
    assert options["Tuy Chon Mua Them"].max_row == 2

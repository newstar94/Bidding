from __future__ import annotations

from copy import copy
from decimal import Decimal
from io import BytesIO
import json
import sqlite3
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import pytest

from backend.documents import award_result_excel_service as service
from backend.documents.workbook_preservation import archive_manifest


def _workbook_bytes(headers, row, *, header_row=1):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách nhà thầu"
    if header_row == 2:
        sheet.append(["Hướng dẫn nhập dữ liệu", *([""] * (len(headers) - 1))])
    sheet.append(headers)
    sheet.append(row)
    for cell in sheet[header_row]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
        fill = copy(cell.fill)
        fill.fill_type = "solid"
        fill.fgColor.rgb = "EEEEEE"
        cell.fill = fill
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[service._get_column_letter(column) if hasattr(service, "_get_column_letter") else chr(64 + column)].width = 18
    status_column = headers.index("Kết quả") + 1
    validation = DataValidation(type="list", formula1='"Trúng thầu,Không trúng thầu"')
    validation.add(sheet.cell(header_row + 1, status_column))
    sheet.add_data_validation(validation)
    sheet.freeze_panes = sheet.cell(header_row + 1, 1)
    sheet.print_area = f"A1:{service._get_column_letter(len(headers)) if hasattr(service, '_get_column_letter') else chr(64 + len(headers))}{header_row + 1}"
    sheet.row_dimensions[header_row + 1].height = 27
    sheet.column_dimensions["B"].hidden = True
    sheet.cell(header_row + 1, status_column).number_format = "@"
    sheet.cell(header_row + 1, status_column).comment = Comment(
        "Chỉ cập nhật trường kết quả", "Test"
    )
    hidden = workbook.create_sheet("Metadata")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=1+1"
    hidden.merge_cells("B2:C2")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def standard_workbook_bytes(*, permuted=False):
    headers = list(service.EXPECTED_HEADERS)
    row = [
        "L01", "Lô 01", "vn001", "001", "Nhà thầu 01", 1_000,
        "", "", "", "", "", "", "", "", "",
    ]
    if permuted:
        order = [4, 0, 2, 1, 3, 5, 7, 6, 8, 9, 10, 11, 12, 13, 14]
        headers = [headers[index] for index in order]
        row = [row[index] for index in order]
    return _workbook_bytes(headers, row)


def standard_multi_lot_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách nhà thầu"
    sheet.append(list(service.EXPECTED_HEADERS))
    rows = [
        ["L01", "Lô 01", "vn001", "001", "Nhà thầu A", 1_000],
        ["L02", "Lô 02", "vn001", "001", "Nhà thầu A", 1_100],
        ["L03", "Lô 03", "vn002", "002", "Nhà thầu B", 1_200],
        ["L04", "Lô 04", "vn003", "003", "Nhà thầu C", 1_300],
        ["L99", "Không khớp", "vn999", "999", "Nhà thầu Z", 9_999],
    ]
    for source in rows:
        sheet.append([*source, *("" for _ in range(9))])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def medicine_workbook_bytes():
    row = [
        1, "THUOC-01", "Atropin sulfat", "vn001", "001", "Nhà thầu 01",
        1_014_000, "", "", "", "", "", "", 5, "", "", "", "", "",
    ]
    return _workbook_bytes(
        list(service.MEDICINE_EXPECTED_HEADERS), row, header_row=2
    )


def medicine_multi_item_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách nhà thầu"
    sheet.append(["Mẫu kết quả thuốc"])
    sheet.append(list(service.MEDICINE_EXPECTED_HEADERS))
    for row in (
        [1, "THUOC-01", "Atropin sulfat", "vn001", "001", "Nhà thầu 01", 1_014_000],
        [2, "THUOC-01", "Natri clorid", "vn001", "001", "Nhà thầu 01", 7_800],
    ):
        sheet.append([*row, *([""] * 6), 5, *([""] * 5)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _record(**overrides):
    values = {
        "opening_id": "opening-1",
        "lot_code": "L01",
        "bidder_identifier": "vn001",
        "tax_code": "001",
        "bidder_name": "Nhà thầu 01",
        "status": "Trúng thầu",
        "corrected_price": 950,
        "award_price": 900,
        "package_duration": "30 ngày",
        "contract_duration": "60 ngày",
    }
    values.update(overrides)
    return service.AwardRecord(**values)


def test_inspection_finds_source_and_output_columns_after_reordering():
    inspection = service.inspect_award_result_workbook(
        standard_workbook_bytes(permuted=True)
    )

    assert inspection["templateType"] == "standard"
    assert inspection["templateVersion"] == "muasamcong-standard-v1"
    assert len(inspection["templateFingerprint"]) == 64
    assert inspection["columnMap"]["source"] == [2, 4, 3, 5, 1, 6]
    assert inspection["columnMap"]["output"][:2] == [8, 7]
    assert inspection["rows"][0]["lotCode"] == "L01"
    assert inspection["rows"][0]["bidderIdentifier"] == "vn001"
    assert inspection["blockingErrors"] == []


def test_inspection_rejects_invalid_workbook_and_missing_template_sheet():
    with pytest.raises(service.AwardResultExcelError) as invalid:
        service.inspect_award_result_workbook(b"not-an-xlsx")
    assert invalid.value.code == "WORKBOOK_INVALID"

    workbook = Workbook()
    workbook.active["A1"] = "Dữ liệu không liên quan"
    output = BytesIO()
    workbook.save(output)
    inspection = service.inspect_award_result_workbook(output.getvalue())

    assert {item["code"] for item in inspection["blockingErrors"]} == {
        "UNSUPPORTED_TEMPLATE_VERSION"
    }


def test_template_registry_rejects_declared_unsupported_ooxml_parts():
    source = standard_workbook_bytes()
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as before, zipfile.ZipFile(output, "w") as after:
        for item in before.infolist():
            after.writestr(item, before.read(item.filename))
        after.writestr("xl/connections.xml", "<connections />")

    inspection = service.inspect_award_result_workbook(output.getvalue())

    issue = next(
        item
        for item in inspection["blockingErrors"]
        if item["code"] == "UNSUPPORTED_TEMPLATE_PART"
    )
    assert issue["parts"] == ["xl/connections.xml"]


def test_inspection_reports_missing_required_result_header():
    headers = list(service.EXPECTED_HEADERS)
    headers[10] = ""
    content = _workbook_bytes(
        headers,
        ["L01", "Lô 01", "vn001", "001", "Nhà thầu 01", 1_000, *("" for _ in range(9))],
    )

    inspection = service.inspect_award_result_workbook(content)

    assert any(
        item["code"] == "REQUIRED_HEADER_MISSING"
        and item["expectedHeader"] == "Giá trúng thầu"
        for item in inspection["blockingErrors"]
    )


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("  L01  ", "l01"),
        ("001", "001"),
        ("001.0", "001"),
        (1.0, "1"),
        (Decimal("12.50"), "12.50"),
        ("  Ｌ０１\u00a0", "l01"),
        ("Straße", "strasse"),
        ("L01\u2003 A", "l01 a"),
        (None, ""),
    ],
)
def test_identifier_normalization_is_stable(value, expected):
    assert service.normalize_code(value) == expected
    assert service.normalize_tax_code(value) == expected


def test_decimal_mapping_never_uses_float_arithmetic():
    assert service._number(Decimal("123.456789012345678")) == {
        "decimal": "123.456789012345678"
    }
    assert service._number("123.456789012345678") == {
        "decimal": "123.456789012345678"
    }


@pytest.mark.parametrize(
    "unsafe_text",
    [
        "=1+1",
        "+SUM(A1:A2)",
        "-1+1",
        "@SUM(A1:A2)",
        "\t=1+1",
        "\r=1+1",
        "\n=1+1",
        " =1+1",
        "\uff1d1+1",
    ],
)
def test_award_result_export_neutralizes_formula_injection(unsafe_text):
    content = standard_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    match = service.match_award_result_rows(
        inspection,
        [_record(other_content=unsafe_text, rejection_reason=unsafe_text)],
    )

    output = service.write_award_result_workbook(
        content, service.export_updates_from_match(match)
    )
    sheet = load_workbook(BytesIO(output), data_only=False).worksheets[0]
    other_content_column = inspection["columnMap"]["output"][-1]
    cell = sheet.cell(2, other_content_column)

    assert cell.data_type != "f"
    assert cell.value == f"'{unsafe_text}"


def test_matching_blocks_export_when_no_approved_result_is_writable():
    content = standard_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    match = service.match_award_result_rows(
        inspection, [_record(status=None, award_price=None)]
    )

    assert match["totalRows"] == 1
    assert match["matchedRows"] == 1
    assert match["approvedRows"] == 0
    assert match["writableRows"] == 0
    assert match["updatedRows"] == 0
    assert match["canExport"] is False
    assert "NO_APPROVED_RESULT_TO_EXPORT" in {
        item["code"] for item in match["blockingErrors"]
    }


def test_matching_blocks_when_approved_values_are_already_identical():
    content = standard_workbook_bytes()
    record = _record()
    first_inspection = service.inspect_award_result_workbook(content)
    first_match = service.match_award_result_rows(first_inspection, [record])
    filled = service.write_award_result_workbook(
        content, service.export_updates_from_match(first_match)
    )

    second_match = service.match_award_result_rows(
        service.inspect_award_result_workbook(filled), [record]
    )

    assert second_match["approvedRows"] == 1
    assert second_match["writableRows"] == 0
    assert second_match["updatedRows"] == 0
    assert second_match["canExport"] is False


def test_matching_uses_primary_then_fallback_and_never_name_only():
    inspection = {
        "templateType": "standard",
        "sheetName": "Danh sách nhà thầu",
        "totalRows": 3,
        "existingResultRows": 0,
        "blockingErrors": [],
        "warnings": [],
        "rows": [
            {"excelRow": 2, "lotCode": "L01", "bidderIdentifier": "vn001", "taxCode": "001", "bidderName": "Tên khác", "sourceFingerprint": "a"},
            {"excelRow": 3, "lotCode": "L02", "bidderIdentifier": "", "taxCode": "002", "bidderName": "Nhà thầu 02", "sourceFingerprint": "b"},
            {"excelRow": 4, "lotCode": "L03", "bidderIdentifier": "", "taxCode": "", "bidderName": "Nhà thầu 03", "sourceFingerprint": "c"},
        ],
    }
    records = [
        _record(),
        _record(opening_id="opening-2", lot_code="L02", bidder_identifier="vn002", tax_code="002"),
        _record(opening_id="opening-3", lot_code="L03", bidder_identifier="vn003", tax_code="003", bidder_name="Nhà thầu 03"),
    ]

    result = service.match_award_result_rows(inspection, records)

    assert result["exactMatches"] == 1
    assert result["fallbackMatches"] == 1
    assert result["unmatchedRows"] == 1
    assert result["missingBidderIdentityRows"] == 1
    assert result["rows"][0]["matchMethod"] == "lot_code_and_bidder_identifier"
    assert any(item["code"] == "BIDDER_NAME_DIFFERS" for item in result["rows"][0]["warnings"])


def test_matching_blocks_duplicate_and_identifier_tax_conflicts():
    inspection = {
        "templateType": "standard",
        "sheetName": "Danh sách nhà thầu",
        "totalRows": 2,
        "existingResultRows": 0,
        "blockingErrors": [],
        "warnings": [],
        "rows": [
            {"excelRow": 2, "lotCode": "L01", "bidderIdentifier": "vn001", "taxCode": "001", "bidderName": "", "sourceFingerprint": "a"},
            {"excelRow": 3, "lotCode": "L02", "bidderIdentifier": "vn002", "taxCode": "009", "bidderName": "", "sourceFingerprint": "b"},
        ],
    }
    records = [
        _record(),
        _record(opening_id="duplicate", bidder_name="Nhà thầu trùng"),
        _record(opening_id="identifier", lot_code="L02", bidder_identifier="vn002", tax_code="002"),
        _record(opening_id="tax", lot_code="L02", bidder_identifier="vn009", tax_code="009"),
    ]

    result = service.match_award_result_rows(inspection, records)

    assert result["duplicateRows"] == 1
    assert result["conflictRows"] == 1
    assert {item["code"] for item in result["blockingErrors"]} == {
        "DUPLICATE_MATCH_KEY", "IDENTIFIER_TAX_CONFLICT",
    }
    assert result["canExport"] is False


def test_standard_export_preserves_source_order_styles_validation_and_hidden_sheet():
    content = standard_workbook_bytes(permuted=True)
    inspection = service.inspect_award_result_workbook(content)
    match = service.match_award_result_rows(inspection, [_record()])

    output = service.write_award_result_workbook(
        content, service.export_updates_from_match(match)
    )

    before = load_workbook(BytesIO(content), data_only=False)
    after = load_workbook(BytesIO(output), data_only=False)
    assert before.sheetnames == after.sheetnames
    assert after["Metadata"].sheet_state == "hidden"
    assert after["Metadata"]["A1"].value == "=1+1"
    assert str(after["Metadata"].merged_cells.ranges) == str(
        before["Metadata"].merged_cells.ranges
    )
    assert after.worksheets[0].freeze_panes == before.worksheets[0].freeze_panes
    assert str(after.worksheets[0].print_area) == str(before.worksheets[0].print_area)
    assert after.worksheets[0].row_dimensions[2].height == 27
    assert after.worksheets[0].column_dimensions["B"].hidden is True
    source_columns = inspection["columnMap"]["source"]
    assert [before.worksheets[0].cell(2, column).value for column in source_columns] == [
        after.worksheets[0].cell(2, column).value for column in source_columns
    ]
    output_columns = inspection["columnMap"]["output"]
    assert after.worksheets[0].cell(2, output_columns[0]).value == "Trúng thầu"
    assert after.worksheets[0].cell(2, output_columns[4]).value == 900
    assert after.worksheets[0].cell(2, output_columns[0]).style_id == (
        before.worksheets[0].cell(2, output_columns[0]).style_id
    )
    assert after.worksheets[0].cell(2, output_columns[0]).comment.text == (
        before.worksheets[0].cell(2, output_columns[0]).comment.text
    )
    assert len(after.worksheets[0].data_validations.dataValidation) == 1
    before_manifest = archive_manifest(content)
    after_manifest = archive_manifest(output)
    assert set(before_manifest) == set(after_manifest)
    changed_parts = {
        name
        for name in before_manifest
        if before_manifest[name]["sha256"] != after_manifest[name]["sha256"]
    }
    assert changed_parts == {"xl/worksheets/sheet1.xml"}


def test_medicine_export_writes_quantity_and_unit_price_but_preserves_discount():
    content = medicine_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    record = _record(
        lot_code="THUOC-01",
        goods_item_id="requirement-1",
        goods_sequence="1",
        goods_name="Atropin sulfat",
        corrected_price=1_014_000,
        award_quantity=1_300,
        award_unit_price=780,
        award_price=1_014_000,
    )
    match = service.match_award_result_rows(inspection, [record])

    output = service.write_award_result_workbook(
        content, service.export_updates_from_match(match)
    )
    sheet = load_workbook(BytesIO(output)).worksheets[0]

    assert inspection["templateType"] == "medicine"
    assert sheet["H3"].value == "Trúng thầu"
    assert sheet["L3"].value == 1_300
    assert sheet["M3"].value == 780
    assert sheet["N3"].value == 5
    assert sheet["O3"].value == 1_014_000


def test_medicine_matching_uses_item_sequence_and_rejects_reusing_an_item():
    content = medicine_multi_item_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    records = [
        _record(
            lot_code="THUOC-01",
            goods_item_id="requirement-1",
            goods_sequence="1",
            goods_name="Atropin sulfat",
            award_quantity=1_300,
            award_unit_price=780,
            award_price=1_014_000,
        ),
        _record(
            lot_code="THUOC-01",
            goods_item_id="requirement-2",
            goods_sequence="2",
            goods_name="Natri clorid",
            award_quantity=10,
            award_unit_price=780,
            award_price=7_800,
        ),
    ]

    match = service.match_award_result_rows(inspection, records)

    assert [row["goodsSequence"] for row in inspection["rows"]] == [1, 2]
    assert match["exactMatches"] == 2
    assert match["blockingErrors"] == []
    output = service.write_award_result_workbook(
        content, service.export_updates_from_match(match)
    )
    sheet = load_workbook(BytesIO(output)).worksheets[0]
    assert [sheet.cell(row, 12).value for row in (3, 4)] == [1_300, 10]
    assert [sheet.cell(row, 15).value for row in (3, 4)] == [1_014_000, 7_800]

    duplicated = service.inspect_award_result_workbook(content)
    duplicated["rows"][1]["goodsSequence"] = 1
    duplicate_match = service.match_award_result_rows(duplicated, records)
    assert "MEDICINE_GOODS_MATCHED_MULTIPLE_ROWS" in {
        issue["code"] for issue in duplicate_match["blockingErrors"]
    }


def test_medicine_matching_allows_two_bidders_for_the_same_required_goods():
    inspection = service.inspect_award_result_workbook(
        medicine_multi_item_workbook_bytes()
    )
    inspection["rows"][1].update(
        goodsSequence=1,
        goodsName="Atropin sulfat",
        bidderIdentifier="vn002",
        taxCode="002",
        bidderName="Nhà thầu 02",
    )
    records = [
        _record(
            opening_id="opening-1",
            lot_code="THUOC-01",
            goods_item_id="requirement-1",
            goods_sequence="1",
            goods_name="Atropin sulfat",
        ),
        _record(
            opening_id="opening-2",
            lot_code="THUOC-01",
            bidder_identifier="vn002",
            tax_code="002",
            bidder_name="Nhà thầu 02",
            goods_item_id="requirement-1",
            goods_sequence="1",
            goods_name="Atropin sulfat",
            status="Không trúng thầu",
            award_price=None,
            rejection_reason="Xếp hạng sau nhà thầu khác",
        ),
    ]

    match = service.match_award_result_rows(inspection, records)

    assert match["exactMatches"] == 2
    assert match["duplicateRows"] == 0
    assert match["blockingErrors"] == []


def test_multi_lot_pipeline_keeps_each_bidder_lot_independent_and_leaves_unmatched_row():
    content = standard_multi_lot_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    records = [
        _record(
            opening_id="l01-a",
            lot_code="L01",
            bidder_identifier="vn001",
            tax_code="001",
            bidder_name="Nhà thầu A",
            status="Trúng thầu",
            corrected_price=900,
            technical_score=90,
            evaluated_price=None,
            award_price=850,
        ),
        _record(
            opening_id="l02-a",
            lot_code="L02",
            bidder_identifier="vn001",
            tax_code="001",
            bidder_name="Nhà thầu A",
            status="Không trúng thầu",
            corrected_price=1_050,
            award_price=None,
            rejection_reason="Xếp hạng sau nhà thầu khác",
        ),
        _record(
            opening_id="l03-b",
            lot_code="L03",
            bidder_identifier="vn002",
            tax_code="002",
            bidder_name="Nhà thầu B",
            status="Không trúng thầu",
            corrected_price=1_200,
            award_price=None,
            rejection_reason="Lô bị hủy theo quyết định phê duyệt",
            lot_cancelled=True,
        ),
        _record(
            opening_id="l04-c",
            lot_code="L04",
            bidder_identifier="vn003",
            tax_code="003",
            bidder_name="Nhà thầu C",
            status="Trúng thầu",
            corrected_price=1_100,
            evaluated_price=1_080,
            award_price=1_050,
        ),
    ]

    match = service.match_award_result_rows(
        inspection,
        records,
        known_lot_codes=["L01", "L02", "L03", "L04"],
    )
    output = service.write_award_result_workbook(
        content, service.export_updates_from_match(match)
    )
    before = load_workbook(BytesIO(content)).worksheets[0]
    after = load_workbook(BytesIO(output)).worksheets[0]

    assert match["exactMatches"] == 4
    assert match["unmatchedRows"] == 1
    assert match["canExport"] is True
    assert [after.cell(row, 1).value for row in range(2, 7)] == [
        "L01", "L02", "L03", "L04", "L99"
    ]
    assert after["G2"].value == "Trúng thầu"
    assert after["I2"].value == 90
    assert after["J2"].value is None
    assert after["K2"].value == 850
    assert after["G3"].value == "Không trúng thầu"
    assert after["K3"].value is None
    assert after["L3"].value == "Xếp hạng sau nhà thầu khác"
    assert after["G4"].value == "Không trúng thầu"
    assert after["L4"].value == "Lô bị hủy theo quyết định phê duyệt"
    assert after["H5"].value == 1_100
    assert after["K5"].value == 1_050
    assert [after.cell(6, column).value for column in range(7, 16)] == [
        before.cell(6, column).value for column in range(7, 16)
    ]
    assert [
        service._row_fingerprint(after[row][0:6]) for row in range(2, 7)
    ] == [service._row_fingerprint(before[row][0:6]) for row in range(2, 7)]


class _Database:
    def __init__(self, connection):
        self.connection = connection

    def get_connection(self):
        return _Connection(self.connection)


class _Connection:
    def __init__(self, connection):
        self.connection = connection

    def cursor(self):
        return self.connection.cursor()

    def close(self):
        pass


def test_medicine_dataset_uses_official_lot_and_bidder_goods_without_n_plus_one():
    connection = sqlite3.connect(":memory:")
    try:
        connection.row_factory = sqlite3.Row
        statements = []
        connection.set_trace_callback(statements.append)
        connection.executescript(
            """
        CREATE TABLE goi_thau (id TEXT, organization_id TEXT, ma_goi_thau TEXT, ten_goi_thau TEXT, phan_lo TEXT, trang_thai TEXT, nha_thau_trung_thau_id TEXT, gia_trung_thau INTEGER, thoi_gian_goi_thau TEXT, thoi_gian_hop_dong TEXT, phuong_phap_danh_gia TEXT, is_thuoc INTEGER, so_quyet_dinh_ket_qua TEXT, ngay_quyet_dinh_ket_qua TEXT);
        CREATE TABLE goi_thau_phan_lo (id TEXT, organization_id TEXT, goi_thau_id TEXT, ma_phan_lo TEXT, ten_phan_lo TEXT, nha_thau_trung_thau_id TEXT, gia_trung_thau INTEGER, thoi_gian_goi_thau TEXT, thoi_gian_hop_dong TEXT, archived_at TEXT, sort_order INTEGER);
        CREATE TABLE dot_xu_ly_phan_lo_chi_tiet (organization_id TEXT, lot_id TEXT, batch_id TEXT, current_stage TEXT, outcome TEXT);
        CREATE TABLE dot_xu_ly_phan_lo (organization_id TEXT, id TEXT, sequence_no INTEGER);
        CREATE TABLE nha_thau (id TEXT, organization_id TEXT, ten_nha_thau TEXT, ma_nha_thau TEXT, ma_so_thue TEXT);
        CREATE TABLE thong_tin_mo_thau (id TEXT, organization_id TEXT, goi_thau_id TEXT, nha_thau_id TEXT, ma_phan_lo TEXT, ma_dinh_danh TEXT, ten_nha_thau TEXT, gia_du_thau INTEGER, gia_sau_giam_gia INTEGER, gia_danh_gia_sau_uu_dai INTEGER, thoi_gian_thuc_hien TEXT, archived_at TEXT);
        CREATE TABLE ket_qua_danh_gia_nha_thau (organization_id TEXT, thong_tin_mo_thau_id TEXT, diem REAL, ly_do_loai TEXT, danh_gia_ket_luan TEXT);
        CREATE TABLE goi_thau_hang_hoa (id TEXT, organization_id TEXT, goi_thau_id TEXT, ma_hang_hoa TEXT, ten_hang_hoa TEXT, don_vi_tinh TEXT);
        CREATE TABLE hang_hoa_du_thau_nha_thau (id TEXT, organization_id TEXT, goi_thau_id TEXT, thong_tin_mo_thau_id TEXT, khoi_luong REAL, don_gia_du_thau INTEGER, gia_tri_co_so_sau_giam_gia INTEGER, goi_thau_hang_hoa_id TEXT, stt_nguon TEXT, danh_muc_hang_hoa TEXT, don_vi_tinh TEXT, mapping_status TEXT, sort_order INTEGER, is_draft INTEGER);
        INSERT INTO goi_thau VALUES ('pkg','org','IB-01','Thuốc','Có','AWARDED',NULL,1014000,NULL,NULL,NULL,1,'QD-01','2026-08-04');
        INSERT INTO goi_thau_phan_lo VALUES ('lot','org','pkg','THUOC-01','Atropin','bidder',1014000,'220 ngày','220 ngày + nghĩa vụ',NULL,1);
        INSERT INTO dot_xu_ly_phan_lo VALUES ('org','batch',1);
        INSERT INTO dot_xu_ly_phan_lo_chi_tiet VALUES ('org','lot','batch','RESULT_APPROVED','AWARDED');
        INSERT INTO nha_thau VALUES ('bidder','org','Nhà thầu 01','vn001','001');
        INSERT INTO thong_tin_mo_thau VALUES ('opening','org','pkg','bidder','THUOC-01','vn001','Nhà thầu 01',1014000,1014000,NULL,'220 ngày',NULL);
        INSERT INTO ket_qua_danh_gia_nha_thau VALUES ('org','opening',NULL,NULL,'Đạt');
        INSERT INTO goi_thau_hang_hoa VALUES ('requirement','org','pkg','TH-01','Atropin sulfat','Ống');
        INSERT INTO goi_thau_hang_hoa VALUES ('requirement-2','org','pkg','TH-02','Natri clorid','Chai');
        INSERT INTO hang_hoa_du_thau_nha_thau VALUES ('goods','org','pkg','opening',1300,780,1014000,'requirement','1','Atropin sulfat','Ống','matched',1,0);
            """
        )
        statements.clear()

        dataset = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )

        record = dataset["records"][0]
        assert dataset["package"]["is_thuoc"] == 1
        assert record.status == "Trúng thầu"
        assert record.award_quantity == 1_300
        assert record.award_unit_price == 780
        assert record.award_price == 1_014_000
        assert record.goods_item_id == "requirement"
        assert record.goods_sequence == "1"
        assert dataset["blockingErrors"] == []
        assert len(
            [sql for sql in statements if sql.lstrip().upper().startswith("SELECT")]
        ) == 4

        connection.execute(
            "UPDATE goi_thau SET so_quyet_dinh_ket_qua = NULL"
        )
        missing_decision = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "AWARD_DECISION_NOT_READY" in {
            item["code"] for item in missing_decision["blockingErrors"]
        }
        assert next(
            item
            for item in missing_decision["blockingErrors"]
            if item["code"] == "AWARD_DECISION_NOT_READY"
        )["editPath"] == "/packages/pkg/award-result"
        connection.execute(
            "UPDATE goi_thau SET so_quyet_dinh_ket_qua = 'QD-01'"
        )

        connection.execute(
            "UPDATE goi_thau_phan_lo SET nha_thau_trung_thau_id = 'missing'"
        )
        invalid_winner = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "AWARD_WINNER_NOT_READY" in {
            item["code"] for item in invalid_winner["blockingErrors"]
        }
        connection.execute(
            "UPDATE goi_thau_phan_lo SET nha_thau_trung_thau_id = 'bidder'"
        )

        connection.execute(
            "UPDATE hang_hoa_du_thau_nha_thau SET goi_thau_hang_hoa_id = NULL"
        )
        missing_goods_id = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "MEDICINE_GOODS_ID_MISSING" in {
            item["code"] for item in missing_goods_id["blockingErrors"]
        }
        connection.execute(
            """UPDATE hang_hoa_du_thau_nha_thau
               SET goi_thau_hang_hoa_id = 'requirement'"""
        )

        connection.execute(
            """INSERT INTO hang_hoa_du_thau_nha_thau VALUES
               ('goods-2','org','pkg','opening',10,780,7800,'requirement-2','2','Natri clorid','Chai','matched',2,0)"""
        )
        connection.execute(
            "UPDATE goi_thau_phan_lo SET gia_trung_thau = 1021800"
        )
        multiple_items = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert multiple_items["blockingErrors"] == []
        assert [record.goods_item_id for record in multiple_items["records"]] == [
            "requirement", "requirement-2"
        ]
        assert [record.award_price for record in multiple_items["records"]] == [
            1_014_000, 7_800
        ]

        connection.execute(
            "UPDATE goi_thau_phan_lo SET gia_trung_thau = 1021801"
        )
        total_conflict = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "MEDICINE_AWARD_VALUE_CONFLICT" in {
            item["code"] for item in total_conflict["blockingErrors"]
        }
        connection.execute(
            "UPDATE goi_thau_phan_lo SET gia_trung_thau = 1021800"
        )

        connection.execute(
            "UPDATE hang_hoa_du_thau_nha_thau SET don_vi_tinh = 'Hộp' WHERE id = 'goods-2'"
        )
        unit_mismatch = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "MEDICINE_GOODS_UNIT_MISMATCH" in {
            item["code"] for item in unit_mismatch["blockingErrors"]
        }
        connection.execute(
            "UPDATE hang_hoa_du_thau_nha_thau SET don_vi_tinh = 'Chai' WHERE id = 'goods-2'"
        )

        connection.execute(
            "UPDATE hang_hoa_du_thau_nha_thau SET stt_nguon = '1' WHERE id = 'goods-2'"
        )
        duplicate_stt = service.load_award_result_dataset(
            "pkg", "org", database_obj=_Database(connection)
        )
        assert "MEDICINE_GOODS_SEQUENCE_DUPLICATED" in {
            item["code"] for item in duplicate_stt["blockingErrors"]
        }
    finally:
        connection.close()


def test_validation_token_is_bound_to_user_org_package_hash_and_expiry(tmp_path, monkeypatch):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    monkeypatch.setenv("AWARD_RESULT_EXCEL_TOKEN_KEY", "x" * 32)
    content = standard_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    token, _metadata = service.create_validation_artifact(
        content,
        inspection,
        user_id="user",
        organization_id="org",
        package_id="pkg",
        original_filename="input.xlsx",
        now=100,
    )

    metadata, stored = service.load_validation_artifact(
        token, user_id="user", organization_id="org", package_id="pkg", now=101
    )
    assert stored == content
    assert metadata["sha256"]
    service.consume_validation_artifact(f"{token}tampered")
    _metadata_after_tamper, stored_after_tamper = service.load_validation_artifact(
        token, user_id="user", organization_id="org", package_id="pkg", now=101
    )
    assert stored_after_tamper == content
    with pytest.raises(service.AwardResultExcelError) as wrong_user:
        service.load_validation_artifact(
            token, user_id="other", organization_id="org", package_id="pkg", now=101
        )
    assert wrong_user.value.code == "VALIDATION_TOKEN_SCOPE_MISMATCH"
    with pytest.raises(service.AwardResultExcelError) as wrong_org:
        service.load_validation_artifact(
            token, user_id="user", organization_id="other", package_id="pkg", now=101
        )
    assert wrong_org.value.code == "VALIDATION_TOKEN_SCOPE_MISMATCH"
    with pytest.raises(service.AwardResultExcelError) as wrong_package:
        service.load_validation_artifact(
            token, user_id="user", organization_id="org", package_id="other", now=101
        )
    assert wrong_package.value.code == "VALIDATION_TOKEN_SCOPE_MISMATCH"
    with pytest.raises(service.AwardResultExcelError) as expired:
        service.load_validation_artifact(
            token,
            user_id="user",
            organization_id="org",
            package_id="pkg",
            now=100 + service.VALIDATION_TTL_SECONDS + 1,
        )
    assert expired.value.code == "VALIDATION_TOKEN_EXPIRED"


def test_validation_artifact_detects_changed_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    monkeypatch.setenv("AWARD_RESULT_EXCEL_TOKEN_KEY", "x" * 32)
    content = standard_workbook_bytes()
    token, metadata = service.create_validation_artifact(
        content,
        service.inspect_award_result_workbook(content),
        user_id="user",
        organization_id="org",
        package_id="pkg",
        original_filename="input.xlsx",
        now=100,
    )
    validation_dir = (
        tmp_path
        / "document-worker-temp"
        / "award-result-validations"
        / f"validation-{metadata['validationId']}"
    )
    (validation_dir / "workbook.xlsx").write_bytes(content + b"changed")

    with pytest.raises(service.AwardResultExcelError) as changed:
        service.load_validation_artifact(
            token,
            user_id="user",
            organization_id="org",
            package_id="pkg",
            now=101,
        )

    assert changed.value.code == "WORKBOOK_CHANGED_AFTER_VALIDATION"


def test_cleanup_does_not_starve_expired_artifact_after_scan_limit(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    root = service._validation_root()
    for index in range(128):
        path = root / f"validation-{index:032x}"
        path.mkdir()
        (path / "metadata.json").write_text(
            json.dumps({"expiresAt": 10_000 + index}), encoding="utf-8"
        )
    expired = root / f"validation-{128:032x}"
    expired.mkdir()
    (expired / "metadata.json").write_text(
        json.dumps({"expiresAt": 1}), encoding="utf-8"
    )

    removed = service.cleanup_expired_validation_artifacts(now=100, limit=128)

    assert removed == 1
    assert not expired.exists()


def test_cleanup_removes_old_corrupt_metadata_safely(tmp_path, monkeypatch):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    path = service._validation_root() / f"validation-{'f' * 32}"
    path.mkdir()
    (path / "metadata.json").write_text("{broken", encoding="utf-8")
    service.os.utime(path, (1, 1))

    assert service.cleanup_expired_validation_artifacts(now=10_000) == 1
    assert not path.exists()


@pytest.mark.parametrize(
    ("limit_name", "second_user", "second_org", "limit_value"),
    [
        ("AWARD_RESULT_ARTIFACT_MAX_PER_USER", "user", "org", "1"),
        ("AWARD_RESULT_ARTIFACT_MAX_PER_ORGANIZATION", "other", "org", "1"),
        ("AWARD_RESULT_ARTIFACT_MAX_GLOBAL_BYTES", "other", "other-org", "1"),
    ],
)
def test_validation_artifact_enforces_user_org_and_global_quotas(
    tmp_path,
    monkeypatch,
    limit_name,
    second_user,
    second_org,
    limit_value,
):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    monkeypatch.setenv("AWARD_RESULT_EXCEL_TOKEN_KEY", "x" * 32)
    if limit_name != "AWARD_RESULT_ARTIFACT_MAX_GLOBAL_BYTES":
        monkeypatch.setenv(limit_name, limit_value)
    content = standard_workbook_bytes()
    inspection = service.inspect_award_result_workbook(content)
    service.create_validation_artifact(
        content,
        inspection,
        user_id="user",
        organization_id="org",
        package_id="pkg",
        original_filename="input.xlsx",
        now=100,
    )
    if limit_name == "AWARD_RESULT_ARTIFACT_MAX_GLOBAL_BYTES":
        monkeypatch.setenv(limit_name, str(len(content)))

    with pytest.raises(service.AwardResultExcelError) as rejected:
        service.create_validation_artifact(
            content,
            inspection,
            user_id=second_user,
            organization_id=second_org,
            package_id="pkg-2",
            original_filename="input.xlsx",
            now=101,
        )

    assert rejected.value.code == "VALIDATION_ARTIFACT_QUOTA_EXCEEDED"


def test_partial_artifact_write_is_never_published(tmp_path, monkeypatch):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    monkeypatch.setattr(service.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("crash")))
    content = standard_workbook_bytes()

    with pytest.raises(OSError):
        service.create_validation_artifact(
            content,
            service.inspect_award_result_workbook(content),
            user_id="user",
            organization_id="org",
            package_id="pkg",
            original_filename="input.xlsx",
            now=100,
        )

    root = service._validation_root()
    assert list(root.glob("validation-*")) == []
    assert list(root.glob(".tmp-validation-*")) == []


def test_claim_prevents_concurrent_export_and_cleanup(tmp_path, monkeypatch):
    monkeypatch.setattr(
        service,
        "resolve_runtime_path",
        lambda _name: tmp_path / "document-worker-temp",
    )
    monkeypatch.setenv("AWARD_RESULT_EXCEL_TOKEN_KEY", "x" * 32)
    content = standard_workbook_bytes()
    token, _metadata = service.create_validation_artifact(
        content,
        service.inspect_award_result_workbook(content),
        user_id="user",
        organization_id="org",
        package_id="pkg",
        original_filename="input.xlsx",
        now=100,
    )
    service.load_validation_artifact(
        token,
        user_id="user",
        organization_id="org",
        package_id="pkg",
        now=999,
        claim=True,
    )

    with pytest.raises(service.AwardResultExcelError) as in_use:
        service.load_validation_artifact(
            token,
            user_id="user",
            organization_id="org",
            package_id="pkg",
            now=999,
            claim=True,
        )
    assert in_use.value.code == "VALIDATION_TOKEN_IN_USE"
    assert service.cleanup_expired_validation_artifacts(
        now=1_001
    ) == 0
    service.release_validation_artifact(token)
    assert service.cleanup_expired_validation_artifacts(
        now=1_001
    ) == 1


def test_output_filename_is_sanitized():
    assert service.output_filename('../35:"bad".xlsx') == "35__bad_da_dien_ket_qua.xlsx"


def test_multi_replica_local_artifact_store_fails_closed():
    with pytest.raises(RuntimeError):
        service.validate_artifact_store_configuration(
            {"APP_ENV": "production", "APP_INSTANCE_COUNT": "2"}
        )
    service.validate_artifact_store_configuration(
        {
            "APP_ENV": "production",
            "APP_INSTANCE_COUNT": "2",
            "AWARD_RESULT_ARTIFACT_SHARED_STORAGE_CONFIRMED": "true",
        }
    )


def test_validation_preview_is_bounded_stable_and_filterable():
    rows = [
        {
            "excelRow": index + 2,
            "status": "matched" if index % 2 == 0 else "unmatched",
            "matchMethod": "lot_code_and_bidder_identifier" if index % 2 == 0 else None,
            "writable": index % 2 == 0,
            "warnings": ([{"code": "RESULT_NOT_FOUND"}] if index % 2 else []),
        }
        for index in range(10_000)
    ]
    match = {
        "totalRows": 10_000,
        "rows": rows,
        "warnings": [
            {"code": "RESULT_NOT_FOUND", "excelRow": index + 2}
            for index in range(5_000)
        ],
        "blockingErrors": [],
        "canExport": True,
    }

    first = service.public_validation_result(match)
    filtered = service.public_validation_result(
        match, page=2, page_size=50, writable=True
    )

    assert len(first["rows"]) == 100
    assert len(first["warnings"]) == 100
    assert first["remainingRows"] == 9_900
    assert first["totalPages"] == 100
    assert first["hasPreviousPage"] is False
    assert first["hasNextPage"] is True
    assert first["warningSummary"] == {"RESULT_NOT_FOUND": 5_000}
    assert [item["excelRow"] for item in filtered["rows"][:2]] == [102, 104]
    assert filtered["filteredRows"] == 5_000
    assert filtered["page"] == 2
    assert filtered["totalPages"] == 100

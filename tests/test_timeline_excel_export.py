import asyncio
from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
from starlette.requests import Request

from backend.documents import routes_excel
from backend.documents.excel_workbook_builder import create_timeline_excel


def _timeline_context():
    return {
        "goi_thau": {
            "ma_goi_thau": "GT-01",
            "ten_goi_thau": "Gói thầu thử nghiệm",
        },
        "ke_hoach": {
            "ma_ke_hoach": "KH-01",
            "ten_ke_hoach": "Kế hoạch lựa chọn nhà thầu",
        },
        "to_chuc": {"ten_to_chuc": "Đơn vị thử nghiệm"},
        "timeline_template_version": 3,
        "generated_date": "03/08/2026",
        "timeline_sections": [
            {
                "code": "I",
                "title": "Chuẩn bị",
                "items": [
                    {
                        "display_code": "1.1",
                        "cong_viec": "Phê duyệt kế hoạch",
                        "don_vi_ban_hanh": "Chủ đầu tư",
                        "so_van_ban": "01/QĐ",
                        "ngay_thuc_te": "2026-08-03",
                        "ngay_du_kien": "",
                        "trang_thai": "DONE",
                        "source_mode": "AUTO",
                        "is_planned_date": False,
                    },
                    {
                        "display_code": "1.2",
                        "cong_viec": "=HYPERLINK(\"https://example.test\")",
                        "don_vi_ban_hanh": "+cmd",
                        "so_van_ban": "@SUM(A1:A2)",
                        "ngay_thuc_te": "",
                        "ngay_du_kien": "2026-08-10",
                        "trang_thai": "PENDING",
                        "source_mode": "MANUAL",
                        "is_planned_date": True,
                    },
                ],
            }
        ],
    }


def test_timeline_excel_contains_metadata_sections_editable_rows_and_safe_text():
    workbook = create_timeline_excel(_timeline_context())
    output = BytesIO()
    workbook.save(output)

    exported = load_workbook(BytesIO(output.getvalue()))
    sheet = exported["Timeline"]

    assert sheet["A1"].value == "TIMELINE GÓI THẦU"
    assert sheet["B3"].value == "GT-01"
    assert sheet["B4"].value == "Gói thầu thử nghiệm"
    assert sheet["A10"].value == "I. Chuẩn bị"
    assert [sheet.cell(9, column).value for column in range(1, 8)] == [
        "STT",
        "Công việc",
        "Đơn vị ban hành",
        "Số văn bản",
        "Thời gian",
        "Trạng thái",
        "Nguồn",
    ]
    assert sheet["A11"].value == "1.1"
    assert isinstance(sheet["E11"].value, datetime)
    assert sheet["E11"].number_format == "dd/mm/yyyy"
    assert sheet["F11"].value == "Đã hoàn thành"
    assert sheet["G11"].value == "Tự động"
    assert sheet["B12"].value.startswith("'")
    assert sheet["C12"].value.startswith("'")
    assert sheet["D12"].value.startswith("'")
    assert sheet["B12"].data_type == "s"
    assert sheet.freeze_panes == "A10"
    assert sheet.column_dimensions["A"].width >= 18
    assert len(sheet.data_validations.dataValidation) == 1


def test_timeline_excel_worker_supports_the_data_only_builder(monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)

    from backend.documents.document_worker import run_document_job

    content = run_document_job(
        "export_excel",
        {"function": "create_timeline_excel", "args": [_timeline_context()]},
        timeout_seconds=15,
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook["Timeline"]["B3"].value == "GT-01"


def test_timeline_export_route_keeps_snapshot_and_access_guards(monkeypatch):
    calls = []
    role = SimpleNamespace(user_id="user-1")
    monkeypatch.setattr(routes_excel, "verify_session", lambda _request: (True, role))
    monkeypatch.setattr(routes_excel, "get_active_org", lambda *_args: "org-1")
    monkeypatch.setattr(
        routes_excel,
        "_timeline_export_entitlement_response",
        lambda *_args: None,
    )
    monkeypatch.setattr(
        routes_excel,
        "_validate_export_snapshot",
        lambda _request, organization_id: (
            calls.append(("snapshot", organization_id)) or 21,
            None,
        ),
    )
    monkeypatch.setattr(
        routes_excel,
        "_can_export_package",
        lambda *_args: calls.append(("access", "gt-1")) or True,
    )

    async def fake_database_read(function, *args, **_kwargs):
        calls.append(("context", function.__name__, *args))
        return _timeline_context()

    async def fake_export(function_name, context):
        calls.append(("worker", function_name, context["goi_thau"]["ma_goi_thau"]))
        workbook = create_timeline_excel(context)
        output = BytesIO()
        workbook.save(output)
        return BytesIO(output.getvalue())

    monkeypatch.setattr(routes_excel, "run_database_read", fake_database_read)
    monkeypatch.setattr(routes_excel, "_export_excel", fake_export)
    monkeypatch.setattr(
        routes_excel,
        "_ensure_export_snapshot_unchanged",
        lambda organization_id, version: calls.append(
            ("snapshot-unchanged", organization_id, version)
        ) or None,
    )
    monkeypatch.setattr(routes_excel, "log_audit", lambda *_args, **_kwargs: None)

    request = Request({
        "type": "http",
        "method": "GET",
        "path": "/api/export-timeline/gt-1",
        "path_params": {"package_id": "gt-1"},
        "query_string": b"snapshotVersion=21",
        "headers": [],
        "client": ("127.0.0.1", 1234),
        "server": ("testserver", 80),
        "scheme": "http",
    })

    async def invoke():
        response = await routes_excel.export_timeline_api(request)
        body = b"".join([chunk async for chunk in response.body_iterator])
        return response, body

    response, body = asyncio.run(invoke())

    assert response.status_code == 200
    assert response.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "Timeline_goi_thau_GT-01.xlsx" in response.headers["content-disposition"]
    assert load_workbook(BytesIO(body))["Timeline"]["B3"].value == "GT-01"
    assert ("snapshot", "org-1") in calls
    assert ("access", "gt-1") in calls
    assert ("snapshot-unchanged", "org-1", 21) in calls

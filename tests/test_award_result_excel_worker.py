from __future__ import annotations

from io import BytesIO
import json
import zipfile

from openpyxl import Workbook, load_workbook
import pytest

from backend.documents import award_result_excel_service as service
from backend.documents.document_worker import DocumentWorkerInputError, run_document_job


def _workbook_with_unrelated_formula():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách nhà thầu"
    sheet.append(list(service.EXPECTED_HEADERS))
    sheet.append(
        [
            "L01",
            "Lô 01",
            "vn001",
            "001",
            "Nhà thầu 01",
            1_000,
            *("" for _ in range(9)),
        ]
    )
    metadata = workbook.create_sheet("Metadata")
    metadata["A1"] = "=SUM(1,2)"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_worker_allows_and_preserves_unrelated_formulas(monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)
    content = _workbook_with_unrelated_formula()

    inspection = run_document_job(
        "inspect_award_result_excel",
        {"content": content},
        timeout_seconds=15,
    )
    output = run_document_job(
        "export_award_result_excel",
        {
            "content": content,
            "updates": [
                {
                    "excelRow": 2,
                    "sourceFingerprint": inspection["rows"][0]["sourceFingerprint"],
                    "values": ["Trúng thầu", 950, 90, None, 900, None, None, None, None],
                }
            ],
        },
        timeout_seconds=15,
    )

    workbook = load_workbook(BytesIO(output), data_only=False)
    assert workbook["Danh sách nhà thầu"]["G2"].value == "Trúng thầu"
    assert workbook["Metadata"]["A1"].value == "=SUM(1,2)"


def test_worker_rejects_archive_path_traversal(monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)
    output = BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr("../escape.xml", "unsafe")
        archive.writestr("[Content_Types].xml", "<Types />")
        archive.writestr("xl/workbook.xml", "<workbook />")

    with pytest.raises(DocumentWorkerInputError, match="đường dẫn nội bộ"):
        run_document_job(
            "inspect_award_result_excel",
            {"content": output.getvalue()},
            timeout_seconds=15,
        )


def test_worker_builds_formula_safe_reconciliation_report(monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)

    output = run_document_job(
        "build_award_result_reconciliation",
        {
            "reportJson": json.dumps({
                "metadata": {
                    "sourceSha256": "a" * 64,
                    "packageCode": "=DANGEROUS",
                    "userId": "user",
                    "generatedAt": "2026-08-04T00:00:00Z",
                },
                "summary": {"totalRows": 1, "updatedRows": 1},
                "rows": [{
                    "excelRow": 2,
                    "lotCode": "L01",
                    "bidderName": "+SUM(A1:A2)",
                    "matchMethod": "lot_code_and_bidder_identifier",
                    "writable": True,
                    "changes": [{
                        "field": "award_price",
                        "oldValue": 900,
                        "newValue": 850,
                        "source": "approved_result.award_price",
                    }],
                    "warnings": [],
                }],
            }, ensure_ascii=False).encode("utf-8")
        },
        timeout_seconds=15,
    )

    workbook = load_workbook(BytesIO(output), data_only=False)
    assert workbook.sheetnames == ["Tổng quan", "Đối chiếu"]
    assert workbook["Tổng quan"]["B3"].data_type != "f"
    assert workbook["Đối chiếu"]["C2"].data_type != "f"
    assert workbook["Đối chiếu"]["G2"].value == 900
    assert workbook["Đối chiếu"]["H2"].value == 850

import asyncio
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest
from starlette.requests import Request

from backend.documents import excel_service, routes_excel
from backend.documents.document_worker import DocumentWorkerInputError, run_document_job


def test_document_worker_builds_excel_from_data_only_spec(monkeypatch):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)

    content = run_document_job(
        "export_excel",
        {
            "function": "create_excel_from_spec",
            "args": [
                {
                    "title": "Danh gia HSDT",
                    "headers": ["Ma nha thau", "Ket qua"],
                    "rows": [["NT-01", "Dat"]],
                    "options_map": {"Ket qua": ["Dat", "Khong dat"]},
                }
            ],
        },
        timeout_seconds=15,
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Danh gia HSDT"]
    assert sheet["A1"].value == "Ma nha thau"
    assert sheet["A2"].value == "NT-01"
    assert sheet["B2"].value == "Dat"


@pytest.mark.parametrize("formula", ["=1+1", "+1+1", "-1+1", "@SUM(A1:A2)"])
def test_document_worker_neutralizes_formula_prefixed_text(monkeypatch, formula):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("DOCUMENT_WORKER_DATABASE_URL", raising=False)

    content = run_document_job(
        "export_excel",
        {
            "function": "create_excel_from_spec",
            "args": [
                {
                    "title": "Export",
                    "headers": ["Contractor"],
                    "rows": [[formula]],
                    "options_map": {"Contractor": [formula]},
                }
            ],
        },
        timeout_seconds=15,
    )

    workbook = load_workbook(BytesIO(content))
    row_cell = workbook["Export"]["A2"]
    dropdown_cell = workbook["Dropdowns"]["A1"]
    assert row_cell.value == f"'{formula}"
    assert row_cell.data_type == "s"
    assert dropdown_cell.value == f"'{formula}"
    assert dropdown_cell.data_type == "s"


@pytest.mark.parametrize(
    ("function_name", "args"),
    [
        ("create_danhgiahsdt_template", ["gt-1", "org-1", "technical"]),
        ("create_opening_fin_template", ["gt-1", "org-1"]),
        ("create_ketquaqd_template", ["gt-1", "org-1"]),
    ],
)
def test_document_worker_rejects_database_bound_excel_functions(
    monkeypatch,
    function_name,
    args,
):
    monkeypatch.delenv("DATABASE_URL", raising=False)
    with pytest.raises(DocumentWorkerInputError, match="không được hỗ trợ"):
        run_document_job(
            "export_excel",
            {"function": function_name, "args": args},
            timeout_seconds=15,
        )


class _EvaluationCursor:
    def __init__(self):
        self._one = None
        self._rows = []

    def execute(self, sql, _params=()):
        normalized = " ".join(str(sql).split())
        self._one = None
        self._rows = []
        if normalized.startswith("SELECT linh_vuc"):
            self._one = ("Hàng hóa", "Một giai đoạn hai túi hồ sơ", "Không")
        elif "FROM goi_thau_phan_lo" in normalized:
            self._rows = []
        elif "FROM thong_tin_mo_thau_lien_danh_thanh_vien" in normalized:
            self._rows = []
        elif "FROM thong_tin_mo_thau m" in normalized:
            self._rows = [[
                "Độc lập", "", "", "NT-01", "Nhà thầu 01",
                1000, 0, 1000, 90, 100, 120, 30,
                "Đạt", "Đạt", "Đạt", "", "", "", "", "", "", "", "",
                950, 940, 0, "opening-1",
            ]]
        return self

    def fetchone(self):
        return self._one

    def fetchall(self):
        return list(self._rows)


class _Connection:
    def __init__(self, cursor):
        self._cursor = cursor

    def cursor(self):
        return self._cursor

    def close(self):
        pass


class _LottedEvaluationCursor(_EvaluationCursor):
    def execute(self, sql, params=()):
        normalized = " ".join(str(sql).split())
        super().execute(sql, params)
        if normalized.startswith("SELECT linh_vuc"):
            self._one = ("Hàng hóa", "Một giai đoạn hai túi hồ sơ", "Có")
        elif "FROM goi_thau_phan_lo" in normalized:
            self._rows = [("PL1",), ("PL2",)]
        elif "FROM thong_tin_mo_thau m" in normalized:
            first = list(self._rows[0])
            second = list(first)
            first[1:5] = ["PL1", "Lô 1", "NT-01", "Nhà thầu 01"]
            first[26] = "opening-1"
            second[1:5] = ["PL2", "Lô 2", "NT-02", "Nhà thầu 02"]
            second[26] = "opening-2"
            self._rows = [first, second]
        return self


def test_evaluation_excel_template_is_limited_to_selected_lot_codes(monkeypatch):
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_LottedEvaluationCursor()),
    )

    spec = excel_service.prepare_danhgiahsdt_template_spec(
        "gt-1", "org-1", "technical", ["PL1"]
    )

    assert spec["options_map"]["Mã phần lô"] == ["PL1"]
    assert [row[1] for row in spec["rows"]] == ["PL1"]


def test_1g2t_evaluation_export_prefetches_database_data_before_worker(monkeypatch):
    monkeypatch.setattr(
        routes_excel,
        "verify_session",
        lambda _request: (True, SimpleNamespace(user_id="user-1")),
    )
    monkeypatch.setattr(routes_excel, "get_active_org", lambda *_args: "org-1")
    monkeypatch.setattr(routes_excel, "_can_export_package", lambda *_args: True)
    monkeypatch.setattr(routes_excel, "_export_entitlement_response", lambda *_args: None)
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_EvaluationCursor()),
    )

    async def isolated_worker(operation, payload, **kwargs):
        return await asyncio.to_thread(
            run_document_job,
            operation,
            payload,
            timeout_seconds=kwargs.get("timeout_seconds", 15),
        )

    monkeypatch.setattr(routes_excel, "run_document_job_async", isolated_worker)
    request = Request({
        "type": "http",
        "method": "GET",
        "path": "/api/export-danhgiahsdt-template",
        "query_string": (
            b"package_id=gt-1&package_name=IB-1&eval_type=technical"
        ),
        "headers": [],
        "client": ("127.0.0.1", 1234),
        "server": ("testserver", 80),
        "scheme": "http",
    })

    async def invoke():
        response = await routes_excel.export_danhgiahsdt_template_api(request)
        if hasattr(response, "body_iterator"):
            body = b"".join([chunk async for chunk in response.body_iterator])
        else:
            body = response.body
        return response, body

    response, body = asyncio.run(invoke())

    assert response.status_code == 200
    assert response.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(body))
    assert workbook["Danh gia HSDT"]["B2"].value == "NT-01"


def test_financial_evaluation_excel_contains_ranking_and_proposed_award_prices(monkeypatch):
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_EvaluationCursor()),
    )
    spec = excel_service.prepare_danhgiahsdt_template_spec("gt-1", "org-1", "financial")
    assert "Giá xếp hạng (VND)" in spec["headers"]
    assert "Giá đề nghị trúng thầu (VND)" in spec["headers"]
    assert "Xử lý giá đề nghị trúng thầu dưới 50%" in spec["headers"]
    assert "Đánh giá tài chính (Điểm hoặc Xếp hạng)" not in spec["headers"]
    ranking_index = spec["headers"].index("Giá xếp hạng (VND)")
    proposed_index = spec["headers"].index("Giá đề nghị trúng thầu (VND)")
    assert spec["rows"][0][ranking_index] == 950
    assert spec["rows"][0][proposed_index] == 940
    decision_index = spec["headers"].index("Xử lý giá đề nghị trúng thầu dưới 50%")
    assert spec["rows"][0][decision_index] == "Không chấp thuận"
    assert spec["options_map"]["Xử lý giá đề nghị trúng thầu dưới 50%"] == [
        "Chấp thuận", "Không chấp thuận"
    ]
    assert spec["formats_map"]["Giá xếp hạng (VND)"] == "currency"
    assert spec["formats_map"]["Giá đề nghị trúng thầu (VND)"] == "currency"


class _JointVentureEvaluationCursor(_EvaluationCursor):
    def execute(self, sql, _params=()):
        normalized = " ".join(str(sql).split())
        if "FROM thong_tin_mo_thau_lien_danh_thanh_vien" in normalized:
            self._one = None
            self._rows = [
                ("opening-1", "Đứng đầu liên danh", "Nhà thầu đứng đầu", "0101"),
                ("opening-1", "Thành viên liên danh", "Nhà thầu thành viên", "0102"),
            ]
            return self
        result = super().execute(sql, _params)
        if "FROM thong_tin_mo_thau m" in normalized:
            self._rows[0][0] = "Liên danh"
            self._rows[0][3] = "JV-01"
            self._rows[0][4] = "Liên danh thử nghiệm"
        return result


def test_evaluation_excel_preserves_joint_venture_members(monkeypatch):
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_JointVentureEvaluationCursor()),
    )
    spec = excel_service.prepare_danhgiahsdt_template_spec("gt-1", "org-1", "financial")
    member_index = spec["headers"].index("Thành viên liên danh")
    assert "Nhà thầu đứng đầu (0101)" in spec["rows"][0][member_index]
    assert "Nhà thầu thành viên (0102)" in spec["rows"][0][member_index]


class _OpeningFinancialCursor(_EvaluationCursor):
    def execute(self, sql, _params=()):
        normalized = " ".join(str(sql).split())
        self._one = None
        self._rows = []
        if normalized.startswith("SELECT m.ma_dinh_danh"):
            assert "LEFT JOIN ket_qua_danh_gia_nha_thau k" in normalized
            assert "k.danh_gia_hop_le" in normalized
            assert "k.danh_gia_ket_luan" in normalized
            self._rows = [[
                "NT-01", "Nhà thầu 01", 1000, 0, 1000, 90, 30,
                "Đạt", "Đạt", "Đạt", "Đạt",
            ]]
        return self


def test_1g2t_financial_opening_export_prefetches_database_data(monkeypatch):
    monkeypatch.setattr(
        routes_excel,
        "verify_session",
        lambda _request: (True, SimpleNamespace(user_id="user-1")),
    )
    monkeypatch.setattr(routes_excel, "get_active_org", lambda *_args: "org-1")
    monkeypatch.setattr(routes_excel, "_can_export_package", lambda *_args: True)
    monkeypatch.setattr(routes_excel, "_export_entitlement_response", lambda *_args: None)
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_OpeningFinancialCursor()),
    )

    async def isolated_worker(operation, payload, **kwargs):
        return await asyncio.to_thread(
            run_document_job,
            operation,
            payload,
            timeout_seconds=kwargs.get("timeout_seconds", 15),
        )

    monkeypatch.setattr(routes_excel, "run_document_job_async", isolated_worker)
    request = Request({
        "type": "http",
        "method": "GET",
        "path": "/api/export-opening-fin-template",
        "query_string": b"package_id=gt-1&package_name=IB-1",
        "headers": [],
        "client": ("127.0.0.1", 1234),
        "server": ("testserver", 80),
        "scheme": "http",
    })

    async def invoke():
        response = await routes_excel.export_opening_fin_template_api(request)
        if hasattr(response, "body_iterator"):
            body = b"".join([chunk async for chunk in response.body_iterator])
        else:
            body = response.body
        return response, body

    response, body = asyncio.run(invoke())

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(body))
    assert workbook["Mo De Xuat Tai Chinh"]["A2"].value == "NT-01"


class _AwardResultCursor(_EvaluationCursor):
    def execute(self, sql, _params=()):
        normalized = " ".join(str(sql).split())
        self._one = None
        self._rows = []
        if normalized.startswith("SELECT nha_thau_trung_thau_id"):
            self._one = ("NT-01", 900, 30, 60)
        elif "FROM thong_tin_mo_thau_lien_danh_thanh_vien" in normalized:
            self._rows = []
        elif normalized.startswith("SELECT m.loai_nha_thau"):
            assert "danh_gia_tai_chinh" not in normalized
            assert "LEFT JOIN ket_qua_danh_gia_nha_thau k" in normalized
            assert "k.ly_do_loai" in normalized
            self._rows = [[
                "Độc lập", "", "", "NT-01", "Nhà thầu 01", 1000, 0, 1000,
                "Nhà thầu có giá đề nghị trúng thầu nhỏ hơn 50% giá gói thầu.", 1, "opening-1",
            ]]
        return self


def test_award_result_export_prefetches_database_data_before_worker(monkeypatch):
    monkeypatch.setattr(
        routes_excel,
        "verify_session",
        lambda _request: (True, SimpleNamespace(user_id="user-1")),
    )
    monkeypatch.setattr(routes_excel, "get_active_org", lambda *_args: "org-1")
    monkeypatch.setattr(routes_excel, "_can_export_package", lambda *_args: True)
    monkeypatch.setattr(routes_excel, "_export_entitlement_response", lambda *_args: None)
    monkeypatch.setattr(
        excel_service.database,
        "get_connection",
        lambda: _Connection(_AwardResultCursor()),
    )

    async def isolated_worker(operation, payload, **kwargs):
        return await asyncio.to_thread(
            run_document_job,
            operation,
            payload,
            timeout_seconds=kwargs.get("timeout_seconds", 15),
        )

    monkeypatch.setattr(routes_excel, "run_document_job_async", isolated_worker)
    request = Request({
        "type": "http",
        "method": "GET",
        "path": "/api/export-ketquaqd-template",
        "query_string": b"package_id=gt-1&package_name=IB-1",
        "headers": [],
        "client": ("127.0.0.1", 1234),
        "server": ("testserver", 80),
        "scheme": "http",
    })

    async def invoke():
        response = await routes_excel.export_ketquaqd_template_api(request)
        if hasattr(response, "body_iterator"):
            body = b"".join([chunk async for chunk in response.body_iterator])
        else:
            body = response.body
        return response, body

    response, body = asyncio.run(invoke())

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(body))
    assert workbook["Ket Qua LCNT"]["B2"].value == "NT-01"
    assert workbook["Ket Qua LCNT"]["E2"].value == "Trúng thầu"
    assert workbook["Ket Qua LCNT"]["F2"].value == "Chấp thuận"
